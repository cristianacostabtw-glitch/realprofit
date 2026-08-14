"""
ProfitFlow (app independiente) — sirve SOLO el dashboard, EN BLANCO.

Fase 1: dashboard limpio, sin métricas (todo en cero). Pensado para subir a la web
y que cada usuario conecte SU PROPIO MercadoPago (Fase 2: OAuth de MP → llena los datos).

Corre en su propio puerto (8010), sin nada de METAFY/VisionPure.
"""
from __future__ import annotations

import datetime as _dt
import json as _json
import secrets as _secrets
import threading
import urllib.parse as _url
from pathlib import Path

import requests
from flask import Flask, Response, jsonify, redirect, request, session, send_file

import os as _os

from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.security import generate_password_hash, check_password_hash

RAIZ = Path(__file__).resolve().parent
# Carpeta de datos persistentes (usuarios + tokens de MP). En Render apuntamos DATA_DIR a un
# disco persistente (ej. /var/data) → las cuentas NO se borran en los deploys. Local: la raíz.
DATA_DIR = Path(_os.getenv("DATA_DIR", str(RAIZ)))
try:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    DATA_DIR = RAIZ
app = Flask(__name__)
app.secret_key = _os.getenv("SECRET_KEY", "profitflow-dev-key-cambiar-en-produccion")

# Rate limiting: frena a quien intente martillar los endpoints sensibles (OAuth) para tirar
# la app o abusar. NO limita el dashboard (así no se rompe la carga normal).
limiter = Limiter(get_remote_address, app=app, default_limits=[], storage_uri="memory://")


@app.after_request
def _headers_seguridad(resp):
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["X-Frame-Options"] = "SAMEORIGIN"
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return resp

# ---------------- MercadoPago OAuth (conectar con un click) ----------------
MP_SECRETS = RAIZ / "mp_secrets.json"     # tu Client ID + Secret (los pega el dueño de la app)
MP_TOKENS = DATA_DIR / "mp_tokens.json"   # tokens de cada usuario que se conecta (persistente)


def _mp_cfg() -> dict:
    # En la web (Render) usamos variables de entorno; en local, el archivo mp_secrets.json.
    import os
    try:
        c = _json.loads(MP_SECRETS.read_text(encoding="utf-8"))
    except Exception:
        c = {}
    return {"client_id": os.getenv("MP_CLIENT_ID") or c.get("client_id", ""),
            "client_secret": os.getenv("MP_CLIENT_SECRET") or c.get("client_secret", ""),
            "redirect_uri": os.getenv("MP_REDIRECT_URI") or c.get("redirect_uri",
                                                                    "http://127.0.0.1:8010/mp/callback")}


def _mp_tokens() -> dict:
    try:
        return _json.loads(MP_TOKENS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _mp_save_token(key, data) -> None:
    d = _mp_tokens()
    d[str(key)] = data
    MP_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


# ---------------- Meta (Facebook / Instagram Ads) OAuth ----------------
META_SECRETS = RAIZ / "meta_secrets.json"     # App ID + App Secret (dueño de la app)
META_TOKENS = DATA_DIR / "meta_tokens.json"   # tokens por usuario (persistente)
META_API = "v21.0"


def _meta_cfg() -> dict:
    import os
    try:
        c = _json.loads(META_SECRETS.read_text(encoding="utf-8"))
    except Exception:
        c = {}
    return {"app_id": os.getenv("META_APP_ID") or c.get("app_id", ""),
            "app_secret": os.getenv("META_APP_SECRET") or c.get("app_secret", ""),
            "redirect_uri": os.getenv("META_REDIRECT_URI") or c.get("redirect_uri",
                                                                    "http://127.0.0.1:8010/meta/callback")}


def _meta_tokens() -> dict:
    try:
        return _json.loads(META_TOKENS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _meta_save_token(key, data) -> None:
    d = _meta_tokens()
    d[str(key)] = data
    META_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


# ---------------- Shopify OAuth (conectar con un click) ----------------
SHOPIFY_SECRETS = RAIZ / "shopify_secrets.json"   # tu Client ID + Secret (dueño de la app)
SHOPIFY_TOKENS = DATA_DIR / "shopify_tokens.json"  # tokens por usuario (persistente)
SHOPIFY_SCOPES = ("read_orders,write_orders,write_order_edits,read_fulfillments,write_fulfillments,"
                  "read_merchant_managed_fulfillment_orders,write_merchant_managed_fulfillment_orders,"
                  "read_assigned_fulfillment_orders,read_third_party_fulfillment_orders,"
                  "read_shipping,write_shipping,read_products,read_inventory,read_customers,"
                  "write_customers,read_locations,read_checkouts,write_draft_orders,write_price_rules")


def _shop_cfg() -> dict:
    import os
    try:
        c = _json.loads(SHOPIFY_SECRETS.read_text(encoding="utf-8"))
    except Exception:
        c = {}
    return {"client_id": os.getenv("SHOPIFY_CLIENT_ID") or c.get("client_id", ""),
            "client_secret": os.getenv("SHOPIFY_CLIENT_SECRET") or c.get("client_secret", ""),
            "redirect_uri": os.getenv("SHOPIFY_REDIRECT_URI") or c.get("redirect_uri",
                                                                       "http://127.0.0.1:8010/shopify/callback"),
            "scopes": os.getenv("SHOPIFY_SCOPES") or c.get("scopes", SHOPIFY_SCOPES)}


def _shop_tokens() -> dict:
    try:
        return _json.loads(SHOPIFY_TOKENS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _shop_save_token(key, data) -> None:
    d = _shop_tokens()
    d[str(key)] = data
    SHOPIFY_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


def _shop_normalizar(shop: str) -> str:
    """Acepta 'mitienda', 'mitienda.myshopify.com' o con http(s):// y devuelve el dominio limpio."""
    shop = (shop or "").strip().lower()
    shop = shop.replace("https://", "").replace("http://", "")
    shop = shop.split("/")[0].split("?")[0]
    if shop and "." not in shop:
        shop = shop + ".myshopify.com"
    return shop


def _shop_valido(shop: str) -> bool:
    import re
    return bool(shop and re.match(r"^[a-z0-9][a-z0-9\-]*\.myshopify\.com$", shop))


# ---------------- Cuentas de usuario (email + contraseña) ----------------
USERS = DATA_DIR / "users.json"   # cada usuario de la app (persistente); su MP se guarda por email


def _users() -> dict:
    try:
        return _json.loads(USERS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _users_save(d: dict) -> None:
    USERS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


def _user_actual():
    """Email del usuario logueado (o None)."""
    return session.get("email")


# Deja SOLO "Dashboard" e "Integraciones" en el menú, y saca el logo. Como pf.html es React
# compilado, no lo editamos: ocultamos con CSS/JS inyectado + MutationObserver (aguanta los
# re-render del SPA sin romper nada).
_SOLO_DASH = r"""
<style>
 aside > a[aria-label="Ir al Dashboard"]{display:none!important}
 /* barra lateral MAS ANGOSTA (colapsada) + hover-expand un poco menos, sin romper nada */
 @media(min-width:768px){
  aside[class*="group/sidebar"]{width:62px!important;padding-left:6px!important;padding-right:6px!important}
  aside[class*="group/sidebar"]:hover{width:206px!important}
 }
 aside nav{gap:2px!important}
 aside nav a{padding-top:8px!important;padding-bottom:8px!important}
 .rp-pill{position:fixed;left:0;z-index:100002;display:flex;align-items:center;justify-content:center;height:46px;box-sizing:border-box;cursor:pointer;text-decoration:none;color:#e2e8f0;font-family:system-ui,-apple-system,sans-serif}
 .rp-pill.rp-open{justify-content:flex-start;padding-left:16px}
 .rp-pill .rp-ic{width:34px;height:34px;flex:none;display:flex;align-items:center;justify-content:center;font-size:16px;border-radius:9px;background:rgba(255,255,255,.06)}
 .rp-pill .rp-lbl{max-width:0;opacity:0;overflow:hidden;white-space:nowrap;transition:max-width .25s ease,opacity .2s ease,margin .25s ease;font-size:13px;font-weight:600;line-height:1.25}
 .rp-pill.rp-open .rp-lbl{max-width:150px;opacity:1;margin-left:10px}
 .rp-pill.rp-active .rp-ic{background:#137fec;color:#fff}
 .rp-pill.rp-active .rp-lbl{color:#fff}
 .rp-pill .rp-lbl a{color:#94a3b8;font-size:11px;font-weight:500;text-decoration:none;display:block}
 .rp-pill .rp-lbl .em{display:block;max-width:150px;overflow:hidden;text-overflow:ellipsis}
</style>
<script>
(function(){
 function strip(){
  try{
   var aside=document.querySelector('aside'); if(!aside)return;
   var nav=aside.querySelector('nav'); if(!nav)return;
   var kids=nav.querySelectorAll(':scope > *');
   for(var i=0;i<kids.length;i++){ var ch=kids[i]; ch.style.display = (ch.querySelector('a[href="/dashboard"]')||ch.id==='rp-prod-nav'||ch.id==='rp-comis-nav'||ch.id==='rp-desp-nav'||ch.id==='rp-fact-nav'||ch.id==='rp-mov-nav'||ch.id==='rp-ads-nav'||ch.id==='rp-stock-nav') ? '' : 'none'; }
   Array.prototype.forEach.call(aside.children,function(c){ if(c.tagName!=='NAV' && !c.querySelector('nav') && !(c.tagName==='A' && c.getAttribute('aria-label')) && !c.classList.contains('rp-pill')) c.style.display='none'; });
   // Agregar "Productos" en la barra: clon del item de Dashboard (queda idéntico y nativo).
   if(!nav.querySelector('#rp-prod-nav')){
    var di=null, kk=nav.querySelectorAll(':scope > *');
    for(var qi=0;qi<kk.length;qi++){ if(kk[qi].querySelector('a[href="/dashboard"]')){ di=kk[qi]; break; } }
    if(di){ var cl=di.cloneNode(true); cl.id='rp-prod-nav'; cl.style.display='';
     var ael=cl.querySelector('a'); if(ael){ ael.setAttribute('href','#'); ael.removeAttribute('aria-current'); ael.classList.remove('bg-white/[0.08]'); ael.classList.remove('text-primary'); ael.addEventListener('click',function(e){ e.preventDefault(); e.stopPropagation(); window.rpProd(true); }); }
     var icn=cl.querySelector('.material-symbols-outlined'); if(icn) icn.textContent='inventory_2';
     var sp2=cl.querySelectorAll('span'); for(var sj=0;sj<sp2.length;sj++){ var s2=sp2[sj]; if(!s2.classList.contains('material-symbols-outlined') && s2.children.length===0 && (s2.textContent||'').trim()){ s2.textContent='Productos'; } }
     di.parentNode.insertBefore(cl, di.nextSibling);
    }
   }
   // Agregar "Comisiones" en la barra (debajo de Productos).
   if(!nav.querySelector('#rp-comis-nav')){
    var pn=nav.querySelector('#rp-prod-nav');
    if(pn){ var cc=pn.cloneNode(true); cc.id='rp-comis-nav'; cc.style.display='';
     var ac=cc.querySelector('a'); if(ac){ ac.setAttribute('href','#'); ac.removeAttribute('aria-current'); ac.classList.remove('bg-white/[0.08]'); ac.classList.remove('text-primary');
      var newac=ac.cloneNode(true); ac.parentNode.replaceChild(newac,ac); newac.addEventListener('click',function(e){ e.preventDefault(); e.stopPropagation(); window.rpComis(true); }); ac=newac; }
     var ico=cc.querySelector('.material-symbols-outlined'); if(ico) ico.textContent='percent';
     var sp3=cc.querySelectorAll('span'); for(var sk=0;sk<sp3.length;sk++){ var s3=sp3[sk]; if(!s3.classList.contains('material-symbols-outlined') && s3.children.length===0 && (s3.textContent||'').trim()){ s3.textContent='Comisiones'; } }
     pn.parentNode.insertBefore(cc, pn.nextSibling);
    }
   }
   // Agregar "Despachos" en la barra (debajo de Comisiones).
   if(!nav.querySelector('#rp-desp-nav')){
    var cn0=nav.querySelector('#rp-comis-nav')||nav.querySelector('#rp-prod-nav');
    if(cn0){ var cd=cn0.cloneNode(true); cd.id='rp-desp-nav'; cd.style.display='';
     var ad=cd.querySelector('a'); if(ad){ ad.setAttribute('href','#'); ad.removeAttribute('aria-current'); ad.classList.remove('bg-white/[0.08]'); ad.classList.remove('text-primary');
      var nad=ad.cloneNode(true); ad.parentNode.replaceChild(nad,ad); nad.addEventListener('click',function(e){ e.preventDefault(); e.stopPropagation(); window.rpDesp(true); }); ad=nad; }
     var icd=cd.querySelector('.material-symbols-outlined'); if(icd) icd.textContent='local_shipping';
     var sp4=cd.querySelectorAll('span'); for(var sm=0;sm<sp4.length;sm++){ var s4=sp4[sm]; if(!s4.classList.contains('material-symbols-outlined') && s4.children.length===0 && (s4.textContent||'').trim()){ s4.textContent='Despachos'; } }
     cn0.parentNode.insertBefore(cd, cn0.nextSibling);
    }
   }
   // Agregar "Facturación" en la barra (debajo de Despachos).
   if(!nav.querySelector('#rp-fact-nav')){
    var dn0=nav.querySelector('#rp-desp-nav')||nav.querySelector('#rp-comis-nav')||nav.querySelector('#rp-prod-nav');
    if(dn0){ var cff=dn0.cloneNode(true); cff.id='rp-fact-nav'; cff.style.display='';
     var afx=cff.querySelector('a'); if(afx){ afx.setAttribute('href','#'); afx.removeAttribute('aria-current'); afx.classList.remove('bg-white/[0.08]'); afx.classList.remove('text-primary');
      var nafx=afx.cloneNode(true); afx.parentNode.replaceChild(nafx,afx); nafx.addEventListener('click',function(e){ e.preventDefault(); e.stopPropagation(); window.rpFact(true); }); afx=nafx; }
     var icff=cff.querySelector('.material-symbols-outlined'); if(icff) icff.textContent='receipt_long';
     var spff=cff.querySelectorAll('span'); for(var sq=0;sq<spff.length;sq++){ var s6=spff[sq]; if(!s6.classList.contains('material-symbols-outlined') && s6.children.length===0 && (s6.textContent||'').trim()){ s6.textContent='Facturación'; } }
     dn0.parentNode.insertBefore(cff, dn0.nextSibling);
    }
   }
   // Agregar "Movimientos" en la barra (debajo de Facturación).
   if(!nav.querySelector('#rp-mov-nav')){
    var fn0=nav.querySelector('#rp-fact-nav')||nav.querySelector('#rp-desp-nav')||nav.querySelector('#rp-comis-nav');
    if(fn0){ var cmv=fn0.cloneNode(true); cmv.id='rp-mov-nav'; cmv.style.display='';
     var amv=cmv.querySelector('a'); if(amv){ amv.setAttribute('href','#'); amv.removeAttribute('aria-current'); amv.classList.remove('bg-white/[0.08]'); amv.classList.remove('text-primary');
      var namv=amv.cloneNode(true); amv.parentNode.replaceChild(namv,amv); namv.addEventListener('click',function(e){ e.preventDefault(); e.stopPropagation(); window.rpMov(true); }); amv=namv; }
     var icmv=cmv.querySelector('.material-symbols-outlined'); if(icmv) icmv.textContent='swap_vert';
     var spmv=cmv.querySelectorAll('span'); for(var sv=0;sv<spmv.length;sv++){ var s7=spmv[sv]; if(!s7.classList.contains('material-symbols-outlined') && s7.children.length===0 && (s7.textContent||'').trim()){ s7.textContent='Movimientos'; } }
     fn0.parentNode.insertBefore(cmv, fn0.nextSibling);
    }
   }
   // Agregar "Subir ADS" en la barra (debajo de Movimientos).
   if(!nav.querySelector('#rp-ads-nav')){
    var mn0=nav.querySelector('#rp-mov-nav')||nav.querySelector('#rp-fact-nav')||nav.querySelector('#rp-desp-nav');
    if(mn0){ var cad=mn0.cloneNode(true); cad.id='rp-ads-nav'; cad.style.display='';
     var aad=cad.querySelector('a'); if(aad){ aad.setAttribute('href','#'); aad.removeAttribute('aria-current'); aad.classList.remove('bg-white/[0.08]'); aad.classList.remove('text-primary');
      var naad=aad.cloneNode(true); aad.parentNode.replaceChild(naad,aad); naad.addEventListener('click',function(e){ e.preventDefault(); e.stopPropagation(); window.rpAds(true); }); aad=naad; }
     var icad=cad.querySelector('.material-symbols-outlined'); if(icad) icad.textContent='rocket_launch';
     var spad=cad.querySelectorAll('span'); for(var sw=0;sw<spad.length;sw++){ var s8=spad[sw]; if(!s8.classList.contains('material-symbols-outlined') && s8.children.length===0 && (s8.textContent||'').trim()){ s8.textContent='Subir ADS'; } }
     mn0.parentNode.insertBefore(cad, mn0.nextSibling);
    }
   }
   // Agregar "Stock" en la barra (debajo de Subir ADS).
   if(!nav.querySelector('#rp-stock-nav')){
    var an0=nav.querySelector('#rp-ads-nav')||nav.querySelector('#rp-mov-nav')||nav.querySelector('#rp-fact-nav');
    if(an0){ var cst=an0.cloneNode(true); cst.id='rp-stock-nav'; cst.style.display='';
     var ast=cst.querySelector('a'); if(ast){ ast.setAttribute('href','#'); ast.removeAttribute('aria-current'); ast.classList.remove('bg-white/[0.08]'); ast.classList.remove('text-primary');
      var nast=ast.cloneNode(true); ast.parentNode.replaceChild(nast,ast); nast.addEventListener('click',function(e){ e.preventDefault(); e.stopPropagation(); window.rpStock(true); }); ast=nast; }
     var icst=cst.querySelector('.material-symbols-outlined'); if(icst) icst.textContent='warehouse';
     var spst=cst.querySelectorAll('span'); for(var sx=0;sx<spst.length;sx++){ var s9=spst[sx]; if(!s9.classList.contains('material-symbols-outlined') && s9.children.length===0 && (s9.textContent||'').trim()){ s9.textContent='Stock'; } }
     an0.parentNode.insertBefore(cst, an0.nextSibling);
    }
   }
   // Al tocar Dashboard (o el logo), cerrar los overlays abiertos (Productos/Integraciones).
   var dls=aside.querySelectorAll('a[href="/dashboard"]');
   for(var dz=0;dz<dls.length;dz++){ if(!dls[dz]._rpc){ dls[dz]._rpc=1; dls[dz].addEventListener('click',function(){ try{window.rpProd(false);}catch(e){} try{window.rpInteg(false);}catch(e){} try{window.rpComis(false);}catch(e){} try{window.rpDesp(false);}catch(e){} try{window.rpFact(false);}catch(e){} try{window.rpMov(false);}catch(e){} try{window.rpAds(false);}catch(e){} try{window.rpStock(false);}catch(e){} }); } }
   // Ocultar TODAS las secciones demo "Top productos" (hardcodeadas del pf.html, una por panel).
   var tops=document.querySelectorAll('h1,h2,h3,h4');
   for(var ti=0;ti<tops.length;ti++){ if((tops[ti].textContent||'').indexOf('Top productos')>-1){ var nd=tops[ti];
     for(var up=0; up<8 && nd.parentElement; up++){ nd=nd.parentElement; var cn=(typeof nd.className==='string')?nd.className:'';
      if(/rounded/.test(cn) && /border/.test(cn)){ nd.style.display='none'; break; } } } }
   // Meter los pills DENTRO del aside: al pasarles el mouse cuenta como hover de la barra y NO se cierra.
   var _pp=document.querySelectorAll('.rp-pill');
   for(var pi=0;pi<_pp.length;pi++){ if(_pp[pi].parentNode!==aside){ try{ aside.appendChild(_pp[pi]); }catch(e){} } }
   // El aside crea su propio stacking-context → subo su z-index POR ENCIMA de los overlays de
   // sección (100000) para que los pills (Integraciones + Cerrar sesión) SIEMPRE sean clickeables,
   // pero POR DEBAJO de los modales full-screen (100002) que sí deben tapar todo.
   try{ aside.style.zIndex='100001'; if(getComputedStyle(aside).position==='static') aside.style.position='relative'; }catch(e){}
   // Los pills (Integraciones + cuenta) se abren/cierran A LA PAR de la barra lateral.
   if(!aside._rpSync){ aside._rpSync=1;
    var expW=220;
    var apply=function(open,w){ var ps=document.querySelectorAll('.rp-pill'); for(var k=0;k<ps.length;k++){ ps[k].style.width=w+'px'; ps[k].classList.toggle('rp-open',open); } };
    var sync=function(){ var w=Math.round(aside.getBoundingClientRect().width); if(w>110)expW=w; apply(w>110,w); var ov=document.getElementById('rp-integ-ov'); if(ov) ov.style.left=w+'px'; var ov2=document.getElementById('rp-prod-ov'); if(ov2) ov2.style.left=w+'px'; var ov3=document.getElementById('rp-comis-ov'); if(ov3) ov3.style.left=w+'px'; var ov4=document.getElementById('rp-desp-ov'); if(ov4) ov4.style.left=w+'px'; var ov5=document.getElementById('rp-fact-ov'); if(ov5) ov5.style.left=w+'px'; var ov6=document.getElementById('rp-mov-ov'); if(ov6) ov6.style.left=w+'px'; var ov7=document.getElementById('rp-ads-ov'); if(ov7) ov7.style.left=w+'px'; var ov8=document.getElementById('rp-stock-ov'); if(ov8) ov8.style.left=w+'px'; var lk=document.getElementById('rpf-lock'); if(lk) lk.style.left=w+'px'; };
    try{ new ResizeObserver(sync).observe(aside); }catch(e){}
    var ps=document.querySelectorAll('.rp-pill');
    for(var k=0;k<ps.length;k++){ (function(p){ p.addEventListener('mouseenter',function(){ apply(true,expW); }); p.addEventListener('mouseleave',function(){ setTimeout(sync,40); }); })(ps[k]); }
    sync();
   }
  }catch(e){}
 }
 function boot(){ strip(); try{ new MutationObserver(strip).observe(document.body,{childList:true,subtree:true}); }catch(e){} }
 if(document.readyState!=='loading') boot(); else document.addEventListener('DOMContentLoaded', boot);
})();
</script>
<a class="rp-pill" id="rp-integ-btn" href="#" onclick="rpInteg(true);return false;" style="bottom:74px"><span class="rp-ic"><span class="material-symbols-outlined" style="font-size:20px">extension</span></span><span class="rp-lbl">Integraciones</span></a>
<div id="rp-integ-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;background:#0b1220;display:none;overflow:auto;transition:left .18s ease;font-family:system-ui,-apple-system,sans-serif">
 <div style="max-width:960px;margin:0 auto;padding:26px 32px 60px">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:16px">
   <div><h1 style="margin:0;font-size:22px;color:#f1f5f9">Integraciones</h1>
    <div style="color:#94a3b8;font-size:13px;margin-top:5px">Conect&aacute; tus ventas, pagos y anuncios para ver, en un solo lugar, si tu negocio gana o pierde plata.</div></div>
   <button onclick="rpInteg(false)" title="Cerrar" style="flex:none;background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;width:38px;height:38px;border-radius:10px;font-size:16px;cursor:pointer">&#10005;</button>
  </div>
  <div style="font-size:11px;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;margin:24px 0 13px;font-weight:700">Plataformas disponibles</div>
  <div id="rp-integ-cards"></div>
 </div>
</div>
<div id="rp-prod-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;background:#0a111e;display:none;overflow:auto;transition:left .18s ease;font-family:system-ui,-apple-system,sans-serif">
 <div style="max-width:1120px;margin:0 auto;padding:26px 30px 60px">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:16px">
   <div style="display:flex;align-items:center;gap:12px">
    <div style="width:42px;height:42px;border-radius:11px;background:#101c2e;border:1px solid #1e2b3d;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="color:#60a5fa">description</span></div>
    <div><h1 style="margin:0;font-size:22px;color:#f1f5f9">Costos</h1><div style="color:#94a3b8;font-size:13px;margin-top:3px">Defin&iacute; el costo de cada producto. Alimenta el margen real del Dashboard.</div></div>
   </div>
   <button onclick="rpProd(false)" title="Cerrar" style="flex:none;background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;width:38px;height:38px;border-radius:10px;font-size:16px;cursor:pointer">&#10005;</button>
  </div>
  <div style="background:#0f1826;border:1px solid #1e2b3d;border-radius:14px;margin-top:20px">
   <div style="display:flex;align-items:center;gap:9px;padding:15px 18px;border-bottom:1px solid #1a2636;font-weight:700;color:#f1f5f9;font-size:14.5px"><span class="material-symbols-outlined" style="color:#60a5fa;font-size:19px">inventory_2</span>Productos</div>
   <div style="padding:18px">
    <div style="display:flex;justify-content:space-between;align-items:center"><h2 style="font-size:16px;color:#f1f5f9;margin:0">Costos de productos</h2><div id="rp-prod-ref" style="color:#94a3b8;font-size:12px;cursor:pointer">&#8635; Actualizar</div></div>
    <div id="rp-prod-chip" style="margin-top:14px"></div>
    <div id="rp-prod-warn" style="margin-top:16px"></div>
    <div id="rp-prod-body" style="margin-top:8px"></div>
   </div>
  </div>
 </div>
</div>
<div id="rp-comis-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;background:#0a111e;display:none;overflow:auto;transition:left .18s ease;font-family:system-ui,-apple-system,sans-serif">
 <div style="max-width:900px;margin:0 auto;padding:26px 30px 60px">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:16px">
   <div style="display:flex;align-items:center;gap:12px">
    <div style="width:42px;height:42px;border-radius:11px;background:#101c2e;border:1px solid #1e2b3d;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="color:#60a5fa">percent</span></div>
    <div><h1 style="margin:0;font-size:22px;color:#f1f5f9">Costos</h1><div style="color:#94a3b8;font-size:13px;margin-top:3px">Comisiones e impuestos. Restan del margen real del Dashboard.</div></div>
   </div>
   <button onclick="rpComis(false)" title="Cerrar" style="flex:none;background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;width:38px;height:38px;border-radius:10px;font-size:16px;cursor:pointer">&#10005;</button>
  </div>
  <div style="display:flex;align-items:center;gap:11px;margin:22px 0 4px"><span class="material-symbols-outlined" style="color:#60a5fa;font-size:22px">account_balance_wallet</span><div style="font-weight:800;color:#f1f5f9;font-size:16px">Comisiones e impuestos</div></div>
  <div style="color:#94a3b8;font-size:12.5px;margin-bottom:6px">Carg&aacute; una vez lo que se lleva cada venta. Se descuenta de tu ganancia real.</div>

  <div style="background:#0f1826;border:1px solid #1e2b3d;border-radius:14px;padding:18px 20px;margin-top:14px;display:flex;gap:15px">
   <div style="width:44px;height:44px;border-radius:11px;flex:none;background:#0d1b30;border:1px solid #1c3350;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="color:#5aa2f5;font-size:22px">credit_card</span></div>
   <div style="flex:1;min-width:0">
    <div style="display:flex;align-items:center;gap:9px;flex-wrap:wrap"><div style="font-weight:700;color:#f1f5f9;font-size:15px">Mercado Pago</div><span style="background:#0d1b30;border:1px solid #1c3350;color:#7db3f5;font-size:10.5px;font-weight:700;border-radius:20px;padding:2px 9px">COBROS</span></div>
    <div style="color:#94a3b8;font-size:12.5px;margin-top:5px">Se usa la plata <b style="color:#7db3f5">REAL que entra</b> a tu MercadoPago. La comisi&oacute;n y las cuotas ya vienen descontadas por MP &mdash; no cargás nada ac&aacute;.</div>
    <input id="rp-c-mp" type="hidden" value="0"><input id="rp-c-cuotas" type="hidden" value="0">
   </div>
  </div>

  <div style="background:#0f1826;border:1px solid #1e2b3d;border-radius:14px;padding:18px 20px;margin-top:14px;display:flex;gap:15px">
   <div style="width:44px;height:44px;border-radius:11px;flex:none;background:#0e2a1c;border:1px solid #17492f;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="color:#34d399;font-size:22px">storefront</span></div>
   <div style="flex:1;min-width:0">
    <div style="font-weight:700;color:#f1f5f9;font-size:15px">Comisi&oacute;n de tienda</div>
    <div style="color:#94a3b8;font-size:12.5px;margin-top:5px">Lo que cobra tu plataforma (Shopify / Tiendanube). <b style="color:#34d399">Fijo 1%</b> por venta, no editable.</div>
    <div style="margin-top:14px"><div style="display:inline-flex;align-items:center;background:#0b1220;border:1px solid #1e2b3d;border-radius:10px;overflow:hidden"><span style="padding:9px 11px;color:#64748b;font-size:13px;border-right:1px solid #1e2b3d">%</span><input id="rp-c-tienda" type="number" step="0.01" placeholder="0" style="border:none;background:transparent;color:#f1f5f9;padding:9px 12px;width:120px;font-size:13px;text-align:right;outline:none"></div></div>
   </div>
  </div>

  <input id="rp-c-iva" type="hidden" value="0">

  <div style="background:#0f1826;border:1px solid #1e2b3d;border-radius:14px;padding:18px 20px;margin-top:14px;display:flex;gap:15px">
   <div style="width:44px;height:44px;border-radius:11px;flex:none;background:#1a1533;border:1px solid #322a5a;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="color:#a78bfa;font-size:22px">account_balance</span></div>
   <div style="flex:1;min-width:0">
    <div style="font-weight:700;color:#f1f5f9;font-size:15px">Ingresos Brutos</div>
    <div style="color:#94a3b8;font-size:12.5px;margin-top:5px">IIBB provincial sobre la venta. <b style="color:#a78bfa">Fijo 3,5%</b>, no editable. A este NO se le suma IVA.</div>
    <div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap;margin-top:14px"><div style="display:inline-flex;align-items:center;background:#0b1220;border:1px solid #1e2b3d;border-radius:10px;overflow:hidden"><span style="padding:9px 11px;color:#64748b;font-size:13px;border-right:1px solid #1e2b3d">%</span><input id="rp-c-iibb" type="number" step="0.01" placeholder="0" style="border:none;background:transparent;color:#f1f5f9;padding:9px 12px;width:120px;font-size:13px;text-align:right;outline:none"></div><span style="color:#5b6b82;font-size:12px">Fijo 3,5%</span></div>
   </div>
  </div>

  <div style="background:#0f1826;border:1px solid #1e2b3d;border-radius:14px;padding:18px 20px;margin-top:14px;display:flex;gap:15px">
   <div style="width:44px;height:44px;border-radius:11px;flex:none;background:#10233a;border:1px solid #1c3f63;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="color:#5aa2f5;font-size:22px">local_shipping</span></div>
   <div style="flex:1;min-width:0">
    <div style="font-weight:700;color:#f1f5f9;font-size:15px">Env&iacute;o</div>
    <div style="color:#94a3b8;font-size:12.5px;margin-top:5px">Se resta autom&aacute;tico por pedido. Si ten&eacute;s <b style="color:#5aa2f5">Envialo</b> conectado, usa el <b style="color:#34d399">costo REAL</b> de cada env&iacute;o; si el pedido a&uacute;n no est&aacute; en Envialo, usa promedio (Domicilio $9.000 &middot; Sucursal $6.000).</div>
   </div>
  </div>

  <div style="display:flex;justify-content:space-between;align-items:center;margin-top:20px;gap:14px;flex-wrap:wrap;background:#0c1521;border:1px solid #1e2b3d;border-radius:14px;padding:16px 20px">
   <div><div style="color:#94a3b8;font-size:12px">Impuestos que sumamos por venta (aparte del neto de MP)</div><b id="rp-c-total" style="color:#34d399;font-size:24px;font-weight:800">0%</b></div>
   <div style="display:flex;gap:12px;align-items:center"><span id="rp-c-msg" style="font-size:12.5px;font-weight:600;display:none"></span><button id="rp-c-go" onclick="rpComisSave()" style="background:#137fec;border:none;color:#fff;border-radius:10px;padding:11px 26px;font-weight:700;font-size:13.5px;cursor:pointer">Guardar</button></div>
  </div>
 </div>
</div>
<div id="rp-desp-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;background:#080c15;display:none;overflow:auto;transition:left .18s ease;font-family:system-ui,-apple-system,sans-serif;color:#f1f5f9">
 <div style="max-width:1240px;margin:0 auto;padding:24px 30px 64px">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:16px">
   <div style="display:flex;align-items:center;gap:13px">
    <div style="width:46px;height:46px;border-radius:12px;background:linear-gradient(160deg,#12233b,#0c1626);border:1px solid #1d3350;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="color:#5aa2f5;font-size:24px">local_shipping</span></div>
    <div><h1 style="margin:0;font-size:23px;color:#f1f5f9">Despachos</h1><div style="color:#8493a8;font-size:13px;margin-top:4px;max-width:640px;line-height:1.4">Eleg&iacute; los pedidos que vas a despachar, export&aacute; el Excel de Andreani y marc&aacute; los que ya enviaste. Solo aparecen las ventas <b style="color:#cbd5e1">pagadas</b>.</div></div>
   </div>
   <button onclick="rpDesp(false)" title="Cerrar" style="flex:none;background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;width:38px;height:38px;border-radius:10px;font-size:16px;cursor:pointer">&#10005;</button>
  </div>

  <div id="rp-d-cards" style="display:grid;grid-template-columns:repeat(4,1fr);gap:13px;margin:20px 0">
   <button class="rp-dc rp-on" data-f="empaquetar" onclick="rpDFilt('empaquetar')" style="background:#0e1521;border:1px solid #5a4a1e;border-radius:16px;padding:16px 18px;cursor:pointer;text-align:left;min-height:100px">
    <div style="display:flex;align-items:center;justify-content:space-between"><span style="width:31px;height:31px;border-radius:9px;background:#241c0b;color:#f0b429;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="font-size:17px">inventory_2</span></span><span class="rp-dl" style="color:#f0b429;font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.7px">Por empaquetar</span></div>
    <div style="display:flex;align-items:flex-end;justify-content:space-between;margin-top:15px"><span id="rp-d-n-empaquetar" style="font-size:27px;font-weight:800;line-height:1">0</span><span id="rp-d-m-empaquetar" style="color:#5b6b82;font-size:13px;font-weight:600">$0</span></div>
   </button>
   <button class="rp-dc" data-f="exportada" onclick="rpDFilt('exportada')" style="background:#0e1521;border:1px solid #1a2333;border-radius:16px;padding:16px 18px;cursor:pointer;text-align:left;min-height:100px">
    <div style="display:flex;align-items:center;justify-content:space-between"><span style="width:31px;height:31px;border-radius:9px;background:#0a2434;color:#38bdf8;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="font-size:17px">task_alt</span></span><span class="rp-dl" style="color:#8493a8;font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.7px">Exportadas</span></div>
    <div style="display:flex;align-items:flex-end;justify-content:space-between;margin-top:15px"><span id="rp-d-n-exportada" style="font-size:27px;font-weight:800;line-height:1">0</span><span id="rp-d-m-exportada" style="color:#5b6b82;font-size:13px;font-weight:600">$0</span></div>
   </button>
   <button class="rp-dc" data-f="enviada" onclick="rpDFilt('enviada')" style="background:#0e1521;border:1px solid #1a2333;border-radius:16px;padding:16px 18px;cursor:pointer;text-align:left;min-height:100px">
    <div style="display:flex;align-items:center;justify-content:space-between"><span style="width:31px;height:31px;border-radius:9px;background:#0e2a1c;color:#34d399;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="font-size:17px">local_shipping</span></span><span class="rp-dl" style="color:#8493a8;font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.7px">Enviadas</span></div>
    <div style="display:flex;align-items:flex-end;justify-content:space-between;margin-top:15px"><span id="rp-d-n-enviada" style="font-size:27px;font-weight:800;line-height:1">0</span><span id="rp-d-m-enviada" style="color:#5b6b82;font-size:13px;font-weight:600">$0</span></div>
   </button>
   <button class="rp-dc" data-f="todas" onclick="rpDFilt('todas')" style="background:#0e1521;border:1px solid #1a2333;border-radius:16px;padding:16px 18px;cursor:pointer;text-align:left;min-height:100px">
    <div style="display:flex;align-items:center;justify-content:space-between"><span style="width:31px;height:31px;border-radius:9px;background:#141d2c;color:#93a3b8;display:flex;align-items:center;justify-content:center"><span class="material-symbols-outlined" style="font-size:17px">receipt_long</span></span><span class="rp-dl" style="color:#8493a8;font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.7px">Todas &middot; facturaci&oacute;n</span></div>
    <div style="display:flex;align-items:flex-end;justify-content:space-between;margin-top:15px"><span id="rp-d-n-todas" style="font-size:27px;font-weight:800;line-height:1">0</span><span id="rp-d-m-todas" style="color:#5b6b82;font-size:13px;font-weight:600">$0</span></div>
   </button>
  </div>

  <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:13px">
   <div id="rp-d-segs" style="display:flex;background:#0b111c;border:1px solid #1a2333;border-radius:11px;padding:3px">
    <button class="rp-seg rp-son" data-p="todas" onclick="rpDPer('todas')" style="border:0;background:#2563eb;color:#fff;font-size:13px;font-weight:700;padding:7px 15px;border-radius:8px;cursor:pointer">Todas</button>
    <button class="rp-seg" data-p="hoy" onclick="rpDPer('hoy')" style="border:0;background:transparent;color:#8493a8;font-size:13px;font-weight:700;padding:7px 15px;border-radius:8px;cursor:pointer">Hoy</button>
    <button class="rp-seg" data-p="ayer" onclick="rpDPer('ayer')" style="border:0;background:transparent;color:#8493a8;font-size:13px;font-weight:700;padding:7px 15px;border-radius:8px;cursor:pointer">Ayer</button>
    <button class="rp-seg" data-p="7" onclick="rpDPer('7')" style="border:0;background:transparent;color:#8493a8;font-size:13px;font-weight:700;padding:7px 15px;border-radius:8px;cursor:pointer">7 d&iacute;as</button>
   </div>
   <div style="display:inline-flex;align-items:center;gap:7px;background:#0b111c;border:1px solid #1a2333;border-radius:10px;padding:6px 11px">
    <input id="rp-d-desde" type="date" onchange="rpDRango()" style="background:transparent;border:0;color:#c7d2e0;font-size:12.5px;outline:none;color-scheme:dark">
    <span style="color:#5b6b82;font-size:12px">a</span>
    <input id="rp-d-hasta" type="date" onchange="rpDRango()" style="background:transparent;border:0;color:#c7d2e0;font-size:12.5px;outline:none;color-scheme:dark">
   </div>
   <div style="flex:1;min-width:170px;display:flex;align-items:center;gap:9px;background:#0b111c;border:1px solid #1a2333;border-radius:11px;padding:0 13px"><span class="material-symbols-outlined" style="color:#5b6b82;font-size:18px">search</span><input id="rp-d-q" oninput="rpDRender()" placeholder="Buscar pedido, cliente, localidad&hellip;" style="flex:1;border:0;background:transparent;color:#f1f5f9;padding:10px 0;font-size:13.5px;outline:none"></div>
  </div>

  <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;justify-content:flex-end;margin-bottom:8px">
   <label style="display:inline-flex;align-items:center;gap:8px;background:#0b111c;border:1px solid #1a2333;color:#c7d2e0;border-radius:11px;padding:10px 14px;font-size:13px;font-weight:700;cursor:pointer"><input type="checkbox" id="rp-d-all" onclick="rpDAll(this)" style="width:16px;height:16px;accent-color:#3b82f6;cursor:pointer">Todas</label>
   <button id="rp-d-sync" onclick="rpDLoad()" style="display:inline-flex;align-items:center;gap:8px;background:#0b111c;border:1px solid #1a2333;color:#c7d2e0;border-radius:11px;padding:10px 16px;font-size:13px;font-weight:700;cursor:pointer"><span class="material-symbols-outlined" style="font-size:17px">sync</span>Sincronizar</button>
   <button onclick="rpDExcel()" style="display:inline-flex;align-items:center;gap:8px;background:linear-gradient(160deg,#4c3a8f,#3a2c73);border:1px solid #4a3a86;color:#e5ddff;border-radius:11px;padding:10px 16px;font-size:13px;font-weight:700;cursor:pointer"><span class="material-symbols-outlined" style="font-size:17px">download</span>Generar Excel Andreani</button>
   <button onclick="rpDActSku()" style="display:inline-flex;align-items:center;gap:8px;background:#0b111c;border:1px solid #1a2333;color:#c7d2e0;border-radius:11px;padding:10px 16px;font-size:13px;font-weight:700;cursor:pointer"><span class="material-symbols-outlined" style="font-size:17px">barcode</span>Actualizar SKUs</button>
   <button onclick="rpDOpenSku()" style="display:inline-flex;align-items:center;gap:8px;background:linear-gradient(160deg,#3b3a8f,#2c2b6b);border:1px solid #3a3a86;color:#dcdcff;border-radius:11px;padding:10px 16px;font-size:13px;font-weight:700;cursor:pointer"><span class="material-symbols-outlined" style="font-size:17px">qr_code_2</span>Insertar SKU</button>
   <button onclick="rpDOpenSeg()" style="display:inline-flex;align-items:center;gap:8px;background:linear-gradient(160deg,#b23a55,#8f2c44);border:1px solid #a23650;color:#ffe0e7;border-radius:11px;padding:10px 16px;font-size:13px;font-weight:700;cursor:pointer"><span class="material-symbols-outlined" style="font-size:17px">local_shipping</span>Enviar seguimiento</button>
  </div>
  <div id="rp-d-status" style="color:#34d399;font-size:12.5px;font-weight:600;min-height:18px;margin:2px 2px 10px"></div>

  <div style="background:#0e1521;border:1px solid #1a2333;border-radius:16px;overflow:hidden">
   <div style="overflow-x:auto">
    <table style="width:100%;border-collapse:collapse;min-width:760px">
     <thead><tr>
      <th style="width:36px;padding:14px 13px;border-bottom:1px solid #1a2333"></th>
      <th style="color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:14px 13px;border-bottom:1px solid #1a2333;text-align:left">Pedido</th>
      <th style="color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:14px 13px;border-bottom:1px solid #1a2333;text-align:left">Cliente</th>
      <th style="color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:14px 13px;border-bottom:1px solid #1a2333;text-align:left">Env&iacute;o</th>
      <th style="color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:14px 13px;border-bottom:1px solid #1a2333;text-align:left">Localidad</th>
      <th style="color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:14px 13px;border-bottom:1px solid #1a2333;text-align:left">CP</th>
      <th style="color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:14px 13px;border-bottom:1px solid #1a2333;text-align:right">U</th>
      <th style="color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:14px 13px;border-bottom:1px solid #1a2333;text-align:right">Total</th>
     </tr></thead>
     <tbody id="rp-d-body"></tbody>
    </table>
   </div>
   <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;padding:13px 15px;color:#8493a8;font-size:12.5px;border-top:1px solid #1a2333"><span id="rp-d-cnt">&mdash;</span><span style="color:#5b6b82">Solo ventas pagadas &middot; no despachadas.</span></div>
  </div>
 </div>
</div>
<div id="rp-d-skuov" style="position:fixed;inset:0;z-index:100002;background:rgba(4,8,14,.72);display:none;align-items:center;justify-content:center;padding:20px;font-family:system-ui,-apple-system,sans-serif" onclick="if(event.target===this)rpDCloseSku()">
 <div style="width:100%;max-width:560px;background:#0e1521;border:1px solid #1a2333;border-radius:18px;padding:22px;box-shadow:0 24px 60px rgba(0,0,0,.6)">
  <div style="display:flex;align-items:flex-start;gap:12px">
   <div style="width:40px;height:40px;border-radius:11px;background:#1c1636;border:1px solid #3a2f6b;display:flex;align-items:center;justify-content:center;flex:none"><span class="material-symbols-outlined" style="color:#a78bfa;font-size:20px">qr_code_2</span></div>
   <div style="flex:1;min-width:0"><div style="font-size:16px;font-weight:800;color:#f1f5f9">Insertar SKU en r&oacute;tulos</div><div style="color:#8493a8;font-size:12.5px;margin-top:3px;line-height:1.45">Sub&iacute; el PDF de etiquetas: detectamos el <b style="color:#cbd5e1">formato de cada una</b> (Andreani domicilio/sucursal o Envialo) e insertamos el SKU en el hueco libre correcto, sin tocar el QR ni el c&oacute;digo de barras.</div></div>
   <button onclick="rpDCloseSku()" style="flex:none;background:#111c2b;border:1px solid #1a2333;color:#cbd5e1;width:32px;height:32px;border-radius:9px;cursor:pointer">&#10005;</button>
  </div>
  <label style="display:block;margin-top:18px;border:1.5px dashed #2b3a52;border-radius:14px;padding:34px 18px;text-align:center;cursor:pointer"><input type="file" accept="application/pdf" style="display:none" onchange="rpDUpSku(this)"><span class="material-symbols-outlined" style="color:#5b6b82;font-size:30px;display:block">upload_file</span><div style="color:#e7edf5;font-size:14px;font-weight:700;margin-top:6px">Arrastr&aacute; el PDF de etiquetas o hac&eacute; clic para elegirlo</div><div style="color:#5b6b82;font-size:12px;margin-top:5px">Te devuelve el PDF con los SKU insertados (Zebra o A4)</div></label>
  <div id="rp-d-skures" style="margin-top:14px"></div>
 </div>
</div>
<div id="rp-d-segov" style="position:fixed;inset:0;z-index:100002;background:rgba(4,8,14,.72);display:none;align-items:center;justify-content:center;padding:20px;font-family:system-ui,-apple-system,sans-serif" onclick="if(event.target===this)rpDCloseSeg()">
 <div style="width:100%;max-width:560px;background:#0e1521;border:1px solid #1a2333;border-radius:18px;padding:22px;box-shadow:0 24px 60px rgba(0,0,0,.6)">
  <div style="display:flex;align-items:flex-start;gap:12px">
   <div style="width:40px;height:40px;border-radius:11px;background:#2a1119;border:1px solid #5a2333;display:flex;align-items:center;justify-content:center;flex:none"><span class="material-symbols-outlined" style="color:#fb7185;font-size:20px">local_shipping</span></div>
   <div style="flex:1;min-width:0"><div style="font-size:16px;font-weight:800;color:#f1f5f9">Enviar seguimiento</div><div style="color:#8493a8;font-size:12.5px;margin-top:3px;line-height:1.45">Sub&iacute; el PDF de Andreani: leemos el <b style="color:#cbd5e1">N&deg; Interno + seguimiento</b>, te lo mostramos y al enviar le avisamos a tu tienda (Shopify/Tiendanube) &mdash; al cliente le llega el mail con el tracking.</div></div>
   <button onclick="rpDCloseSeg()" style="flex:none;background:#111c2b;border:1px solid #1a2333;color:#cbd5e1;width:32px;height:32px;border-radius:9px;cursor:pointer">&#10005;</button>
  </div>
  <label style="display:block;margin-top:18px;border:1.5px dashed #2b3a52;border-radius:14px;padding:34px 18px;text-align:center;cursor:pointer"><input type="file" accept="application/pdf" style="display:none" onchange="rpDUpSeg(this)"><span class="material-symbols-outlined" style="color:#5b6b82;font-size:30px;display:block">upload_file</span><div style="color:#e7edf5;font-size:14px;font-weight:700;margin-top:6px">Arrastr&aacute; el PDF o hac&eacute; clic para elegirlo</div><div style="color:#5b6b82;font-size:12px;margin-top:5px">Leemos cada r&oacute;tulo y cargamos el seguimiento</div></label>
  <div id="rp-d-segres" style="margin-top:14px"></div>
 </div>
</div>
<div id="rp-fact-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;background:#090b14;display:none;overflow:auto;transition:left .18s ease;font-family:system-ui,-apple-system,sans-serif;color:#e8edf4">
<style>
#rp-fact-ov .wrap{max-width:1560px;margin:0 auto;padding:22px 40px 60px}
#rp-fact-ov .topbar{display:flex;align-items:center;justify-content:flex-end;gap:10px;margin-bottom:20px;flex-wrap:wrap}
#rp-fact-ov .rangebox,#rp-fact-ov .emitbox{display:inline-flex;align-items:center;gap:8px;background:#0e1626;border:1px solid #1c2637;border-radius:10px;padding:7px 12px;font-size:12.5px;color:#c7d2e0}
#rp-fact-ov .rangebox svg{width:15px;height:15px;color:#7d8ba0;flex:none}
#rp-fact-ov .rangebox input,#rp-fact-ov .emitbox input{background:transparent;border:0;color:#dbeafe;font-size:12.5px;font-family:inherit;outline:none;color-scheme:dark}
#rp-fact-ov .emitbox{border-color:#274a7a;background:#0d1a2c}
#rp-fact-ov .emitbox .lab{color:#7fb0e6;font-weight:700;font-size:10.5px;text-transform:uppercase;letter-spacing:.6px}
#rp-fact-ov .chip{display:inline-flex;align-items:center;gap:7px;background:#0e1626;border:1px solid #1c2637;color:#c7d2e0;border-radius:10px;padding:8px 12px;font-size:12.5px;font-weight:600}
#rp-fact-ov .head{display:flex;align-items:flex-start;gap:13px;margin-bottom:16px}
#rp-fact-ov .head .hico{width:44px;height:44px;border-radius:12px;background:linear-gradient(160deg,#12233b,#0c1626);border:1px solid #1d3350;display:flex;align-items:center;justify-content:center;flex:none}
#rp-fact-ov .head .hico svg{width:22px;height:22px;color:#5aa2f5}
#rp-fact-ov .head h1{margin:0;font-size:24px;font-weight:800;color:#f4f7fb}
#rp-fact-ov .head p{margin:5px 0 0;color:#8b97a8;font-size:13px;max-width:560px;line-height:1.45}
#rp-fact-ov .head .x{margin-left:auto;flex:none;background:#0e1626;border:1px solid #1c2637;color:#aeb8c6;width:36px;height:36px;border-radius:10px;font-size:14px;cursor:pointer}
#rp-fact-ov .arca{display:flex;align-items:center;gap:10px;flex-wrap:wrap;background:#0c1626;border:1px solid #16324f;border-radius:12px;padding:11px 16px;margin-bottom:16px;font-size:13px;color:#9db0c4}
#rp-fact-ov .arca .dot{width:9px;height:9px;border-radius:50%;background:#34d399;box-shadow:0 0 0 4px rgba(52,211,153,.15);flex:none}
#rp-fact-ov .arca b{color:#e8edf4;font-weight:700}
#rp-fact-ov .arca .lbl{color:#5b6678;font-size:10.5px;text-transform:uppercase;letter-spacing:.6px;font-weight:800}
#rp-fact-ov .arca .cuit{background:rgba(19,127,236,.12);border:1px solid rgba(45,110,180,.5);color:#7fbaff;border-radius:7px;padding:3px 10px;font-weight:800}
#rp-fact-ov .arca .sep{color:#334155}
#rp-fact-ov .pills{display:flex;gap:9px;flex-wrap:wrap;margin-bottom:20px}
#rp-fact-ov .pill{display:inline-flex;align-items:center;gap:7px;background:#0d1524;border:1px solid #1a2436;color:#9aa6b6;border-radius:20px;padding:8px 15px;font-size:12.5px;font-weight:600;cursor:pointer}
#rp-fact-ov .pill.on{background:rgba(19,127,236,.14);border-color:#1e4f8a;color:#bcd7f7}
#rp-fact-ov .pill .d{width:7px;height:7px;border-radius:50%}
#rp-fact-ov .cont{background:#0b111e;border:1px solid #18212f;border-radius:20px;padding:20px 22px;margin-bottom:20px}
#rp-fact-ov .cont-h{display:flex;align-items:center;gap:10px;margin-bottom:6px}
#rp-fact-ov .cont-h .ci{width:26px;height:26px;border-radius:8px;background:#0e2036;color:#5aa2f5;display:flex;align-items:center;justify-content:center}
#rp-fact-ov .cont-h .ci svg{width:15px;height:15px}
#rp-fact-ov .cont-h b{font-size:15px;font-weight:700;color:#f4f7fb}
#rp-fact-ov .sec-lbl{display:flex;align-items:center;gap:8px;margin:16px 2px 12px}
#rp-fact-ov .sec-lbl .si{width:18px;height:18px;color:#34d399}
#rp-fact-ov .sec-lbl .si svg{width:16px;height:16px}
#rp-fact-ov .sec-lbl span{font-size:11px;font-weight:800;letter-spacing:1px;text-transform:uppercase;color:#7d8ba0}
#rp-fact-ov .sec-lbl small{color:#3f4a5a;font-size:12px;font-weight:700}
#rp-fact-ov .kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}
#rp-fact-ov .kcard{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;padding:17px 19px;position:relative;overflow:hidden}
#rp-fact-ov .kcard .kt{display:flex;align-items:center;justify-content:space-between}
#rp-fact-ov .kico{width:31px;height:31px;border-radius:9px;display:flex;align-items:center;justify-content:center}
#rp-fact-ov .kico svg{width:17px;height:17px}
#rp-fact-ov .kcard.v .kico{background:rgba(90,162,245,.13);color:#5aa2f5}
#rp-fact-ov .kcard.f .kico{background:rgba(52,211,153,.13);color:#34d399}
#rp-fact-ov .kcard.p .kico{background:rgba(251,191,36,.13);color:#fbbf24}
#rp-fact-ov .klabel{font-size:11px;font-weight:800;letter-spacing:.7px;text-transform:uppercase;color:#8b97a8}
#rp-fact-ov .knum{font-size:31px;font-weight:800;line-height:1;margin-top:15px}
#rp-fact-ov .kcard.v .knum{color:#f4f7fb}
#rp-fact-ov .kcard.f .knum{color:#34d399}
#rp-fact-ov .kcard.p .knum{color:#fbbf24}
#rp-fact-ov .ksub{font-size:12px;color:#5b6678;margin-top:6px}
#rp-fact-ov .acts{display:flex;gap:10px;flex-wrap:wrap;margin:18px 0 14px}
#rp-fact-ov .btn{display:inline-flex;align-items:center;gap:8px;border-radius:11px;padding:10px 16px;font-size:13px;font-weight:700;cursor:pointer;border:1px solid}
#rp-fact-ov .btn svg{width:16px;height:16px}
#rp-fact-ov .btn-primary{background:#137fec;border-color:#137fec;color:#fff}
#rp-fact-ov .btn-green{background:rgba(52,211,153,.12);border-color:#1f7a52;color:#34d399}
#rp-fact-ov .btn-ghost{background:#0d1524;border-color:#1a2436;color:#aeb8c6}
#rp-fact-ov .sld{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;padding:17px 20px;margin-bottom:20px}
#rp-fact-ov .sld-head{cursor:pointer;user-select:none;display:flex;align-items:center;gap:9px;font-size:13.5px;font-weight:700;color:#f4f7fb}
#rp-fact-ov .sld-head small{color:#5b6678;font-weight:500;font-size:12px}
#rp-fact-ov .sld-head .chev{width:17px;height:17px;margin-left:auto;color:#8b97a8;transition:transform .2s}
#rp-fact-ov .sld.open .sld-head .chev{transform:rotate(90deg)}
#rp-fact-ov .sld-body{display:none;margin-top:16px}
#rp-fact-ov .sld.open .sld-body{display:block}
#rp-fact-ov .sld .row{display:flex;align-items:center;gap:16px}
#rp-fact-ov .rng{flex:1;-webkit-appearance:none;appearance:none;height:8px;border-radius:20px;background:#182234;outline:none;cursor:pointer}
#rp-fact-ov .rng::-webkit-slider-thumb{-webkit-appearance:none;width:18px;height:18px;border-radius:50%;background:#fff;cursor:pointer}
#rp-fact-ov .pctv{font-size:26px;font-weight:800;color:#4a9bf0}
#rp-fact-ov .obj{color:#8b97a8;font-size:13px;margin-top:14px;display:flex;align-items:center;gap:6px;flex-wrap:wrap}
#rp-fact-ov .obj b{color:#e8edf4}
#rp-fact-ov .obj .hl{color:#4aa8ff;font-weight:800;font-size:15px;background:rgba(19,127,236,.14);border:1px solid rgba(45,110,180,.55);border-radius:9px;padding:4px 11px}
#rp-fact-ov .obj .sep{color:#3a4657}
#rp-fact-ov .autocard{display:flex;align-items:center;gap:16px;flex-wrap:wrap;background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;padding:15px 20px;margin-bottom:20px}
#rp-fact-ov .autocard .ac-ico{width:40px;height:40px;border-radius:11px;background:rgba(19,127,236,.13);color:#4aa8ff;display:flex;align-items:center;justify-content:center;flex:none}
#rp-fact-ov .autocard .ac-ico svg{width:22px;height:22px}
#rp-fact-ov .autocard .ac-txt{flex:1;min-width:220px}
#rp-fact-ov .autocard .ac-txt b{color:#e8edf4;font-size:14px;font-weight:700}
#rp-fact-ov .autocard .ac-txt .st{color:#8b97a8;font-size:12.5px;margin-top:3px;line-height:1.4}
#rp-fact-ov .ac-sel{background:#0d1524;border:1px solid #1a2436;color:#c7d2e0;border-radius:9px;padding:8px 11px;font-size:12.5px;cursor:pointer;color-scheme:dark;font-family:inherit}
#rp-fact-ov .switch{position:relative;display:inline-block;width:48px;height:27px;flex:none}
#rp-fact-ov .switch input{opacity:0;width:0;height:0}
#rp-fact-ov .switch .slider{position:absolute;inset:0;background:#243149;border-radius:20px;transition:.2s;cursor:pointer}
#rp-fact-ov .switch .slider:before{content:"";position:absolute;height:21px;width:21px;left:3px;top:3px;background:#fff;border-radius:50%;transition:.2s}
#rp-fact-ov .switch input:checked+.slider{background:#137fec}
#rp-fact-ov .switch input:checked+.slider:before{transform:translateX(21px)}
#rp-fact-ov .tcard{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;overflow:hidden}
#rp-fact-ov table{width:100%;border-collapse:collapse;font-size:13px;min-width:820px}
#rp-fact-ov thead th{text-align:left;color:#5b6678;font-size:10.5px;font-weight:800;letter-spacing:.7px;text-transform:uppercase;padding:14px 16px;border-bottom:1px solid #1b2536}
#rp-fact-ov tbody td{padding:13px 16px;border-bottom:1px solid #141d2e;color:#e8edf4}
#rp-fact-ov tbody tr:last-child td{border-bottom:none}
#rp-fact-ov .chk{width:15px;height:15px;border-radius:4px;border:1.6px solid #33507a;background:#0a1322;display:inline-block;vertical-align:middle;cursor:pointer}
#rp-fact-ov .chk.on{background:#137fec;border-color:#137fec;position:relative}
#rp-fact-ov .chk.on:after{content:"";position:absolute;left:4px;top:1px;width:4px;height:8px;border:solid #fff;border-width:0 2px 2px 0;transform:rotate(45deg)}
#rp-fact-ov .cust{display:flex;align-items:center;gap:10px}
#rp-fact-ov .av2{width:28px;height:28px;border-radius:50%;flex:none;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800;color:#fff}
#rp-fact-ov .medio{display:inline-flex;align-items:center;gap:6px;font-size:12px;color:#b8c6da}
#rp-fact-ov .midot{width:7px;height:7px;border-radius:50%}
#rp-fact-ov .comp{color:#5aa2f5;font-size:12.5px}
#rp-fact-ov .est{display:inline-flex;align-items:center;gap:5px;font-size:11px;font-weight:800;padding:4px 10px;border-radius:20px}
#rp-fact-ov .e-ok{background:rgba(52,211,153,.12);color:#34d399;border:1px solid #1f5a3d}
#rp-fact-ov .e-pd{background:rgba(251,191,36,.1);color:#fbbf24;border:1px solid #5a4410}
#rp-fact-ov .foot{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;padding:13px 16px;color:#8b97a8;font-size:12.5px;border-bottom:1px solid #1b2536}
#rp-fact-ov .pager{display:flex;align-items:center;justify-content:center;gap:6px;padding:16px;flex-wrap:wrap}
#rp-fact-ov .pbtn{min-width:34px;height:34px;padding:0 11px;border-radius:9px;background:#0d1524;border:1px solid #1a2436;color:#aeb8c6;font-size:13px;font-weight:700;cursor:pointer}
#rp-fact-ov .pbtn.on{background:#137fec;border-color:#137fec;color:#fff}
#rp-fact-ov .pbtn:disabled{opacity:.35;cursor:default}
#rp-fact-ov .pager .info{color:#5b6678;font-size:12px;margin:0 8px}
#rp-fact-ov .lk-in{background:#0a1322;border:1px solid #22324a;color:#e8edf4;border-radius:10px;padding:11px 13px;font-size:13.5px;font-family:inherit;outline:none;width:100%}
#rp-fact-ov .lk-in::placeholder{color:#5b6678}
#rp-fact-ov .lk-lb{font-size:11px;font-weight:700;color:#8b97a8;text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px;display:block}
#rp-fact-ov .lk-step{display:flex;gap:11px;align-items:flex-start;margin-bottom:11px}
#rp-fact-ov .lk-step .n{width:22px;height:22px;border-radius:50%;background:rgba(19,127,236,.15);color:#4aa8ff;font-size:12px;font-weight:800;display:flex;align-items:center;justify-content:center;flex:none;margin-top:1px}
#rp-fact-ov .lk-step div{color:#c7d2e0;font-size:13px;line-height:1.45}
#rp-fact-ov .lk-step b{color:#e8edf4}
</style>
 <div id="rpf-lock" style="position:fixed;top:0;left:72px;right:0;bottom:0;z-index:30;background:#090b14;display:none;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:30px">
  <div style="width:86px;height:86px;border-radius:24px;background:linear-gradient(160deg,#12233b,#0c1626);border:1px solid #1d3350;display:flex;align-items:center;justify-content:center;margin-bottom:22px"><svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="#5aa2f5" stroke-width="1.6"><rect x="4" y="10" width="16" height="11" rx="2"/><path d="M8 10V7a4 4 0 0 1 8 0v3"/><circle cx="12" cy="15.5" r="1.4"/></svg></div>
  <h2 style="margin:0;font-size:22px;font-weight:800;color:#f4f7fb">Facturaci&oacute;n en configuraci&oacute;n</h2>
  <p style="margin:10px 0 0;color:#8b97a8;font-size:14px;max-width:440px;line-height:1.5">Estamos terminando de conectar <b style="color:#cbd5e1">ARCA</b> para que puedas emitir tus facturas. Esta secci&oacute;n va a estar disponible <b style="color:#cbd5e1">muy pronto</b>.</p>
  <div style="margin-top:22px;display:inline-flex;align-items:center;gap:9px;background:rgba(251,191,36,.12);border:1px solid #5a4410;color:#fbd67e;border-radius:20px;padding:10px 18px;font-size:13.5px;font-weight:700"><svg width="17" height="17" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14.7 6.3a4 4 0 0 0-5.4 5.4l-6 6 2.7 2.7 6-6a4 4 0 0 0 5.4-5.4l-2.3 2.3-2.7-.7-.7-2.7z"/></svg>Disponible en breve &middot; en configuraci&oacute;n</div>
 </div>
 <div id="rpf-linkmodal" style="position:fixed;inset:0;z-index:40;background:rgba(4,8,14,.74);display:none;align-items:center;justify-content:center;padding:20px" onclick="if(event.target===this)rpFCloseLink()">
  <div style="width:100%;max-width:560px;background:#0c1420;border:1px solid #1b2536;border-radius:18px;padding:24px;box-shadow:0 24px 60px rgba(0,0,0,.6);max-height:90vh;overflow:auto">
   <div style="display:flex;align-items:center;gap:12px;margin-bottom:4px">
    <div style="width:40px;height:40px;border-radius:11px;background:rgba(19,127,236,.13);color:#4aa8ff;display:flex;align-items:center;justify-content:center;flex:none"><svg width="21" height="21" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><rect x="4" y="10" width="16" height="11" rx="2"/><path d="M8 10V7a4 4 0 0 1 8 0v3"/></svg></div>
    <div style="flex:1"><div style="font-size:17px;font-weight:800;color:#f4f7fb">Vincular ARCA</div><div style="color:#8b97a8;font-size:12.5px">Conect&aacute; tu facturaci&oacute;n electr&oacute;nica para empezar a emitir.</div></div>
    <button onclick="rpFCloseLink()" style="flex:none;background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;width:32px;height:32px;border-radius:9px;cursor:pointer">&#10005;</button>
   </div>
   <div style="display:flex;gap:8px;margin:16px 0 14px">
    <button id="rpf-m-serv" class="btn btn-primary" style="flex:1;justify-content:center" onclick="rpFSetMetodo('servicio')">Servicio (token)</button>
    <button id="rpf-m-cert" class="btn btn-ghost" style="flex:1;justify-content:center" onclick="rpFSetMetodo('cert')">Certificado AFIP</button>
   </div>
   <div id="rpf-steps" style="background:#0a1322;border:1px solid #17233a;border-radius:12px;padding:15px 16px;margin-bottom:15px"></div>
   <div style="display:grid;gap:12px">
    <div><span class="lk-lb">CUIT (11 d&iacute;gitos)</span><input id="rpf-in-cuit" class="lk-in" inputmode="numeric" placeholder="20402615483"></div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
     <div><span class="lk-lb">Punto de venta</span><input id="rpf-in-pv" class="lk-in" inputmode="numeric" placeholder="5"></div>
     <div><span class="lk-lb">Raz&oacute;n social</span><input id="rpf-in-nombre" class="lk-in" placeholder="Tu nombre / empresa"></div>
    </div>
    <div><span class="lk-lb" id="rpf-tk-lb">Token del servicio</span><input id="rpf-in-token" class="lk-in" placeholder="Peg&aacute; tu API token"></div>
   </div>
   <button class="btn btn-primary" style="width:100%;justify-content:center;margin-top:16px;padding:12px" onclick="rpFVincular()">Vincular y desbloquear</button>
   <div id="rpf-link-status" style="min-height:16px;margin-top:10px;font-size:12.5px;font-weight:600;color:#fb7185;text-align:center"></div>
  </div>
 </div>
 <div class="wrap">
  <div class="topbar">
   <div class="rangebox"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="4" width="18" height="18" rx="2"/><path d="M3 9h18M8 2v4M16 2v4"/></svg><input id="rpf-desde" type="date" onchange="rpFRango()"><span style="color:#5b6678">&rarr;</span><input id="rpf-hasta" type="date" onchange="rpFRango()"></div>
   <div class="emitbox"><span class="lab">Fecha de emisi&oacute;n</span><input id="rpf-emit" type="date"></div>
   <span class="chip">&#127462;&#127479; ARS</span>
  </div>
  <div class="head">
   <div class="hico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M6 3h9l3 3v15l-2.2-1.4L13.6 21 12 19.6 10.4 21 8.2 19.6 6 21Z"/><path d="M9 8h6M9 12h6M9 16h4"/></svg></div>
   <div><h1>Facturaci&oacute;n</h1><p>Factur&aacute; contra ARCA y export&aacute; el informe para el contador, sin salir de RealProfit.</p></div>
   <button class="x" onclick="rpFact(false)">&#10005;</button>
  </div>
  <div class="arca" id="rpf-arca"><span class="dot"></span><span>Conectado a ARCA</span></div>
  <div class="pills">
   <span class="pill on" data-f="todas" onclick="rpFFilt('todas')"><span class="d" style="background:#137fec"></span>Todas</span>
   <span class="pill" data-f="facturadas" onclick="rpFFilt('facturadas')"><span class="d" style="background:#34d399"></span>Facturadas</span>
   <span class="pill" data-f="pendientes" onclick="rpFFilt('pendientes')"><span class="d" style="background:#fbbf24"></span>No facturadas</span>
  </div>
  <div class="cont">
   <div class="cont-h"><span class="ci"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 3v18h18"/><path d="m7 14 4-4 3 3 5-6"/></svg></span><b>Resumen del per&iacute;odo</b></div>
   <div class="sec-lbl"><span class="si"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M6 3h9l3 3v15H6Z"/><path d="M9 9h6M9 13h6"/></svg></span><span>Facturaci&oacute;n</span><small>3</small></div>
   <div class="kpis">
    <div class="kcard v"><div class="kt"><span class="kico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 3v18h18"/><path d="m19 9-5 5-3-3-4 4"/></svg></span><span class="klabel">Ventas</span></div><div class="knum" id="rpf-k-ventas">0</div><div class="ksub">ventas pagadas del per&iacute;odo</div></div>
    <div class="kcard f"><div class="kt"><span class="kico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 12l2 2 4-4"/><circle cx="12" cy="12" r="9"/></svg></span><span class="klabel">Facturadas</span></div><div class="knum" id="rpf-k-fact">0</div><div class="ksub" id="rpf-ks-fact">emitidas</div></div>
    <div class="kcard p"><div class="kt"><span class="kico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 2"/></svg></span><span class="klabel">Pendientes</span></div><div class="knum" id="rpf-k-pend">0</div><div class="ksub">sin facturar</div></div>
   </div>
  </div>
  <div class="acts">
   <span class="btn btn-ghost" onclick="rpFSelectAll()"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="18" height="18" rx="3"/><path d="m8 12 3 3 5-6"/></svg>Seleccionar todas</span>
   <span class="btn btn-ghost" onclick="rpFDeselect()"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="18" height="18" rx="3"/><path d="M9 9l6 6M15 9l-6 6"/></svg>Deseleccionar</span>
   <span class="btn btn-green" onclick="rpFInforme()"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 3v12m0 0 4-4m-4 4-4-4"/><path d="M4 17v3h16v-3"/></svg>Exportar informe (contador)</span>
   <span class="btn btn-primary" onclick="rpFFacturarSel()"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M6 3h9l3 3v15H6Z"/><path d="M9 9h6M9 13h6"/></svg>Facturar seleccionadas</span>
  </div>
  <div id="rpf-status" style="min-height:18px;margin:0 2px 12px;font-size:12.5px;font-weight:600;color:#34d399"></div>
  <div class="sld" id="rpf-sld">
   <div class="sld-head" onclick="rpFToggleSld()"><svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="#137fec" stroke-width="2"><rect x="3" y="3" width="18" height="18" rx="2"/><path d="M3 9h18"/></svg>Facturar un % de las ventas<small>&mdash; herramienta aparte: factura por porcentaje, no toca lo seleccionado</small><svg class="chev" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4"><path d="m9 6 6 6-6 6"/></svg></div>
   <div class="sld-body">
    <div class="row"><input id="rpf-pct" class="rng" type="range" min="0" max="100" step="10" value="0" oninput="rpFPctInput()"><div class="pctv" id="rpf-pctv">0%</div><span class="btn btn-primary" style="padding:9px 15px" onclick="rpFFacturarPct()">Facturar este %</span></div>
    <div class="obj" id="rpf-pctobj"></div>
   </div>
  </div>
  <div class="autocard">
   <div class="ac-ico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><rect x="4" y="8" width="16" height="12" rx="2"/><path d="M12 8V4M9 4h6"/><circle cx="9" cy="14" r="1"/><circle cx="15" cy="14" r="1"/></svg></div>
   <div class="ac-txt"><b>Facturaci&oacute;n autom&aacute;tica</b><div class="st" id="rpf-autostate">Desactivada. Las ventas nuevas quedan <b style="color:#cbd5e1">pendientes</b> hasta que la actives.</div></div>
   <select id="rpf-autodelay" class="ac-sel" onchange="if(document.getElementById('rpf-autotoggle').checked)rpFAuto(document.getElementById('rpf-autotoggle'))"><option value="5">a los 5 min</option><option value="10" selected>a los 10 min</option><option value="15">a los 15 min</option><option value="30">a los 30 min</option></select>
   <label class="switch"><input type="checkbox" id="rpf-autotoggle" onchange="rpFAuto(this)"><span class="slider"></span></label>
  </div>
  <div class="tcard">
   <div class="foot"><label style="display:inline-flex;align-items:center;gap:8px;color:#c7d2e0;font-size:12.5px;font-weight:700;cursor:pointer"><input type="checkbox" id="rpf-chkall" onchange="rpFToggleAll(this)" style="width:15px;height:15px;accent-color:#137fec;cursor:pointer">Seleccionar la p&aacute;gina</label><span id="rpf-cnt" style="color:#8b97a8">&mdash;</span></div>
   <div style="overflow-x:auto"><table><thead><tr><th style="width:36px"></th><th>Fecha</th><th>Pedido</th><th>Cliente</th><th>Medio</th><th>Comprobante</th><th style="text-align:right">Total</th><th>Estado</th></tr></thead><tbody id="rpf-body"></tbody></table></div>
   <div class="pager" id="rpf-pager"></div>
  </div>
 </div>
</div>
<div id="rp-mov-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;background:#090b14;display:none;overflow:auto;transition:left .18s ease;font-family:system-ui,-apple-system,sans-serif;color:#e8edf4">
 <style>
  #rp-mov-ov .mv-wrap{max-width:1360px;margin:0 auto;padding:24px 34px 60px}
  #rp-mov-ov .mv-top{display:flex;align-items:flex-start;justify-content:space-between;gap:16px;margin-bottom:18px}
  #rp-mov-ov .mv-head{display:flex;align-items:flex-start;gap:13px}
  #rp-mov-ov .mv-hico{width:44px;height:44px;border-radius:12px;background:linear-gradient(160deg,#12233b,#0c1626);border:1px solid #1d3350;display:flex;align-items:center;justify-content:center;flex:none}
  #rp-mov-ov .mv-hico svg{width:22px;height:22px;color:#5aa2f5}
  #rp-mov-ov h1{margin:0;font-size:24px;font-weight:800;letter-spacing:-.3px;color:#f4f7fb}
  #rp-mov-ov .mv-head p{margin:5px 0 0;color:#8b97a8;font-size:13px;max-width:660px;line-height:1.45}
  #rp-mov-ov .mv-chips{display:flex;gap:10px;flex:none}
  #rp-mov-ov .mv-chip{display:inline-flex;align-items:center;gap:7px;background:#0e1626;border:1px solid #1c2637;color:#c7d2e0;border-radius:10px;padding:8px 12px;font-size:12.5px;font-weight:600}
  #rp-mov-ov .mv-x{flex:none;background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;width:38px;height:38px;border-radius:10px;font-size:16px;cursor:pointer}
  #rp-mov-ov .mv-kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:6px 0 22px}
  #rp-mov-ov .kcard{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;padding:17px 19px;position:relative;overflow:hidden}
  #rp-mov-ov .kcard::before{content:"";position:absolute;top:-40px;right:-30px;width:120px;height:120px;border-radius:50%;opacity:.13;filter:blur(24px)}
  #rp-mov-ov .kcard.i::before{background:#34d399}#rp-mov-ov .kcard.e::before{background:#f87171}#rp-mov-ov .kcard.c::before{background:#5aa2f5}#rp-mov-ov .kcard.a::before{background:#a78bfa}
  #rp-mov-ov .kt{display:flex;align-items:center;justify-content:space-between;position:relative}
  #rp-mov-ov .kico{width:31px;height:31px;border-radius:9px;display:flex;align-items:center;justify-content:center}
  #rp-mov-ov .kico svg{width:17px;height:17px}
  #rp-mov-ov .kcard.i .kico{background:rgba(52,211,153,.13);color:#34d399}#rp-mov-ov .kcard.e .kico{background:rgba(248,113,113,.13);color:#f87171}#rp-mov-ov .kcard.c .kico{background:rgba(90,162,245,.13);color:#5aa2f5}#rp-mov-ov .kcard.a .kico{background:rgba(167,139,250,.13);color:#a78bfa}
  #rp-mov-ov .klabel{font-size:11px;font-weight:800;letter-spacing:.7px;text-transform:uppercase;color:#8b97a8}
  #rp-mov-ov .knum{font-size:26px;font-weight:800;letter-spacing:-.5px;line-height:1;margin-top:14px;font-variant-numeric:tabular-nums;position:relative}
  #rp-mov-ov .kcard.i .knum{color:#34d399}#rp-mov-ov .kcard.e .knum{color:#f87171}#rp-mov-ov .kcard.c .knum{color:#f4f7fb}#rp-mov-ov .kcard.a .knum{color:#a78bfa}
  #rp-mov-ov .ksub{font-size:12px;color:#5b6678;margin-top:6px;position:relative}
  #rp-mov-ov .mv-sec{font-size:12px;font-weight:800;letter-spacing:1px;text-transform:uppercase;color:#7d8ba0;margin:8px 2px 12px;display:flex;align-items:center;gap:9px}
  #rp-mov-ov .mv-sec .bar{height:1px;background:#1a2436;flex:1}
  #rp-mov-ov .socios{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px;margin-bottom:14px}
  #rp-mov-ov .socio{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;padding:18px 20px}
  #rp-mov-ov .socio .top{display:flex;align-items:center;gap:11px;margin-bottom:14px}
  #rp-mov-ov .socio .av{width:38px;height:38px;border-radius:11px;display:flex;align-items:center;justify-content:center;font-size:15px;font-weight:800;color:#0a1322}
  #rp-mov-ov .socio .nm{font-size:15px;font-weight:700;color:#f4f7fb}#rp-mov-ov .socio .rol{font-size:11.5px;color:#5b6678}
  #rp-mov-ov .socio .grid{display:grid;grid-template-columns:1fr 1fr;gap:10px}
  #rp-mov-ov .socio .box{background:#0a1322;border:1px solid #17233a;border-radius:11px;padding:10px 12px}
  #rp-mov-ov .socio .box .l{font-size:10.5px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#5b6678}
  #rp-mov-ov .socio .box .v{font-size:17px;font-weight:800;margin-top:3px;font-variant-numeric:tabular-nums}
  #rp-mov-ov .socio .saldo{margin-top:12px;background:rgba(52,211,153,.09);border:1px solid #1f5a3d;border-radius:11px;padding:11px 13px;display:flex;align-items:center;justify-content:space-between}
  #rp-mov-ov .socio .saldo .l{font-size:12px;color:#9be8c6;font-weight:600}#rp-mov-ov .socio .saldo .v{font-size:19px;font-weight:800;color:#34d399;font-variant-numeric:tabular-nums}
  #rp-mov-ov .resumen{background:#0b111e;border:1px solid #18212f;border-radius:16px;padding:15px 20px;display:flex;align-items:center;gap:22px;flex-wrap:wrap;margin-bottom:22px}
  #rp-mov-ov .resumen .r{display:flex;flex-direction:column;gap:2px}
  #rp-mov-ov .resumen .r .l{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#5b6678}
  #rp-mov-ov .resumen .r .v{font-size:19px;font-weight:800;font-variant-numeric:tabular-nums}
  #rp-mov-ov .resumen .sepv{width:1px;align-self:stretch;background:#1a2436}
  #rp-mov-ov .resumen .note{margin-left:auto;color:#8b97a8;font-size:12.5px;max-width:320px;line-height:1.4}
  #rp-mov-ov .row2{display:grid;grid-template-columns:1.7fr 1fr;gap:14px;margin-bottom:22px}
  #rp-mov-ov .panel{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;padding:18px 20px}
  #rp-mov-ov .panel .ph{display:flex;align-items:center;justify-content:space-between;margin-bottom:6px}
  #rp-mov-ov .panel .ph b{font-size:14.5px;font-weight:700;color:#f4f7fb}#rp-mov-ov .panel .ph .sub{font-size:12px;color:#5b6678}
  #rp-mov-ov .panel .big{font-size:24px;font-weight:800;color:#f4f7fb;font-variant-numeric:tabular-nums}
  #rp-mov-ov .panel .up{font-size:12.5px;font-weight:700}
  #rp-mov-ov .xlab{display:flex;justify-content:space-between;color:#5b6678;font-size:10.5px;margin-top:6px}
  #rp-mov-ov .gastos .g{display:flex;align-items:center;gap:11px;margin-bottom:14px}#rp-mov-ov .gastos .g:last-child{margin-bottom:0}
  #rp-mov-ov .gastos .gi{width:30px;height:30px;border-radius:8px;display:flex;align-items:center;justify-content:center;flex:none;font-size:13px}
  #rp-mov-ov .gastos .gm{flex:1;min-width:0}
  #rp-mov-ov .gastos .gt{display:flex;justify-content:space-between;font-size:12.5px}
  #rp-mov-ov .gastos .gt .n{color:#c7d2e0;font-weight:600}#rp-mov-ov .gastos .gt .v{color:#f4f7fb;font-weight:800;font-variant-numeric:tabular-nums}
  #rp-mov-ov .gastos .gb{height:6px;border-radius:20px;background:#182234;margin-top:6px;overflow:hidden}#rp-mov-ov .gastos .gb i{display:block;height:100%;border-radius:20px}
  #rp-mov-ov .bar-f{display:flex;align-items:center;gap:9px;flex-wrap:wrap;margin-bottom:16px}
  #rp-mov-ov .mpill{background:#0d1524;border:1px solid #1a2436;color:#9aa6b6;border-radius:20px;padding:8px 15px;font-size:12.5px;font-weight:600;cursor:pointer;display:inline-flex;align-items:center;gap:7px}
  #rp-mov-ov .mpill.on{background:rgba(19,127,236,.14);border-color:#1e4f8a;color:#bcd7f7}
  #rp-mov-ov .mbtn{display:inline-flex;align-items:center;gap:8px;border-radius:11px;padding:10px 16px;font-size:13px;font-weight:700;cursor:pointer;border:1px solid #137fec;background:#137fec;color:#fff}
  #rp-mov-ov .mbtn svg{width:16px;height:16px}
  #rp-mov-ov .tcard{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;overflow:hidden}
  #rp-mov-ov .scroll{overflow-x:auto}
  #rp-mov-ov table{width:100%;border-collapse:collapse;font-size:13px;min-width:840px}
  #rp-mov-ov thead th{text-align:left;color:#5b6678;font-size:10.5px;font-weight:800;letter-spacing:.7px;text-transform:uppercase;padding:14px 16px;border-bottom:1px solid #1b2536}
  #rp-mov-ov tbody td{padding:13px 16px;border-bottom:1px solid #141d2e;color:#e8edf4}
  #rp-mov-ov tbody tr:last-child td{border-bottom:none}#rp-mov-ov tbody tr:hover{background:rgba(255,255,255,.02)}
  #rp-mov-ov .num{font-variant-numeric:tabular-nums}
  #rp-mov-ov .tag{display:inline-flex;align-items:center;gap:6px;font-size:11px;font-weight:800;padding:4px 10px;border-radius:20px}
  #rp-mov-ov .t-in{background:rgba(52,211,153,.12);color:#34d399;border:1px solid #1f5a3d}#rp-mov-ov .t-eg{background:rgba(248,113,113,.1);color:#f87171;border:1px solid #5a2a2a}
  #rp-mov-ov .cat{color:#b8c6da;font-size:12.5px}
  #rp-mov-ov .fue{display:inline-flex;align-items:center;gap:7px}
  #rp-mov-ov .fdot{width:22px;height:22px;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:800;color:#0a1322}
  #rp-mov-ov .afuera{font-size:10px;font-weight:700;margin-left:5px}
  #rp-mov-ov .del{opacity:0;color:#fb7185;cursor:pointer;font-weight:800;padding:2px 7px;border-radius:6px}
  #rp-mov-ov tr:hover .del{opacity:1}#rp-mov-ov .del:hover{background:rgba(251,113,133,.12)}
  #rp-mov-ov .foot{padding:12px 16px;color:#8b97a8;font-size:12.5px;border-top:1px solid #1b2536}
  #rp-mov-ov .mv-modal{position:fixed;inset:0;z-index:120;background:rgba(4,8,14,.74);display:none;align-items:center;justify-content:center;padding:20px}
  #rp-mov-ov .mv-modal .box{width:100%;max-width:520px;background:#0c1420;border:1px solid #1b2536;border-radius:18px;padding:24px;box-shadow:0 24px 60px rgba(0,0,0,.6)}
  #rp-mov-ov .lk-lb{font-size:11px;font-weight:700;color:#8b97a8;text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px;display:block}
  #rp-mov-ov .lk-in{background:#0a1322;border:1px solid #22324a;color:#e8edf4;border-radius:10px;padding:11px 13px;font-size:13.5px;font-family:inherit;outline:none;width:100%;color-scheme:dark}
  #rp-mov-ov .lk-in::placeholder{color:#5b6678}
  #rp-mov-ov .seg{display:flex;gap:7px;flex-wrap:wrap}
  #rp-mov-ov .seg .s{flex:1;min-width:calc(50% - 4px);text-align:center;background:#0d1524;border:1px solid #1a2436;color:#aeb8c6;border-radius:10px;padding:10px;font-size:12.5px;font-weight:700;cursor:pointer}
  #rp-mov-ov .seg .s.on{background:rgba(19,127,236,.14);border-color:#2b6fd0;color:#bcd7f7}
  @media(max-width:900px){#rp-mov-ov .mv-kpis{grid-template-columns:1fr 1fr}#rp-mov-ov .row2{grid-template-columns:1fr}#rp-mov-ov .mv-wrap{padding:16px 14px}}
 </style>
 <div class="mv-wrap">
  <div class="mv-top">
   <div class="mv-head">
    <div class="mv-hico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><path d="M3 7h18M3 12h18M3 17h18"/></svg></div>
    <div><h1>Movimientos</h1><p>Ingresos, egresos y <b style="color:#cbd5e1">aportes de socios</b>. Separ&aacute; la <b style="color:#cbd5e1">reinversi&oacute;n de la marca</b> de la <b style="color:#cbd5e1">plata que pusieron de afuera</b> &mdash; para dividir y devolver todo bien cuando se recupere.</p></div>
   </div>
   <div class="mv-chips">
    <button class="mv-chip" style="cursor:pointer;border-color:#1f5a3d;color:#8be6bd" onclick="rpMovBase()" title="Fija tu saldo actual de Mercado Pago (para que el total cuadre)">💰 Fijar saldo MP</button>
    <span class="mv-chip">🇦🇷 ARS</span>
    <button class="mv-x" onclick="rpMov(false)" title="Cerrar">&#10005;</button>
   </div>
  </div>
  <div class="mv-kpis">
   <div class="kcard i"><div class="kt"><span class="kico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 19V5m0 0-6 6m6-6 6 6"/></svg></span><span class="klabel">Ingresos</span></div><div class="knum" id="mv-k-ing">$0</div><div class="ksub">cobros + aportes del mes</div></div>
   <div class="kcard e"><div class="kt"><span class="kico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 5v14m0 0 6-6m-6 6-6-6"/></svg></span><span class="klabel">Egresos</span></div><div class="knum" id="mv-k-egr">$0</div><div class="ksub">ads &middot; stock &middot; env&iacute;os &middot; dise&ntilde;o</div></div>
   <div class="kcard c"><div class="kt"><span class="kico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="6" width="18" height="13" rx="2"/><path d="M3 10h18M7 15h4"/></svg></span><span class="klabel">Caja</span></div><div class="knum" id="mv-k-caja">$0</div><div class="ksub">saldo actual</div></div>
   <div class="kcard a"><div class="kt"><span class="kico"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/></svg></span><span class="klabel">A devolver a socios</span></div><div class="knum" id="mv-k-dev">$0</div><div class="ksub">aportes de afuera sin recuperar</div></div>
  </div>
  <div class="mv-sec"><span>Socios</span><span class="bar"></span></div>
  <div class="socios" id="mv-socios"></div>
  <div class="resumen">
   <div class="r"><span class="l">Total aportado de afuera</span><span class="v" id="mv-r-ap" style="color:#e8edf4">$0</span></div>
   <div class="sepv"></div>
   <div class="r"><span class="l">Ya recuperado</span><span class="v" id="mv-r-rec" style="color:#34d399">$0</span></div>
   <div class="sepv"></div>
   <div class="r"><span class="l">Falta devolver</span><span class="v" id="mv-r-fal" style="color:#fbbf24">$0</span></div>
   <div class="note">Cuando entre ganancia, se usa primero para <b style="color:#cbd5e1">devolver los aportes de afuera</b>. Reci&eacute;n ah&iacute; se reparte 50/50.</div>
  </div>
  <div class="row2">
   <div class="panel">
    <div class="ph"><b>Flujo de caja</b><span class="sub">Este mes</span></div>
    <div class="big" id="mv-fc-big">$0</div><div class="up" id="mv-fc-up"></div>
    <svg viewBox="0 0 420 120" preserveAspectRatio="none" style="width:100%;height:110px;margin-top:8px;display:block">
     <defs><linearGradient id="mvCajaG" x1="0" y1="0" x2="0" y2="1"><stop offset="0" stop-color="#4a9bf0" stop-opacity=".34"/><stop offset="1" stop-color="#4a9bf0" stop-opacity="0"/></linearGradient></defs>
     <path id="mv-fc-area" fill="url(#mvCajaG)"></path>
     <polyline id="mv-fc-line" fill="none" stroke="#4a9bf0" stroke-width="2.6" vector-effect="non-scaling-stroke"></polyline>
     <circle id="mv-fc-dot" r="4.5" fill="#4a9bf0"></circle>
    </svg>
    <div class="xlab"><span>1</span><span>8</span><span>15</span><span>22</span><span>31</span></div>
   </div>
   <div class="panel gastos"><div class="ph"><b>Gastos por categor&iacute;a</b><span class="sub" id="mv-g-tot">$0</span></div><div id="mv-gastos" style="margin-top:12px"></div></div>
  </div>
  <div class="bar-f" id="mv-pills">
   <span class="mpill on" data-f="todos" onclick="rpMovF('todos')">Todos</span>
   <span class="mpill" data-f="ingreso" onclick="rpMovF('ingreso')"><span style="color:#34d399">&#9679;</span> Ingresos</span>
   <span class="mpill" data-f="egreso" onclick="rpMovF('egreso')"><span style="color:#f87171">&#9679;</span> Egresos</span>
   <span class="mpill" data-f="afuera" onclick="rpMovF('afuera')"><span style="color:#fbbf24">&#9679;</span> Aportes de afuera</span>
   <span class="mbtn" style="margin-left:auto" onclick="rpMovOpen()"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 5v14M5 12h14"/></svg>Agregar movimiento</span>
  </div>
  <div class="tcard"><div class="scroll">
   <table><thead><tr><th>Fecha</th><th>Tipo</th><th>Categor&iacute;a</th><th>Descripci&oacute;n</th><th>Fuente / Socio</th><th style="text-align:right">Monto</th><th></th></tr></thead><tbody id="mv-body"></tbody></table>
  </div><div class="foot" id="mv-cnt">&mdash;</div></div>
 </div>
 <div class="mv-modal" id="mv-modal" onclick="if(event.target===this)rpMovClose()">
  <div class="box">
   <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px"><div style="width:40px;height:40px;border-radius:11px;background:rgba(19,127,236,.13);color:#4aa8ff;display:flex;align-items:center;justify-content:center;flex:none"><svg width="21" height="21" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 5v14M5 12h14"/></svg></div><div style="flex:1"><div style="font-size:17px;font-weight:800;color:#f4f7fb">Agregar movimiento</div><div style="color:#8b97a8;font-size:12.5px">Se guarda y recalcula caja, socios y gastos.</div></div><button onclick="rpMovClose()" style="background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;width:32px;height:32px;border-radius:9px;cursor:pointer">&#10005;</button></div>
   <span class="lk-lb">Tipo de movimiento</span>
   <div class="seg" id="mv-mtipo" style="margin-bottom:14px">
    <div class="s on" data-t="ingreso" onclick="rpMovT(this)">&uarr; Ingreso (venta)</div>
    <div class="s" data-t="egreso" onclick="rpMovT(this)">&darr; Gasto de la marca</div>
    <div class="s" data-t="aporte" onclick="rpMovT(this)">+ Aporte de socio</div>
    <div class="s" data-t="devolucion" onclick="rpMovT(this)">&#8629; Devoluci&oacute;n a socio</div>
   </div>
   <div id="mv-msocio-wrap" style="display:none;margin-bottom:12px"><span class="lk-lb">Socio</span>
    <div class="seg" id="mv-msocio"><div class="s on" data-s="cristian" onclick="rpMovS(this)">Cristian</div><div class="s" data-s="socio" onclick="rpMovS(this)">Socio</div></div></div>
   <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px">
    <div><span class="lk-lb">Monto ($)</span><input id="mv-m-monto" class="lk-in" inputmode="numeric" placeholder="50000"></div>
    <div><span class="lk-lb">Fecha</span><input id="mv-m-fecha" class="lk-in" type="date"></div>
   </div>
   <div style="margin-bottom:12px"><span class="lk-lb">Categor&iacute;a</span><input id="mv-m-cat" class="lk-in" placeholder="Ads, Stock, Env&iacute;os, Dise&ntilde;o…"></div>
   <div><span class="lk-lb">Descripci&oacute;n</span><input id="mv-m-desc" class="lk-in" placeholder="Detalle del movimiento"></div>
   <div id="mv-m-err" style="display:none;color:#fb7185;font-size:12.5px;margin-top:10px"></div>
   <button class="mbtn" style="width:100%;justify-content:center;margin-top:16px;padding:12px" onclick="rpMovSave(this)">Guardar movimiento</button>
  </div>
 </div>
</div>
<script>
(function(){
  var SOC={cristian:{nm:'Cristian',rol:'Socio · 50%',c:'#93c5fd'},socio:{nm:'Socio',rol:'Socio · 50%',c:'#c4b5fd'},marca:{nm:'Marca',c:'#6ee7b7'}};
  var CATC={'Ads':'#5aa2f5','Ads Meta':'#5aa2f5','Stock':'#fbbf24','Mercadería':'#fbbf24','Envíos':'#38bdf8','Diseño':'#a78bfa','Ventas':'#34d399','Aporte':'#93c5fd','Devolución':'#f472b6'};
  var MOV=[], filtro='todos', mtipo='ingreso', msocio='cristian', _loaded=false;
  function fmt(n){return '$'+Math.round(n||0).toLocaleString('es-AR');}
  function esIn(m){return m.clase==='ingreso'||m.clase==='aporte';}
  function esEg(m){return m.clase==='egreso'||m.clase==='devolucion';}
  function esAfuera(m){return m.clase==='aporte'||m.clase==='devolucion';}
  function fueTag(m){ if(m.socio==='marca') return m.clase==='egreso'?'<span class="afuera" style="color:#8b97a8">reinversión</span>':''; return m.clase==='aporte'?'<span class="afuera" style="color:#fbbf24">de afuera</span>':'<span class="afuera" style="color:#34d399">recupero</span>'; }
  function esc(s){return (''+s).replace(/[&<>]/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;'}[c];});}
  function $(id){return document.getElementById(id);}
  window.rpMovBase=function(){ var v=prompt('Pegá tu saldo ACTUAL de Mercado Pago (lo que tenés disponible ahora):'); if(v===null||v==='')return;
    fetch('/pf-movimientos-base',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({saldo:v})}).then(function(r){return r.json();}).then(function(j){
      if(j&&j.ok){ if(typeof rpMovLoad==='function')rpMovLoad(); else location.reload(); } else { alert('No se pudo fijar el saldo.'); } }).catch(function(){ alert('Error de conexión.'); }); };
  window.rpMovF=function(f){ filtro=f; var ps=document.querySelectorAll('#mv-pills .mpill'); for(var i=0;i<ps.length;i++){ ps[i].classList.toggle('on',ps[i].getAttribute('data-f')===f); } rpMovRender(); };
  function filtered(){ return MOV.filter(function(m){ if(filtro==='todos')return true; if(filtro==='afuera')return esAfuera(m); if(filtro==='ingreso')return esIn(m); if(filtro==='egreso')return esEg(m); return true; }); }
  window.rpMov=function(open){ var o=$('rp-mov-ov'); if(!o)return;
    if(open){ ['rp-prod-ov','rp-comis-ov','rp-integ-ov','rp-desp-ov','rp-fact-ov','rp-ads-ov','rp-stock-ov'].forEach(function(id){ var x=$(id); if(x)x.style.display='none'; });
      var _lk=$('rpf-lock'); if(_lk)_lk.style.display='none';
      try{rpProdSetActive(false);}catch(e){} try{rpComisSetActive(false);}catch(e){} var ib=$('rp-integ-btn'); if(ib)ib.classList.remove('rp-active');
      _rpNavActive('rp-mov-nav');
    } else { _rpNavActive(null); }
    o.style.display=open?'block':'none';
    if(open && !_loaded) rpMovLoad(); };
  window.rpMovLoad=function(){ fetch('/pf-movimientos').then(function(r){return r.json();}).then(function(j){
      MOV=(j&&j.ok&&j.rows)?j.rows:[]; _loaded=true; rpMovRender();
    }).catch(function(){ MOV=[]; _loaded=true; rpMovRender(); }); };
  window.rpMovRender=function(){
    var ing=0,egr=0; MOV.forEach(function(m){ if(esIn(m))ing+=m.monto; else egr+=m.monto; });
    var caja=ing-egr;
    var ap={cristian:0,socio:0}, rec={cristian:0,socio:0};
    MOV.forEach(function(m){ if(m.clase==='aporte')ap[m.socio]=(ap[m.socio]||0)+m.monto; if(m.clase==='devolucion')rec[m.socio]=(rec[m.socio]||0)+m.monto; });
    var sal={cristian:ap.cristian-rec.cristian, socio:ap.socio-rec.socio};
    var totAp=ap.cristian+ap.socio, totRec=rec.cristian+rec.socio, dev=Math.max(0,sal.cristian)+Math.max(0,sal.socio);
    $('mv-k-ing').textContent=fmt(ing); $('mv-k-egr').textContent=fmt(egr); $('mv-k-caja').textContent=fmt(caja); $('mv-k-dev').textContent=fmt(dev);
    $('mv-r-ap').textContent=fmt(totAp); $('mv-r-rec').textContent=fmt(totRec); $('mv-r-fal').textContent=fmt(dev);
    $('mv-fc-big').textContent=fmt(caja);
    var upel=$('mv-fc-up'); upel.textContent=(caja>=0?'▲ ':'▼ ')+fmt(Math.abs(caja))+' de saldo'; upel.style.color=caja>=0?'#34d399':'#f87171';
    var reinv=0; MOV.forEach(function(m){ if(m.socio==='marca'&&m.clase==='egreso')reinv+=m.monto; });
    var soc=$('mv-socios'); var h='';
    ['cristian','socio'].forEach(function(k){ var s=SOC[k];
      h+='<div class="socio"><div class="top"><span class="av" style="background:'+s.c+'">'+s.nm[0]+'</span><div><div class="nm">'+s.nm+'</div><div class="rol">'+s.rol+'</div></div></div>'
        +'<div class="grid"><div class="box"><div class="l">Aportó de afuera</div><div class="v" style="color:#e8edf4">'+fmt(ap[k])+'</div></div><div class="box"><div class="l">Recuperó</div><div class="v" style="color:#8b97a8">'+fmt(rec[k])+'</div></div></div>'
        +'<div class="saldo"><span class="l">Saldo a favor (a devolver)</span><span class="v">'+fmt(sal[k])+'</span></div></div>';
    });
    h+='<div class="socio"><div class="top"><span class="av" style="background:'+SOC.marca.c+'">M</span><div><div class="nm">Marca</div><div class="rol">Caja / reinversión con ganancia</div></div></div>'
      +'<div class="grid"><div class="box"><div class="l">Reinvertido</div><div class="v" style="color:#e8edf4">'+fmt(reinv)+'</div></div><div class="box"><div class="l">Caja disponible</div><div class="v" style="color:#34d399">'+fmt(caja)+'</div></div></div>'
      +'<div class="saldo" style="background:rgba(90,162,245,.08);border-color:#1e4f8a"><span class="l" style="color:#9dc3f5">Es de la marca — no se devuelve</span><span class="v" style="color:#5aa2f5">—</span></div></div>';
    soc.innerHTML=h;
    var gc={}; MOV.forEach(function(m){ if(esEg(m)){ gc[m.cat]=(gc[m.cat]||0)+m.monto; } });
    var arr=Object.keys(gc).map(function(k){return {c:k,v:gc[k]};}).sort(function(a,b){return b.v-a.v;}); var mx=arr.length?arr[0].v:1;
    $('mv-g-tot').textContent=fmt(egr);
    $('mv-gastos').innerHTML=arr.slice(0,6).map(function(g){ var col=CATC[g.c]||'#8b97a8';
      return '<div class="g"><span class="gi" style="background:'+col+'22;color:'+col+'">'+(g.c[0]||'·')+'</span><div class="gm"><div class="gt"><span class="n">'+esc(g.c)+'</span><span class="v">'+fmt(g.v)+'</span></div><div class="gb"><i style="width:'+Math.max(6,Math.round(g.v/mx*100))+'%;background:'+col+'"></i></div></div></div>';
    }).join('')||'<div style="color:#5b6678;font-size:12.5px">Sin gastos aún.</div>';
    var ord=MOV.slice().sort(function(a,b){return (a.d||'').localeCompare(b.d||'');}); var run=0, pts=[];
    ord.forEach(function(m){ run+= esIn(m)?m.monto:-m.monto; pts.push(run); });
    if(pts.length<2)pts=[0,caja];
    var mn=Math.min.apply(null,pts.concat([0])), mx2=Math.max.apply(null,pts), rng=(mx2-mn)||1;
    var W=420,H=110, step=W/(pts.length-1);
    var xy=pts.map(function(p,i){ return [Math.round(i*step), Math.round(H-8-(p-mn)/rng*(H-20))]; });
    $('mv-fc-line').setAttribute('points',xy.map(function(a){return a[0]+','+a[1];}).join(' '));
    $('mv-fc-area').setAttribute('d','M'+xy.map(function(a){return a[0]+','+a[1];}).join(' L')+' L'+W+','+H+' L0,'+H+' Z');
    var last=xy[xy.length-1]; var dot=$('mv-fc-dot'); dot.setAttribute('cx',last[0]); dot.setAttribute('cy',last[1]);
    var data=filtered().slice().reverse();
    $('mv-body').innerHTML=data.map(function(m){
      var esI=esIn(m); var av=SOC[m.socio]||SOC.marca;
      return '<tr><td class="num">'+esc(m.d)+'</td><td><span class="tag '+(esI?'t-in':'t-eg')+'">'+(esI?'↑ Ingreso':'↓ Egreso')+'</span></td>'
        +'<td class="cat">'+esc(m.cat)+'</td><td>'+esc(m.desc)+'</td>'
        +'<td><span class="fue"><span class="fdot" style="background:'+av.c+'">'+av.nm[0]+'</span>'+av.nm+' '+fueTag(m)+'</span></td>'
        +'<td class="num" style="text-align:right;font-weight:700;color:'+(esI?'#34d399':'#f87171')+'">'+(esI?'+':'−')+fmt(m.monto)+'</td>'
        +'<td style="text-align:right"><span class="del" onclick="rpMovDel('+m.id+')">✕</span></td></tr>';
    }).join('')||'<tr><td colspan="7" style="padding:36px;text-align:center;color:#5b6678">No hay movimientos en este filtro.</td></tr>';
    $('mv-cnt').textContent=data.length+' movimiento'+(data.length!==1?'s':'')+' · caja '+fmt(caja);
  };
  window.rpMovDel=function(id){ if(!confirm('¿Borrar este movimiento?'))return;
    fetch('/pf-movimientos-del',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:id})})
     .then(function(r){return r.json();}).then(function(j){ if(j&&j.ok){ MOV=MOV.filter(function(m){return m.id!==id;}); rpMovRender(); } }); };
  window.rpMovT=function(el){ mtipo=el.getAttribute('data-t'); var ss=document.querySelectorAll('#mv-mtipo .s'); for(var i=0;i<ss.length;i++)ss[i].classList.toggle('on',ss[i]===el); $('mv-msocio-wrap').style.display=(mtipo==='aporte'||mtipo==='devolucion')?'block':'none'; };
  window.rpMovS=function(el){ msocio=el.getAttribute('data-s'); var ss=document.querySelectorAll('#mv-msocio .s'); for(var i=0;i<ss.length;i++)ss[i].classList.toggle('on',ss[i]===el); };
  window.rpMovOpen=function(){ var m=$('mv-modal'); m.style.display='flex'; var f=$('mv-m-fecha'); if(!f.value){var t=new Date();f.value=t.getFullYear()+'-'+String(t.getMonth()+1).padStart(2,'0')+'-'+String(t.getDate()).padStart(2,'0');} var e=$('mv-m-err'); if(e)e.style.display='none'; };
  window.rpMovClose=function(){ $('mv-modal').style.display='none'; };
  window.rpMovSave=function(btn){
    var monto=parseFloat((($('mv-m-monto').value||'').replace(/[^\d.]/g,'')))||0;
    var err=$('mv-m-err');
    if(monto<=0){ $('mv-m-monto').style.borderColor='#fb7185'; if(err){err.textContent='Poné un monto válido.';err.style.display='block';} return; }
    $('mv-m-monto').style.borderColor='';
    var f=$('mv-m-fecha').value||''; var dd=f?(f.slice(8,10)+'/'+f.slice(5,7)):'—';
    var cat=($('mv-m-cat').value||'').trim() || (mtipo==='ingreso'?'Ventas':mtipo==='aporte'?'Aporte':mtipo==='devolucion'?'Devolución':'Gasto');
    var soc=(mtipo==='aporte'||mtipo==='devolucion')?msocio:'marca';
    var mv={d:dd, clase:mtipo, cat:cat, desc:($('mv-m-desc').value||'').trim()||cat, socio:soc, monto:monto};
    if(btn){btn.disabled=true;btn.style.opacity='.6';}
    fetch('/pf-movimientos-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(mv)})
     .then(function(r){return r.json();}).then(function(j){
       if(btn){btn.disabled=false;btn.style.opacity='';}
       if(j&&j.ok){ mv.id=j.id; MOV.push(mv); $('mv-m-monto').value=''; $('mv-m-cat').value=''; $('mv-m-desc').value=''; rpMovClose(); rpMovRender(); }
       else if(err){ err.textContent='No se pudo guardar.'; err.style.display='block'; }
     }).catch(function(){ if(btn){btn.disabled=false;btn.style.opacity='';} if(err){err.textContent='Error de conexión.';err.style.display='block';} });
  };
})();
</script>
<script>
(function(){
  var L={envialo:`<svg width="22" height="22" viewBox="0 0 24 24" fill="#ff6b35"><path d="M3 4a1 1 0 0 0-1 1v11a1 1 0 0 0 1 1h1a3 3 0 0 0 6 0h4a3 3 0 0 0 6 0h1a1 1 0 0 0 1-1v-4a1 1 0 0 0-.29-.71l-3-3A1 1 0 0 0 18 8h-2V5a1 1 0 0 0-1-1H3zm13 6h1.59L20 12.41V13h-4v-3zM7 16.5A1.5 1.5 0 1 1 5.5 15 1.5 1.5 0 0 1 7 16.5zm10 0A1.5 1.5 0 1 1 15.5 15a1.5 1.5 0 0 1 1.5 1.5z"/></svg>`,mp:`<svg width="21" height="21" viewBox="0 0 24 24" fill="#00b1ea"><path d="M11.115 16.479a.93.927 0 0 1-.939-.886c-.002-.042-.006-.155-.103-.155-.04 0-.074.023-.113.059-.112.103-.254.206-.46.206a.816.814 0 0 1-.305-.066c-.535-.214-.542-.578-.521-.725.006-.038.007-.08-.02-.11l-.032-.03h-.034c-.027 0-.055.012-.093.039a.788.786 0 0 1-.454.16.7.699 0 0 1-.253-.05c-.708-.27-.65-.928-.617-1.126.005-.041-.005-.072-.03-.092l-.05-.04-.047.043a.728.726 0 0 1-.505.203.73.728 0 0 1-.732-.725c0-.4.328-.722.732-.722.364 0 .675.27.721.63l.026.195.11-.165c.01-.018.307-.46.852-.46.102 0 .21.016.316.05.434.13.508.52.519.68.008.094.075.1.09.1.037 0 .064-.024.083-.045a.746.744 0 0 1 .54-.225c.128 0 .263.03.402.09.69.293.379 1.158.374 1.167-.058.144-.061.207-.005.244l.027.013h.02c.03 0 .07-.014.134-.035.093-.032.235-.08.367-.08a.944.942 0 0 1 .94.93.936.934 0 0 1-.94.928zm7.302-4.171c-1.138-.98-3.768-3.24-4.481-3.77-.406-.302-.685-.462-.928-.533a1.559 1.554 0 0 0-.456-.07c-.182 0-.376.032-.58.095-.46.145-.918.505-1.362.854l-.023.018c-.414.324-.84.66-1.164.73a1.986 1.98 0 0 1-.43.049c-.362 0-.687-.104-.81-.258-.02-.025-.007-.066.04-.125l.008-.008 1-1.067c.783-.774 1.525-1.506 3.23-1.545h.085c1.062 0 2.12.469 2.24.524a7.03 7.03 0 0 0 3.056.724c1.076 0 2.188-.263 3.354-.795a9.135 9.11 0 0 0-.405-.317c-1.025.44-2.003.66-2.946.66-.962 0-1.925-.229-2.858-.68-.05-.022-1.22-.567-2.44-.57-.032 0-.065 0-.096.002-1.434.033-2.24.536-2.782.976-.528.013-.982.138-1.388.25-.361.1-.673.186-.979.185-.125 0-.35-.01-.37-.012-.35-.01-2.115-.437-3.518-.962-.143.1-.28.203-.415.31 1.466.593 3.25 1.053 3.812 1.089.157.01.323.027.491.027.372 0 .744-.103 1.104-.203.213-.059.446-.123.692-.17l-.196.194-1.017 1.087c-.08.08-.254.294-.14.557a.705.703 0 0 0 .268.292c.243.162.677.27 1.08.271.152 0 .297-.015.43-.044.427-.095.874-.448 1.349-.82.377-.296.913-.672 1.323-.782a1.494 1.49 0 0 1 .37-.05.611.61 0 0 1 .095.005c.27.034.533.125 1.003.472.835.62 4.531 3.815 4.566 3.846.002.002.238.203.22.537-.007.186-.11.352-.294.466a.902.9 0 0 1-.484.15.804.802 0 0 1-.428-.124c-.014-.01-1.28-1.157-1.746-1.543-.074-.06-.146-.115-.22-.115a.122.122 0 0 0-.096.045c-.073.09.01.212.105.294l1.48 1.47c.002 0 .184.17.204.395.012.244-.106.447-.35.606a.957.955 0 0 1-.526.171.766.764 0 0 1-.42-.127l-.214-.206a21.035 20.978 0 0 0-1.08-1.009c-.072-.058-.148-.112-.221-.112a.127.127 0 0 0-.094.038c-.033.037-.056.103.028.212a.698.696 0 0 0 .075.083l1.078 1.198c.01.01.222.26.024.511l-.038.048a1.18 1.178 0 0 1-.1.096c-.184.15-.43.164-.527.164a.8.798 0 0 1-.147-.012c-.106-.018-.178-.048-.212-.089l-.013-.013c-.06-.06-.602-.609-1.054-.98-.059-.05-.133-.11-.21-.11a.128.128 0 0 0-.096.042c-.09.096.044.24.1.293l.92 1.003a.204.204 0 0 1-.033.062c-.033.044-.144.155-.479.196a.91.907 0 0 1-.122.007c-.345 0-.712-.164-.902-.264a1.343 1.34 0 0 0 .13-.576 1.368 1.365 0 0 0-1.42-1.357c.024-.342-.025-.99-.697-1.274a1.455 1.452 0 0 0-.575-.125c-.146 0-.287.025-.42.075a1.153 1.15 0 0 0-.671-.564 1.52 1.515 0 0 0-.494-.085c-.28 0-.537.08-.767.242a1.168 1.165 0 0 0-.903-.43 1.173 1.17 0 0 0-.82.335c-.287-.217-1.425-.93-4.467-1.613a17.39 17.344 0 0 1-.692-.189 4.822 4.82 0 0 0-.077.494l.67.157c3.108.682 4.136 1.391 4.309 1.525a1.145 1.142 0 0 0-.09.442 1.16 1.158 0 0 0 1.378 1.132c.096.467.406.821.879 1.003a1.165 1.162 0 0 0 .415.08c.09 0 .179-.012.266-.034.086.22.282.493.722.668a1.233 1.23 0 0 0 .457.094c.122 0 .241-.022.355-.063a1.373 1.37 0 0 0 1.269.841c.37.002.726-.147.985-.41.221.121.688.341 1.163.341.06 0 .118-.002.175-.01.47-.059.689-.24.789-.382a.571.57 0 0 0 .048-.078c.11.032.234.058.373.058.255 0 .501-.086.75-.265.244-.174.418-.424.444-.637v-.01c.083.017.167.026.251.026.265 0 .527-.082.773-.242.48-.31.562-.715.554-.98a1.28 1.279 0 0 0 .978-.194 1.04 1.04 0 0 0 .502-.808 1.088 1.085 0 0 0-.16-.653c.804-.342 2.636-1.003 4.795-1.483a4.734 4.721 0 0 0-.067-.492 27.742 27.667 0 0 0-5.049 1.62zm5.123-.763c0 4.027-5.166 7.293-11.537 7.293-6.372 0-11.538-3.266-11.538-7.293 0-4.028 5.165-7.293 11.539-7.293 6.371 0 11.537 3.265 11.537 7.293zm.46.004c0-4.272-5.374-7.755-12-7.755S.002 7.277.002 11.55L0 12.004c0 4.533 4.695 8.203 11.999 8.203 7.347 0 12-3.67 12-8.204z"/></svg>`,meli:`<svg width="21" height="21" viewBox="0 0 24 24" fill="#ffe600"><path d="M11.115 16.479a.93.927 0 0 1-.939-.886c-.002-.042-.006-.155-.103-.155-.04 0-.074.023-.113.059-.112.103-.254.206-.46.206a.816.814 0 0 1-.305-.066c-.535-.214-.542-.578-.521-.725.006-.038.007-.08-.02-.11l-.032-.03h-.034c-.027 0-.055.012-.093.039a.788.786 0 0 1-.454.16.7.699 0 0 1-.253-.05c-.708-.27-.65-.928-.617-1.126.005-.041-.005-.072-.03-.092l-.05-.04-.047.043a.728.726 0 0 1-.505.203.73.728 0 0 1-.732-.725c0-.4.328-.722.732-.722.364 0 .675.27.721.63l.026.195.11-.165c.01-.018.307-.46.852-.46.102 0 .21.016.316.05.434.13.508.52.519.68.008.094.075.1.09.1.037 0 .064-.024.083-.045a.746.744 0 0 1 .54-.225c.128 0 .263.03.402.09.69.293.379 1.158.374 1.167-.058.144-.061.207-.005.244l.027.013h.02c.03 0 .07-.014.134-.035.093-.032.235-.08.367-.08a.944.942 0 0 1 .94.93.936.934 0 0 1-.94.928zm7.302-4.171c-1.138-.98-3.768-3.24-4.481-3.77-.406-.302-.685-.462-.928-.533a1.559 1.554 0 0 0-.456-.07c-.182 0-.376.032-.58.095-.46.145-.918.505-1.362.854l-.023.018c-.414.324-.84.66-1.164.73a1.986 1.98 0 0 1-.43.049c-.362 0-.687-.104-.81-.258-.02-.025-.007-.066.04-.125l.008-.008 1-1.067c.783-.774 1.525-1.506 3.23-1.545h.085c1.062 0 2.12.469 2.24.524a7.03 7.03 0 0 0 3.056.724c1.076 0 2.188-.263 3.354-.795a9.135 9.11 0 0 0-.405-.317c-1.025.44-2.003.66-2.946.66-.962 0-1.925-.229-2.858-.68-.05-.022-1.22-.567-2.44-.57-.032 0-.065 0-.096.002-1.434.033-2.24.536-2.782.976-.528.013-.982.138-1.388.25-.361.1-.673.186-.979.185-.125 0-.35-.01-.37-.012-.35-.01-2.115-.437-3.518-.962-.143.1-.28.203-.415.31 1.466.593 3.25 1.053 3.812 1.089.157.01.323.027.491.027.372 0 .744-.103 1.104-.203.213-.059.446-.123.692-.17l-.196.194-1.017 1.087c-.08.08-.254.294-.14.557a.705.703 0 0 0 .268.292c.243.162.677.27 1.08.271.152 0 .297-.015.43-.044.427-.095.874-.448 1.349-.82.377-.296.913-.672 1.323-.782a1.494 1.49 0 0 1 .37-.05.611.61 0 0 1 .095.005c.27.034.533.125 1.003.472.835.62 4.531 3.815 4.566 3.846.002.002.238.203.22.537-.007.186-.11.352-.294.466a.902.9 0 0 1-.484.15.804.802 0 0 1-.428-.124c-.014-.01-1.28-1.157-1.746-1.543-.074-.06-.146-.115-.22-.115a.122.122 0 0 0-.096.045c-.073.09.01.212.105.294l1.48 1.47c.002 0 .184.17.204.395.012.244-.106.447-.35.606a.957.955 0 0 1-.526.171.766.764 0 0 1-.42-.127l-.214-.206a21.035 20.978 0 0 0-1.08-1.009c-.072-.058-.148-.112-.221-.112a.127.127 0 0 0-.094.038c-.033.037-.056.103.028.212a.698.696 0 0 0 .075.083l1.078 1.198c.01.01.222.26.024.511l-.038.048a1.18 1.178 0 0 1-.1.096c-.184.15-.43.164-.527.164a.8.798 0 0 1-.147-.012c-.106-.018-.178-.048-.212-.089l-.013-.013c-.06-.06-.602-.609-1.054-.98-.059-.05-.133-.11-.21-.11a.128.128 0 0 0-.096.042c-.09.096.044.24.1.293l.92 1.003a.204.204 0 0 1-.033.062c-.033.044-.144.155-.479.196a.91.907 0 0 1-.122.007c-.345 0-.712-.164-.902-.264a1.343 1.34 0 0 0 .13-.576 1.368 1.365 0 0 0-1.42-1.357c.024-.342-.025-.99-.697-1.274a1.455 1.452 0 0 0-.575-.125c-.146 0-.287.025-.42.075a1.153 1.15 0 0 0-.671-.564 1.52 1.515 0 0 0-.494-.085c-.28 0-.537.08-.767.242a1.168 1.165 0 0 0-.903-.43 1.173 1.17 0 0 0-.82.335c-.287-.217-1.425-.93-4.467-1.613a17.39 17.344 0 0 1-.692-.189 4.822 4.82 0 0 0-.077.494l.67.157c3.108.682 4.136 1.391 4.309 1.525a1.145 1.142 0 0 0-.09.442 1.16 1.158 0 0 0 1.378 1.132c.096.467.406.821.879 1.003a1.165 1.162 0 0 0 .415.08c.09 0 .179-.012.266-.034.086.22.282.493.722.668a1.233 1.23 0 0 0 .457.094c.122 0 .241-.022.355-.063a1.373 1.37 0 0 0 1.269.841c.37.002.726-.147.985-.41.221.121.688.341 1.163.341.06 0 .118-.002.175-.01.47-.059.689-.24.789-.382a.571.57 0 0 0 .048-.078c.11.032.234.058.373.058.255 0 .501-.086.75-.265.244-.174.418-.424.444-.637v-.01c.083.017.167.026.251.026.265 0 .527-.082.773-.242.48-.31.562-.715.554-.98a1.28 1.279 0 0 0 .978-.194 1.04 1.04 0 0 0 .502-.808 1.088 1.085 0 0 0-.16-.653c.804-.342 2.636-1.003 4.795-1.483a4.734 4.721 0 0 0-.067-.492 27.742 27.667 0 0 0-5.049 1.62zm5.123-.763c0 4.027-5.166 7.293-11.537 7.293-6.372 0-11.538-3.266-11.538-7.293 0-4.028 5.165-7.293 11.539-7.293 6.371 0 11.537 3.265 11.537 7.293zm.46.004c0-4.272-5.374-7.755-12-7.755S.002 7.277.002 11.55L0 12.004c0 4.533 4.695 8.203 11.999 8.203 7.347 0 12-3.67 12-8.204z"/></svg>`,tn:`<img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAADAAAAAwCAYAAABXAvmHAAAHq0lEQVR4nM1aeVRUVRj/3TczDCAyIAKBGKTWgIhWFi7IGJ52FXHJorB9OcfMUsuOp90MPVJpaXVaTsvJIssC2ss8pkMnAy01FMItFyBl34ZtZm7n3sc8ZnDmvQdME79/eI93373fd++3/L7vDYE3MG93FGxCGggmA3YjgFEACQMQ1D2iBaC1AI6DkjKA/AaNfSe+MFUNdGnS7zdv2h2OLuE2gC4CcHk/Z9kLYAt09GN8Zqr2jQKzCy+Exv4YQO4BEADvoA0g74DaclAw/fR/o8D9e3U417YYwBon0/A22gCsR4h1Hd5Pa/eeArMLE6GlW0GRCN+gBJTejALTYaWBguJUGeZMaGiRD4VnGAdCipBRuBADUiCjcAl3MiAQvscQgH6CDPOK/pnQHPNiELyGwQBKl6PAtEG9AsxsxJ1XNjHfwA6QTORP+1RZAeawzOb/H7ORQysoTe7t2K47fMO3eh5tBp/wDENAhE9w0yE/eFRAH7zSx9Gmj6BJ6Kpf4d6EeIalZV7MrhyTLxmK9OQwpCUZMDJMj4gQHepbrDjb0Il9x1rwZXEdfvijHpYOu9opLRCsRnyRdsZVgTnmzSB40FuCXzXOgOysOEwxDlUce66xC6u3nsJbP/6DLhtVnpziVRSkPtyjACdm5GRfdt8QqEVCTAAiQ/xQ32rF6ZoOnDjbDq2GYP3tcViWPsJlfEOrFYdOW1Dd2IVhQVpcGK5HXIS/yxh2IhlrS3GmtkNp+TboaCwjgFp+24ksEHXCpycPw8OzomEaa+DCOqO8so0LmnyxuOuUAvm/1WLj15X4pbQJNrvr7saPCMAdMyKxdGY0AvUCJo4OQlHOBFz7bAlKTlnkxAhAF24BsEmUIMO8T4kSDw/WIXe5EVdPCFGjJ2qbrch8qQzbDzQojo0e5sfnNiUa+D07yeSVB1DT1CXzFi1GvimZ8GLETirlFhgR5gfzC+NxUaR45Gwjt++vx08HG1BZ14mhARq+ewumDEdokHiozJYX5pTxE1ADPy3Bx8uNmD9lOL//Zl8dZq2R5XIUOms0wdzC20Apy7pu4a8TYM4ejyvGBEl2euer5W6POMhfg+ysWDw0M5rfd1opUlYdwN6jLaqUCNQLKMwej8tGiWtd/UwJdhyUOUFCMwVQTJKbdMnMKEl48+EmmJ446NE+W9ptWPrOcaz84G9pV19/YAwElVWHpcOOxW8e477D8NTCkfIvUDJZAGi8p+dMgFXzxUkaLVYsWF+qKl7n5J/BV8V1/PrKMUG4/vJQdRoA2FPejB/31/Pr1LEGRBh0MgrAyDLxaE/P05JCeMhj2PRNFY/XavF0LovKIhamhKMv2Par6Dfs5GSVJ3QMU8BjWEmJD5au8/aoc0YH9p9oxbF/xKpw0iXKycwZLOQ6MPoC11zRC6GCXH3Loo8DR6pYudo3HO1+Z6TTPGpQWdcpXUeHyr1Lhor24QF2p8RDiOfEVrBqLL++e/MRvLfjrPTs+tWH+iA23K6lQCwoOwGPMa6qvsfmL45yn6hrmqzSdZTsbqlH9LCeearqe07DDZqZAh4DbWFpo3Q9dzJrtJ2PQ6csPLExTEvo8ZmBIMXJ945WyXVXaAMLo0c9Pf65pJFTAoYlN0YhMuT8kMbC6x/HxUOckWSQotZAsGCqmI0Zd/r+dzGkugc5Ioi9SvdgmXTt56cl9rltZQLPlr2xZdc5/levE/D4vJgBCT81PhjXdPMtljirZfkQygXeaJXB5m+rUNxNBZiJsFQ/IW6Iy5h3d5yVFlo6K5rzov4gUC/g9QdGS068+tNT8i9Q8quDzFXItVhYON39wniMciJzjKP8dEAkc8y0Fl0VISnG+LzpiT85q1QLPy1B7op4zOv2NZbJ07MVyJxdF+Wg06xLPFGJTn+0zIhrL1VHp9mJ3PryX1xJJcSE6ZG7wigFAXV0GkXIT53kMOgPlRZhk133XAnfFbb7VjelX1lFm2Ru4cE6bH92HM8RM5JCzit+GBJiArFuURz+em2iJDwLm7OzDysJzwzmwwGXlPExAbigu6Q8ea4dJ6s7uKAv3XURr7KcwSo1piAr5kOHaBEb4Y/YcL3LGKb83HWHUVErG/sZLLBZY/FVWk3PtmTs3gQQ1gv1Clg0WZsVK1VZcqhp6sKLBRXY8GUFj3zKoBuRb1rWqyuxaySIUObtphYjcnOSwzA9MZgX8hEGP95WYaby+/EWFBTVYvv+BrR1qm6rtMKqMeLrqSzw9Io8GeYnATyPwQxCViFv2jrHrWtW6mjO4R8XBisIDkIb+rLzv1wV+O7GDhDMZyQJgw+tgP0WfJbo4uHn84K81HKA3Cu2tAcN7KD0TuRNL+39wH3/n/fhqdcikhewHAWmbe4eeP6AkW96AwRL/ueToAAeRX7qK54GKDc8+Ic2+q74zcqnaAGld3naeQfUdWzm7koA1Wzl/XlfRRtBczM+n+qR6jug7hsYc54If0b2HpErQb0AC4Dn0N6crEb4/v3UYN7OGNi0j4HgPi9+DLEAeBtWTY4jw/rmxx5WkglKswByZT8dtBigW2Cz5TJi1h8x+q+AM9L3RELTmcZ6lazdB8J+bgPWjnM4PjO7alCcAKFloMIeCGQn8lLEWnQA+BetL81v6u4ZVgAAAABJRU5ErkJggg==" width="26" height="26" style="border-radius:6px;display:block">`,shopify:`<svg width="22" height="22" width="256px" height="292px" viewBox="0 0 256 292" version="1.1" xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" preserveAspectRatio="xMidYMid"> <g> <path d="M223.773626,57.3402078 C223.572932,55.8793405 222.29409,55.0718963 221.236945,54.9832175 C220.182133,54.8945386 197.853734,53.2399781 197.853734,53.2399781 C197.853734,53.2399781 182.346604,37.8448639 180.64537,36.1412966 C178.941803,34.4377293 175.616346,34.9558004 174.325836,35.336186 C174.134476,35.3921937 170.937371,36.3793293 165.646977,38.0152206 C160.466266,23.1101737 151.325344,9.41162582 135.241802,9.41162582 C134.798408,9.41162582 134.341011,9.43029505 133.883615,9.45596525 C129.309654,3.40713457 123.643542,0.779440373 118.74987,0.779440373 C81.285392,0.779440373 63.3862673,47.6135387 57.7738299,71.414474 C43.2164974,75.9254268 32.8737437,79.1318671 31.5528956,79.5472575 C23.4271131,82.0956074 23.1704111,82.3523094 22.1039313,90.0090275 C21.2988208,95.8058236 0.0369009009,260.235071 0.0369009009,260.235071 L165.714653,291.277334 L255.485648,271.856667 C255.485648,271.856667 223.971987,58.8010751 223.773626,57.3402078 L223.773626,57.3402078 Z M156.48972,40.8482763 C152.328815,42.1364532 147.598499,43.5996542 142.471461,45.1865388 C142.476129,44.1994032 142.480796,43.2262696 142.480796,42.1644571 C142.480796,32.8998514 141.194953,25.4414939 139.132003,19.5280151 C147.418807,20.5688247 152.937899,29.9967861 156.48972,40.8482763 L156.48972,40.8482763 Z M128.852258,21.3646006 C131.155574,27.1380602 132.65378,35.4225312 132.65378,46.6030666 C132.65378,47.1748118 132.649112,47.6975503 132.644445,48.2272897 C123.52686,51.0510108 113.620499,54.1174319 103.690802,57.1931876 C109.265901,35.6768995 119.716003,25.2851391 128.852258,21.3646006 L128.852258,21.3646006 Z M117.720729,10.8281537 C119.337951,10.8281537 120.966841,11.3765623 122.525722,12.4500431 C110.519073,18.099819 97.6489725,32.3304399 92.2138928,60.7473424 C84.2701352,63.2070135 76.506069,65.6106769 69.3277499,67.834649 C75.6939575,46.1596724 90.8113669,10.8281537 117.720729,10.8281537 L117.720729,10.8281537 Z" fill="#95BF46"></path> <path d="M221.236945,54.9832175 C220.182133,54.8945386 197.853734,53.2399781 197.853734,53.2399781 C197.853734,53.2399781 182.346604,37.8448639 180.64537,36.1412966 C180.008283,35.5065427 179.149498,35.1821649 178.251042,35.0421456 L165.723988,291.275001 L255.485648,271.856667 C255.485648,271.856667 223.971987,58.8010751 223.773626,57.3402078 C223.572932,55.8793405 222.29409,55.0718963 221.236945,54.9832175" fill="#5E8E3E"></path> <path d="M135.241802,104.585029 L124.173282,137.510551 C124.173282,137.510551 114.474617,132.334507 102.586984,132.334507 C85.1592573,132.334507 84.2818035,143.272342 84.2818035,146.028387 C84.2818035,161.066452 123.48252,166.828244 123.48252,202.052414 C123.48252,229.764553 105.90544,247.610004 82.2048516,247.610004 C53.7646126,247.610004 39.2212821,229.90924 39.2212821,229.90924 L46.8359944,204.750118 C46.8359944,204.750118 61.7853808,217.585214 74.4011133,217.585214 C82.6435785,217.585214 85.9970391,211.095323 85.9970391,206.353338 C85.9970391,186.736644 53.8369559,185.861524 53.8369559,153.629098 C53.8369559,126.500372 73.3089633,100.246767 112.614694,100.246767 C127.760108,100.246767 135.241802,104.585029 135.241802,104.585029" fill="#FFFFFF"></path> </g> </svg>`,meta:`<svg width="22" height="22" width="256px" height="171px" viewBox="0 0 256 171" version="1.1" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid"> <title>Meta</title> <defs> <linearGradient x1="13.8784354%" y1="55.9337491%" x2="89.143574%" y2="58.6936324%" id="linearGradient-1"> <stop stop-color="#0064E1" offset="0%"></stop> <stop stop-color="#0064E1" offset="40%"></stop> <stop stop-color="#0073EE" offset="83%"></stop> <stop stop-color="#0082FB" offset="100%"></stop> </linearGradient> <linearGradient x1="54.3150272%" y1="82.782443%" x2="54.3150272%" y2="39.3067715%" id="linearGradient-2"> <stop stop-color="#0082FB" offset="0%"></stop> <stop stop-color="#0064E0" offset="100%"></stop> </linearGradient> </defs> <g> <path d="M27.6511337,112.135763 C27.6511337,121.910697 29.7966337,129.415496 32.6009181,133.955766 C36.2776464,139.902629 41.7615802,142.422034 47.3523439,142.422034 C54.5633607,142.422034 61.1601057,140.632633 73.8728613,123.050216 C84.0573098,108.957574 96.0578662,89.1762415 104.132425,76.775073 L117.806649,55.7651968 C127.305606,41.1740159 138.300181,24.9536792 150.906107,13.9591042 C161.197385,4.98539435 172.29879,0 183.471415,0 C202.228961,0 220.096258,10.8699402 233.770483,31.2566421 C248.735568,53.5840868 256,81.7070524 256,110.72917 C256,127.982195 252.599249,140.659341 246.81263,150.674642 C241.221867,160.360551 230.325219,170.037557 211.994992,170.037557 L211.994992,142.422034 C227.690082,142.422034 231.607178,128 231.607178,111.494784 C231.607178,87.9744053 226.123244,61.8723049 214.042565,43.2215885 C205.469467,29.9924885 194.35916,21.9090277 182.136041,21.9090277 C168.915844,21.9090277 158.277368,31.8798164 146.321324,49.6580887 C139.964946,59.1036305 133.439421,70.61455 126.112672,83.6032828 L118.047016,97.8917791 C101.844485,126.620114 97.7404368,133.163444 89.639171,143.962164 C75.4396995,162.871053 63.3145083,170.037557 47.3523439,170.037557 C28.4167478,170.037557 16.4428989,161.838364 9.02712477,149.481708 C2.97343163,139.412992 0,126.201697 0,111.147587 L27.6511337,112.135763 Z" fill="#0081FB"></path> <path d="M21.8021978,33.2062874 C34.4793434,13.665322 52.7739602,0 73.7571289,0 C85.9090277,0 97.9897065,3.59660593 110.604535,13.8967868 C124.403394,25.1584365 139.110307,43.702323 157.458339,74.2645709 L164.037279,85.2324384 C179.919321,111.690638 188.955348,125.302546 194.243427,131.721241 C201.04493,139.964946 205.807762,142.422034 211.994992,142.422034 C227.690082,142.422034 231.607178,128 231.607178,111.494784 L256,110.72917 C256,127.982195 252.599249,140.659341 246.81263,150.674642 C241.221867,160.360551 230.325219,170.037557 211.994992,170.037557 C200.599805,170.037557 190.504382,167.562665 179.340659,157.03102 C170.758659,148.947559 160.725553,134.587843 153.007094,121.679232 L130.047573,83.3273056 C118.527751,64.0801224 107.960495,49.7293087 101.844485,43.230491 C95.2655446,36.2420364 86.8081792,27.802476 73.3120045,27.802476 C62.3886493,27.802476 53.1122548,35.4675198 45.3492836,47.192099 L21.8021978,33.2062874 Z" fill="url(#linearGradient-1)"></path> <path d="M73.3120045,27.802476 C62.3886493,27.802476 53.1122548,35.4675198 45.3492836,47.192099 C34.3725136,63.7596328 27.6511337,88.4373348 27.6511337,112.135763 C27.6511337,121.910697 29.7966337,129.415496 32.6009181,133.955766 L9.02712477,149.481708 C2.97343163,139.412992 0,126.201697 0,111.147587 C0,83.7724301 7.51370149,55.2399499 21.8021978,33.2062874 C34.4793434,13.665322 52.7739602,0 73.7571289,0 L73.3120045,27.802476 Z" fill="url(#linearGradient-2)"></path> </g> </svg>`,gads:`<svg width="22" height="22" width="256px" height="230px" viewBox="0 0 256 230" version="1.1" xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" preserveAspectRatio="xMidYMid"> <g> <path d="M5.888,166.405103 L90.88,20.9 C101.676138,27.2558621 156.115862,57.3844138 164.908138,63.1135172 L79.9161379,208.627448 C70.6206897,220.906621 -5.888,185.040138 5.888,166.396276 L5.888,166.405103 Z" fill="#FBBC04"></path> <path d="M250.084224,166.401789 L165.092224,20.9055131 C153.210293,1.13172 127.619121,-6.05393517 106.600638,5.62496138 C85.582155,17.3038579 79.182155,42.4624786 91.0640861,63.1190303 L176.056086,208.632961 C187.938017,228.397927 213.52919,235.583582 234.547672,223.904686 C254.648086,212.225789 261.966155,186.175582 250.084224,166.419444 L250.084224,166.401789 Z" fill="#4285F4"></path> <ellipse fill="#34A853" cx="42.6637241" cy="187.924414" rx="42.6637241" ry="41.6044138"></ellipse> </g> </svg>`,tiktok:`<svg width="22" height="22" width="256px" height="290px" viewBox="0 0 256 290" version="1.1" xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" preserveAspectRatio="xMidYMid"> <title>TikTok</title> <g> <path d="M189.720224,104.421475 C208.398189,117.766281 231.279538,125.618095 255.992548,125.618095 L255.992548,78.0872726 C251.315611,78.0882654 246.650588,77.6008156 242.074913,76.6318726 L242.074913,114.045382 C217.363889,114.045382 194.485518,106.193568 175.80259,92.8497541 L175.80259,189.846306 C175.80259,238.368905 136.447224,277.701437 87.902784,277.701437 C69.7897057,277.701437 52.9543216,272.228299 38.9691786,262.841664 C54.9309256,279.153859 77.1908018,289.273158 101.81744,289.273158 C150.364858,289.273158 189.72221,249.940626 189.72221,201.416041 L189.72221,104.421475 L189.720224,104.421475 Z M206.889179,56.4687254 C197.343701,46.0456391 191.076347,32.5757434 189.720224,17.6842019 L189.720224,11.5707278 L176.531282,11.5707278 C179.851103,30.497877 191.174632,46.6681056 206.889179,56.4687254 L206.889179,56.4687254 Z M69.6735517,225.606854 C64.3403943,218.617757 61.4583846,210.068027 61.4712906,201.277053 C61.4712906,179.084685 79.472186,161.090739 101.680438,161.090739 C105.819294,161.089747 109.933331,161.723134 113.877603,162.974023 L113.877603,114.380938 C109.268175,113.749536 104.616057,113.481488 99.9659254,113.579773 L99.9659254,151.402303 C96.0186741,150.151413 91.9026521,149.516041 87.7628035,149.520012 C65.5545513,149.520012 47.5546487,167.511972 47.5546487,189.707318 C47.5546487,205.401018 56.552118,218.98806 69.6735517,225.606854 Z" fill="#FF004F"></path> <path d="M175.80259,92.8487613 C194.485518,106.192575 217.363889,114.044389 242.074913,114.044389 L242.074913,76.6308799 C228.281375,73.6942679 216.070311,66.4897401 206.889179,56.4687254 C191.173639,46.6671128 179.851103,30.4968842 176.531282,11.5707278 L141.8876,11.5707278 L141.8876,201.414056 C141.809172,223.545865 123.839052,241.466346 101.678453,241.466346 C88.6195635,241.466346 77.0180599,235.24466 69.6705734,225.606854 C56.5501325,218.98806 47.5526631,205.400025 47.5526631,189.708311 C47.5526631,167.51495 65.5525657,149.521004 87.760818,149.521004 C92.0158278,149.521004 96.1169583,150.183182 99.9639399,151.403295 L99.9639399,113.580765 C52.272289,114.565593 13.9166419,153.513923 13.9166419,201.415048 C13.9166419,225.326893 23.4680767,247.004014 38.9701714,262.842657 C52.9553144,272.228299 69.7906985,277.70243 87.9037768,277.70243 C136.449209,277.70243 175.803582,238.367912 175.803582,189.846306 L175.803582,92.8487613 L175.80259,92.8487613 Z" fill="#000000"></path> <path d="M242.074913,76.6308799 L242.074913,66.5145593 C229.636505,66.5334219 217.442318,63.0517795 206.889179,56.4677326 C216.231139,66.6902795 228.532545,73.7389425 242.074913,76.6308799 Z M176.531282,11.5707278 C176.214589,9.76190185 175.971361,7.9411627 175.80259,6.11347418 L175.80259,0 L127.968973,0 L127.968973,189.845313 C127.89253,211.974144 109.923403,229.894625 87.760818,229.894625 C81.2542071,229.894625 75.1109499,228.350869 69.6705734,225.607847 C77.0180599,235.24466 88.6195635,241.465353 101.678453,241.465353 C123.837066,241.465353 141.810164,223.546857 141.8876,201.415048 L141.8876,11.5707278 L176.531282,11.5707278 Z M99.9659254,113.580765 L99.9659254,102.811203 C95.9690357,102.265179 91.9393845,101.991175 87.9047695,101.99315 C39.3553659,101.99315 0,141.326686 0,189.845313 C0,220.263769 15.4673478,247.071522 38.9711641,262.840672 C23.4690694,247.003021 13.9176347,225.324907 13.9176347,201.414056 C13.9176347,153.513923 52.272289,114.565593 99.9659254,113.580765 Z" fill="#00F2EA"></path> </g> </svg>`};
 var PLAT=[
  {key:'mp',logo:'mp',nm:'Mercado Pago',tag:'Pagos',desc:'Trae tus pagos y movimientos para calcular tu ganancia real.',real:true},
  {key:'shopify',logo:'shopify',nm:'Shopify',tag:'Ventas',desc:'Conecta tu tienda para centralizar ventas y costos por orden.'},
  {key:'envialo',logo:'envialo',nm:'Envialo',tag:'Envios',desc:'Trae el costo real de cada envio para restarlo de la ganancia.'},
  {key:'tn',logo:'tn',nm:'Tiendanube',tag:'Ventas',desc:'Trae tus ventas online de tu tienda para el beneficio real.'},
  {key:'meli',logo:'meli',nm:'Mercado Libre',tag:'Ventas',desc:'Incluye tus ventas de marketplace en el calculo de beneficio.'},
  {key:'meta',logo:'meta',nm:'Meta Ads',tag:'Anuncios',desc:'Trae tus campanas de Facebook e Instagram Ads.'},
  {key:'gads',logo:'gads',nm:'Google Ads',tag:'Anuncios',desc:'Trae tus campanas de Google Ads (Search, Performance Max, Shopping).'},
  {key:'tiktok',logo:'tiktok',nm:'TikTok Ads',tag:'Anuncios',desc:'Trae tus campanas de TikTok Ads para medir su rentabilidad real.',soon:true}
 ];
 function esc(s){return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}
 function chip(txt,color,bg,bd){return '<span style="display:inline-flex;align-items:center;gap:7px;background:'+bg+';border:1px solid '+bd+';color:'+color+';border-radius:20px;padding:7px 14px;font-size:12.5px;font-weight:600"><span style="width:7px;height:7px;border-radius:50%;background:'+color+'"></span>'+txt+'</span>';}
 function cards(mpOn, shopOn, metaOn){
  var h='';
  var bs='background:#137fec;color:#fff;border-radius:10px;padding:9px 17px;font-weight:700;font-size:13px;text-decoration:none';
  var ds='background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;border-radius:10px;padding:9px 15px;font-weight:600;font-size:13px;text-decoration:none';
  PLAT.forEach(function(p){
   var right, on=(p.key==='mp'&&mpOn)||(p.key==='shopify'&&shopOn)||(p.key==='meta'&&metaOn)||(p.key==='envialo'&&window._rpEnv)||(p.key==='tn'&&window._rpTn);
   if(p.soon){ right='<span style="display:inline-flex;align-items:center;gap:6px;background:#241a10;border:1px solid #4a3a1a;color:#ffb35a;border-radius:20px;padding:7px 14px;font-size:12.5px;font-weight:700">&#128336; Proximamente</span>'; }
   else if(on){ var du=(p.key==='shopify')?'/desconectar-shopify':(p.key==='meta')?'/desconectar-meta':(p.key==='envialo')?'/desconectar-envialo':(p.key==='tn')?'/desconectar-tiendanube':'/desconectar-mp'; right=chip('Conectado','#34d399','#0e2a1c','#17492f')+'<a href="'+du+'" onclick="window.location.assign(\''+du+'\');return false;" style="'+ds+'">Desconectar</a>'; }
   else if((p.key==='shopify'&&window._rpTn)||(p.key==='tn'&&window._rpShop)){ var otra=(p.key==='shopify')?'Tiendanube':'Shopify'; right='<span style="display:inline-flex;align-items:center;gap:7px;background:#1c150c;border:1px solid #3d2e14;color:#e0a83b;border-radius:20px;padding:7px 13px;font-size:12px;font-weight:600">&#128274; Ya conectaste '+otra+'</span>'; }
   else { var b;
    if(p.key==='mp'){ b='<a href="/conectar-mp" onclick="window.location.assign(\'/conectar-mp\');return false;" style="'+bs+'">&#9889; Conectar</a>'; }
    else if(p.key==='shopify'){ b='<a href="#" onclick="rpShopToggle();return false;" style="'+bs+'">&#9889; '+(window._rpShopOpen?'Cerrar':'Conectar')+'</a>'; }
    else if(p.key==='meta'){ b='<a href="#" onclick="rpMetaTokToggle();return false;" style="'+bs+'">&#9889; '+(window._rpMetaTokOpen?'Cerrar':'Conectar')+'</a>'; }
    else if(p.key==='envialo'){ b='<a href="#" onclick="rpEnvToggle();return false;" style="'+bs+'">&#9889; '+(window._rpEnvOpen?'Cerrar':'Conectar')+'</a>'; }
    else if(p.key==='tn'){ b='<a href="#" onclick="rpTnToggle();return false;" style="'+bs+'">&#9889; '+(window._rpTnOpen?'Cerrar':'Conectar')+'</a>'; }
    else { b='<a href="#" onclick="alert(\'Muy pronto podes conectar \'+esc(p.nm)+\'.\');return false;" style="'+bs+'">&#9889; Conectar</a>'; }
    right=chip('No conectado','#94a3b8','#141d2c','#1e2b3d')+b; }
   var row='<div style="display:flex;align-items:center;gap:13px;padding:13px 17px">'
     +'<div style="width:44px;height:44px;border-radius:11px;flex:none;display:flex;align-items:center;justify-content:center;background:#131c2b">'+L[p.logo]+'</div>'
     +'<div style="flex:1;min-width:0"><div style="font-weight:700;font-size:14.5px;color:#f1f5f9;display:flex;align-items:center;gap:8px">'+esc(p.nm)+' <span style="font-size:10px;color:#94a3b8;background:#111c2b;border:1px solid #1e2b3d;padding:2px 8px;border-radius:16px;font-weight:600">'+p.tag+'</span></div>'
     +'<div style="color:#94a3b8;font-size:12.5px;margin-top:2px">'+esc(p.desc)+'</div></div>'
     +'<div style="display:flex;align-items:center;gap:11px;flex:none">'+right+'</div></div>';
   var panel='';
   if(p.key==='shopify' && window._rpShopOpen && !shopOn) panel=rpShopPanel();
   else if(p.key==='meta' && metaOn) panel=rpMetaPanel();
   else if(p.key==='meta' && !metaOn && window._rpMetaTokOpen) panel=rpMetaTokenPanel();
   else if(p.key==='envialo' && window._rpEnvOpen && !window._rpEnv) panel=rpEnvPanel();
   else if(p.key==='tn' && window._rpTnOpen && !window._rpTn) panel=rpTnPanel();
   h+='<div style="background:#0f1826;border:1px solid #1e2b3d;border-radius:12px;margin-bottom:9px;overflow:hidden">'+row+panel+'</div>';
  });
  document.getElementById('rp-integ-cards').innerHTML=h;
  if(metaOn){ try{ rpMetaLoad(); }catch(e){} }
 }
 window.rpMetaTokToggle=function(){ window._rpMetaTokOpen=!window._rpMetaTokOpen; cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); };
 function rpMetaTokenPanel(){ return ''
  +'<div style="border-top:1px solid #1e2b3d;padding:16px 17px 18px;background:#0c1521">'
  +'<div style="font-weight:700;color:#e2e8f0;font-size:13px">Peg&aacute; tu <b>token de Meta</b> (sin &laquo;Vincular&raquo;, sin App Review)</div>'
  +'<div style="color:#94a3b8;font-size:11.5px;margin-top:5px;line-height:1.6"><b style="color:#cbd5e1">1)</b> Abr&iacute; Meta con el bot&oacute;n de abajo y gener&aacute; un <b style="color:#cbd5e1">System User token</b> (caducidad <b style="color:#cbd5e1">Nunca</b>, permisos <b style="color:#cbd5e1">ads_read</b> + <b style="color:#cbd5e1">ads_management</b>). <b style="color:#cbd5e1">2)</b> Peg&aacute;lo abajo y toc&aacute; Conectar.</div>'
  +'<div style="display:flex;gap:9px;flex-wrap:wrap;margin-top:11px">'
  +'<a href="https://business.facebook.com/settings/system-users" target="_blank" rel="noopener" style="display:inline-flex;align-items:center;gap:7px;background:#0d1b30;border:1px solid #1c3350;color:#7db3f5;border-radius:8px;padding:9px 14px;font-size:12.5px;font-weight:700;text-decoration:none">&#128279; Sacar mi token (System User)</a>'
  +'<a href="https://developers.facebook.com/tools/explorer/" target="_blank" rel="noopener" style="display:inline-flex;align-items:center;gap:7px;background:#111c2b;border:1px solid #1e2b3d;color:#94a3b8;border-radius:8px;padding:9px 14px;font-size:12.5px;font-weight:600;text-decoration:none">&#9889; Token r&aacute;pido (Graph Explorer)</a>'
  +'</div>'
  +'<textarea id="rp-meta-tok" placeholder="Peg&aacute; ac&aacute; el token (empieza con EAAG...)" style="width:100%;height:74px;margin-top:11px;background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:12px;box-sizing:border-box;resize:vertical;font-family:monospace"></textarea>'
  +'<div style="display:flex;gap:11px;align-items:center;margin-top:10px"><button onclick="rpMetaTokSave(this)" style="background:#137fec;border:none;color:#fff;border-radius:8px;padding:10px 22px;font-weight:700;font-size:13px;cursor:pointer">Conectar</button><span id="rp-meta-tok-msg" style="font-size:12px;font-weight:600"></span></div>'
  +'</div>'; }
 window.rpMetaTokSave=function(btn){ var ta=document.getElementById('rp-meta-tok'), m=document.getElementById('rp-meta-tok-msg'); var t=(ta&&ta.value||'').trim();
  if(t.length<30){ if(m){m.style.color='#fb7185';m.textContent='Peg&aacute; un token v&aacute;lido.';} return; }
  if(btn){btn.disabled=true;} if(m){m.style.color='#94a3b8';m.textContent='Verificando con Meta...';}
  fetch('/meta/token-manual',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({token:t})}).then(function(r){return r.json();}).then(function(j){
   if(btn){btn.disabled=false;}
   if(j&&j.ok){ if(m){m.style.color='#34d399';m.textContent='✓ Conectado &middot; '+((j.cuentas||[]).length)+' cuenta(s)';}
    window._rpMeta=true; window._rpMetaTokOpen=false; window._rpMetaCuentas={cuentas:j.cuentas||[],elegida:j.elegida};
    setTimeout(function(){ cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); },800); }
   else { if(m){m.style.color='#fb7185';m.textContent=(j&&j.error)||'No se pudo conectar.';} }
  }).catch(function(){ if(btn){btn.disabled=false;} if(m){m.style.color='#fb7185';m.textContent='Error de red, reintent&aacute;.';} }); };
 function rpMetaPanel(){ return '<div style="border-top:1px solid #1e2b3d;padding:14px 17px;background:#0c1521"><div style="font-weight:700;color:#e2e8f0;font-size:12.5px;margin-bottom:8px">Cuenta publicitaria <span style="color:#94a3b8;font-weight:400">(de ac&aacute; sale el gasto de ads)</span></div><div id="rp-meta-cuentas" style="color:#94a3b8;font-size:12.5px">Cargando cuentas...</div></div>'; }
 function rpMetaRender(cuentas,elegida){ var c=document.getElementById('rp-meta-cuentas'); if(!c)return;
  if(!cuentas.length){ c.innerHTML='<span style="color:#94a3b8;font-size:12px">No encontramos cuentas publicitarias en tu Meta (revis&aacute; permisos).</span>'; return; }
  var h='<select onchange="rpMetaSave(this.value)" style="background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:9px;padding:9px 12px;font-size:13px;min-width:280px;max-width:100%"><option value="">Eleg&iacute; una cuenta...</option>';
  cuentas.forEach(function(a){ var sel=(String(a.id)===String(elegida))?' selected':''; h+='<option value="'+a.id+'"'+sel+'>'+esc(a.name)+' (act_'+a.id+')</option>'; });
  h+='</select>'+(elegida?'<span style="color:#34d399;font-size:12px;margin-left:10px;font-weight:600">&#10003; guardada</span>':'');
  c.innerHTML=h; }
 window.rpMetaLoad=function(force){ var c=document.getElementById('rp-meta-cuentas'); if(!c)return;
  if(window._rpMetaCuentas && !force){ rpMetaRender(window._rpMetaCuentas.cuentas, window._rpMetaCuentas.elegida); return; }
  fetch('/meta/cuentas').then(function(r){return r.json();}).then(function(j){ window._rpMetaCuentas={cuentas:(j&&j.cuentas)||[],elegida:j&&j.elegida}; rpMetaRender(window._rpMetaCuentas.cuentas, window._rpMetaCuentas.elegida); }).catch(function(){ var e=document.getElementById('rp-meta-cuentas'); if(e)e.innerHTML='<span style="color:#f87171;font-size:12px">No se pudieron cargar las cuentas.</span>'; }); };
 window.rpMetaSave=function(cid){ if(!cid)return; fetch('/meta/cuenta',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({cuenta:cid})}).then(function(r){return r.json();}).then(function(){ if(window._rpMetaCuentas)window._rpMetaCuentas.elegida=cid; rpMetaRender((window._rpMetaCuentas||{}).cuentas||[],cid); }).catch(function(){}); };
 function load(){ cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta);
  fetch('/mp/estado').then(function(r){return r.json();}).then(function(j){ window._rpMp=!!(j&&j.conectado); cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); }).catch(function(){});
  fetch('/shopify/estado').then(function(r){return r.json();}).then(function(j){ window._rpShop=!!(j&&j.conectado); cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); }).catch(function(){});
  fetch('/meta/estado').then(function(r){return r.json();}).then(function(j){ window._rpMeta=!!(j&&j.conectado); cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); }).catch(function(){});
  fetch('/envialo/estado').then(function(r){return r.json();}).then(function(j){ window._rpEnv=!!(j&&j.conectado); cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); }).catch(function(){});
  fetch('/tiendanube/estado').then(function(r){return r.json();}).then(function(j){ window._rpTn=!!(j&&j.conectado); cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); }).catch(function(){}); }
 function rpShopPanel(){ var CB='https://www.realprofitapp.com/shopify/callback'; return ''
  +'<div style="border-top:1px solid #1e2b3d;padding:16px 17px 18px;background:#0c1521">'
  +'<div style="font-weight:700;color:#e2e8f0;font-size:13px">1) Dominio de tu tienda</div>'
  +'<input id="rp-shop-dom" placeholder="mitienda.myshopify.com" style="width:100%;margin-top:6px;background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:13px;box-sizing:border-box">'
  +'<div style="font-weight:700;color:#e2e8f0;font-size:13px;margin-top:14px">2) En tu app de Shopify, peg&aacute; esta <b>URL de redireccionamiento</b>:</div>'
  +'<div style="position:relative;margin-top:6px">'
  +'<input id="rp-shop-redirect" readonly value="'+CB+'" style="width:100%;background:#0b1220;border:1px solid #1e2b3d;color:#93c5fd;border-radius:8px;padding:9px 64px 9px 11px;font-size:12px;font-family:ui-monospace,monospace;box-sizing:border-box">'
  +'<button onclick="rpCopy(\'rp-shop-redirect\',this)" style="position:absolute;top:6px;right:6px;background:#137fec;color:#fff;border:none;border-radius:6px;padding:5px 10px;font-size:11.5px;font-weight:600;cursor:pointer">Copiar</button>'
  +'</div>'
  +'<div style="color:#94a3b8;font-size:11.5px;margin-top:6px">Pon&eacute; la app en <b style="color:#cbd5e1">Distribuci&oacute;n personalizada</b> (NO p&uacute;blica) y tild&aacute; estos permisos:</div>'
  +'<div style="position:relative;margin-top:6px">'
  +'<textarea id="rp-shop-scopes" readonly rows="4" style="width:100%;background:#0b1220;border:1px solid #1e2b3d;color:#93c5fd;border-radius:8px;padding:9px 40px 9px 11px;font-size:11.5px;font-family:ui-monospace,monospace;box-sizing:border-box;resize:none">read_orders,write_orders,write_order_edits,read_fulfillments,write_fulfillments,read_merchant_managed_fulfillment_orders,write_merchant_managed_fulfillment_orders,read_assigned_fulfillment_orders,read_third_party_fulfillment_orders,read_shipping,write_shipping,read_products,read_inventory,read_customers,write_customers,read_locations,read_checkouts,write_draft_orders,write_price_rules</textarea>'
  +'<button onclick="rpCopy(\'rp-shop-scopes\',this)" style="position:absolute;top:7px;right:7px;background:#137fec;color:#fff;border:none;border-radius:6px;padding:4px 10px;font-size:11.5px;font-weight:600;cursor:pointer">Copiar</button>'
  +'</div>'
  +'<div style="font-weight:700;color:#e2e8f0;font-size:13px;margin-top:14px">3) Peg&aacute; el <b>Client ID</b> y el <b>Client Secret</b> de tu app:</div>'
  +'<input id="rp-shop-cid" placeholder="Client ID" style="width:100%;margin-top:6px;background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:13px;box-sizing:border-box">'
  +'<input id="rp-shop-secret" placeholder="Client Secret (shpss_...)" style="width:100%;margin-top:7px;background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:13px;box-sizing:border-box">'
  +'<div style="color:#94a3b8;font-size:11.5px;margin-top:6px">Al tocar Conectar te lleva a Shopify a <b style="color:#cbd5e1">aprobar</b>, y vuelve conectado.</div>'
  +'<div id="rp-shop-msg" style="margin-top:12px;font-size:12.5px;display:none;font-weight:600"></div>'
  +'<div style="display:flex;gap:9px;justify-content:flex-end;margin-top:14px">'
  +'<button onclick="rpShopToggle()" style="background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;border-radius:8px;padding:9px 15px;font-weight:600;font-size:12.5px;cursor:pointer">Cancelar</button>'
  +'<button id="rp-shop-go" onclick="rpShopGo()" style="background:#137fec;border:none;color:#fff;border-radius:8px;padding:9px 20px;font-weight:700;font-size:12.5px;cursor:pointer">Conectar</button>'
  +'</div></div>'; }
 window.rpShopToggle=function(){ window._rpShopOpen=!window._rpShopOpen; cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); };
 window.rpEnvToggle=function(){ window._rpEnvOpen=!window._rpEnvOpen; cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); };
 function rpEnvPanel(){ return ''
  +'<div style="border-top:1px solid #1e2b3d;padding:16px 17px 18px;background:#0c1521">'
  +'<div style="font-weight:700;color:#e2e8f0;font-size:13px">Peg&aacute; tu <b>API Key</b> de Envialo</div>'
  +'<div style="color:#94a3b8;font-size:11.5px;margin-top:5px">La sac&aacute;s de Envialo &rarr; Configuraci&oacute;n &rarr; API. Con eso traemos el costo real de cada env&iacute;o.</div>'
  +'<input id="rp-env-key" placeholder="rl_..." style="width:100%;margin-top:9px;background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:13px;box-sizing:border-box;font-family:ui-monospace,monospace">'
  +'<div id="rp-env-msg" style="margin-top:12px;font-size:12.5px;display:none;font-weight:600"></div>'
  +'<div style="display:flex;gap:9px;justify-content:flex-end;margin-top:14px">'
  +'<button onclick="rpEnvToggle()" style="background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;border-radius:8px;padding:9px 15px;font-weight:600;font-size:12.5px;cursor:pointer">Cancelar</button>'
  +'<button id="rp-env-go" onclick="rpEnvGo()" style="background:#137fec;border:none;color:#fff;border-radius:8px;padding:9px 20px;font-weight:700;font-size:12.5px;cursor:pointer">Vincular</button>'
  +'</div></div>'; }
 window.rpEnvGo=function(){ var k=(document.getElementById('rp-env-key').value||'').trim(); var msg=document.getElementById('rp-env-msg'); var go=document.getElementById('rp-env-go');
  function show(txt,ok){ msg.style.display='block'; msg.style.color=ok?'#34d399':'#f87171'; msg.textContent=txt; }
  if(!k){ show('Pegá tu API key de Envialo.',false); return; }
  go.disabled=true; go.textContent='Vinculando...';
  fetch('/envialo/conectar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({api_key:k})})
   .then(function(r){return r.json();})
   .then(function(j){ if(j&&j.ok){ window._rpEnv=true; window._rpEnvOpen=false; show('¡Vinculado!',true); cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); } else { go.disabled=false; go.textContent='Vincular'; show((j&&j.error)||'No se pudo vincular.',false); } })
   .catch(function(){ go.disabled=false; go.textContent='Vincular'; show('Error de conexión. Probá de nuevo.',false); }); };
 window.rpTnToggle=function(){ window._rpTnOpen=!window._rpTnOpen; cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); };
 function rpTnPanel(){ return ''
  +'<div style="border-top:1px solid #1e2b3d;padding:16px 17px 18px;background:#0c1521">'
  +'<div style="font-weight:700;color:#e2e8f0;font-size:13px">Conect&aacute; tu Tiendanube</div>'
  +'<a href="/conectar-tiendanube" onclick="window.location.assign(\'/conectar-tiendanube\');return false;" style="display:block;text-align:center;background:#137fec;color:#fff;border-radius:8px;padding:11px;font-weight:700;font-size:13px;text-decoration:none;margin-top:10px">&#9889; Conectar con un clic (autorizar tu tienda)</a>'
  +'<div style="text-align:center;color:#5b6b82;font-size:11px;margin:12px 0 2px">&mdash; o peg&aacute; el token (para VisionPure) &mdash;</div>'
  +'<input id="rp-tn-store" placeholder="Store ID (ej: 762180)" style="width:100%;margin-top:9px;background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:13px;box-sizing:border-box">'
  +'<input id="rp-tn-token" placeholder="Access Token" style="width:100%;margin-top:8px;background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:9px 11px;font-size:13px;box-sizing:border-box;font-family:ui-monospace,monospace">'
  +'<div id="rp-tn-msg" style="margin-top:12px;font-size:12.5px;display:none;font-weight:600"></div>'
  +'<div style="display:flex;gap:9px;justify-content:flex-end;margin-top:14px">'
  +'<button onclick="rpTnToggle()" style="background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;border-radius:8px;padding:9px 15px;font-weight:600;font-size:12.5px;cursor:pointer">Cancelar</button>'
  +'<button id="rp-tn-go" onclick="rpTnGo()" style="background:#137fec;border:none;color:#fff;border-radius:8px;padding:9px 20px;font-weight:700;font-size:12.5px;cursor:pointer">Conectar</button>'
  +'</div></div>'; }
 window.rpTnGo=function(){ var s=(document.getElementById('rp-tn-store').value||'').trim(); var t=(document.getElementById('rp-tn-token').value||'').trim(); var msg=document.getElementById('rp-tn-msg'); var go=document.getElementById('rp-tn-go');
  function show(txt,ok){ msg.style.display='block'; msg.style.color=ok?'#34d399':'#f87171'; msg.textContent=txt; }
  if(!s||!t){ show('Pegá el Store ID y el token.',false); return; }
  go.disabled=true; go.textContent='Conectando...';
  fetch('/tiendanube/guardar-token',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({store_id:s,token:t})})
   .then(function(r){return r.json();})
   .then(function(j){ if(j&&j.ok){ window._rpTn=true; window._rpTnOpen=false; show('¡Conectado!',true); cards(!!window._rpMp,!!window._rpShop,!!window._rpMeta); } else { go.disabled=false; go.textContent='Conectar'; show((j&&j.msg)||'No se pudo conectar.',false); } })
   .catch(function(){ go.disabled=false; go.textContent='Conectar'; show('Error de conexión. Probá de nuevo.',false); }); };
 window.rpCopy=function(id,btn){ var t=document.getElementById(id); if(!t)return; var v=(t.value!==undefined&&t.value!=='')?t.value:t.textContent; try{ if(t.select)t.select(); document.execCommand('copy'); }catch(e){} try{ navigator.clipboard.writeText(v); }catch(e){} if(btn){ var o=btn.textContent; btn.textContent='¡Copiado!'; setTimeout(function(){ btn.textContent=o; },1200); } };
 window.rpShopGo=function(){ var d=document.getElementById('rp-shop-dom').value.trim(); var cid=document.getElementById('rp-shop-cid').value.trim(); var sec=document.getElementById('rp-shop-secret').value.trim(); var msg=document.getElementById('rp-shop-msg'); var go=document.getElementById('rp-shop-go');
  function show(txt,ok){ msg.style.display='block'; msg.style.color=ok?'#34d399':'#f87171'; msg.textContent=txt; }
  if(!d){ show('Poné el dominio de tu tienda (paso 1).',false); return; }
  if(!cid||!sec){ show('Pegá el Client ID y el Client Secret (paso 3).',false); return; }
  go.disabled=true; go.textContent='Conectando...';
  fetch('/shopify/byoa-start',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({shop:d,client_id:cid,client_secret:sec})})
   .then(function(r){return r.json();})
   .then(function(j){ if(j&&j.ok&&j.url){ show('Redirigiendo a Shopify...',true); window.location.assign(j.url); } else { go.disabled=false; go.textContent='Conectar'; show((j&&j.error)||'No se pudo iniciar la conexión.',false); } })
   .catch(function(){ go.disabled=false; go.textContent='Conectar'; show('Error de conexión. Probá de nuevo.',false); }); };
 window.rpInteg=function(open){ var o=document.getElementById('rp-integ-ov'); if(!o)return; if(open){ var op=document.getElementById('rp-prod-ov'); if(op) op.style.display='none'; try{rpProdSetActive(false);}catch(e){} var oc=document.getElementById('rp-comis-ov'); if(oc) oc.style.display='none'; try{rpComisSetActive(false);}catch(e){} var od=document.getElementById('rp-desp-ov'); if(od)od.style.display='none'; var _of=document.getElementById('rp-fact-ov'); if(_of)_of.style.display='none'; var _om=document.getElementById('rp-mov-ov'); if(_om)_om.style.display='none'; var _oa=document.getElementById('rp-ads-ov'); if(_oa)_oa.style.display='none'; var _osk=document.getElementById('rp-stock-ov'); if(_osk)_osk.style.display='none'; var _olk=document.getElementById('rpf-lock'); if(_olk)_olk.style.display='none'; } o.style.display=open?'block':'none'; var b=document.getElementById('rp-integ-btn'); if(b) b.classList.toggle('rp-active',!!open); if(open) load(); };
 window._rpNavActive=function(id){ try{
   ['rp-prod-nav','rp-comis-nav','rp-desp-nav','rp-fact-nav','rp-mov-nav','rp-ads-nav','rp-stock-nav'].forEach(function(nid){ var a=document.querySelector('#'+nid+' a'); if(a){ a.classList.toggle('bg-white/[0.08]', nid===id); a.classList.toggle('text-primary', nid===id); } });
   var das=document.querySelectorAll('aside nav a[href="/dashboard"]'), da=null; for(var i=0;i<das.length;i++){ if(das[i].querySelector('.material-symbols-outlined')){ da=das[i]; break; } }
   if(da){ da.classList.toggle('bg-white/[0.08]', !id); da.classList.toggle('text-primary', !id); }
 }catch(e){} }
 function rpProdSetActive(on){ _rpNavActive(on?'rp-prod-nav':null); }
 window.rpProd=function(open){ var o=document.getElementById('rp-prod-ov'); if(!o)return; if(open){ var oi=document.getElementById('rp-integ-ov'); if(oi) oi.style.display='none'; var ib=document.getElementById('rp-integ-btn'); if(ib) ib.classList.remove('rp-active'); var oc=document.getElementById('rp-comis-ov'); if(oc) oc.style.display='none'; try{rpComisSetActive(false);}catch(e){} var od=document.getElementById('rp-desp-ov'); if(od)od.style.display='none'; var _of=document.getElementById('rp-fact-ov'); if(_of)_of.style.display='none'; var _om=document.getElementById('rp-mov-ov'); if(_om)_om.style.display='none'; var _oa=document.getElementById('rp-ads-ov'); if(_oa)_oa.style.display='none'; var _osk=document.getElementById('rp-stock-ov'); if(_osk)_osk.style.display='none'; var _olk=document.getElementById('rpf-lock'); if(_olk)_olk.style.display='none'; } o.style.display=open?'block':'none'; rpProdSetActive(!!open); if(open) rpProdLoad(); };
 function rpComisSetActive(on){ _rpNavActive(on?'rp-comis-nav':null); }
 function rpComisTotal(){ var g=function(id){var el=document.getElementById(id); return el?(parseFloat(el.value||'0')||0):0;};
   var ti=g('rp-c-tienda'), iva=g('rp-c-iva'), iibb=g('rp-c-iibb'), t;
   if(window._rpMpReal!=null){ t=window._rpMpReal + ti + iibb; }
   else { t=(g('rp-c-mp')+g('rp-c-cuotas'))*(1+iva/100) + ti + iibb; }
   var el=document.getElementById('rp-c-total'); if(el) el.textContent=(Math.round(t*100)/100)+'%'; }
 window.rpComis=function(open){ var o=document.getElementById('rp-comis-ov'); if(!o)return; if(open){ var op=document.getElementById('rp-prod-ov'); if(op)op.style.display='none'; try{rpProdSetActive(false);}catch(e){} var oi=document.getElementById('rp-integ-ov'); if(oi){oi.style.display='none'; var ib=document.getElementById('rp-integ-btn'); if(ib)ib.classList.remove('rp-active');} var od=document.getElementById('rp-desp-ov'); if(od)od.style.display='none'; var _of=document.getElementById('rp-fact-ov'); if(_of)_of.style.display='none'; var _om=document.getElementById('rp-mov-ov'); if(_om)_om.style.display='none'; var _oa=document.getElementById('rp-ads-ov'); if(_oa)_oa.style.display='none'; var _osk=document.getElementById('rp-stock-ov'); if(_osk)_osk.style.display='none'; var _olk=document.getElementById('rpf-lock'); if(_olk)_olk.style.display='none'; } o.style.display=open?'block':'none'; rpComisSetActive(!!open); if(open) rpComisLoad(); };

 // ===================== DESPACHOS =====================
 var _dRows=[], _dFilt='empaquetar', _dDesde=null, _dHasta=null, _dLoaded=false;
 var _DCOL=['#5aa2f5','#34d399','#f0b429','#a78bfa','#fb7185','#38bdf8','#f472b6'];
 function _dEsc(s){ return String(s==null?'':s).replace(/[&<>"']/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];}); }
 function _dFmt(n){ try{return '$'+Math.round(n||0).toLocaleString('es-AR');}catch(e){return '$'+Math.round(n||0);} }
 function _dIni(nm){ var p=(nm||'').trim().split(/\s+/); return ((p[0]||'?')[0]+((p[1]||'')[0]||'')).toUpperCase(); }
 function _dColor(nm){ var s=0; for(var i=0;i<(nm||'').length;i++) s+=nm.charCodeAt(i); return _DCOL[s%_DCOL.length]; }
 function _dStat(msg,color){ var s=document.getElementById('rp-d-status'); if(s){ s.textContent=msg||''; s.style.color=color||'#34d399'; } }
 window.rpDesp=function(open){ var o=document.getElementById('rp-desp-ov'); if(!o)return;
   if(open){ ['rp-prod-ov','rp-comis-ov','rp-integ-ov','rp-fact-ov','rp-mov-ov','rp-ads-ov','rp-stock-ov'].forEach(function(id){ var x=document.getElementById(id); if(x)x.style.display='none'; });
     var _oa=document.getElementById('rp-ads-ov'); if(_oa)_oa.style.display='none'; var _osk=document.getElementById('rp-stock-ov'); if(_osk)_osk.style.display='none'; var _olk=document.getElementById('rpf-lock'); if(_olk)_olk.style.display='none';
     try{rpProdSetActive(false);}catch(e){} try{rpComisSetActive(false);}catch(e){} var ib=document.getElementById('rp-integ-btn'); if(ib)ib.classList.remove('rp-active');
     _rpNavActive('rp-desp-nav');
   } else { _rpNavActive(null); }
   o.style.display=open?'block':'none';
   if(open && !_dLoaded) rpDLoad(); };
 window.rpDLoad=function(){ var b=document.getElementById('rp-d-sync'); var bh=b?b.innerHTML:''; if(b){b.style.opacity='.6';}
   _dStat('Trayendo pedidos de tu tienda…','#38bdf8');
   var qs=[]; if(_dDesde)qs.push('desde='+_dDesde); if(_dHasta)qs.push('hasta='+_dHasta);
   fetch('/pf-despachos'+(qs.length?('?'+qs.join('&')):'')).then(function(r){return r.json();}).then(function(j){
     if(b)b.style.opacity='';
     if(!j||!j.ok){ _dStat('No se pudo cargar.', '#fb7185'); return; }
     if(j.shopify===false){ _dRows=[]; _dLoaded=true; rpDRender(); _dStat('Conectá tu tienda (Shopify) en Integraciones para ver los despachos.', '#f0b429'); return; }
     _dRows=j.rows||[]; _dLoaded=true;
     var R=j.resumen||{};
     [['empaquetar',R.empaquetar],['exportada',R.exportada],['enviada',R.enviada],['todas',R.todas]].forEach(function(x){
       var e1=document.getElementById('rp-d-n-'+x[0]), e2=document.getElementById('rp-d-m-'+x[0]);
       if(e1)e1.textContent=(x[1]&&x[1].n)||0; if(e2)e2.textContent=_dFmt((x[1]&&x[1].monto)||0);
     });
     rpDRender(); _dStat('');
   }).catch(function(){ if(b)b.style.opacity=''; _dStat('Error de conexión.', '#fb7185'); }); };
 window.rpDVisibles=function(){ var q=((document.getElementById('rp-d-q')||{}).value||'').toLowerCase().trim();
   var base=_dFilt==='todas'?_dRows:_dRows.filter(function(r){return r.estado===_dFilt;});
   if(!q) return base;
   return base.filter(function(r){ return (r.num+' '+r.nombre+' '+r.localidad+' '+r.cp).toLowerCase().indexOf(q)>=0; }); };
 window.rpDRender=function(){ var rows=rpDVisibles(), tb=document.getElementById('rp-d-body'); if(!tb)return;
   var td='padding:13px;border-bottom:1px solid #141c2a;font-size:13px';
   if(!rows.length){ tb.innerHTML='<tr><td colspan="8" style="padding:40px 14px;text-align:center;color:#5b6b82;font-size:13.5px">No hay pedidos en este estado.</td></tr>'; }
   else tb.innerHTML=rows.map(function(r){
     var env = r.tipo==='sucursal'
       ? '<span style="background:#0e1d33;border:1px solid #1e3a5f;color:#8fbdf5;font-size:11.5px;font-weight:700;border-radius:16px;padding:3px 10px">🏤 Sucursal</span>'
       : '<span style="background:#0e2a1c;border:1px solid #17492f;color:#34d399;font-size:11.5px;font-weight:700;border-radius:16px;padding:3px 10px">🏠 Domicilio</span>';
     var est = r.estado==='exportada' ? ' <span style="background:#0a2434;border:1px solid #155066;color:#38bdf8;font-size:11px;font-weight:700;border-radius:14px;padding:2px 8px;margin-left:6px">exportada</span>'
             : r.estado==='enviada' ? ' <span style="background:#1c1636;border:1px solid #3a2f6b;color:#a78bfa;font-size:11px;font-weight:700;border-radius:14px;padding:2px 8px;margin-left:6px">enviada</span>' : '';
     var w = r.incompleta ? ' <span title="Dirección incompleta" style="color:#fb7185;font-size:12px;cursor:help">⚠</span>' : '';
     return '<tr onmouseover="this.style.background=&#39;#0d1622&#39;" onmouseout="this.style.background=&#39;&#39;">'
       +'<td style="'+td+';width:36px"><input type="checkbox" class="rp-d-chk" value="'+_dEsc(r.num)+'" onclick="rpDCnt()" style="width:16px;height:16px;accent-color:#3b82f6;cursor:pointer"></td>'
       +'<td style="'+td+';color:#cbd5e1;font-weight:700">#'+_dEsc(r.num)+w+'</td>'
       +'<td style="'+td+'"><span style="display:inline-flex;align-items:center;gap:9px"><span style="width:27px;height:27px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:10.5px;font-weight:800;color:#fff;background:'+_dColor(r.nombre)+'">'+_dEsc(_dIni(r.nombre))+'</span>'+_dEsc(r.nombre)+'</span></td>'
       +'<td style="'+td+'">'+env+est+'</td>'
       +'<td style="'+td+';color:#8493a8">'+_dEsc(r.localidad)+'</td>'
       +'<td style="'+td+';color:#8493a8">'+_dEsc(r.cp)+'</td>'
       +'<td style="'+td+';text-align:right">'+r.unidades+'</td>'
       +'<td style="'+td+';text-align:right;font-weight:700">'+_dFmt(r.total)+'</td></tr>';
   }).join('');
   var sa=document.getElementById('rp-d-all'); if(sa)sa.checked=false; rpDCnt(); };
 window.rpDFilt=function(f){ _dFilt=f;
   var cs=document.querySelectorAll('#rp-d-cards .rp-dc');
   for(var i=0;i<cs.length;i++){ var on=cs[i].getAttribute('data-f')===f; cs[i].style.borderColor=on?'#5a4a1e':'#1a2333'; cs[i].style.background=on?'linear-gradient(165deg,rgba(240,180,41,.07),transparent 60%)':'#0e1521'; var lb=cs[i].querySelector('.rp-dl'); if(lb)lb.style.color=on?'#f0b429':'#8493a8'; }
   rpDRender(); };   // los botones son los mismos en las 4 vistas
 window.rpDPer=function(p){ var ss=document.querySelectorAll('#rp-d-segs .rp-seg');
   for(var i=0;i<ss.length;i++){ var on=ss[i].getAttribute('data-p')===p; ss[i].style.background=on?'#2563eb':'transparent'; ss[i].style.color=on?'#fff':'#8493a8'; }
   var di=document.getElementById('rp-d-desde'), ha=document.getElementById('rp-d-hasta'); if(di)di.value=''; if(ha)ha.value='';
   var now=new Date(); var off=now.getTimezoneOffset(); // usar fecha local AR aprox
   function iso(d){ return d.toISOString().slice(0,10); }
   if(p==='todas'){ _dDesde=null; _dHasta=null; }
   else if(p==='hoy'){ _dDesde=iso(now); _dHasta=iso(now); }
   else if(p==='ayer'){ var y=new Date(now.getTime()-86400000); _dDesde=iso(y); _dHasta=iso(y); }
   else if(p==='7'){ var w=new Date(now.getTime()-6*86400000); _dDesde=iso(w); _dHasta=iso(now); }
   rpDLoad(); };
 window.rpDRango=function(){ var di=document.getElementById('rp-d-desde').value, ha=document.getElementById('rp-d-hasta').value;
   if(!di&&!ha) return; _dDesde=di||ha; _dHasta=ha||di;
   var ss=document.querySelectorAll('#rp-d-segs .rp-seg'); for(var i=0;i<ss.length;i++){ ss[i].style.background='transparent'; ss[i].style.color='#8493a8'; }
   rpDLoad(); };
 window.rpDSel=function(){ return [].slice.call(document.querySelectorAll('.rp-d-chk:checked')).map(function(c){return c.value;}); };
 window.rpDCnt=function(){ var n=rpDSel().length, tot=rpDVisibles().length; var e=document.getElementById('rp-d-cnt'); if(e)e.textContent=n?(n+' de '+tot+' seleccionada'+(n>1?'s':'')):(tot+' pedido'+(tot!==1?'s':'')); };
 window.rpDAll=function(c){ document.querySelectorAll('.rp-d-chk').forEach(function(x){x.checked=c.checked;}); rpDCnt(); };
 window.rpDMarcar=function(accion){ var sel=rpDSel(); if(!sel.length){ _dStat('Tildá al menos un pedido.', '#fb7185'); return; }
   _dStat('Guardando…','#38bdf8');
   fetch('/pf-despachos-marcar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nums:sel,accion:accion})})
    .then(function(r){return r.json();}).then(function(j){ if(j&&j.ok){ _dLoaded=false; rpDLoad(); _dStat('✓ '+sel.length+' pedido(s) → '+(accion==='exportada'?'Exportadas':'Enviadas')+'.'); } else _dStat('No se pudo guardar.', '#fb7185'); })
    .catch(function(){ _dStat('Error de conexión.', '#fb7185'); }); };
 window.rpDExcel=function(){ var sel=rpDSel(); if(!sel.length){ _dStat('Tildá los pedidos para el Excel.', '#fb7185'); return; }
   _dStat('⏳ Generando Excel Andreani…','#a78bfa');
   fetch('/pf-despachos-excel',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nums:sel})})
    .then(function(r){ if(!r.ok) return r.json().then(function(e){throw (e&&e.msg)||'error';}); return r.blob(); })
    .then(function(b){ var u=URL.createObjectURL(b); var a=document.createElement('a'); a.href=u; a.download='Andreani.xlsx'; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(u);
      // al exportar, los pasa a "Exportadas"
      fetch('/pf-despachos-marcar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nums:sel,accion:'exportada'})}).then(function(){ _dLoaded=false; rpDLoad(); });
      _dStat('✅ Excel descargado ('+sel.length+' pedidos) → Exportadas. Subílo a Andreani.'); })
    .catch(function(e){ _dStat('No se pudo generar el Excel'+(typeof e==='string'?': '+e:'')+'.', '#fb7185'); }); };
 window.rpDActSku=function(){ _dStat('↻ Actualizando SKUs desde tu tienda…','#38bdf8');
   fetch('/pf-despachos-sku-sync',{method:'POST'}).then(function(r){return r.json();}).then(function(j){ _dStat(j&&j.ok?('✓ '+(j.n||0)+' SKUs actualizados desde tu tienda.'):'No se pudo actualizar.'); }).catch(function(){ _dStat('Actualización de SKUs: pendiente de conectar.', '#f0b429'); }); };
 // ---- Modal Insertar SKU ----
 window.rpDOpenSku=function(){ var m=document.getElementById('rp-d-skuov'); if(m){ m.style.display='flex'; var r=document.getElementById('rp-d-skures'); if(r)r.innerHTML=''; } };
 window.rpDCloseSku=function(){ var m=document.getElementById('rp-d-skuov'); if(m)m.style.display='none'; };
 function rpDBarra(pct,msg){ return '<div style="font-size:12.5px;color:#c9b8ff;font-weight:600;margin-bottom:8px">'+msg+'</div>'
    +'<div style="height:12px;background:#0b1220;border:1px solid #2b2350;border-radius:20px;overflow:hidden">'
    +'<div style="height:100%;width:'+pct+'%;background:linear-gradient(90deg,#7c3aed,#a78bfa);border-radius:20px;transition:width .3s ease"></div></div>'
    +'<div style="font-size:11px;color:#7a6ca8;margin-top:5px;text-align:right">'+pct+'%</div>'; }
 window.rpDUpSku=function(inp){ var f=inp.files&&inp.files[0]; if(!f)return; var res=document.getElementById('rp-d-skures');
   res.innerHTML=rpDBarra(2,'⏳ Subiendo el PDF…');
   var fd=new FormData(); fd.append('pdf',f);
   fetch('/pf-despachos-sku',{method:'POST',body:fd}).then(function(r){return r.json();}).then(function(j){
     if(!j||!j.ok){ throw (j&&j.msg)||'error'; }
     var job=j.job;
     var poll=setInterval(function(){
       fetch('/pf-despachos-sku-progreso?job='+job).then(function(r){return r.json();}).then(function(p){
         if(!p||!p.ok){ return; }
         if(p.error){ clearInterval(poll); res.innerHTML='<div style="color:#fb7185;font-size:12.5px">No se pudo procesar: '+p.error+'</div>'; return; }
         var pct=p.total?Math.round(p.done/p.total*100):5; if(pct<2)pct=2; if(!p.listo&&pct>98)pct=98;
         res.innerHTML=rpDBarra(pct,p.msg||'Procesando…');
         if(p.listo){ clearInterval(poll);
           var s=p.stats||{};
           fetch('/pf-despachos-sku-descargar?job='+job).then(function(r){return r.blob();}).then(function(b){
             var u=URL.createObjectURL(b); var a=document.createElement('a'); a.href=u; a.download='etiquetas-con-sku.pdf'; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(u);
             var extra=''; if(s.conflicto)extra+=' · ⚠️ '+s.conflicto+' sin estampar (nombre no coincide)'; if(s.sin_pedido)extra+=' · '+s.sin_pedido+' sin pedido en la tienda';
             res.innerHTML='<div style="background:#0e2a1c;border:1px solid #17492f;border-radius:12px;padding:13px 15px;color:#34d399;font-size:13px;font-weight:700">✅ '+(s.estampadas||0)+' de '+(s.total||0)+' etiquetas con SKU — PDF descargado.'+extra+'</div>';
           });
         }
       }).catch(function(){});
     },400);
   }).catch(function(e){ res.innerHTML='<div style="color:#fb7185;font-size:12.5px">No se pudo procesar'+(typeof e==='string'?': '+e:'')+'.</div>'; }); inp.value=''; };
 // ---- Modal Enviar seguimiento ----
 window.rpDOpenSeg=function(){ var m=document.getElementById('rp-d-segov'); if(m){ m.style.display='flex'; var r=document.getElementById('rp-d-segres'); if(r)r.innerHTML=''; } };
 window.rpDCloseSeg=function(){ var m=document.getElementById('rp-d-segov'); if(m)m.style.display='none'; };
 var _dSeg=[];
 window.rpDUpSeg=function(inp){ var f=inp.files&&inp.files[0]; if(!f)return; var res=document.getElementById('rp-d-segres');
   res.innerHTML='<div style="color:#fb7185;font-size:12.5px">⏳ Leyendo el PDF (N° Interno + seguimiento)…</div>';
   var fd=new FormData(); fd.append('pdf',f);
   fetch('/pf-despachos-seg-leer',{method:'POST',body:fd}).then(function(r){return r.json();}).then(function(j){
     if(!j||!j.ok||!(j.pedidos&&j.pedidos.length)){ res.innerHTML='<div style="color:#fb7185;font-size:12.5px">'+((j&&j.msg)||'No pude leer pedidos del PDF')+'.</div>'; return; }
     _dSeg=j.pedidos;
     var filas=_dSeg.map(function(o){ return '<tr><td style="padding:9px 8px;border-top:1px solid #141c2a;color:#cbd5e1;font-weight:700">#'+_dEsc(o.num)+'</td><td style="padding:9px 8px;border-top:1px solid #141c2a">'+_dEsc(o.nombre||'')+'</td><td style="padding:9px 8px;border-top:1px solid #141c2a;color:#8493a8;font-family:ui-monospace,monospace;font-size:12px">'+_dEsc(o.track||'')+'</td></tr>'; }).join('');
     res.innerHTML='<div style="background:#0b111c;border:1px solid #1a2333;border-radius:12px;overflow:hidden"><div style="padding:11px 14px;color:#e7edf5;font-size:13px;font-weight:700;border-bottom:1px solid #1a2333">✅ '+_dSeg.length+' pedidos leídos</div>'
       +'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:12.5px"><thead><tr><th style="text-align:left;padding:8px;color:#5b6b82;font-size:10px;text-transform:uppercase">Pedido</th><th style="text-align:left;padding:8px;color:#5b6b82;font-size:10px;text-transform:uppercase">Cliente</th><th style="text-align:left;padding:8px;color:#5b6b82;font-size:10px;text-transform:uppercase">Seguimiento</th></tr></thead><tbody>'+filas+'</tbody></table></div>'
       +'<div style="padding:12px 14px;display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap;border-top:1px solid #1a2333"><span style="color:#5b6b82;font-size:11.5px">Al enviar, tu tienda le manda el mail con el tracking al cliente.</span><button onclick="rpDEnviarSeg()" style="display:inline-flex;align-items:center;gap:7px;background:linear-gradient(160deg,#b23a55,#8f2c44);border:1px solid #a23650;color:#ffe0e7;border-radius:10px;padding:9px 15px;font-size:12.5px;font-weight:700;cursor:pointer">Enviar '+_dSeg.length+' seguimientos</button></div></div>';
   }).catch(function(){ res.innerHTML='<div style="color:#fb7185;font-size:12.5px">Error leyendo el PDF.</div>'; }); inp.value=''; };
 window.rpDEnviarSeg=function(){ if(!_dSeg.length)return; var res=document.getElementById('rp-d-segres');
   res.innerHTML='<div style="color:#fb7185;font-size:12.5px">⏳ Enviando seguimientos y avisando a tu tienda…</div>';
   fetch('/pf-despachos-seg-enviar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pedidos:_dSeg})}).then(function(r){return r.json();}).then(function(j){
     if(j&&j.ok){ res.innerHTML='<div style="background:#0e2a1c;border:1px solid #17492f;border-radius:12px;padding:13px 15px;color:#34d399;font-size:13px;font-weight:700">📲 '+(j.enviados||0)+' seguimientos enviados — la tienda le avisó por mail a cada cliente.</div>'; _dLoaded=false; rpDLoad(); setTimeout(rpDCloseSeg,1600); }
     else res.innerHTML='<div style="color:#fb7185;font-size:12.5px">'+((j&&j.msg)||'No se pudo enviar')+'.</div>';
   }).catch(function(){ res.innerHTML='<div style="color:#fb7185;font-size:12.5px">Error de conexión.</div>'; }); };
 // ===================== FACTURACIÓN =====================
 var _fRows=[], _fFilter='todas', _fPage=1, _fSel={}, _fPer=50, _fAutoOn=false;
 function _fStat(m,c){ var s=document.getElementById('rpf-status'); if(s){ s.textContent=m||''; s.style.color=c||'#34d399'; } }
 function _fEmit(){ var v=(document.getElementById('rpf-emit')||{}).value||''; if(!v)return ''; var p=v.split('-'); return p[2]+'/'+p[1]+'/'+p[0]; }
 window.rpFact=function(open){ var o=document.getElementById('rp-fact-ov'); if(!o)return;
   if(open){ ['rp-prod-ov','rp-comis-ov','rp-integ-ov','rp-desp-ov','rp-mov-ov','rp-ads-ov','rp-stock-ov'].forEach(function(id){ var x=document.getElementById(id); if(x)x.style.display='none'; });
     try{rpProdSetActive(false);}catch(e){} try{rpComisSetActive(false);}catch(e){} var ib=document.getElementById('rp-integ-btn'); if(ib)ib.classList.remove('rp-active');
     _rpNavActive('rp-fact-nav');
   } else { _rpNavActive(null); }
   o.style.display=open?'block':'none';
   if(open){ var _lk0=document.getElementById('rpf-lock'); if(_lk0)_lk0.style.display='flex';  /* BLOQUEADO en configuración: candado al instante */
     var e=document.getElementById('rpf-emit'); if(e&&!e.value){ var t=new Date(); e.value=t.getFullYear()+'-'+String(t.getMonth()+1).padStart(2,'0')+'-'+String(t.getDate()).padStart(2,'0'); } rpFLoad(); } };
 window.rpFLoad=function(){ _fStat('Trayendo ventas de tu tienda…','#38bdf8');
   var d=(document.getElementById('rpf-desde')||{}).value, h=(document.getElementById('rpf-hasta')||{}).value, qs=[];
   if(d)qs.push('desde='+d); if(h)qs.push('hasta='+h);
   fetch('/pf-facturacion'+(qs.length?('?'+qs.join('&')):'')).then(function(r){return r.json();}).then(function(j){
     if(!j||!j.ok){ _fStat('No se pudo cargar.', '#fb7185'); return; }
     var _lka=j.arca||{}; window._fArca=_lka; var _lk=document.getElementById('rpf-lock'); if(_lk)_lk.style.display='flex';  /* BLOQUEADO: en configuración hasta conectar ARCA */
     if(j.shopify===false){ _fRows=[]; rpFRender(); if(_lka.vinculado)_fStat('Conectá tu tienda en Integraciones para ver las ventas.', '#f0b429'); return; }
     _fRows=j.rows||[];
     var a=j.arca||{}, ab=document.getElementById('rpf-arca');
     if(ab) ab.innerHTML='<span class="dot"></span><span>Conectado a ARCA</span><span class="sep">·</span><span class="lbl">Facturás como</span><b>'+_dEsc(a.nombre||'—')+'</b>'+(a.cuit?('<span class="cuit">CUIT '+_dEsc(a.cuit)+'</span>'):'<span class="cuit" style="color:#7d8ba0;background:#141d2e;border-color:#26324a">CUIT sin configurar</span>')+'<span class="sep">·</span><span>'+_dEsc(a.tipo||'Factura C')+'</span><span class="sep">·</span><span>Punto de Venta '+_dEsc(a.pv||'5')+'</span>';
     var au=j.auto||{}; _fAutoOn=!!au.on; var t=document.getElementById('rpf-autotoggle'); if(t)t.checked=_fAutoOn; var dl=document.getElementById('rpf-autodelay'); if(dl&&au.delay)dl.value=String(au.delay); _fAutoState();
     rpFRender(); _fStat(''); }).catch(function(){ _fStat('Error de conexión.', '#fb7185'); }); };
 function _fFiltered(){ return _fFilter==='todas'?_fRows:(_fFilter==='facturadas'?_fRows.filter(function(o){return o.facturada;}):_fRows.filter(function(o){return !o.facturada;})); }
 window.rpFRender=function(){
   var data=_fFiltered(), pages=Math.max(1,Math.ceil(data.length/_fPer)); if(_fPage>pages)_fPage=pages;
   var start=(_fPage-1)*_fPer, rows=data.slice(start,start+_fPer), tb=document.getElementById('rpf-body'); if(!tb)return;
   tb.innerHTML=rows.map(function(o){
     var on=!!_fSel[o.num];
     var est=o.facturada?'<span class="est e-ok">✓ Emitida</span>':'<span class="est e-pd">Pendiente</span>';
     var mc=o.medio==='Tarjeta'?'#34d399':(o.medio==='Transferencia'?'#fbbf24':'#38bdf8');
     var comp=o.comprobante?('<span class="comp">'+_dEsc(o.comprobante)+'</span>'+(o.emit?('<div style="color:#5b6678;font-size:10.5px;margin-top:2px">emisión '+_dEsc(o.emit)+'</div>'):'')):'<span style="color:#5b6b82">—</span>';
     return '<tr>'
       +'<td><span class="chk '+(on?'on':'')+'" onclick="rpFToggle(&#39;'+_dEsc(o.num)+'&#39;)"></span></td>'
       +'<td>'+_dEsc(o.fecha_dmy||(o.fecha||'').slice(0,10))+'</td>'
       +'<td style="color:#cbd5e1;font-weight:700">#'+_dEsc(o.num)+'</td>'
       +'<td><div class="cust"><span class="av2" style="background:'+_dColor(o.nombre)+'">'+_dEsc(_dIni(o.nombre))+'</span>'+_dEsc(o.nombre)+'</div></td>'
       +'<td><span class="medio"><span class="midot" style="background:'+mc+'"></span>'+_dEsc(o.medio||'—')+'</span></td>'
       +'<td>'+comp+'</td>'
       +'<td style="text-align:right;font-weight:700">'+_dFmt(o.total)+'</td>'
       +'<td>'+est+'</td></tr>';
   }).join('')||'<tr><td colspan="8" style="padding:40px;text-align:center;color:#5b6b82">No hay ventas en este período/filtro.</td></tr>';
   var fac=_fRows.filter(function(o){return o.facturada;}), fm=fac.reduce(function(a,o){return a+o.total;},0);
   var setT=function(id,v){var e=document.getElementById(id);if(e)e.textContent=v;};
   setT('rpf-k-ventas',_fRows.length); setT('rpf-k-fact',fac.length); setT('rpf-k-pend',_fRows.length-fac.length); setT('rpf-ks-fact','emitidas · '+_dFmt(fm));
   var nsel=Object.keys(_fSel).length;
   setT('rpf-cnt',data.length+' venta'+(data.length!==1?'s':'')+(nsel?(' · '+nsel+' seleccionada'+(nsel>1?'s':'')):''));
   var ca=document.getElementById('rpf-chkall'); if(ca)ca.checked=rows.length>0&&rows.every(function(o){return _fSel[o.num];});
   _fPager(pages,data.length,start,rows.length); rpFPctInput();
 };
 function _fPager(pages,total,start,shown){ var p=document.getElementById('rpf-pager'); if(!p)return; var h='';
   h+='<button class="pbtn" '+(_fPage<=1?'disabled':'')+' onclick="rpFGoto('+(_fPage-1)+')">‹ Anterior</button>';
   var lo=Math.max(1,_fPage-2),hi=Math.min(pages,_fPage+2);
   if(lo>1){h+='<button class="pbtn" onclick="rpFGoto(1)">1</button>';if(lo>2)h+='<span class="info">…</span>';}
   for(var i=lo;i<=hi;i++)h+='<button class="pbtn '+(i===_fPage?'on':'')+'" onclick="rpFGoto('+i+')">'+i+'</button>';
   if(hi<pages){if(hi<pages-1)h+='<span class="info">…</span>';h+='<button class="pbtn" onclick="rpFGoto('+pages+')">'+pages+'</button>';}
   h+='<button class="pbtn" '+(_fPage>=pages?'disabled':'')+' onclick="rpFGoto('+(_fPage+1)+')">Siguiente ›</button>';
   h+='<span class="info">'+(total?(start+1):0)+'–'+(start+shown)+' de '+total+'</span>'; p.innerHTML=h; }
 window.rpFGoto=function(n){ _fPage=n; rpFRender(); var o=document.getElementById('rp-fact-ov'); if(o)o.scrollTop=0; };
 window.rpFToggle=function(id){ id=String(id); if(_fSel[id])delete _fSel[id]; else _fSel[id]=1; rpFRender(); };
 window.rpFToggleAll=function(c){ _fFiltered().slice((_fPage-1)*_fPer,(_fPage-1)*_fPer+_fPer).forEach(function(o){if(c.checked)_fSel[o.num]=1;else delete _fSel[o.num];}); rpFRender(); };
 window.rpFSelectAll=function(){ _fFiltered().forEach(function(o){_fSel[o.num]=1;}); rpFRender(); _fStat('✓ '+Object.keys(_fSel).length+' ventas seleccionadas.'); };
 window.rpFDeselect=function(){ _fSel={}; rpFRender(); _fStat('Selección limpia.'); };
 window.rpFFilt=function(f){ _fFilter=f; _fPage=1; var ps=document.querySelectorAll('#rp-fact-ov .pill'); for(var i=0;i<ps.length;i++){ps[i].classList.toggle('on',ps[i].getAttribute('data-f')===f);} rpFRender(); };
 window.rpFRango=function(){ _fPage=1; rpFLoad(); };
 window.rpFToggleSld=function(){ var s=document.getElementById('rpf-sld'); if(s)s.classList.toggle('open'); rpFPctInput(); };
 window.rpFPctInput=function(){ var el=document.getElementById('rpf-pct'); if(!el)return; var v=+el.value;
   var pv=document.getElementById('rpf-pctv'); if(pv)pv.textContent=v+'%';
   el.style.background='linear-gradient(90deg,#137fec '+v+'%,#182234 '+v+'%)';
   var per=_fRows, target=Math.round(v/100*per.length), yf=per.filter(function(o){return o.facturada;}).length, need=Math.max(0,target-yf);
   var pend=per.filter(function(o){return !o.facturada;}).sort(function(a,b){return (a.prio||2)-(b.prio||2);}).slice(0,need);
   var monto=pend.reduce(function(a,o){return a+o.total;},0), ob=document.getElementById('rpf-pctobj');
   if(ob)ob.innerHTML=need>0?('<span>Al <b>'+v+'%</b> facturarías</span><span class="hl">'+need+' ventas</span><span class="sep">·</span><span class="hl">'+_dFmt(monto)+'</span>'):('<span>Al <b>'+v+'%</b> ya está todo facturado —</span><span class="hl">0 ventas</span><span class="sep">·</span><span class="hl">$0</span>'); };
 function _fMarcar(nums,accion,cb){ fetch('/pf-facturacion-marcar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nums:nums,accion:accion,emit:_fEmit()})}).then(function(r){return r.json();}).then(function(j){ if(j&&j.ok){cb&&cb(j);} else _fStat('No se pudo guardar.','#fb7185'); }).catch(function(){ _fStat('Error de conexión.','#fb7185'); }); }
 window.rpFFacturarSel=function(){ var ids=Object.keys(_fSel); if(!ids.length){ _fStat('Tildá las ventas que querés facturar.','#fb7185'); return; } _fStat('Facturando…','#38bdf8'); _fMarcar(ids,'facturada',function(){ _fSel={}; rpFLoad(); _fStat('🧾 '+ids.length+' factura(s) emitida(s) con fecha de emisión '+(_fEmit()||'hoy')+'.'); }); };
 window.rpFFacturarPct=function(){ var v=+(document.getElementById('rpf-pct')||{}).value; if(!v){ _fStat('Elegí un % con la barra.','#fb7185'); return; }
   var per=_fRows, target=Math.round(v/100*per.length), yf=per.filter(function(o){return o.facturada;}).length, need=target-yf;
   if(need<=0){ _fStat('Ya facturaste el '+v+'% o más de este período.','#fbbf24'); return; }
   var pend=per.filter(function(o){return !o.facturada;}).sort(function(a,b){return (a.prio||2)-(b.prio||2);}).slice(0,need).map(function(o){return o.num;});
   _fStat('Facturando '+v+'%…','#38bdf8'); _fMarcar(pend,'facturada',function(){ rpFLoad(); _fStat('🧾 Facturado el '+v+'% — '+pend.length+' factura(s) nuevas.'); }); };
 window.rpFInforme=function(){ _fStat('⏳ Generando informe del contador…','#34d399');
   var d=(document.getElementById('rpf-desde')||{}).value, h=(document.getElementById('rpf-hasta')||{}).value;
   fetch('/pf-facturacion-informe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({desde:d,hasta:h})})
    .then(function(r){ if(!r.ok) return r.json().then(function(e){throw (e&&e.msg)||'error';}); return r.blob(); })
    .then(function(b){ var u=URL.createObjectURL(b); var a=document.createElement('a'); a.href=u; a.download='Informe-Facturacion.xlsx'; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(u); _fStat('✅ Informe descargado.'); })
    .catch(function(e){ _fStat('No se pudo generar'+(typeof e==='string'?': '+e:'')+'.', '#fb7185'); }); };
 function _fAutoState(){ var m=(document.getElementById('rpf-autodelay')||{}).value||'10', el=document.getElementById('rpf-autostate'); if(!el)return;
   el.innerHTML=_fAutoOn?('<span style="color:#34d399;font-weight:800">● ACTIVADA</span> — cada venta nueva se factura sola <b>a los '+m+' min</b>. Solo las que entran desde que la activaste.'):'Desactivada. Las ventas nuevas quedan <b style="color:#cbd5e1">pendientes</b> hasta que la actives.'; }
 window.rpFAuto=function(el){ _fAutoOn=el.checked;
   fetch('/pf-facturacion-auto',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({on:_fAutoOn,delay:+(document.getElementById('rpf-autodelay')||{}).value||10,emit:_fEmit()})})
    .then(function(r){return r.json();}).then(function(){ _fAutoState(); _fStat(_fAutoOn?'🤖 Facturación automática activada.':'Facturación automática desactivada.'); })
    .catch(function(){ _fStat('No se pudo guardar.','#fb7185'); }); };
 // ---- Vincular ARCA (candado) ----
 var _fMetodo='servicio';
 window.rpFSetMetodo=function(m){ _fMetodo=m;
   var s=document.getElementById('rpf-m-serv'), c=document.getElementById('rpf-m-cert');
   if(s){s.className='btn '+(m==='servicio'?'btn-primary':'btn-ghost');} if(c){c.className='btn '+(m==='cert'?'btn-primary':'btn-ghost');}
   var steps=document.getElementById('rpf-steps'), tk=document.getElementById('rpf-tk-lb'), inp=document.getElementById('rpf-in-token');
   if(m==='servicio'){
     if(tk)tk.textContent='Token del servicio'; if(inp)inp.placeholder='Pegá tu API token';
     steps.innerHTML='<div class="lk-step"><span class="n">1</span><div>Creá tu cuenta en <a href="https://www.tusfacturas.app" target="_blank" rel="noopener" style="color:#4aa8ff;font-weight:700">TusFacturas.app</a> o <a href="https://www.facturante.com" target="_blank" rel="noopener" style="color:#4aa8ff;font-weight:700">Facturante</a>.</div></div>'
       +'<div class="lk-step"><span class="n">2</span><div>En el panel, generá tu <b>API token</b> y copialo.</div></div>'
       +'<div class="lk-step"><span class="n">3</span><div>Pegá abajo tu <b>CUIT</b>, <b>punto de venta</b> y el <b>token</b>. Listo, sin manejar certificados.</div></div>';
   } else {
     if(tk)tk.textContent='Certificado (.crt / .pem)'; if(inp)inp.placeholder='Pegá el contenido del certificado';
     steps.innerHTML='<div class="lk-step"><span class="n">1</span><div>Entrá a <a href="https://www.afip.gob.ar" target="_blank" rel="noopener" style="color:#4aa8ff;font-weight:700">AFIP</a> con tu Clave Fiscal → <b>Administración de Certificados Digitales</b> y generá un certificado.</div></div>'
       +'<div class="lk-step"><span class="n">2</span><div>Asociale el web service <b>Facturación Electrónica (wsfe)</b>.</div></div>'
       +'<div class="lk-step"><span class="n">3</span><div>Configurá tu <b>Punto de Venta</b> como factura electrónica.</div></div>'
       +'<div class="lk-step"><span class="n">4</span><div>Pegá abajo el <b>certificado</b> + tu CUIT y punto de venta. <b style="color:#fbbf24">Nunca compartas tu Clave Fiscal</b>, solo el certificado.</div></div>';
   } };
 window.rpFOpenLink=function(){ var m=document.getElementById('rpf-linkmodal'); if(!m)return; m.style.display='flex';
   var a=window._fArca||{}; var g=function(id,v){var e=document.getElementById(id); if(e&&v)e.value=v;};
   g('rpf-in-cuit',(a.cuit||'').replace(/\D/g,'')); g('rpf-in-pv',a.pv||''); g('rpf-in-nombre',a.nombre||'');
   document.getElementById('rpf-link-status').textContent=''; rpFSetMetodo(_fMetodo); };
 window.rpFCloseLink=function(){ var m=document.getElementById('rpf-linkmodal'); if(m)m.style.display='none'; };
 window.rpFVincular=function(){ var st=document.getElementById('rpf-link-status');
   var cuit=(document.getElementById('rpf-in-cuit')||{}).value||'', pv=(document.getElementById('rpf-in-pv')||{}).value||'';
   var nombre=(document.getElementById('rpf-in-nombre')||{}).value||'', token=(document.getElementById('rpf-in-token')||{}).value||'';
   if(cuit.replace(/\D/g,'').length!==11){ st.style.color='#fb7185'; st.textContent='El CUIT debe tener 11 dígitos.'; return; }
   st.style.color='#38bdf8'; st.textContent='Verificando conexión…';
   fetch('/pf-facturacion-arca',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({cuit:cuit,pv:pv,nombre:nombre,token:token,metodo:_fMetodo})})
    .then(function(r){return r.json();}).then(function(j){ if(j&&j.ok){ st.style.color='#34d399'; st.textContent='✓ ARCA vinculado. Desbloqueando…';
      setTimeout(function(){ rpFCloseLink(); var lk=document.getElementById('rpf-lock'); if(lk)lk.style.display='none'; rpFLoad(); },700); }
      else { st.style.color='#fb7185'; st.textContent=(j&&j.msg)||'No se pudo vincular.'; } })
    .catch(function(){ st.style.color='#fb7185'; st.textContent='Error de conexión.'; }); };
 function rpComisLoad(){ fetch('/pf-comisiones').then(function(r){return r.json();}).then(function(j){ var c=(j&&j.comis)||{};
    var set=function(id,v){var el=document.getElementById(id); if(el)el.value=(!v||v===0)?'':v;};
    // 1% tienda y 3,5% Ingresos Brutos: FIJOS y no editables (pedido del usuario).
    var fix=function(id,v){var el=document.getElementById(id); if(!el)return; el.value=v; el.readOnly=true; el.style.opacity='.65'; el.style.cursor='not-allowed'; el.oninput=null; el.title='Fijo, no editable';};
    fix('rp-c-tienda',1); fix('rp-c-iibb',3.5);
    rpComisTotal();
   }).catch(function(){ rpComisTotal(); }); }
 window.rpComisSave=function(){ var g=function(id){var el=document.getElementById(id); return el?(parseFloat(el.value||'0')||0):0;};
   var body={mp_comision:g('rp-c-mp'),mp_cuotas:g('rp-c-cuotas'),comision_tienda:g('rp-c-tienda'),iva:g('rp-c-iva'),ingresos_brutos:g('rp-c-iibb')};
   var go=document.getElementById('rp-c-go'),msg=document.getElementById('rp-c-msg'); if(go){go.disabled=true; go.textContent='Guardando...';}
   fetch('/pf-comisiones',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(function(r){return r.json();}).then(function(j){ if(go){go.disabled=false; go.textContent='Guardar';} if(msg){msg.style.display='inline'; msg.style.color='#34d399'; msg.textContent='¡Guardado!'; setTimeout(function(){msg.style.display='none';},1500);} rpComisTotal(); }).catch(function(){ if(go){go.disabled=false; go.textContent='Guardar';} if(msg){msg.style.display='inline'; msg.style.color='#f87171'; msg.textContent='Error, probá de nuevo';} }); };
 function rpProdWarn(){ var warn=document.getElementById('rp-prod-warn'); if(!warn)return; var n=document.querySelectorAll('#rp-prod-body .rp-sincosto').length;
  warn.innerHTML = n>0 ? '<div style="display:flex;align-items:center;gap:14px;background:#1c1608;border:1px solid #4a3a1a;border-radius:12px;padding:13px 16px"><span style="font-size:18px">&#9888;&#65039;</span><div style="flex:1"><b style="color:#f1f5f9;font-size:13.5px">'+n+' '+(n===1?'producto':'productos')+' sin costo cargado</b><div style="color:#c9a35b;font-size:12.5px;margin-top:2px">Su ganancia se calcula de m&aacute;s. Carg&aacute; el costo para que el margen sea real.</div></div></div>' : ''; }
 window.rpSkuSave=function(id){
  var t=(document.getElementById('rp-sku-t-'+id)||{}).value||'xn';
  var bi=document.getElementById('rp-sku-b-'+id); var b=bi?(bi.value||'').trim():'';
  if(bi){ bi.disabled=(t==='spray'); bi.style.opacity=(t==='spray')?'.4':'1'; bi.placeholder=(t==='fijo'?'código fijo':'nombre (ej: pote)'); }
  fetch('/pf-sku-set',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pid:id,tipo:t,base:b})})
   .then(function(r){return r.json();}).then(function(j){ var ej=document.getElementById('rp-sku-ej-'+id);
     if(ej&&j&&j.ejemplos){ ej.textContent='1→'+(j.ejemplos['1']||'—')+' · 2→'+(j.ejemplos['2']||'—')+' · 3→'+(j.ejemplos['3']||'—'); ej.style.color='#5b6b82'; } }).catch(function(){}); };
 window.rpSaveCosto=function(inp,id){ var v=parseFloat(String(inp.value||'').replace(/\./g,'').replace(',','.'))||0;
  fetch('/pf-guardar-costo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:id,costo:v})}).catch(function(){});
  var row=inp.closest('tr'); var b=row&&row.querySelector('.rp-badge');
  if(b){ if(v>0){ b.className='rp-badge rp-concosto'; b.style.cssText='background:#0e2a1c;border:1px solid #17492f;color:#34d399;border-radius:7px;padding:3px 9px;font-size:11.5px;font-weight:600'; b.innerHTML='&#10003; Cargado'; } else { b.className='rp-badge rp-sincosto'; b.style.cssText='background:#2a2210;border:1px solid #4a3a1a;color:#f0c674;border-radius:7px;padding:3px 9px;font-size:11.5px;font-weight:600'; b.innerHTML='&#9888;&#65039; Sin costo'; } }
  rpProdWarn(); };
 // Cambio de tipo (Unitario/Variable/Fijo): guarda el tipo y recarga para mostrar el editor correcto.
 window.rpSkuTipoChg=function(id){
  var t=(document.getElementById('rp-sku-t-'+id)||{}).value||'unitario';
  var be=document.getElementById('rp-sku-b-'+id); var base=be?be.value:'';
  fetch('/pf-sku-set',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pid:id,tipo:t,base:base})}).then(function(){ rpProdLoad(); }).catch(function(){ rpProdLoad(); }); };
 // Variable: guarda el SKU de cada cantidad (1 a 4).
 window.rpSkuVarSave=function(id){
  var map={}; for(var q=1;q<=4;q++){ var e=document.getElementById('rp-skum-'+id+'-'+q); if(e&&e.value.trim()) map[q]=e.value.trim(); }
  fetch('/pf-sku-set',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pid:id,tipo:'variable',map:map})}).catch(function(){}); };
 // Variable: guarda el COSTO de cada cantidad (1 a 4). El margen usa el costo exacto de esa cantidad.
 window.rpCostoVarSave=function(id){
  var g=function(x){ var e=document.getElementById(x); return e?(parseFloat(String(e.value||'').replace(/\./g,'').replace(',','.'))||0):0; };
  var costos={}, algo=false; for(var q=1;q<=4;q++){ var v=g('rp-cvar-'+id+'-'+q); if(v>0){ costos[q]=v; algo=true; } }
  fetch('/pf-guardar-costo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:id,variable:1,costos:costos})}).catch(function(){});
  var el=document.getElementById('rp-cvar-'+id+'-1'), row=el&&el.closest('tr'), b=row&&row.querySelector('.rp-badge');
  if(b){ if(algo){ b.className='rp-badge rp-concosto'; b.style.cssText='background:#0e2a1c;border:1px solid #17492f;color:#34d399;border-radius:7px;padding:3px 9px;font-size:11.5px;font-weight:600'; b.innerHTML='&#10003; Cargado'; } else { b.className='rp-badge rp-sincosto'; b.style.cssText='background:#2a2210;border:1px solid #4a3a1a;color:#f0c674;border-radius:7px;padding:3px 9px;font-size:11.5px;font-weight:600'; b.innerHTML='&#9888;&#65039; Sin costo'; } }
  rpProdWarn(); };
 function rpProdLoad(){ var chip=document.getElementById('rp-prod-chip'), warn=document.getElementById('rp-prod-warn'), body=document.getElementById('rp-prod-body'); if(!body)return;
  body.innerHTML='<div style="color:#94a3b8;font-size:13px;padding:24px 4px">Cargando...</div>'; warn.innerHTML=''; chip.innerHTML='';
  fetch('/pf-productos').then(function(r){return r.json();}).then(function(j){
   var ps=(j&&j.productos)||[]; var tienda=j&&j.tienda;
   if(!tienda){ chip.innerHTML=''; warn.innerHTML=''; body.innerHTML='<div style="text-align:center;color:#94a3b8;font-size:13.5px;padding:46px 12px;border:1px dashed #24354c;border-radius:12px;margin-top:8px">A&uacute;n no conectaste una tienda.<br><span style="font-size:12.5px;line-height:1.9">Conect&aacute; <b style="color:#cbd5e1">Shopify</b> o <b style="color:#cbd5e1">Tiendanube</b> en Integraciones y ac&aacute; van a aparecer tus productos.</span></div>'; return; }
   chip.innerHTML='<span style="display:inline-flex;align-items:center;gap:9px;background:#0d1826;border:1px solid #29527e;color:#e2e8f0;border-radius:12px;padding:9px 14px;font-weight:600;font-size:13.5px">'+(tienda==='Shopify'?L.shopify:L.tn)+'<span>'+tienda+'</span><span style="background:#152238;border:1px solid #23344d;color:#93c5fd;border-radius:20px;padding:1px 9px;font-size:12px;font-weight:700">'+ps.length+'</span><span style="width:8px;height:8px;border-radius:50%;background:#22c55e"></span></span>';
   if(!ps.length){ warn.innerHTML=''; body.innerHTML='<div style="color:#94a3b8;font-size:13px;padding:24px 4px">Tu tienda no tiene productos cargados todav&iacute;a.</div>'; return; }
   var h='<table style="width:100%;border-collapse:collapse;margin-top:14px"><thead><tr><th style="text-align:left;color:#94a3b8;font-size:12px;font-weight:600;padding:9px 10px;border-bottom:1px solid #1a2636">Producto</th><th style="text-align:right;color:#94a3b8;font-size:12px;font-weight:600;padding:9px 10px;border-bottom:1px solid #1a2636">Precio de venta</th><th style="text-align:right;color:#94a3b8;font-size:12px;font-weight:600;padding:9px 10px;border-bottom:1px solid #1a2636">SKU</th><th style="text-align:right;color:#94a3b8;font-size:12px;font-weight:600;padding:9px 10px;border-bottom:1px solid #1a2636">Costo</th></tr></thead><tbody>';
   ps.forEach(function(p){ var img=p.img?'<img src="'+p.img+'" style="width:100%;height:100%;object-fit:cover">':'&#128247;';
    var esVar=(p.sku_tipo==='variable');
    var cmap=(p.costo&&typeof p.costo==='object')?p.costo:{};
    var smap=p.sku_map||{};
    var _cnum=(p.costo&&typeof p.costo==='object')?(Number(p.costo['1'])||0):(Number(p.costo)||0);
    var tiene=esVar?['1','2','3','4'].some(function(k){return Number(cmap[k])>0;}):(_cnum>0);
    var badge= tiene ? '<span class="rp-badge rp-concosto" style="background:#0e2a1c;border:1px solid #17492f;color:#34d399;border-radius:7px;padding:3px 9px;font-size:11.5px;font-weight:600">&#10003; Cargado</span>' : '<span class="rp-badge rp-sincosto" style="background:#2a2210;border:1px solid #4a3a1a;color:#f0c674;border-radius:7px;padding:3px 9px;font-size:11.5px;font-weight:600">&#9888;&#65039; Sin costo</span>';
    var sIn='background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:9px;padding:9px 12px;font-size:13px;text-align:right';
    var sSm='background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:8px;padding:7px 9px;font-size:12.5px';
    var skuC='<div style="display:flex;flex-direction:column;gap:6px;align-items:flex-end">'
      +'<select id="rp-sku-t-'+p.id+'" onchange="rpSkuTipoChg(\''+p.id+'\')" style="background:#0b1220;border:1px solid #1e2b3d;color:#f1f5f9;border-radius:9px;padding:8px 8px;font-size:12.5px;cursor:pointer">'
       +'<option value="unitario"'+(p.sku_tipo==='unitario'?' selected':'')+'>Unitario</option>'
       +'<option value="variable"'+(p.sku_tipo==='variable'?' selected':'')+'>Variable</option>'
       +'<option value="fijo"'+(p.sku_tipo==='fijo'?' selected':'')+'>Fijo</option>'
      +'</select>';
    if(esVar){ for(var q=1;q<=4;q++){ var ejs=(q===1?'1 30ml':q===2?'1 60ml':q===3?'1 30ml + 1 60ml':'2 60ml'); skuC+='<div style="display:flex;gap:6px;align-items:center"><span style="color:#5b6b82;font-size:10.5px;width:14px;text-align:right">'+q+'</span><input id="rp-skum-'+p.id+'-'+q+'" value="'+esc(smap[q]||'')+'" placeholder="'+ejs+'" onchange="rpSkuVarSave(\''+p.id+'\')" style="width:132px;'+sSm+';text-align:left"></div>'; } }
    else { skuC+='<input id="rp-sku-b-'+p.id+'" value="'+esc(p.sku_base||'')+'" placeholder="'+(p.sku_tipo==='fijo'?'código fijo':'nombre (ej: Pote)')+'" onchange="rpSkuSave(\''+p.id+'\')" style="width:120px;'+sIn+'"><span id="rp-sku-ej-'+p.id+'" style="color:#5b6b82;font-size:11px">ej: 2&rarr;'+esc(p.sku_ej||'')+'</span>'; }
    skuC+='</div>';
    var costC;
    if(esVar){ costC='<div style="display:flex;flex-direction:column;gap:6px;align-items:flex-end"><div style="height:32px"></div>'; for(var w=1;w<=4;w++){ var cv=Number(cmap[w])||0; costC+='<input id="rp-cvar-'+p.id+'-'+w+'" value="'+(cv?cv.toLocaleString('es-AR'):'')+'" placeholder="0" onchange="rpCostoVarSave(\''+p.id+'\')" style="width:94px;'+sSm+';text-align:right">'; } costC+='</div>'; }
    else { costC='<input value="'+(tiene?_cnum.toLocaleString('es-AR'):'')+'" placeholder="0" onchange="rpSaveCosto(this,\''+p.id+'\')" style="width:110px;'+sIn+'">'; }
    h+='<tr><td style="padding:14px 10px;border-bottom:1px solid #141f2e;vertical-align:top"><div style="display:flex;align-items:center;gap:13px"><div style="width:44px;height:44px;border-radius:9px;background:#101c2e;border:1px solid #1e2b3d;flex:none;overflow:hidden;display:flex;align-items:center;justify-content:center">'+img+'</div><div style="min-width:0"><div style="font-weight:600;color:#f1f5f9;font-size:14px">'+esc(p.nombre)+'</div><div style="margin-top:5px">'+badge+'</div></div></div></td>'
    +'<td style="padding:14px 10px;border-bottom:1px solid #141f2e;text-align:right;color:#f1f5f9;font-weight:600;font-size:14px;vertical-align:top">'+(p.precio?('$ '+Number(p.precio).toLocaleString('es-AR')):'&mdash;')+'</td>'
    +'<td style="padding:14px 10px;border-bottom:1px solid #141f2e;text-align:right;vertical-align:top">'+skuC+'</td>'
    +'<td style="padding:14px 10px;border-bottom:1px solid #141f2e;text-align:right;vertical-align:top">'+costC+'</td></tr>'; });
   h+='</tbody></table>'; body.innerHTML=h; rpProdWarn();
   var rf=document.getElementById('rp-prod-ref'); if(rf) rf.onclick=rpProdLoad;
  }).catch(function(){ body.innerHTML='<div style="color:#f87171;font-size:13px;padding:24px 4px">No se pudieron cargar los productos.</div>'; }); }
 if(new URLSearchParams(location.search).get('integ')==='1'){ try{history.replaceState({},'','/');}catch(e){} var _n=0,_t=setInterval(function(){ _n++; var o=document.getElementById('rp-integ-ov'); if(o){ window.rpInteg(true); o.style.display='block'; } if(_n>50)clearInterval(_t); },300); }
})();
</script>
<script>
/* Parche breakeven: el dashboard compilado calcula mal el Break Even ROAS (0.00x) y CPA
   (costos viejos). Sobreescribo esas 2 tarjetas con los valores reales del backend
   (be_roas = facturación / contribución antes de ads · be_cpa = contribución antes de ads / pedidos). */
(function(){
  var _raw=null, _of=window.fetch;
  window.fetch=function(){ var args=arguments, p=_of.apply(this,args);
    try{ var u=(args[0]&&args[0].url)||args[0];
      if(typeof u==='string' && u.indexOf('/pf-periodo')>-1){
        p.then(function(res){ try{ res.clone().json().then(function(j){ var r=(j&&j.raw)||j;
          if(r && (r.be_cpa!=null || r.be_roas!=null)){ _raw=r; setTimeout(paint,80); setTimeout(paint,450); pedirRecompras(r.desde,r.hasta); } }).catch(function(){}); }catch(e){} });
      } }catch(e){}
    return p; };
  // Recompras: se piden APARTE (el histórico es pesado y frenaba el dashboard). Se rellenan al llegar.
  var _recKey='';
  function pedirRecompras(d,h){ if(!d||!h) return; var k=d+'|'+h; if(k===_recKey) return; _recKey=k;
    fetch('/pf-recompras?desde='+encodeURIComponent(d)+'&hasta='+encodeURIComponent(h)).then(function(r){return r.json();}).then(function(j){
      if(j&&j.ok&&_raw){ _raw.recompras=j.recompras; _raw.fact_recompra=j.fact_recompra; try{paint();}catch(e){} setTimeout(paint,200); }
    }).catch(function(){ _recKey=''; }); }
  function money(n){ try{ return '$'+Math.round(n).toLocaleString('es-AR'); }catch(e){ return '$'+Math.round(n); } }
  function set(label,text){ var all=document.querySelectorAll('span');
    for(var i=0;i<all.length;i++){ if((all[i].textContent||'').trim()===label){ var box=all[i].parentElement; if(!box)continue;
      var v=box.nextElementSibling; while(v && !(/font-bold/.test(v.className||''))) v=v.nextElementSibling;
      if(v && v.textContent!==text) v.textContent=text; } } }
  // Chip de canal ACTIVO (Todas / Shopify / MercadoLibre). El activo lleva la clase text-primary/bg-primary.
  function _canalActivo(){ var bs=document.querySelectorAll('button');
    for(var i=0;i<bs.length;i++){ var t=(bs[i].textContent||'').replace(/\s+/g,' ').trim();
      if((t==='Todas'||t==='Shopify'||t==='MercadoLibre') && /text-primary|bg-primary/.test(bs[i].className||'')) return t; }
    return 'Todas'; }
  function _ceroRaw(){ var z={}; for(var k in _raw){ z[k]=(typeof _raw[k]==='number')?0:_raw[k]; } return z; }
  function paint(){ if(!_raw)return;
    // MercadoLibre no tiene ventas (todo es Shopify) → paso datos en CERO para no pintar los de Shopify.
    var meli=(_canalActivo()==='MercadoLibre'); var save=_raw; if(meli) _raw=_ceroRaw();
    try{ costos4(); }catch(e){}
    try{ metricas(); }catch(e){}
    _raw=save;
    try{ fixFacturacion(); }catch(e){} }
  // El KPI 'Facturación' en prod lee un campo que a veces llega en 0 (aunque tot_facturado esté bien).
  // Lo forzamos SIEMPRE al valor real del resumen. Se re-aplica tras cada poll (paint 80/450ms) → aguanta a React.
  // Busca (UNA vez) el elemento hoja del monto del KPI 'Facturación' VISIBLE. Ignora la pestaña oculta del
  // gráfico killeado (que también dice 'Facturación'). Es el scan caro (lee offsetParent) → se cachea.
  function _findFactEl(){
    var all=document.querySelectorAll('span,p,div');
    for(var i=0;i<all.length;i++){ var e=all[i], tx=(e.textContent||'').replace(/\s+/g,' ').trim();
      if(tx.length>34 || tx.slice(-11)!=='Facturación') continue;
      if(e.offsetParent===null) continue;
      var card=e; for(var k=0;k<9&&card;k++){ card=card.parentElement; if(card&&/rounded/.test(card.className||'')) break; }
      if(!card||!/rounded/.test(card.className||'')||card.offsetParent===null) continue;
      var dvs=card.querySelectorAll('span,div');
      for(var q=0;q<dvs.length;q++){ if(dvs[q].children.length===0){ var vt=(dvs[q].textContent||'').trim();
        if(/^\$\s?-?[\d.,]+$/.test(vt)) return dvs[q]; } } }
    return null; }
  var _factEl=null;
  function fixFacturacion(){ if(!_raw) return;
    if(_canalActivo()==='MercadoLibre') return;   // canal sin ventas → dejar el $0 nativo, no forzar
    var real=_raw.tot_facturado||_raw.facturado||0;
    if(!real && _raw.iibb_monto) real=Math.round(_raw.iibb_monto*100/3.5);   // último recurso: derivar de IIBB (3,5%)
    if(!real) return;
    if(!_factEl || !_factEl.isConnected) _factEl=_findFactEl();              // scan caro SOLO si hace falta
    if(_factEl){ var t=fmt(real); if(_factEl.textContent!==t) _factEl.textContent=t; } }
  function num(n){ return (Math.round((n||0)*100)/100); }
  function num(n){ return (Math.round((n||0)*100)/100); }
  function esHeader(el){ return !!(el && /col-span-full/.test(el.className||'')); }
  // La grilla es PLANA: headers (Finanzas/Publicidad/Costos) son divs col-span-full y las tarjetas son hermanas.
  function findGrid(){ var sp=document.querySelectorAll('span');
    for(var i=0;i<sp.length;i++){ var cn=sp[i].className||''; if((sp[i].textContent||'').trim()==='Publicidad' && /uppercase/.test(cn) && /tracking/.test(cn)){
      var hdr=sp[i]; for(var k=0;k<6 && hdr;k++){ hdr=hdr.parentElement; if(esHeader(hdr)) break; }
      if(esHeader(hdr) && hdr.parentElement) return hdr.parentElement; } }
    return null; }
  function metricas(){ if(!_raw)return; var grid=findGrid(); if(!grid) return;
    // Recolecto las tarjetas de la sección PUBLICIDAD (entre su header y el próximo) y remapeo POR POSICIÓN.
    var inPub=false, cards=[], kids=grid.children;
    for(var i=0;i<kids.length;i++){ var el=kids[i];
      if(esHeader(el)){ inPub = (el.textContent||'').indexOf('Publicidad')>=0; continue; }
      if(inPub){ var c=/rounded-2xl/.test(el.className||'')?el:(el.querySelector?el.querySelector('[class*=\"rounded-2xl\"]'):null); if(c) cards.push(c); } }
    if(!cards.length) return;
    var seq=[['Gasto Ads',money(_raw.publi_ars||0),'Inversión en anuncios'],
             ['Margen',num(_raw.margen)+'%','Ganancia ÷ facturación'],
             ['ROAS',num(_raw.roas)+'x','Recuperás por cada $1 invertido'],
             ['Break Even ROAS',num(_raw.be_roas)+'x','Mínimo para no perder'],
             ['CPA',money(_raw.cpa||0),'Costo por cada venta'],
             ['Break Even CPA',money(_raw.be_cpa||0),'Tope por venta'],
             ['Recompras',String(_raw.recompras||0),'Clientes que recompraron'],
             ['Facturación Recompra',money(_raw.fact_recompra||0),'Ventas de clientes que volvieron']];
    for(var j=0;j<cards.length && j<seq.length;j++) setCard(cards[j], seq[j][0], seq[j][1], seq[j][2]); }
  // ESTRUCTURA (no depende de los datos → corre de entrada, evita el parpadeo):
  // esconde la sección FINANZAS entera y las tarjetas de PUBLICIDAD sobrantes (Reembolsos, etc.).
  function estructura(){ var grid=findGrid(); if(!grid) return; var kids=grid.children, sec='', pub=0;
    for(var i=0;i<kids.length;i++){ var el=kids[i], tgt='';
      if(esHeader(el)){ sec=el.textContent||''; pub=0; tgt=/Finanzas/.test(sec)?'none':''; if(el.style.display!==tgt) el.style.display=tgt; continue; }
      if(/Finanzas/.test(sec)) tgt='none';                       // tarjetas de Finanzas → fuera
      else if(/Publicidad/.test(sec)){ pub++; tgt = (pub>8)?'none':''; }  // más de 8 (Reembolsos) → fuera
      if(el.style.display!==tgt) el.style.display=tgt; }
    // Barrido: cualquier tarjeta 'Reembolsos / cancel.' que haya quedado suelta → ocultar (no va en el diseño).
    var sp=document.querySelectorAll('span,div');
    for(var q=0;q<sp.length;q++){ var tx=(sp[q].textContent||'').replace(/\s+/g,' ').trim();
      if(tx==='Reembolsos / cancel.'){ var c=sp[q]; for(var k=0;k<9&&c;k++){ c=c.parentElement; if(c&&/rounded-2xl/.test(c.className||'')) break; }
        if(c&&/rounded-2xl/.test(c.className||'')&&c.style.display!=='none') c.style.display='none'; } } }
  function cardDe(label){ var sp=document.querySelectorAll('span');
    for(var i=0;i<sp.length;i++){ var s=sp[i], cn=s.className||'';
      // SOLO etiquetas de tarjeta (uppercase + tracking) → evita el sidebar y otros textos.
      if((s.textContent||'').trim()===label && /uppercase/.test(cn) && /tracking/.test(cn)){
        var card=s; for(var k=0;k<9 && card;k++){ card=card.parentElement; if(card && /rounded/.test(card.className||'')) break; }
        return (card && /rounded/.test(card.className||'')) ? card : null; } }
    return null; }
  function cardByAny(labels){ for(var i=0;i<labels.length;i++){ var c=cardDe(labels[i]); if(c) return c; } return null; }
  function costos4(){ if(!_raw)return;
    // SOLO la sección COSTOS: busco cada tarjeta por su etiqueta propia (nunca toca Finanzas).
    var viejo=document.getElementById('rp-costos4'); if(viejo) viejo.remove();
    var neto=(_raw.facturado||0)-(_raw.mp_costo_real||0)-(_raw.tienda_monto||0);
    var slots=[
      [['Costo producto','Productos'],'Productos',_raw.costo_prod,'Costo de los productos vendidos'],
      [['Comisiones','Envíos'],'Envíos',_raw.envio_monto,'Lo que pagás de envío'],
      [['Costo envíos','IIBB'],'IIBB',_raw.iibb_monto,'Impuesto a pagar al mes (3,5%)'],
      [['Logística','Neto por venta'],'Neto por venta',neto,'Lo que te queda tras MercadoPago + 1% tienda']
    ];
    for(var s=0;s<slots.length;s++){ var card=cardByAny(slots[s][0]);
      if(card){ if(card.style.display==='none') card.style.display=''; setCard(card, slots[s][1], money(slots[s][2]||0), slots[s][3]); } } }
  function setCard(card,label,val,sub){
    var sps=card.querySelectorAll('span'), lab=null;
    for(var i=0;i<sps.length;i++){ var cn=sps[i].className||''; if(/uppercase/.test(cn)&&/tracking/.test(cn)){ lab=sps[i]; break; } }
    if(lab && lab.textContent!==label) lab.textContent=label;
    var dvs=card.querySelectorAll('div'), v=null;
    for(var d=0;d<dvs.length;d++){ var cd=dvs[d].className||''; if(/font-bold/.test(cd)&&/(text-2xl|text-xl)/.test(cd)){ v=dvs[d]; break; } }
    if(v && v.textContent!==val) v.textContent=val;
    var ps=card.querySelectorAll('p'); if(ps.length){ var p=ps[ps.length-1]; if(p.textContent!==sub) p.textContent=sub; } }
  // ===== Modal de detalle de orden (al tocar una fila de "Últimas ventas") =====
  function esc(s){ return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }
  function fmt(n){ try{ return '$'+Math.round(n||0).toLocaleString('es-AR'); }catch(e){ return '$'+Math.round(n||0); } }
  function fechaTxt(f){ try{ var d=new Date(f); return d.toLocaleDateString('es-AR')+' '+d.toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit'}); }catch(e){ return f||''; } }
  window.cerrarOrden=function(){ var ov=document.getElementById('rp-orden-ov'); if(ov) ov.style.display='none'; };
  var _ordCache={};   // num -> orden ya traída (reabrir es instantáneo)
  function _ovBox(){ var ov=document.getElementById('rp-orden-ov');
    if(!ov){ ov=document.createElement('div'); ov.id='rp-orden-ov'; ov.style.cssText='position:fixed;inset:0;z-index:100000;background:rgba(3,7,12,.72);display:flex;align-items:center;justify-content:center;padding:20px'; ov.addEventListener('click',function(e){ if(e.target===ov) cerrarOrden(); }); document.body.appendChild(ov); }
    if(!document.getElementById('rp-spin-css')){ var st=document.createElement('style'); st.id='rp-spin-css'; st.textContent='@keyframes rpspin{to{transform:rotate(360deg)}}'; document.head.appendChild(st); }
    return ov; }
  window.abrirOrden=function(num){
    var ov=_ovBox(); ov.setAttribute('data-num',num); ov.style.display='flex';
    // Abre YA (con spinner) y llena cuando llega la data pesada de Shopify/MP.
    ov.innerHTML='<div style=\"width:100%;max-width:440px;background:#0b111b;border:1px solid #1e2b3d;border-radius:18px;padding:40px 18px;box-shadow:0 24px 60px rgba(0,0,0,.6);text-align:center;color:#93a3b8\"><div style=\"width:30px;height:30px;border:3px solid #1e2b3d;border-top-color:#5aa2f5;border-radius:50%;margin:0 auto 14px;animation:rpspin .7s linear infinite\"></div><div style=\"font-size:13px\">Cargando orden #'+esc(num)+'…</div></div>';
    if(_ordCache[num]){ _pintarOrden(num,_ordCache[num]); return; }
    fetch('/pf-orden?num='+encodeURIComponent(num)).then(function(r){return r.json();}).then(function(j){
      if(!j||!j.ok||!j.orden){ var ov2=document.getElementById('rp-orden-ov'); if(ov2&&ov2.getAttribute('data-num')===String(num)) ov2.innerHTML='<div style=\"max-width:420px;background:#0b111b;border:1px solid #1e2b3d;border-radius:18px;padding:28px;text-align:center;color:#93a3b8\">No se pudo cargar la orden. <span onclick=\"cerrarOrden()\" style=\"color:#5aa2f5;cursor:pointer\">Cerrar</span></div>'; return; }
      _ordCache[num]=j.orden; _pintarOrden(num,j.orden);
    }).catch(function(){ var ov2=document.getElementById('rp-orden-ov'); if(ov2&&ov2.getAttribute('data-num')===String(num)) ov2.innerHTML='<div style=\"max-width:420px;background:#0b111b;border:1px solid #1e2b3d;border-radius:18px;padding:28px;text-align:center;color:#93a3b8\">No se pudo cargar la orden. <span onclick=\"cerrarOrden()\" style=\"color:#5aa2f5;cursor:pointer\">Cerrar</span></div>'; }); };
  function _pintarOrden(num,o){
      var ov=document.getElementById('rp-orden-ov'); if(!ov||ov.getAttribute('data-num')!==String(num)) return;   // el usuario abrió otra / cerró
      var items=(o.items||[]).map(function(it){
        var foto = it.foto ? '<img src=\"'+it.foto+'\" style=\"width:46px;height:46px;border-radius:9px;object-fit:cover;background:#101c2e;border:1px solid #1e2b3d\">' : '<div style=\"width:46px;height:46px;border-radius:9px;background:#101c2e;border:1px solid #1e2b3d\"></div>';
        return '<div style=\"display:flex;align-items:center;gap:12px;padding:10px 0;border-top:1px solid #16202e\">'+foto+'<div style=\"flex:1;min-width:0\"><div style=\"color:#e7edf5;font-size:13px;font-weight:600;line-height:1.3\">'+esc(it.nombre)+'</div><div style=\"color:#7a8aa0;font-size:12px;margin-top:2px\">'+fmt(it.precio)+' × '+it.cantidad+'</div></div><div style=\"color:#e7edf5;font-weight:700;font-size:13.5px\">'+fmt(it.precio*it.cantidad)+'</div></div>'; }).join('');
      function fila(l,v,neg,strong){ return '<div style=\"display:flex;justify-content:space-between;align-items:center;padding:11px 0;border-top:1px solid #16202e\"><span style=\"color:'+(strong?'#e7edf5':'#93a3b8')+';font-size:13px'+(strong?';font-weight:700':'')+'\">'+l+'</span><span style=\"color:'+(strong?'#34d399':(neg?'#fb7185':'#e7edf5'))+';font-weight:'+(strong?'800':'600')+';font-size:'+(strong?'16px':'13.5px')+'\">'+v+'</span></div>'; }
      var body='<div style=\"display:flex;align-items:center;justify-content:space-between;gap:10px\">'
        +'<div style=\"display:flex;align-items:center;gap:11px\"><div style=\"width:38px;height:38px;border-radius:11px;background:#0d1b30;border:1px solid #1c3350;display:flex;align-items:center;justify-content:center;font-size:18px\">🛍️</div>'
        +'<div><div style=\"color:#f1f5f9;font-weight:800;font-size:15px\">Venta en '+esc(o.origen)+'</div><div style=\"color:#7a8aa0;font-size:12px\">Orden #'+esc(o.num)+'</div></div></div>'
        +'<div style=\"display:flex;align-items:center;gap:8px\"><span style=\"background:#0e2a1c;border:1px solid #17492f;color:#34d399;font-size:11.5px;font-weight:700;border-radius:20px;padding:3px 11px\">'+esc(o.estado)+'</span>'
        +'<button onclick=\"cerrarOrden()\" style=\"background:#111c2b;border:1px solid #1e2b3d;color:#cbd5e1;border-radius:9px;width:30px;height:30px;cursor:pointer;font-size:14px\">✕</button></div></div>'
        +'<div style=\"margin-top:14px;color:#93a3b8;font-size:12.5px;line-height:1.9\">🕐 '+fechaTxt(o.fecha)+'<br>👤 '+esc(o.cliente)+' <span style=\"color:#5b6b82\">'+esc(o.email)+'</span><br>💳 '+esc(o.medio)+'</div>'
        +'<div style=\"margin-top:14px;background:#0b111b;border:1px solid #1a2636;border-radius:12px;padding:2px 14px 12px\"><div style=\"color:#7a8aa0;font-size:11px;font-weight:700;letter-spacing:.6px;text-transform:uppercase;padding-top:10px\">Productos · '+(o.items||[]).length+' item'+((o.items||[]).length===1?'':'s')+'</div>'+items+'</div>'
        +'<div style=\"margin-top:14px;background:#0b111b;border:1px solid #1a2636;border-radius:12px;padding:2px 14px 12px\"><div style=\"color:#7a8aa0;font-size:11px;font-weight:700;letter-spacing:.6px;text-transform:uppercase;padding-top:10px\">Resumen financiero</div>'
        +fila('Total cobrado',fmt(o.total))
        +(o.fee_mp?fila('Comisión de pago · MercadoPago','-'+fmt(o.fee_mp),true):'')
        +(o.fee_cuotas?fila('Comisión de cuotas ('+o.cuotas+'x)','-'+fmt(o.fee_cuotas),true):'')
        +fila('Fee tienda (1%)','-'+fmt(o.fee_tienda),true)
        +fila('IIBB (3,5%)','-'+fmt(o.iibb),true)
        +(o.costo_prod?fila('Costo de productos','-'+fmt(o.costo_prod),true):fila('Costo de productos','cargá en Productos'))
        +fila('Envío'+(o.envio_real?'':' (aprox.)'),'-'+fmt(o.envio),true)
        +fila('Neto real',fmt(o.neto),false,true)+'</div>';
      ov.innerHTML='<div style=\"width:100%;max-width:440px;max-height:88vh;overflow:auto;background:#0b111b;border:1px solid #1e2b3d;border-radius:18px;padding:18px;box-shadow:0 24px 60px rgba(0,0,0,.6)\">'+body+'</div>';
      ov.style.display='flex';
  }
  // ===== Tabla "Últimas ventas" propia (la del dashboard es placeholder sin cablear) =====
  var _vt={dias:7,per:5,page:1,data:[],loading:false};
  function vtFecha(iso){ try{ var d=new Date(iso); return d.toLocaleDateString('es-AR',{day:'2-digit',month:'short'})+', '+d.toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit'}); }catch(e){ return iso||''; } }
  function cargarVentas(){ _vt.loading=true; renderVentas();
    fetch('/pf-ventas?dias='+_vt.dias).then(function(r){return r.json();}).then(function(j){ _vt.data=(j&&j.ventas)||[]; _vt.loading=false; _vt.page=1; renderVentas(); }).catch(function(){ _vt.loading=false; _vt.data=[]; renderVentas(); }); }
  function renderVentas(){ var tb=document.getElementById('rp-vt-body'); if(!tb) return;
    var all=_vt.data, tot=all.length, pages=Math.max(1,Math.ceil(tot/_vt.per));
    if(_vt.page>pages) _vt.page=pages; var ini=(_vt.page-1)*_vt.per, rows=all.slice(ini,ini+_vt.per);
    var tdb='padding:13px 10px;border-bottom:1px solid #131e2c;font-size:13px';
    if(_vt.loading) tb.innerHTML='<tr><td colspan=\"7\" style=\"color:#5b6b82;padding:20px 10px\">Cargando…</td></tr>';
    else if(!rows.length) tb.innerHTML='<tr><td colspan=\"7\" style=\"color:#5b6b82;padding:20px 10px\">Sin ventas en el período</td></tr>';
    else tb.innerHTML=rows.map(function(v){ var pend=(v.estado||'').toLowerCase().indexOf('pend')>=0;
      return '<tr style=\"cursor:pointer\" onmouseover=\"this.style.background=&#39;#0f1826&#39;\" onmouseout=\"this.style.background=&#39;&#39;\" onclick=\"abrirOrden(&#39;'+v.num+'&#39;)\">'
        +'<td style=\"'+tdb+';color:#cbd5e1;font-weight:700\">#'+esc(v.num)+'</td>'
        +'<td style=\"'+tdb+'\"><span style=\"background:#0d1b30;border:1px solid #1c3350;color:#9cc7f5;font-size:11.5px;font-weight:600;border-radius:16px;padding:3px 9px\">🛍️ '+esc(v.origen)+'</span></td>'
        +'<td style=\"'+tdb+'\"><span style=\"background:'+(pend?'#241a0e':'#0e2a1c')+';border:1px solid '+(pend?'#4a3a1a':'#17492f')+';color:'+(pend?'#f0b429':'#34d399')+';font-size:11px;font-weight:700;border-radius:14px;padding:2px 9px\">'+esc(v.estado)+'</span></td>'
        +'<td style=\"'+tdb+';color:#93a3b8\">'+vtFecha(v.fecha)+'</td>'
        +'<td style=\"'+tdb+';text-align:right\">'+fmt(v.total)+'</td>'
        +'<td style=\"'+tdb+';text-align:right;color:#34d399;font-weight:700\">'+fmt(v.neto)+'</td>'
        +'<td style=\"'+tdb+';text-align:right;color:#5aa2f5;font-size:12.5px\">ver ›</td></tr>'; }).join('');
    var cnt=document.getElementById('rp-vt-cnt'); if(cnt) cnt.textContent= tot? ((ini+1)+'–'+Math.min(ini+_vt.per,tot)+' de '+tot+' órdenes') : '0 órdenes';
    var pgn=document.getElementById('rp-vt-pgn'); if(pgn) pgn.textContent=_vt.page+' / '+pages;
    var db=document.querySelectorAll('#rp-vt-dias button'); for(var i=0;i<db.length;i++){ var on=(+db[i].getAttribute('data-d')===_vt.dias); db[i].style.background=on?'#5aa2f5':'transparent'; db[i].style.color=on?'#04121f':'#93a3b8'; }
    var pb=document.querySelectorAll('#rp-vt-pp button'); for(var j=0;j<pb.length;j++){ var o2=(+pb[j].getAttribute('data-n')===_vt.per); pb[j].style.background=o2?'#5aa2f5':'#0b1220'; pb[j].style.color=o2?'#04121f':'#93a3b8'; pb[j].style.borderColor=o2?'#5aa2f5':'#1e2b3d'; } }
  function vtCSV(){ var rows=[['N','Origen','Estado','Fecha','Total','Neto']].concat(_vt.data.map(function(v){return [v.num,v.origen,v.estado,v.fecha,v.total,v.neto];}));
    var csv=rows.map(function(r){return r.join(',');}).join('\n'); var a=document.createElement('a'); a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv); a.download='ventas.csv'; a.click(); }
  function tablaVentas(){
    var head=leafTxt('Últimas ventas');
    var mine=document.getElementById('rp-ventas-panel');
    if(mine){ if(head){ var pp=head; for(var k=0;k<10&&pp;k++){ pp=pp.parentElement; if(pp&&/rounded-2xl|rounded-xl/.test(pp.className||'')){ if(pp.style.display!=='none')pp.style.display='none'; break; } } } return; }
    if(!head) return;
    var panel=head; for(var k2=0;k2<10&&panel;k2++){ panel=panel.parentElement; if(panel&&/rounded-2xl|rounded-xl/.test(panel.className||'')) break; }
    if(!panel||!/rounded/.test(panel.className||'')) return;
    panel.style.display='none';
    var m=document.createElement('div'); m.id='rp-ventas-panel'; m.style.cssText='background:#0b111b;border:1px solid #1e2b3d;border-radius:18px;padding:18px';
    var TH='color:#5b6b82;font-size:10.5px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;padding:12px 10px;border-bottom:1px solid #1e2b3d';
    m.innerHTML='<div style=\"display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-bottom:6px\">'
      +'<div style=\"display:flex;align-items:center;gap:9px;font-weight:800;font-size:15px;color:#e7edf5\"><span style=\"width:26px;height:26px;border-radius:8px;background:#0d1b30;border:1px solid #1c3350;display:flex;align-items:center;justify-content:center;color:#5aa2f5\">🧾</span>Últimas ventas</div>'
      +'<div style=\"display:flex;align-items:center;gap:8px;flex-wrap:wrap\"><span id=\"rp-vt-csv\" style=\"color:#93a3b8;font-size:12.5px;cursor:pointer\">⬇ CSV</span>'
      +'<div id=\"rp-vt-dias\" style=\"display:flex;background:#0b1220;border:1px solid #1e2b3d;border-radius:10px;overflow:hidden\">'
      +['7','14','30','60','90'].map(function(d){return '<button data-d=\"'+d+'\" style=\"background:transparent;border:0;color:#93a3b8;padding:7px 12px;font-size:12.5px;font-weight:700;cursor:pointer\">'+d+'d</button>';}).join('')+'</div></div></div>'
      +'<div style=\"overflow-x:auto\"><table style=\"width:100%;border-collapse:collapse\"><thead><tr>'
      +'<th style=\"'+TH+';text-align:left\">Nº Venta</th><th style=\"'+TH+';text-align:left\">Origen</th><th style=\"'+TH+';text-align:left\">Estado</th><th style=\"'+TH+';text-align:left\">Fecha</th>'
      +'<th style=\"'+TH+';text-align:right\">Total</th><th style=\"'+TH+';text-align:right\">Neto</th><th style=\"'+TH+';text-align:right\">Acción</th>'
      +'</tr></thead><tbody id=\"rp-vt-body\"></tbody></table></div>'
      +'<div style=\"display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-top:12px;color:#93a3b8;font-size:12.5px\"><span id=\"rp-vt-cnt\">—</span>'
      +'<div style=\"display:flex;gap:16px;align-items:center;flex-wrap:wrap\"><div id=\"rp-vt-pp\" style=\"display:flex;align-items:center;gap:6px\"><span>Por página</span>'
      +['5','10','25','50'].map(function(n){return '<button data-n=\"'+n+'\" style=\"background:#0b1220;border:1px solid #1e2b3d;color:#93a3b8;border-radius:8px;padding:5px 11px;font-size:12.5px;font-weight:700;cursor:pointer\">'+n+'</button>';}).join('')+'</div>'
      +'<div style=\"display:flex;align-items:center;gap:8px\"><button id=\"rp-vt-prev\" style=\"background:#0b1220;border:1px solid #1e2b3d;color:#93a3b8;border-radius:8px;padding:5px 11px;cursor:pointer\">‹</button><span id=\"rp-vt-pgn\">1 / 1</span><button id=\"rp-vt-next\" style=\"background:#0b1220;border:1px solid #1e2b3d;color:#93a3b8;border-radius:8px;padding:5px 11px;cursor:pointer\">›</button></div></div></div>';
    panel.parentElement.insertBefore(m, panel.nextSibling);
    document.getElementById('rp-vt-dias').addEventListener('click',function(e){ var d=e.target.getAttribute&&e.target.getAttribute('data-d'); if(d){ _vt.dias=+d; cargarVentas(); } });
    document.getElementById('rp-vt-pp').addEventListener('click',function(e){ var n=e.target.getAttribute&&e.target.getAttribute('data-n'); if(n){ _vt.per=+n; _vt.page=1; renderVentas(); } });
    document.getElementById('rp-vt-prev').onclick=function(){ if(_vt.page>1){ _vt.page--; renderVentas(); } };
    document.getElementById('rp-vt-next').onclick=function(){ _vt.page++; renderVentas(); };
    document.getElementById('rp-vt-csv').onclick=vtCSV;
    cargarVentas();
  }
  // Layout FIJO: KPI (Ventas/Facturación/Ticket/Ganancia) arriba, después Resumen (Publicidad+Costos)
  // y Últimas ventas. Escondo gráfico, Recomendación y Riesgos. Uso CSS 'order' (no toca el DOM → no pelea con React).
  // Busca el título aunque tenga un ícono adelante (los headers son '<span>icono</span>Texto').
  function leafTxt(t){ var all=document.querySelectorAll('span,h2,h3,div,p'); for(var i=0;i<all.length;i++){ var e=all[i]; var tx=(e.textContent||'').replace(/\s+/g,' ').trim(); if(tx.length<=t.length+24 && tx.slice(-t.length)===t) return e; } return null; }
  function layoutFijo(){
    var rt=leafTxt('Resumen del período'), ut=leafTxt('Últimas ventas'); if(!rt||!ut) return;
    var anc=[], a=rt; while(a){ anc.push(a); a=a.parentElement; }
    var cont=ut; while(cont && anc.indexOf(cont)<0) cont=cont.parentElement; if(!cont) return;
    // Guarda: nunca reordenar un contenedor que incluya el header/chips (romperia la pagina y el chip "Todas").
    if(/Visión operativa|período seleccionado/.test(cont.textContent||'')) return;
    var kids=cont.children;
    for(var i=0;i<kids.length;i++){ var el=kids[i], t=el.textContent||'', ord=90, hide=false;
      if(/Facturaci/.test(t) && /Ganancia/.test(t) && /Ticket/.test(t)) ord=1;              // fila KPI
      else if(t.indexOf('Resumen del per')>=0) ord=2;                                        // Publicidad + Costos
      else if(t.indexOf('Últimas ventas')>=0) ord=3;                                         // tabla
      else if(el.querySelector && el.querySelector('[class*=\"rounded\"]')) hide=true;       // gráfico / Recomendación / Riesgos
      else { var inner=t.replace(/drag_indicator|Mover|visibility_off|visibility_on|visibility/gi,'').replace(/\s+/g,''); if(!inner) hide=true; }   // widget oculto/vacío (solo el handle) → colapsar
      if(el.style.order!==String(ord)) el.style.order=ord;
      var d=hide?'none':''; if(el.style.display!==d) el.style.display=d; }
    if(getComputedStyle(cont).display==='block') cont.style.display='flex', cont.style.flexDirection='column';
  }
  // FIX filtrado: cuando filtrás por tienda/canal, el widget de los 4 KPIs queda al fondo. Lo MOVEMOS
  // físicamente al tope de su lista (más confiable que el CSS order, que en la vista filtrada no engancha).
  // Ancla real = label con icono shopping_bag + "Ventas". Guarda: nunca tocamos el contenedor del header/chips.
  function kpiArriba(){
    var sp=document.getElementsByTagName('span'), cand=[];
    for(var i=0;i<sp.length;i++){ var t=(sp[i].textContent||'').replace(/\s+/g,' ').trim();
      if(t.length<20 && t.slice(-6)==='Ventas' && /shopping_bag/.test(sp[i].innerHTML)) cand.push(sp[i]); }
    if(!cand.length) return;
    var lab=null; for(var c=0;c<cand.length;c++){ if(cand[c].offsetParent!==null){ lab=cand[c]; break; } }
    if(!lab) lab=cand[0];
    var kpi=lab;                                        // subir al widget que agrupa los 4 KPIs
    for(var k=0;k<8;k++){ kpi=kpi.parentElement; if(!kpi) return; var tk=kpi.textContent||'';
      if(/Facturaci/.test(tk) && /Ganancia/.test(tk) && /Ticket/.test(tk)) break; }
    if(!kpi || !/Facturaci/.test(kpi.textContent||'')) return;
    // subir hasta el contenedor que TAMBIÉN contiene "Resumen del período" (= la lista de widgets).
    // Si topamos con el header/chips antes, abortamos (seguridad total).
    var lista=kpi.parentElement;
    while(lista){ var tl=lista.textContent||'';
      if(/período seleccionado|Visión operativa/.test(tl)) return;   // header → abort
      if(/Resumen del per/.test(tl)) break;                          // contenedor con KPI + Resumen
      lista=lista.parentElement; }
    if(!lista) return;
    var item=kpi; while(item && item.parentElement!==lista) item=item.parentElement;  // item del KPI (hijo directo)
    if(!item) return;
    if(lista.firstElementChild!==item){ try{ lista.insertBefore(item, lista.firstElementChild); }catch(e){} }  // KPI al tope
  }
  // Mientras arrastrás para reordenar ("Mover"), NO re-aplico nada (sino cancelo el drag).
  var _busy=false, _bt=null;
  document.addEventListener('pointerdown', function(){ _busy=true; }, true);
  document.addEventListener('pointerup', function(){ clearTimeout(_bt); _bt=setTimeout(function(){ _busy=false; try{ tick(); }catch(e){} }, 500); }, true);
  var _th=null;
  // Saca el chip "Mover / 👁" de arriba (el layout es fijo, no se reordena → ese control confunde). Una sola vez.
  var _movDone=false;
  function sacarMover(){ if(_movDone) return; var all=document.querySelectorAll('button,div,span,a');
    for(var i=0;i<all.length;i++){ var e=all[i], tx=(e.textContent||'').replace(/\s+/g,' ').trim();
      if(tx==='Mover'){ var box=e; for(var k=0;k<4&&box;k++){ if(box.parentElement){ var pt=(box.parentElement.textContent||'').replace(/\s+/g,' ').trim(); if(pt.length>14) break; box=box.parentElement; } else break; }
        if(box&&box.style.display!=='none') box.style.display='none'; _movDone=true; } } }
  // Mata el gráfico "Evolución" (muestra data demo falsa) — INDEPENDIENTE del layoutFijo, por si el pf.html
  // de producción difiere y layoutFijo no llega a esconderlo. Busca la tarjeta que contiene "Evolución".
  var _grafDone=false;
  function matarGrafico(){ if(_grafDone) return; var t=leafTxt('Evolución'); if(!t) return;
    var c=t; for(var k=0;k<10&&c;k++){ c=c.parentElement; if(c&&/rounded-2xl|rounded-xl/.test(c.className||'')) break; }
    if(c&&/rounded/.test(c.className||'')){ if(c.style.display!=='none') c.style.display='none'; _grafDone=true; } }
  function tick(){ if(_busy) return; try{ sacarMover(); }catch(e){} try{ matarGrafico(); }catch(e){} try{ layoutFijo(); }catch(e){} try{ kpiArriba(); }catch(e){} try{ estructura(); }catch(e){} try{ tablaVentas(); }catch(e){} if(_raw){ try{ paint(); }catch(e){} } }
  function schedule(){ if(_busy||_th) return; _th=setTimeout(function(){ _th=null; tick(); }, 220); }   // throttle: no en cada mutación
  try{ new MutationObserver(schedule).observe(document.body,{childList:true,subtree:true}); }catch(e){}
  [0,150,350,700,1300,2600].forEach(function(ms){ setTimeout(tick, ms); });   // arranques rápidos → sin parpadeo de Finanzas
  setInterval(function(){ if(_raw && !_busy){ try{ fixFacturacion(); }catch(e){} } }, 1200);   // Facturación: auto-repara si React la resetea
  // Al tocar un chip de canal (Todas/Shopify/MercadoLibre) React re-renderiza → re-aplico mis parches.
  document.addEventListener('click', function(e){ var el=e.target;
    for(var k=0;k<4&&el;k++){ var t=(el.textContent||'').replace(/\s+/g,' ').trim();
      if(t==='Todas'||t==='Shopify'||t==='MercadoLibre'){ _factEl=null; [120,400,800,1400,2200].forEach(function(ms){ setTimeout(function(){ _busy=false; try{ tick(); }catch(e){} }, ms); }); break; }
      el=el.parentElement; } }, true);
})();
</script>

<script>
/* RealProfit — ÍCONO + TEXTO de tienda del dashboard reflejan lo realmente conectado.
   MÉTODO: SOLO CSS (content:url / font-size:0 + ::after / display:none). NO se muta el DOM de React →
   (1) el chip de canal sigue 100% clickeable (no le tocamos el botón ni su texto), y
   (2) ícono/texto NO vuelven a Shopify en cada re-render (CSS pisa el pixel, React no lo revierte).
   Shopify -> icono+texto Shopify · Tiendanube -> icono+texto Tiendanube · sin tienda -> "Sin tienda".
   Además se ocultan las notificaciones de venta (#mfy_toasts) — RealProfit se actualiza solo, sin avisos. */
(function(){
  var SIG='LjYyIDQuMzR6IiBmaWxsPSIjZmZmIi8+PC9zdmc+';   // trozo del base64 del icono Shopify (los 16 usos)
  var TN_URI='data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAADAAAAAwCAYAAABXAvmHAAAHq0lEQVR4nM1aeVRUVRj/3TczDCAyIAKBGKTWgIhWFi7IGJ52FXHJorB9OcfMUsuOp90MPVJpaXVaTsvJIssC2ss8pkMnAy01FMItFyBl34ZtZm7n3sc8ZnDmvQdME79/eI93373fd++3/L7vDYE3MG93FGxCGggmA3YjgFEACQMQ1D2iBaC1AI6DkjKA/AaNfSe+MFUNdGnS7zdv2h2OLuE2gC4CcHk/Z9kLYAt09GN8Zqr2jQKzCy+Exv4YQO4BEADvoA0g74DaclAw/fR/o8D9e3U417YYwBon0/A22gCsR4h1Hd5Pa/eeArMLE6GlW0GRCN+gBJTejALTYaWBguJUGeZMaGiRD4VnGAdCipBRuBADUiCjcAl3MiAQvscQgH6CDPOK/pnQHPNiELyGwQBKl6PAtEG9AsxsxJ1XNjHfwA6QTORP+1RZAeawzOb/H7ORQysoTe7t2K47fMO3eh5tBp/wDENAhE9w0yE/eFRAH7zSx9Gmj6BJ6Kpf4d6EeIalZV7MrhyTLxmK9OQwpCUZMDJMj4gQHepbrDjb0Il9x1rwZXEdfvijHpYOu9opLRCsRnyRdsZVgTnmzSB40FuCXzXOgOysOEwxDlUce66xC6u3nsJbP/6DLhtVnpziVRSkPtyjACdm5GRfdt8QqEVCTAAiQ/xQ32rF6ZoOnDjbDq2GYP3tcViWPsJlfEOrFYdOW1Dd2IVhQVpcGK5HXIS/yxh2IhlrS3GmtkNp+TboaCwjgFp+24ksEHXCpycPw8OzomEaa+DCOqO8so0LmnyxuOuUAvm/1WLj15X4pbQJNrvr7saPCMAdMyKxdGY0AvUCJo4OQlHOBFz7bAlKTlnkxAhAF24BsEmUIMO8T4kSDw/WIXe5EVdPCFGjJ2qbrch8qQzbDzQojo0e5sfnNiUa+D07yeSVB1DT1CXzFi1GvimZ8GLETirlFhgR5gfzC+NxUaR45Gwjt++vx08HG1BZ14mhARq+ewumDEdokHiozJYX5pTxE1ADPy3Bx8uNmD9lOL//Zl8dZq2R5XIUOms0wdzC20Apy7pu4a8TYM4ejyvGBEl2euer5W6POMhfg+ysWDw0M5rfd1opUlYdwN6jLaqUCNQLKMwej8tGiWtd/UwJdhyUOUFCMwVQTJKbdMnMKEl48+EmmJ446NE+W9ptWPrOcaz84G9pV19/YAwElVWHpcOOxW8e477D8NTCkfIvUDJZAGi8p+dMgFXzxUkaLVYsWF+qKl7n5J/BV8V1/PrKMUG4/vJQdRoA2FPejB/31/Pr1LEGRBh0MgrAyDLxaE/P05JCeMhj2PRNFY/XavF0LovKIhamhKMv2Par6Dfs5GSVJ3QMU8BjWEmJD5au8/aoc0YH9p9oxbF/xKpw0iXKycwZLOQ6MPoC11zRC6GCXH3Loo8DR6pYudo3HO1+Z6TTPGpQWdcpXUeHyr1Lhor24QF2p8RDiOfEVrBqLL++e/MRvLfjrPTs+tWH+iA23K6lQCwoOwGPMa6qvsfmL45yn6hrmqzSdZTsbqlH9LCeearqe07DDZqZAh4DbWFpo3Q9dzJrtJ2PQ6csPLExTEvo8ZmBIMXJ945WyXVXaAMLo0c9Pf65pJFTAoYlN0YhMuT8kMbC6x/HxUOckWSQotZAsGCqmI0Zd/r+dzGkugc5Ioi9SvdgmXTt56cl9rltZQLPlr2xZdc5/levE/D4vJgBCT81PhjXdPMtljirZfkQygXeaJXB5m+rUNxNBZiJsFQ/IW6Iy5h3d5yVFlo6K5rzov4gUC/g9QdGS068+tNT8i9Q8quDzFXItVhYON39wniMciJzjKP8dEAkc8y0Fl0VISnG+LzpiT85q1QLPy1B7op4zOv2NZbJ07MVyJxdF+Wg06xLPFGJTn+0zIhrL1VHp9mJ3PryX1xJJcSE6ZG7wigFAXV0GkXIT53kMOgPlRZhk133XAnfFbb7VjelX1lFm2Ru4cE6bH92HM8RM5JCzit+GBJiArFuURz+em2iJDwLm7OzDysJzwzmwwGXlPExAbigu6Q8ea4dJ6s7uKAv3XURr7KcwSo1piAr5kOHaBEb4Y/YcL3LGKb83HWHUVErG/sZLLBZY/FVWk3PtmTs3gQQ1gv1Clg0WZsVK1VZcqhp6sKLBRXY8GUFj3zKoBuRb1rWqyuxaySIUObtphYjcnOSwzA9MZgX8hEGP95WYaby+/EWFBTVYvv+BrR1qm6rtMKqMeLrqSzw9Io8GeYnATyPwQxCViFv2jrHrWtW6mjO4R8XBisIDkIb+rLzv1wV+O7GDhDMZyQJgw+tgP0WfJbo4uHn84K81HKA3Cu2tAcN7KD0TuRNL+39wH3/n/fhqdcikhewHAWmbe4eeP6AkW96AwRL/ueToAAeRX7qK54GKDc8+Ic2+q74zcqnaAGld3naeQfUdWzm7koA1Wzl/XlfRRtBczM+n+qR6jug7hsYc54If0b2HpErQb0AC4Dn0N6crEb4/v3UYN7OGNi0j4HgPi9+DLEAeBtWTY4jw/rmxx5WkglKswByZT8dtBigW2Cz5TJi1h8x+q+AM9L3RELTmcZ6lazdB8J+bgPWjnM4PjO7alCcAKFloMIeCGQn8lLEWnQA+BetL81v6u4ZVgAAAABJRU5ErkJggg==';   // logo REAL de Tiendanube (favicon oficial de la marca)
  var stEl=document.getElementById('rp-tienda-css'); if(!stEl){ stEl=document.createElement('style'); stEl.id='rp-tienda-css'; (document.head||document.documentElement).appendChild(stEl); }
  function paint(){
    var sel='img[src*="'+SIG+'"]';                       // los 16 <img> del icono Shopify
    var chip='button:has(> img[src*="'+SIG+'"])';        // el chip de canal (único botón con el icono como hijo directo)
    var noNotif='#mfy_toasts{display:none!important}';   // sin notificaciones de venta arriba a la derecha
    if(window._rpShop===true){ stEl.textContent=noNotif; return; }   // Shopify: icono + texto originales
    if(window._rpTn===true){                                          // Tiendanube: ícono Y texto por CSS (no se toca React)
      stEl.textContent=noNotif
        +sel+'{content:url("'+TN_URI+'")!important}'
        +chip+'{font-size:0!important}'                               // oculta el texto "Shopify" del chip
        +chip+'::after{content:"Tiendanube"!important;font-size:12px!important}'; return; }
    if(window._rpShop===false && window._rpTn===false){               // sin tienda: icono oculto + "Sin tienda"
      stEl.textContent=noNotif
        +sel+'{display:none!important}'
        +chip+'{font-size:0!important}'
        +chip+'::after{content:"Sin tienda"!important;font-size:12px!important}'; return; }
    stEl.textContent=noNotif;   // unknown (aún no resolvió estado): al menos sacar las notificaciones
  }
  function estado(k,url){ return fetch(url).then(function(r){return r.json();}).then(function(j){ window[k]=!!(j&&j.conectado); }).catch(function(){}); }
  var ps=[];
  if(window._rpShop===undefined) ps.push(estado('_rpShop','/shopify/estado'));
  if(window._rpTn===undefined) ps.push(estado('_rpTn','/tiendanube/estado'));
  Promise.all(ps).then(paint);
  paint();   // por si los estados ya venían seteados
})();
</script>

<div id="rp-stock-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;display:none;overflow:auto;transition:left .18s ease;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;background:radial-gradient(1200px 560px at 88% -14%,rgba(84,168,240,.06),transparent 60%),radial-gradient(900px 500px at 0% 118%,rgba(180,164,245,.06),transparent 58%),#080c14;color:#eef3f9">
<style>
#rp-stock-ov{--card:#0f1723;--card2:#0b111b;--inset:#0a1119;--line:#1b2635;--line2:#25344a;--hair:rgba(255,255,255,.045);--ink:#eef3f9;--ink2:#93a3ba;--ink3:#5a6a80;--blue:#54a8f0;--green:#37c98d;--warn:#e8b13e;--crit:#f0637f;font-variant-numeric:tabular-nums}
#rp-stock-ov .wrap{max-width:800px;margin:0 auto;padding:30px 22px 70px}
#rp-stock-ov .hdr{display:flex;align-items:center;gap:13px}
#rp-stock-ov .hdr .ic{width:42px;height:42px;border-radius:12px;flex:none;display:grid;place-items:center;background:linear-gradient(150deg,#16324f,#122539);border:1px solid #23415f}
#rp-stock-ov .hdr .ic .material-symbols-outlined{color:var(--blue);font-size:22px}
#rp-stock-ov h1{margin:0;font-size:25px;font-weight:800;letter-spacing:-.5px}
#rp-stock-ov .lead{margin:5px 0 0;color:var(--ink2);font-size:13px;max-width:52ch;line-height:1.5}
#rp-stock-ov .lead b{color:#9fc0e4;font-weight:700}
#rp-stock-ov .card{position:relative;background:linear-gradient(180deg,var(--card),var(--card2));border:1px solid var(--line);border-radius:18px;box-shadow:0 1px 0 var(--hair) inset,0 18px 40px -34px #000}
#rp-stock-ov .hero{padding:22px 24px;margin-top:20px;overflow:hidden}
#rp-stock-ov .hero::before{content:"";position:absolute;top:-90px;right:-70px;width:240px;height:240px;border-radius:50%;background:var(--hc);opacity:.10;filter:blur(46px)}
#rp-stock-ov .htop{display:flex;align-items:center;gap:11px;flex-wrap:wrap}
#rp-stock-ov .hnm{font-weight:800;font-size:16px;letter-spacing:-.2px}
#rp-stock-ov .sku{font-size:10.5px;font-weight:700;letter-spacing:.5px;color:var(--ink2);background:#0b1521;border:1px solid var(--line);border-radius:6px;padding:4px 8px}
#rp-stock-ov .pill{margin-left:auto;display:inline-flex;align-items:center;gap:7px;font-size:11.5px;font-weight:700;padding:6px 12px;border-radius:8px;background:var(--hpb);color:var(--hc)}
#rp-stock-ov .pill .dot{width:7px;height:7px;border-radius:50%;background:var(--hc)}
#rp-stock-ov .hmid{display:flex;align-items:flex-end;justify-content:space-between;gap:18px;margin-top:14px;flex-wrap:wrap}
#rp-stock-ov .num{font-size:40px;font-weight:800;letter-spacing:-1.2px;line-height:1;color:var(--hc)}
#rp-stock-ov .num .u{font-size:13.5px;font-weight:600;color:var(--ink2);letter-spacing:0;margin-left:7px}
#rp-stock-ov .days{text-align:right}
#rp-stock-ov .days .n{font-size:21px;font-weight:800;letter-spacing:-.4px;color:var(--ink)}
#rp-stock-ov .days .n b{color:var(--hc)}
#rp-stock-ov .days .l{font-size:11.5px;color:var(--ink3);font-weight:600;margin-top:1px}
#rp-stock-ov .cov{margin-top:18px}
#rp-stock-ov .covbar{height:10px;border-radius:6px;background:#0c1622;border:1px solid var(--line);overflow:hidden;position:relative}
#rp-stock-ov .covbar>b{display:block;height:100%;border-radius:6px 0 0 6px;background:linear-gradient(90deg,var(--hc),color-mix(in srgb,var(--hc) 60%,#0b111b))}
#rp-stock-ov .covbar>s{position:absolute;top:-4px;bottom:-4px;width:2px;background:var(--ink2);opacity:.55}
#rp-stock-ov .covlb{display:flex;justify-content:space-between;font-size:10.5px;color:var(--ink3);margin-top:7px}
#rp-stock-ov .tiles{display:grid;grid-template-columns:repeat(4,1fr);gap:11px;margin-top:12px}
@media(max-width:680px){#rp-stock-ov .tiles{grid-template-columns:1fr 1fr}}
#rp-stock-ov .tile{padding:14px 15px}
#rp-stock-ov .tile .k{font-size:10px;color:var(--ink3);text-transform:uppercase;letter-spacing:.7px;font-weight:700}
#rp-stock-ov .tile .v{font-size:20px;font-weight:800;letter-spacing:-.5px;margin-top:8px}
#rp-stock-ov .tile .v small{font-size:11.5px;color:var(--ink2);font-weight:600}
#rp-stock-ov .tile.accent .v{color:var(--green)}
#rp-stock-ov .sec{display:flex;align-items:center;gap:10px;margin:26px 0 12px}
#rp-stock-ov .sec .bb{width:14px;height:2px;border-radius:2px;background:#b4a4f5}
#rp-stock-ov .sec h2{margin:0;font-size:12px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;color:var(--ink2)}
#rp-stock-ov .sec .x{margin-left:auto;font-size:11.5px;color:var(--ink3)}
#rp-stock-ov .note{color:var(--ink3);font-size:12px;margin:-4px 0 12px;line-height:1.5}
#rp-stock-ov .note b{color:var(--green)}
#rp-stock-ov .proj{padding:20px 22px}
#rp-stock-ov .pstats{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
@media(max-width:560px){#rp-stock-ov .pstats{grid-template-columns:1fr}}
#rp-stock-ov .ps{background:var(--inset);border:1px solid var(--line);border-radius:12px;padding:12px 14px}
#rp-stock-ov .ps .k{font-size:10px;color:var(--ink3);text-transform:uppercase;letter-spacing:.6px;font-weight:700}
#rp-stock-ov .ps .v{font-size:19px;font-weight:800;margin-top:6px;letter-spacing:-.4px}
#rp-stock-ov .ps .v small{font-size:11px;color:var(--ink2);font-weight:600}
#rp-stock-ov .chart{margin:18px 0 6px}
#rp-stock-ov .chart svg{width:100%;height:60px;display:block}
#rp-stock-ov .chart .lb{display:flex;justify-content:space-between;font-size:10px;color:var(--ink3);margin-top:5px}
#rp-stock-ov .proj-sel{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-top:16px;padding-top:16px;border-top:1px solid var(--line)}
#rp-stock-ov .proj-sel .lab{font-size:12.5px;color:var(--ink2)}
#rp-stock-ov .chips{display:inline-flex;gap:4px;background:#0b1420;border:1px solid var(--line);border-radius:10px;padding:3px}
#rp-stock-ov .chips button{border:none;background:transparent;color:var(--ink2);font-family:inherit;font-weight:700;font-size:12.5px;padding:7px 13px;border-radius:7px;cursor:pointer;transition:.12s}
#rp-stock-ov .chips button.on{background:#18314e;color:#dcebfb}
#rp-stock-ov .cinput{width:66px}
#rp-stock-ov .proj-out{margin-top:15px;background:linear-gradient(100deg,rgba(55,201,141,.08),transparent);border:1px solid rgba(55,201,141,.2);border-radius:12px;padding:14px 17px;display:flex;align-items:baseline;gap:10px;flex-wrap:wrap}
#rp-stock-ov .proj-out .big{font-size:22px;font-weight:800;color:var(--green);letter-spacing:-.5px}
#rp-stock-ov .proj-out .tx{font-size:13px;color:var(--ink2)}
#rp-stock-ov input.f{background:var(--inset);border:1px solid var(--line2);border-radius:11px;color:var(--ink);font-family:inherit;outline:none;font-weight:800;transition:.15s}
#rp-stock-ov input.f:focus{border-color:#356fae;box-shadow:0 0 0 3px rgba(53,111,174,.16)}
#rp-stock-ov input.big{width:100%;font-size:21px;padding:13px 15px}
#rp-stock-ov input.sm{font-size:15px;padding:8px 10px;text-align:center}
#rp-stock-ov .pedir{padding:18px 20px}
#rp-stock-ov .pedir-row{display:flex;align-items:flex-end;gap:12px;flex-wrap:wrap}
#rp-stock-ov .field{flex:1;min-width:180px}
#rp-stock-ov .field .l{font-size:11.5px;color:var(--ink2);font-weight:600;margin-bottom:8px}
#rp-stock-ov .btn{border:none;border-radius:11px;cursor:pointer;font-family:inherit;font-weight:800;color:#fff}
#rp-stock-ov .btn-b{background:linear-gradient(135deg,#2b8ef0,#1668cc);box-shadow:0 8px 20px -9px rgba(43,142,240,.8);padding:13px 22px;font-size:14px;display:inline-flex;gap:8px;align-items:center;white-space:nowrap}
#rp-stock-ov .hint{margin-top:11px;font-size:12px;color:var(--ink3)}
#rp-stock-ov .hint b{color:#b4a4f5}
#rp-stock-ov .addlink{color:var(--blue);cursor:pointer;font-weight:700;font-size:13px}
#rp-stock-ov .plist{display:flex;flex-direction:column;gap:10px}
#rp-stock-ov .empty{color:var(--ink3);font-size:12.5px;text-align:center;padding:20px;border:1px dashed var(--line2);border-radius:13px}
#rp-stock-ov .prow{display:flex;align-items:center;gap:13px;padding:13px 15px}
#rp-stock-ov .prow .tx{flex:1}
#rp-stock-ov .prow .tx .q{font-weight:800;font-size:14.5px}
#rp-stock-ov .prow .tx .d{color:var(--ink3);font-size:11.5px;margin-top:2px}
#rp-stock-ov .badge{font-size:10px;font-weight:800;padding:3px 8px;border-radius:7px;background:rgba(232,177,62,.13);color:var(--warn);margin-left:6px}
#rp-stock-ov .btn-dep{background:#0e2a20;border:1px solid rgba(55,201,141,.3);color:var(--green);padding:9px 13px;font-size:12px;border-radius:10px;white-space:nowrap;font-weight:800;cursor:pointer;font-family:inherit}
#rp-stock-ov .corr{margin-top:12px}
#rp-stock-ov .corr a{color:var(--ink2);cursor:pointer;font-size:11.5px}
</style>
<div class="wrap">
  <div class="hdr"><span class="ic"><span class="material-symbols-outlined">warehouse</span></span><div><h1>Stock</h1><p class="lead">El <b>nombre, SKU y unidad</b> de cada producto salen de tu secci&oacute;n <b>Productos</b>. Ac&aacute; ves cu&aacute;nto ten&eacute;s, proyect&aacute;s la venta y ped&iacute;s reposici&oacute;n.</p></div></div>
  <div id="rp-stock-body"></div>
</div>
<div id="rp-stk-toast" style="position:fixed;left:50%;bottom:26px;transform:translateX(-50%) translateY(18px);opacity:0;background:#0e2a20;border:1px solid rgba(55,201,141,.35);color:#cdf3e2;padding:11px 17px;border-radius:11px;font-size:13px;font-weight:700;transition:.28s;pointer-events:none;z-index:100010"></div>
<script>
(function(){
 function $(id){return document.getElementById(id);}
 function esc(t){return (t||'').replace(/[<>&]/g,function(c){return {'<':'&lt;','>':'&gt;','&':'&amp;'}[c];});}
 function sid(pid){return (pid||'').replace(/[^a-z0-9]/gi,'');}
 function ars(n){return '$'+Math.round(n).toLocaleString('es-AR');}
 function salud(d){ if(d>=10)return {c:'var(--green)',pb:'rgba(55,201,141,.13)',lb:'Saludable'}; if(d>=5)return {c:'var(--warn)',pb:'rgba(232,177,62,.14)',lb:'Reponer pronto'}; return {c:'var(--crit)',pb:'rgba(240,99,127,.14)',lb:'Crítico'}; }
 function prod(pid){return (window._STK||[]).filter(function(x){return x.id===pid;})[0];}
 function tile(k,v,cls){return '<div class="card tile '+cls+'"><div class="k">'+k+'</div><div class="v">'+v+'</div></div>';}
 function tstk(m){var t=$('rp-stk-toast');if(!t)return;t.textContent=m;t.style.opacity='1';t.style.transform='translateX(-50%) translateY(0)';clearTimeout(t._t);t._t=setTimeout(function(){t.style.opacity='0';t.style.transform='translateX(-50%) translateY(18px)';},2500);}
 window.rpStock=function(open){ var o=$('rp-stock-ov'); if(!o)return;
  if(open){ ['rp-prod-ov','rp-comis-ov','rp-integ-ov','rp-desp-ov','rp-fact-ov','rp-mov-ov','rp-ads-ov'].forEach(function(id){var x=$(id);if(x)x.style.display='none';}); var lk=$('rpf-lock'); if(lk)lk.style.display='none'; }
  o.style.display=open?'block':'none'; try{window._rpNavActive(open?'rp-stock-nav':null);}catch(e){} if(open) rpStkLoad(); };
 window.rpStkLoad=function(){ var b=$('rp-stock-body'); if(b)b.innerHTML='<div class="note" style="margin-top:22px">Cargando&hellip;</div>';
  fetch('/pf-stock').then(function(r){return r.json();}).then(function(j){ var P=(j&&j.productos)||[]; window._STK=P; if(!b)return;
   var add='<div style="text-align:center;margin-top:16px"><a class="addlink" onclick="rpStkAgregar()">+ Agregar producto</a></div>';
   if(!P.length){ b.innerHTML='<div class="card" style="margin-top:20px;padding:24px;text-align:center;color:var(--ink2)">Todav&iacute;a no cargaste productos al stock.</div>'+add; return; }
   b.innerHTML=P.map(bloque).join('')+add;
   P.forEach(function(p){ renderProj(p.id); proj(p.id); });
  }).catch(function(){ if(b)b.innerHTML='<div class="card" style="margin-top:20px;padding:22px">No se pudo cargar el stock.</div>'; }); };
 function bloque(p){
   var rate=p.rate||0, d=rate?Math.round(p.stock/rate):0, s=salud(rate?d:99);
   var cov=Math.max(5,Math.min(100,rate?d/20*100:100)), id=sid(p.id), u=esc(p.unidad), usg=esc((p.unidad||'').replace(/s$/,''));
   var pend=(p.pendientes||[]);
   var plist = pend.length? pend.map(function(o){return '<div class="card prow"><span class="tx"><div class="q">'+o.qty.toLocaleString('es-AR')+' '+u+'<span class="badge">en proceso</span></div><div class="d">Pedido '+esc(o.fecha)+' &middot; '+ars(o.qty*(p.costo||0))+'</div></span><button class="btn-dep" onclick="rpStkDep(\''+o.id+'\')">Poner en depósito</button></div>';}).join('') : '<div class="empty">No hay pedidos en proceso.</div>';
   return '<div class="card hero" style="--hc:'+s.c+';--hpb:'+s.pb+'">'+
       '<div class="htop"><span class="hnm">'+esc(p.nombre)+'</span>'+(p.sku?'<span class="sku">'+esc(p.sku)+'</span>':'')+'<span class="pill"><span class="dot"></span>'+s.lb+'</span></div>'+
       '<div class="hmid"><div class="num">'+p.stock.toLocaleString('es-AR')+'<span class="u">'+u+' en depósito</span></div>'+
         '<div class="days"><div class="n"><b>'+d+'</b> días</div><div class="l">te alcanza a este ritmo</div></div></div>'+
       '<div class="cov"><div class="covbar"><b style="width:'+cov+'%"></b><s style="left:50%"></s></div><div class="covlb"><span>0</span><span>10 días &middot; sano</span><span>20+</span></div></div>'+
     '</div>'+
     '<div class="tiles">'+tile('En stock',p.stock.toLocaleString('es-AR')+' <small>'+u+'</small>','')+tile('Valor en stock',ars(p.stock*(p.costo||0)),'accent')+tile('Venta por día',rate+' <small>'+u+'/día</small>','')+tile('Valor por '+usg,ars(p.costo||0),'')+'</div>'+
     '<div class="sec"><span class="bb"></span><h2>Proyección de ventas</h2><span class="x">ritmo actual '+rate+' '+u+'/día</span></div>'+
     '<div class="card proj">'+
       '<div class="pstats"><div class="ps"><div class="k">Últimos 7 días</div><div class="v">'+p.d7+' <small>'+u+'</small></div></div><div class="ps"><div class="k">Últimos 14 días</div><div class="v">'+p.d14+' <small>'+u+'</small></div></div><div class="ps"><div class="k">Promedio por día</div><div class="v">'+rate+' <small>'+u+'</small></div></div></div>'+
       '<div class="chart"><svg id="spark-'+id+'" viewBox="0 0 700 60" preserveAspectRatio="none"></svg><div class="lb"><span>hace 14 días</span><span>hoy</span></div></div>'+
       '<div class="proj-sel"><span class="lab">Proyectar a</span><span class="chips" id="chips-'+id+'"></span><input id="ndias-'+id+'" class="f sm cinput" type="number" value="30" oninput="proj(\''+p.id+'\')"><span class="lab">días</span></div>'+
       '<div class="proj-out" id="pout-'+id+'"></div>'+
     '</div>'+
     '<div class="sec"><span class="bb"></span><h2>Pedir stock</h2></div>'+
     '<div class="note">Poné cuántas unidades vas a pedir. Queda <b>en proceso</b> &mdash; no suma al stock todavía.</div>'+
     '<div class="card pedir"><div class="pedir-row"><div class="field"><div class="l">Unidades a pedir ('+u+')</div><input id="ped-'+id+'" class="f big" type="number" placeholder="0" oninput="rpStkHint(\''+p.id+'\')"></div><button class="btn btn-b" onclick="rpStkPedir(\''+p.id+'\')">Pedir</button></div>'+
       '<div class="hint" id="hint-'+id+'">Usá la proyección de arriba para saber cuánto pedir.</div>'+
       '<div class="corr"><a onclick="rpStkEditToggle(\''+id+'\')">&#9998; Corregir stock (si te confundiste o hubo rotura/ajuste)</a>'+
       '<div id="corr-'+id+'" style="display:none;margin-top:10px;align-items:center;gap:8px;flex-wrap:wrap"><span style="font-size:12px;color:var(--ink2)">Stock real:</span><input id="corrin-'+id+'" class="f sm" type="number" value="'+p.stock+'" style="width:110px"><button class="btn btn-b" style="padding:9px 15px;font-size:13px" onclick="rpStkGuardar(\''+p.id+'\')">Guardar</button><a onclick="rpStkEditToggle(\''+id+'\')" style="color:var(--ink3);cursor:pointer;font-size:12px;margin-left:2px">cancelar</a></div></div>'+
     '</div>'+
     '<div class="sec"><span class="bb"></span><h2>En proceso</h2><span class="x">'+(pend.length?(pend.length+' pedido'+(pend.length>1?'s':'')):'')+'</span></div>'+
     '<div class="plist">'+plist+'</div>';
 }
 function renderProj(pid){ var p=prod(pid); if(!p)return; var id=sid(pid), v=p.ventas14||[];
   var W=700,H=60,n=v.length,gap=7,bw=(W-gap*(n-1))/n,mx=Math.max.apply(null,v),mn=Math.min.apply(null,v),svg='';
   for(var i=0;i<n;i++){var fr=(v[i]-mn)/((mx-mn)||1),h=(H-4)*(0.32+0.68*fr),x=i*(bw+gap),y=H-h,last=i===n-1; svg+='<rect x="'+x.toFixed(1)+'" y="'+y.toFixed(1)+'" width="'+bw.toFixed(1)+'" height="'+h.toFixed(1)+'" rx="3" fill="'+(last?'#54a8f0':'#2a3c54')+'"/>';}
   var sp=$('spark-'+id); if(sp)sp.innerHTML=svg;
   var ch=$('chips-'+id); if(ch)ch.innerHTML=[15,30,60,90].map(function(x){return '<button onclick="rpStkSetN(\''+pid+'\','+x+')">'+x+'</button>';}).join('');
 }
 window.rpStkSetN=function(pid,n){ var el=$('ndias-'+sid(pid)); if(el)el.value=n; proj(pid); };
 window.proj=function(pid){ var p=prod(pid); if(!p)return; var id=sid(pid), n=Math.max(1,Math.round(+($('ndias-'+id)||{}).value||30)), rate=p.rate||0;
   var vender=Math.round(rate*n), faltan=Math.max(0,vender-p.stock), o=$('pout-'+id);
   if(o)o.innerHTML='<span class="big">'+vender.toLocaleString('es-AR')+' '+esc(p.unidad)+'</span><span class="tx">vas a vender en <b style="color:var(--ink)">'+n+' días</b> &middot; '+(faltan>0?('para cubrirlos te faltan <b style="color:var(--ink)">'+faltan.toLocaleString('es-AR')+' '+esc(p.unidad)+'</b> <a onclick="rpStkUsar(\''+pid+'\','+faltan+')" style="color:var(--blue);cursor:pointer;font-weight:700">pedir eso</a>'):'te alcanza el stock &#10003;')+'</span>';
   var ch=$('chips-'+id); if(ch)ch.querySelectorAll('button').forEach(function(b){b.classList.toggle('on',+b.textContent===n);});
 };
 window.rpStkUsar=function(pid,q){ var el=$('ped-'+sid(pid)); if(el){el.value=q; rpStkHint(pid); el.scrollIntoView({behavior:'smooth',block:'center'});} };
 window.rpStkHint=function(pid){ var p=prod(pid); if(!p)return; var id=sid(pid), q=Math.max(0,Math.round(+($('ped-'+id)||{}).value||0)), h=$('hint-'+id), rate=p.rate||0;
   if(!h)return; if(q>0){var dias=rate?Math.round((p.stock+q)/rate):0; h.innerHTML='Al llegar quedarías con <b style="color:var(--ink)">'+(p.stock+q).toLocaleString('es-AR')+' '+esc(p.unidad)+'</b> (~'+dias+' días) &middot; cuesta '+ars(q*(p.costo||0));} else h.innerHTML='Usá la proyección de arriba para saber cuánto pedir.'; };
 window.rpStkPedir=function(pid){ var q=Math.max(0,Math.round(+($('ped-'+sid(pid))||{}).value||0)); if(q<=0)return;
   fetch('/pf-stock-pedir',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pid:pid,qty:q})}).then(function(r){return r.json();}).then(function(){ tstk('Pedido registrado en proceso'); rpStkLoad(); }); };
 window.rpStkDep=function(id){ fetch('/pf-stock-depositar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:id})}).then(function(r){return r.json();}).then(function(){ tstk('Sumado al stock'); rpStkLoad(); }); };
 window.rpStkEditToggle=function(id){ var e=$('corr-'+id); if(!e)return; var vis=(e.style.display==='flex'); e.style.display=vis?'none':'flex'; if(!vis){var inp=$('corrin-'+id); if(inp){inp.focus();inp.select();}} };
 window.rpStkGuardar=function(pid){ var el=$('corrin-'+sid(pid)); if(!el)return; var n=Math.max(0,Math.round(+el.value||0)); fetch('/pf-stock-set',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pid:pid,stock:n})}).then(function(r){return r.json();}).then(function(){ tstk('Stock corregido a '+n); rpStkLoad(); }); };
 window.rpStkAgregar=function(){ fetch('/pf-stock-catalogo').then(function(r){return r.json();}).then(function(j){ var cat=(j&&j.productos)||[]; if(!cat.length){ alert('Conectá tu tienda para traer los productos.'); return; }
   var nombre=prompt('Producto (pegá parte del nombre):'); if(!nombre)return; var m=cat.filter(function(x){return (x.nombre||'').toLowerCase().indexOf(nombre.toLowerCase())>-1;})[0]; if(!m){ alert('No encontré ese producto.'); return; }
   var unidad=prompt('Unidad (potes / sprays / u):','u')||'u'; var stock=parseInt(prompt('Stock actual que tenés hoy en depósito:','0')||'0',10)||0;
   fetch('/pf-stock-set',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pid:m.id,nombre:m.nombre,sku:m.sku,unidad:unidad,stock:stock,costo:m.costo})}).then(function(r){return r.json();}).then(function(){ rpStkLoad(); }); }); };
})();
</script>
</div>
<style>
@media(max-width:820px){
 #rp-stock-ov,#rp-ads-ov,#rp-prod-ov,#rp-comis-ov,#rp-integ-ov,#rp-desp-ov,#rp-fact-ov,#rp-mov-ov,#rpf-lock{left:0 !important}
 #rpf-lock{padding:18px !important}
 #rpf-lock p{max-width:100% !important}
 #rp-moblau{display:flex !important}
}
#rp-moblau{display:none;position:fixed;right:15px;bottom:80px;z-index:100055;width:54px;height:54px;border-radius:16px;background:linear-gradient(135deg,#2b8ef0,#1668cc);box-shadow:0 12px 26px -8px rgba(43,142,240,.75);align-items:center;justify-content:center;cursor:pointer;border:none}
#rp-moblau .material-symbols-outlined{color:#fff;font-size:27px}
/* En celu: que las secciones no queden zoomeadas -> apilar grillas anchas y achicar padding */
@media(max-width:760px){
 #rp-stock-ov .wrap{padding:20px 14px 70px}
 #rp-prod-ov [class*="aw"],#rp-comis-ov [class*="aw"],#rp-desp-ov [class*="aw"],#rp-fact-ov [class*="aw"],#rp-mov-ov [class*="aw"],#rp-ads-ov [class*="aw"],#rp-integ-ov [class*="aw"]{padding-left:14px !important;padding-right:14px !important}
 [id$="-ov"] [style*="repeat(4"]{grid-template-columns:repeat(2,1fr) !important}
 [id$="-ov"] [style*="repeat(3"]{grid-template-columns:1fr !important}
 [id$="-ov"] [style*="minmax(2"]{grid-template-columns:1fr !important}
 [id$="-ov"] [style*="minmax(3"]{grid-template-columns:1fr !important}
 [id$="-ov"] table{font-size:12px !important}
 [id$="-ov"] .row,[id$="-ov"] [style*="display:flex;gap"]{flex-wrap:wrap}
}
@media(max-width:480px){
 [id$="-ov"] [style*="repeat(2"]{grid-template-columns:1fr !important}
 [id$="-ov"] [style*="repeat(4"]{grid-template-columns:1fr !important}
}
#rp-mobsheet{display:none;position:fixed;inset:0;z-index:100065;background:rgba(4,8,14,.6);align-items:flex-end;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif}
#rp-mobsheet.on{display:flex}
#rp-mobsheet .sh{width:100%;background:#0e1521;border-top-left-radius:22px;border-top-right-radius:22px;padding:12px 12px 30px;border-top:1px solid #1e2b3d;max-height:82vh;overflow:auto}
#rp-mobsheet .gr{width:40px;height:4px;border-radius:3px;background:#2b3a52;margin:3px auto 10px}
#rp-mobsheet .ti{font-size:11px;font-weight:800;letter-spacing:.6px;text-transform:uppercase;color:#5b6678;padding:6px 12px 8px}
#rp-mobsheet .mi{display:flex;align-items:center;gap:14px;padding:14px 12px;border-radius:12px;color:#e8edf4;font-size:15.5px;font-weight:600;cursor:pointer}
#rp-mobsheet .mi:active{background:#16273e}
#rp-mobsheet .mi .material-symbols-outlined{color:#54a8f0;font-size:23px}
</style>
<button id="rp-moblau" onclick="document.getElementById('rp-mobsheet').classList.add('on')"><span class="material-symbols-outlined">grid_view</span></button>
<div id="rp-mobsheet" onclick="if(event.target===this)this.classList.remove('on')">
 <div class="sh"><div class="gr"></div>
  <div class="ti">Ir a una secci&oacute;n</div>
  <div class="mi" onclick="rpMobClose()"><span class="material-symbols-outlined">home</span>Inicio (Dashboard)</div>
  <div class="mi" onclick="rpMob('rpStock')"><span class="material-symbols-outlined">warehouse</span>Stock</div>
  <div class="mi" onclick="rpMob('rpMov')"><span class="material-symbols-outlined">swap_vert</span>Movimientos</div>
  <div class="mi" onclick="rpMob('rpFact')"><span class="material-symbols-outlined">receipt_long</span>Facturaci&oacute;n</div>
  <div class="mi" onclick="rpMob('rpDesp')"><span class="material-symbols-outlined">local_shipping</span>Despachos</div>
  <div class="mi" onclick="rpMob('rpProd')"><span class="material-symbols-outlined">inventory_2</span>Productos</div>
  <div class="mi" onclick="rpMob('rpComis')"><span class="material-symbols-outlined">percent</span>Comisiones</div>
  <div class="mi" onclick="rpMob('rpAds')"><span class="material-symbols-outlined">rocket_launch</span>Subir ADS</div>
 </div>
</div>
<script>
 window.rpMob=function(name){ var sh=document.getElementById('rp-mobsheet'); if(sh)sh.classList.remove('on'); try{ if(typeof window[name]==='function') window[name](true); }catch(e){} };
 window.rpMobClose=function(){ var sh=document.getElementById('rp-mobsheet'); if(sh)sh.classList.remove('on'); ['rpProd','rpComis','rpInteg','rpDesp','rpFact','rpMov','rpAds','rpStock'].forEach(function(n){ try{ if(typeof window[n]==='function') window[n](false); }catch(e){} }); };
 // El "Más" y el ☰ nativos del celu estaban de adorno -> los enganchamos para abrir nuestro menú.
 function rpMobOpen(e){ var sh=document.getElementById('rp-mobsheet'); if(sh){ if(e){e.preventDefault();e.stopPropagation();} sh.classList.add('on'); return true; } return false; }
 document.addEventListener('click', function(e){ if(window.innerWidth>820)return; try{
   var el=e.target;
   for(var i=0;i<6 && el;i++){
     var clickable=(el.tagName==='BUTTON'||el.tagName==='A'||(el.getAttribute&&el.getAttribute('role')==='button'));
     if(clickable){
       var raw=(el.textContent||'').trim(), t=raw.toLowerCase(), r=el.getBoundingClientRect();
       var isMas=(t.indexOf('más')>-1||/(^|[^a-z])mas($|[^a-z])/.test(t)) && raw.length<26;
       var isBurger=(r.top<96 && r.left<74 && r.width<66 && r.height<66 && r.width>16);
       if(isMas||isBurger){ rpMobOpen(e); return; }
     }
     el=el.parentElement;
   }
 }catch(err){} }, true);
</script>
<div id="rp-ads-ov" style="position:fixed;top:0;right:0;bottom:0;left:72px;z-index:100000;background:#080c15;display:none;overflow:auto;transition:left .18s ease;font-family:system-ui,-apple-system,sans-serif;color:#e8edf4">
<style>
 #rp-ads-ov .aw{max-width:1180px;margin:0 auto;padding:26px 30px 70px}
 #rp-ads-ov .lb{font-size:11px;font-weight:700;color:#8b97a8;text-transform:uppercase;letter-spacing:.5px;margin:0 0 7px;display:block}
 #rp-ads-ov .in{background:#0a1322;border:1px solid #22324a;color:#e8edf4;border-radius:10px;padding:11px 13px;font-size:13.5px;font-family:inherit;outline:none;width:100%;box-sizing:border-box;color-scheme:dark}
 #rp-ads-ov textarea.in{resize:vertical;min-height:90px;line-height:1.5}
 #rp-ads-ov .card{background:linear-gradient(165deg,#101a2c,#0b1220);border:1px solid #1b2536;border-radius:16px;padding:18px 20px;margin-bottom:16px}
 #rp-ads-ov .ch{display:flex;align-items:center;gap:10px;margin-bottom:14px}
 #rp-ads-ov .cn{width:24px;height:24px;border-radius:7px;background:#0d1524;border:1px solid #2b3a52;color:#9fb4d0;font-size:12px;font-weight:800;display:flex;align-items:center;justify-content:center}
 #rp-ads-ov .ct{font-size:14.5px;font-weight:800;color:#f4f7fb}
 #rp-ads-ov .cs{font-size:12px;color:#5b6678;margin-left:auto}
 #rp-ads-ov .row{display:flex;gap:12px}#rp-ads-ov .row>*{flex:1}
 #rp-ads-ov .seg{display:flex;gap:7px}
 #rp-ads-ov .seg .s{flex:1;text-align:center;background:#0d1524;border:1px solid #1a2436;color:#aeb8c6;border-radius:10px;padding:11px;font-size:13px;font-weight:700;cursor:pointer}
 #rp-ads-ov .seg .s.on{background:rgba(19,127,236,.14);border-color:#2b6fd0;color:#bcd7f7}
 #rp-ads-ov .seg .s small{display:block;font-size:10px;font-weight:600;color:#5b6678;margin-top:2px}
 #rp-ads-ov .hint{font-size:11.5px;color:#5b6678;margin-top:8px;line-height:1.4}
 #rp-ads-ov .step{display:flex;align-items:center;background:#0a1322;border:1px solid #22324a;border-radius:10px;overflow:hidden;width:fit-content}
 #rp-ads-ov .step .b{width:40px;height:42px;border:none;background:transparent;color:#9aa6b6;font-size:20px;font-weight:800;cursor:pointer}
 #rp-ads-ov .step b{min-width:44px;text-align:center;font-size:16px;font-weight:800}
 #rp-ads-ov .vitem{display:flex;align-items:center;gap:10px;background:#0a1322;border:1px solid #17233a;border-radius:9px;padding:8px 11px;font-size:12.5px;margin-top:6px}
 #rp-ads-ov .vitem .vi{width:20px;height:20px;border-radius:6px;background:rgba(52,211,153,.14);border:1px solid #1f5a3d;color:#34d399;display:flex;align-items:center;justify-content:center;font-weight:800;flex:none}
 #rp-ads-ov .res{position:sticky;top:20px;background:linear-gradient(165deg,#141033,#0d0b22);border:1px solid #2b2350;border-radius:18px;padding:20px}
 #rp-ads-ov .rl{display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid #241d47;font-size:13px}
 #rp-ads-ov .rl .k{color:#9a8fc5}#rp-ads-ov .rl .v{color:#f1ecff;font-weight:800;text-align:right}
 #rp-ads-ov .go{width:100%;margin-top:15px;border:none;border-radius:12px;padding:15px;font-size:15px;font-weight:800;cursor:pointer;background:linear-gradient(90deg,#7c3aed,#a78bfa);color:#fff}
 #rp-ads-ov .ptrack{height:12px;background:#0b0a1a;border:1px solid #2b2350;border-radius:20px;overflow:hidden;margin-top:14px}
 #rp-ads-ov .ptrack i{display:block;height:100%;width:0;background:linear-gradient(90deg,#7c3aed,#a78bfa);border-radius:20px;transition:width .4s}
 #rp-ads-ov .sw{display:inline-flex;align-items:center;gap:10px;cursor:pointer}
 #rp-ads-ov .tk{width:44px;height:25px;border-radius:20px;background:#1a2436;border:1px solid #2b3a52;position:relative;flex:none}
 #rp-ads-ov .tk.on{background:rgba(52,211,153,.3);border-color:#1f5a3d}
 #rp-ads-ov .tk i{position:absolute;top:2px;left:2px;width:19px;height:19px;border-radius:50%;background:#8b97a8;transition:.2s}
 #rp-ads-ov .tk.on i{left:21px;background:#34d399}
 @media(max-width:900px){#rp-ads-ov .agrid{grid-template-columns:1fr !important}}
</style>
<div class="aw">
 <div style="display:flex;align-items:flex-start;gap:14px;margin-bottom:20px">
  <div style="width:46px;height:46px;border-radius:13px;background:rgba(124,58,237,.15);border:1px solid #3b2a6b;display:flex;align-items:center;justify-content:center;color:#c4b5fd"><span class="material-symbols-outlined">rocket_launch</span></div>
  <div style="flex:1"><h1 style="margin:0;font-size:26px;color:#f4f7fb">Subir ADS</h1>
   <p style="margin:5px 0 0;color:#8b97a8;font-size:13px">Peg&aacute; el Drive, eleg&iacute; la config y lanz&aacute; la campa&ntilde;a con 1 click.</p></div>
  <button onclick="rpAds(false)" style="background:#111c2b;border:1px solid #1b2536;color:#cbd5e1;border-radius:10px;width:38px;height:38px;font-size:17px;cursor:pointer">&#10005;</button>
 </div>
 <div style="display:flex;align-items:center;gap:13px;background:linear-gradient(90deg,rgba(251,191,36,.09),rgba(19,127,236,.05));border:1px solid #4a3a1a;border-radius:14px;padding:13px 17px;margin-bottom:18px">
  <span style="font-size:20px;flex:none">&#128273;</span>
  <div style="flex:1;font-size:12.5px;color:#cbd5e1;line-height:1.5">Para que pueda leer tus videos, <b style="color:#f4f7fb">compart&iacute; la carpeta de Drive</b> con esta cuenta (permiso <b style="color:#f4f7fb">Lector</b>):<br>
   <span style="color:#fbbf24;font-weight:700;font-family:ui-monospace,monospace;font-size:12px">bot-finanzas-sheets@bot-finanzas-499200.iam.gserviceaccount.com</span></div>
  <button onclick="navigator.clipboard.writeText('bot-finanzas-sheets@bot-finanzas-499200.iam.gserviceaccount.com');var b=this;b.innerHTML='&#10003; copiado';setTimeout(function(){b.innerHTML='Copiar mail';},1500)" style="flex:none;background:#241a10;border:1px solid #4a3a1a;color:#fbbf24;border-radius:9px;padding:10px 14px;font-size:12px;font-weight:800;cursor:pointer;white-space:nowrap">Copiar mail</button>
 </div>
 <div class="agrid" style="display:grid;grid-template-columns:1fr 330px;gap:18px;align-items:start">
  <div>
   <div class="card">
    <div class="ch"><div class="cn">1</div><div class="ct">Cuenta e identidad</div></div>
    <div style="background:rgba(19,127,236,.1);border:1px solid #1e4f8a;border-radius:10px;padding:10px 13px;font-size:13px;font-weight:700;color:#bcd7f7;margin-bottom:12px">CP1 &mdash; NoxaLab &middot; CUENTA 1</div>
    <div class="row"><div><span class="lb">P&aacute;gina</span><select class="in" id="rpa-page"></select></div>
     <div><span class="lb">Instagram</span><select class="in" id="rpa-ig"></select></div></div>
    <div style="margin-top:12px"><span class="lb">Pixel</span><select class="in" id="rpa-pixel"></select></div>
   </div>
   <div class="card">
    <div class="ch"><div class="cn">2</div><div class="ct">Creativos</div><div class="cs" id="rpa-vc">eleg&iacute; de d&oacute;nde</div></div>
    <div class="seg" style="margin-bottom:12px"><div class="s on" id="rpa-fdrive" onclick="rpaFuente('drive')">Google Drive</div><div class="s" id="rpa-farch" onclick="rpaFuente('arch')">Mis archivos</div></div>
    <div id="rpa-srcdrive">
     <span class="lb">Link de carpeta de Google Drive</span>
     <div style="display:flex;gap:9px"><input class="in" id="rpa-drive" style="flex:1" placeholder="https://drive.google.com/drive/folders/…" oninput="rpaReset()">
      <button id="rpa-btnb" onclick="rpaBuscar()" style="flex:none;background:#137fec;border:none;color:#fff;border-radius:10px;padding:0 18px;font-weight:800;cursor:pointer;white-space:nowrap">Buscar</button></div>
    </div>
    <div id="rpa-srcarch" style="display:none">
     <span class="lb">Sub&iacute; tus videos o fotos (.mp4 / .mov / .jpg / .png)</span>
     <input class="in" id="rpa-file" type="file" accept="video/*,image/*,.mp4,.mov,.m4v,.jpg,.jpeg,.png,.webp" multiple style="width:100%;box-sizing:border-box;padding:9px" onchange="rpaSubir()">
     <div style="color:#5b6678;font-size:11.5px;margin-top:6px">Pod&eacute;s elegir varios de una. Se suben directo desde tu compu.</div>
    </div>
    <div id="rpa-vids"></div>
   </div>
   <div class="card">
    <div class="ch"><div class="cn">3</div><div class="ct">Campa&ntilde;a</div></div>
    <div class="seg" style="margin-bottom:14px"><div class="s on" id="rpa-cn" onclick="rpaCmp('nueva')">Campa&ntilde;a nueva</div><div class="s" id="rpa-ce" onclick="rpaCmp('exist')">Usar una existente</div></div>
    <div id="rpa-boxn">
     <div class="row"><div><span class="lb">&Aacute;ngulo (nombre)</span><input class="in" id="rpa-ang" value="UGC RENOVACION" oninput="rpaCalc()"></div>
      <div><span class="lb">Presupuesto diario</span><input class="in" id="rpa-presup" value="35" oninput="rpaCalc()"></div></div>
     <span class="lb" style="margin-top:13px">Presupuesto a nivel</span>
     <div class="seg"><div class="s on" id="rpa-tc" onclick="rpaTipo('cbo')">CBO<small>en la campa&ntilde;a</small></div><div class="s" id="rpa-ta" onclick="rpaTipo('abo')">ABO<small>por conjunto</small></div></div>
    </div>
    <div id="rpa-boxe" style="display:none">
     <span class="lb">Eleg&iacute; la campa&ntilde;a <span style="color:#5b6678;font-weight:500;text-transform:none;letter-spacing:0">(solo activas)</span></span>
     <select class="in" id="rpa-cmp" onchange="rpaCmpChange()"></select>
    </div>
   </div>
   <div class="card">
    <div class="ch"><div class="cn">4</div><div class="ct">Conjuntos y anuncios</div><div class="cs" id="rpa-cjctx"></div></div>
    <div id="rpa-cjmodo" class="seg" style="margin-bottom:14px;display:none">
     <div class="s on" id="rpa-cjdup" onclick="rpaCj('dup')">Agregar conjunto<small>misma config + tus videos</small></div>
     <div class="s" id="rpa-cjusar" onclick="rpaCj('usar')">A&ntilde;adir al mismo<small>suma los ads a uno</small></div></div>
    <div id="rpa-cjlista" style="display:none;margin-bottom:13px"><span class="lb" id="rpa-cjlb">Copiar la config de</span><select class="in" id="rpa-cjsel"></select></div>
    <div id="rpa-cjnom" style="margin-bottom:13px"><span class="lb">Nombre del conjunto</span><input class="in" id="rpa-cjnombre" value="CONJUNTO 1" oninput="rpaCalc()"></div>
    <div id="rpa-cjcant" class="row" style="align-items:flex-end">
     <div><span class="lb" id="rpa-cantlb">Cantidad de conjuntos</span><div class="step"><button class="b" onclick="rpaConj(-1)">&ndash;</button><b id="rpa-nconj">1</b><button class="b" onclick="rpaConj(1)">+</button></div></div>
     <div><span class="lb">Ads por conjunto</span><div style="background:#0a1322;border:1px solid #22324a;border-radius:10px;padding:11px 13px;font-size:13.5px;color:#8fb3e0;font-weight:700">= tus <span id="rpa-adsx">0</span> videos</div></div></div>
    <div class="hint" id="rpa-cjhint">Cada conjunto lleva 1 anuncio por video.</div>
   </div>
   <div class="card">
    <div class="ch"><div class="cn">5</div><div class="ct">Anuncio</div><div class="cs">t&iacute;tulo &middot; copy &middot; destino</div></div>
    <div class="row"><div><span class="lb">T&iacute;tulo</span><input class="in" id="rpa-titulo" placeholder="Titular del anuncio"></div>
     <div><span class="lb">Subt&iacute;tulo</span><input class="in" id="rpa-sub" placeholder="descripci&oacute;n (opcional)"></div></div>
    <span class="lb" style="margin-top:13px">Copy</span><textarea class="in" id="rpa-copy"></textarea>
    <span class="lb" style="margin-top:13px">URL de destino</span><input class="in" id="rpa-url"></div>
   <div class="card">
    <div class="ch"><div class="cn">6</div><div class="ct">Estado y programaci&oacute;n</div></div>
    <label class="sw" onclick="rpaEstado()"><span class="tk on" id="rpa-tk"><i></i></span><span id="rpa-estlb" style="font-size:13.5px;font-weight:700">Programada (se activa sola el d&iacute;a/hora)</span></label>
    <div id="rpa-progbox" class="row" style="margin-top:15px"><div><span class="lb">D&iacute;a de salida</span><input class="in" type="date" id="rpa-fecha"></div><div><span class="lb">Horario</span><input class="in" type="time" id="rpa-hora" value="05:00"></div></div>
   </div>
  </div>
  <div><div class="res">
   <h3 style="margin:0 0 4px;font-size:15px;color:#e9e2ff">Resumen</h3><div style="font-size:12px;color:#8a7fb5;margin-bottom:14px">Lo que se va a crear</div>
   <div class="rl"><span class="k">Campa&ntilde;a</span><span class="v" id="rpa-rcmp">&mdash;</span></div>
   <div class="rl"><span class="k">Tipo</span><span class="v" id="rpa-rtipo">CBO $35</span></div>
   <div class="rl"><span class="k">Conjuntos</span><span class="v" id="rpa-rconj">1</span></div>
   <div class="rl"><span class="k">Videos</span><span class="v" id="rpa-rvids">0</span></div>
   <div class="rl"><span class="k">Estado</span><span class="v" id="rpa-rest">Programada</span></div>
   <div style="text-align:center;margin:13px 0 2px;padding:13px;background:rgba(124,58,237,.1);border:1px solid #3b2a6b;border-radius:12px"><div style="font-size:31px;font-weight:800;color:#c4b5fd;line-height:1" id="rpa-rads">0</div><div style="font-size:11px;color:#8a7fb5;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:3px">anuncios en total</div></div>
   <button class="go" id="rpa-go" onclick="rpaLanzar()">&#128640; Lanzar campa&ntilde;a</button>
   <div id="rpa-prog" style="display:none"><div class="ptrack"><i id="rpa-bar"></i></div><div id="rpa-msg" style="font-size:12px;color:#c4b5fd;font-weight:600;margin-top:8px;text-align:center"></div></div>
  </div></div>
 </div>
</div>
</div>
<script>
(function(){
 var VIDS=0,NCONJ=1,TIPO='cbo',EST='activa',CMP='nueva',CJ='nuevo',CMPS=[],CJS=[],UPLOAD_ID='';
 function $(id){return document.getElementById(id);}
 function opt(a){return a.map(function(o){return '<option value="'+o.v+'">'+o.t+'</option>';}).join('');}
 window.rpAds=function(open){var o=$('rp-ads-ov');if(!o)return;
  if(open){['rp-prod-ov','rp-comis-ov','rp-integ-ov','rp-desp-ov','rp-fact-ov','rp-mov-ov'].forEach(function(id){var x=document.getElementById(id);if(x)x.style.display='none';});var _l=document.getElementById('rpf-lock');if(_l)_l.style.display='none';}
  o.style.display=open?'block':'none';try{window._rpNavActive(open?'rp-ads-nav':null);}catch(e){}if(open)rpaInit();};
 var _inited=false;
 function rpaInit(){ var d=new Date();d.setDate(d.getDate()+1);$('rpa-fecha').value=d.getFullYear()+'-'+('0'+(d.getMonth()+1)).slice(-2)+'-'+('0'+d.getDate()).slice(-2);
  $('rpa-fecha').addEventListener('change',rpaCalc);$('rpa-hora').addEventListener('change',rpaCalc);
  fetch('/pf-ads-cuentas').then(function(r){return r.json();}).then(function(j){if(j&&j.cuentas&&j.cuentas[0]){var c0=j.cuentas[0];$('rpa-copy').value=c0.copy||'';if(!$('rpa-titulo').value)$('rpa-titulo').value=c0.titulo||'';if(!$('rpa-sub').value)$('rpa-sub').value=c0.subtitulo||'';}});
  fetch('/pf-ads-identidad?cuenta=cp1').then(function(r){return r.json();}).then(function(j){if(!j||!j.ok)return;
   $('rpa-page').innerHTML=opt(j.pages.map(function(p){return {v:p.id,t:p.name};}));
   $('rpa-ig').innerHTML=opt(j.igs.map(function(i){return {v:i.id,t:i.name};}).concat([{v:'',t:'Sin IG (page-backed)'}]));
   $('rpa-pixel').innerHTML=opt(j.pixels.map(function(p){return {v:p.id,t:p.name};}));
   if(j.def){if(j.def.page)$('rpa-page').value=j.def.page;if(j.def.pixel)$('rpa-pixel').value=j.def.pixel;$('rpa-ig').value=j.def.ig||'';}});
  fetch('/pf-ads-campanas?cuenta=cp1').then(function(r){return r.json();}).then(function(j){CMPS=(j&&j.campanas)||[];
   $('rpa-cmp').innerHTML=opt(CMPS.map(function(c){return {v:c.id,t:c.name+' · '+(c.cbo?('CBO $'+c.presupuesto):'ABO')};}));});
  rpaCalc();}
 window.rpaTipo=function(t){TIPO=t;$('rpa-tc').classList.toggle('on',t=='cbo');$('rpa-ta').classList.toggle('on',t=='abo');rpaCalc();};
 window.rpaCmp=function(m){CMP=m;$('rpa-cn').classList.toggle('on',m=='nueva');$('rpa-ce').classList.toggle('on',m=='exist');
  $('rpa-boxn').style.display=m=='nueva'?'block':'none';$('rpa-boxe').style.display=m=='exist'?'block':'none';
  $('rpa-cjmodo').style.display=m=='exist'?'flex':'none';$('rpa-cjctx').textContent=m=='exist'?'en la campaña elegida':'';
  rpaCj(m=='exist'?'dup':'nuevo');if(m=='exist')rpaCmpChange();rpaCalc();};
 window.rpaCmpChange=function(){var cid=$('rpa-cmp').value;if(!cid)return;$('rpa-cjsel').innerHTML='<option>cargando…</option>';
  fetch('/pf-ads-conjuntos?campaign_id='+cid).then(function(r){return r.json();}).then(function(j){CJS=(j&&j.conjuntos)||[];
   $('rpa-cjsel').innerHTML=opt(CJS.map(function(c){return {v:c.id,t:c.name+' · '+c.n_ads+' ads'};}));rpaCalc();});};
 window.rpaCj=function(m){CJ=m;var d=$('rpa-cjdup'),u=$('rpa-cjusar');if(d)d.classList.toggle('on',m=='dup');if(u)u.classList.toggle('on',m=='usar');
  $('rpa-cjlista').style.display=(m=='dup'||m=='usar')?'block':'none';$('rpa-cjcant').style.display=(m=='usar')?'none':'flex';$('rpa-cjnom').style.display=(m=='usar')?'none':'block';
  $('rpa-cjlb').textContent=m=='usar'?'Conjunto al que sumar los ads':'Copiar la config de';
  $('rpa-cjhint').innerHTML=(m=='usar')?'Los anuncios se agregan al conjunto elegido (no crea uno nuevo).':(m=='dup')?'Agrega un conjunto con la misma config del elegido + tus videos.':'Cada conjunto lleva 1 anuncio por video.';
  if(m=='usar')NCONJ=1;rpaCalc();};
 window.rpaConj=function(d){NCONJ=Math.max(1,Math.min(20,NCONJ+d));$('rpa-nconj').textContent=NCONJ;rpaCalc();};
 window.rpaEstado=function(){EST=EST=='activa'?'pausada':'activa';$('rpa-tk').classList.toggle('on',EST=='activa');
  $('rpa-estlb').textContent=EST=='activa'?'Programada (se activa sola el día/hora)':'Pausada (para revisar antes)';
  $('rpa-progbox').style.display=EST=='activa'?'flex':'none';rpaCalc();};
 window.rpaReset=function(){$('rpa-vids').innerHTML='';VIDS=0;UPLOAD_ID='';$('rpa-vc').textContent='cargá tus videos';rpaCalc();};
 window.rpaBuscar=function(){var b=$('rpa-btnb');b.textContent='Buscando…';b.disabled=true;
  fetch('/pf-ads-drive-listar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({drive:$('rpa-drive').value})}).then(function(r){return r.json();}).then(function(j){
   b.textContent='Buscar videos';b.disabled=false;
   if(!j||!j.ok){$('rpa-vids').innerHTML='<div style="color:#fb7185;font-size:12.5px;margin-top:10px">'+((j&&j.msg)||'no pude leer el Drive')+'</div>';VIDS=0;rpaCalc();return;}
   VIDS=j.videos.length;$('rpa-vc').textContent=VIDS+' creativos';
   $('rpa-vids').innerHTML=j.videos.map(function(v){return '<div class="vitem"><div class="vi">✓</div><div style="flex:1;color:#e2e8f0;font-weight:600">'+v.name+'</div><div style="color:#5b6678;font-size:11px">'+(v.mb?v.mb+'MB':'')+'</div></div>';}).join('')+'<div style="color:#34d399;font-size:12.5px;font-weight:700;margin-top:8px">✅ '+VIDS+' creativos ubicados en orden.</div>';rpaCalc();
  }).catch(function(){b.textContent='Buscar videos';b.disabled=false;});};
 window.rpaFuente=function(f){ $('rpa-fdrive').classList.toggle('on',f=='drive');$('rpa-farch').classList.toggle('on',f=='arch');
  $('rpa-srcdrive').style.display=f=='drive'?'block':'none';$('rpa-srcarch').style.display=f=='arch'?'block':'none';rpaReset();};
 window.rpaSubir=function(){ var inp=$('rpa-file'); if(!inp.files||!inp.files.length)return;
  var fd=new FormData(),allImg=true,allVid=true; for(var i=0;i<inp.files.length;i++){fd.append('videos',inp.files[i]); if(/\.(jpg|jpeg|png|webp)$/i.test(inp.files[i].name))allVid=false; else allImg=false;} var KIND=allImg?'fotos':(allVid?'videos':'creativos');
  $('rpa-vc').textContent='subiendo…'; $('rpa-vids').innerHTML='<div style=\"color:#c4b5fd;font-size:12.5px;margin-top:10px\">\u23f3 Subiendo tus '+KIND+'… (no cierres esto)</div>';
  fetch('/pf-ads-subir',{method:'POST',body:fd}).then(function(r){return r.json();}).then(function(j){
   if(!j||!j.ok){$('rpa-vids').innerHTML='<div style=\"color:#fb7185;font-size:12.5px;margin-top:10px\">'+((j&&j.msg)||'no pude subir')+'</div>';VIDS=0;UPLOAD_ID='';rpaCalc();return;}
   UPLOAD_ID=j.upload_id;VIDS=j.videos.length;$('rpa-vc').textContent=VIDS+' creativos';
   $('rpa-vids').innerHTML=j.videos.map(function(v){return '<div class=\"vitem\"><div class=\"vi\">\u2713</div><div style=\"flex:1;color:#e2e8f0;font-weight:600\">'+v.name+'</div><div style=\"color:#5b6678;font-size:11px\">'+(v.mb?v.mb+'MB':'')+'</div></div>';}).join('')+'<div style=\"color:#34d399;font-size:12.5px;font-weight:700;margin-top:8px\">\u2705 '+VIDS+' creativos subidos.</div>';rpaCalc();
  }).catch(function(){$('rpa-vids').innerHTML='<div style=\"color:#fb7185;font-size:12.5px;margin-top:10px\">error subiendo (probá de nuevo)</div>';VIDS=0;UPLOAD_ID='';rpaCalc();});};
 function schedTxt(){var f=$('rpa-fecha').value,h=$('rpa-hora').value||'05:00';if(!f)return h;var p=f.split('-');return p[2]+'/'+p[1]+' '+h;}
 window.rpaCalc=function(){var p=$('rpa-presup').value||'35',ang=$('rpa-ang').value||'VARIOS';
  var cmpName=CMP=='nueva'?(( new Date().getDate())+'-'+(new Date().getMonth()+1)+' '+ang):((CMPS.find(function(c){return c.id==$('rpa-cmp').value;})||{}).name||'(existente)');
  $('rpa-rcmp').textContent=cmpName;
  $('rpa-rtipo').textContent=CMP=='exist'?'(la de la campaña)':((TIPO=='cbo'?'CBO':'ABO')+' $'+p);
  $('rpa-rconj').textContent=(CMP=='exist'&&CJ=='usar')?'usar 1':NCONJ;
  $('rpa-rvids').textContent=VIDS;$('rpa-adsx').textContent=VIDS;
  $('rpa-rest').textContent=EST=='activa'?('Prog. '+schedTxt()):'Pausada';
  $('rpa-rads').textContent=(CMP=='exist'&&CJ=='usar')?VIDS:(VIDS*NCONJ);};
 window.rpaLanzar=function(){ if(VIDS<1){alert('Primero cargá tus videos (Drive o Mis archivos).');return;}
  var body={cuenta:'cp1',drive:$('rpa-drive').value,upload_id:UPLOAD_ID,page:$('rpa-page').value,pixel:$('rpa-pixel').value,ig:$('rpa-ig').value,
   modo_campana:CMP=='exist'?'existente':'nueva',campaign_id:$('rpa-cmp').value,angulo:$('rpa-ang').value,tipo:TIPO,presupuesto:$('rpa-presup').value,
   modo_conjunto:CMP=='exist'?CJ:'nuevo',adset_src_id:$('rpa-cjsel').value,conjunto_nombre:$('rpa-cjnombre').value,conjuntos:NCONJ,
   titulo:$('rpa-titulo').value,subtitulo:$('rpa-sub').value,copy:$('rpa-copy').value,url:$('rpa-url').value,
   estado:EST,fecha:$('rpa-fecha').value,hora:$('rpa-hora').value};
  var go=$('rpa-go');go.disabled=true;go.textContent='Lanzando…';var pr=$('rpa-prog'),bar=$('rpa-bar'),msg=$('rpa-msg');pr.style.display='block';
  fetch('/pf-ads-lanzar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(function(r){return r.json();}).then(function(j){
   if(!j||!j.ok){msg.style.color='#fb7185';msg.textContent=(j&&j.msg)||'error';go.disabled=false;go.textContent='🚀 Lanzar campaña';return;}
   var job=j.job;var poll=setInterval(function(){fetch('/pf-ads-progreso?job='+job).then(function(r){return r.json();}).then(function(p){
    if(!p||!p.ok)return; if(p.error){clearInterval(poll);msg.style.color='#fb7185';msg.textContent='Error: '+p.error;go.disabled=false;go.textContent='🚀 Lanzar campaña';return;}
    var pct=p.total?Math.round(p.done/p.total*100):5;if(pct<3)pct=3;if(!p.listo&&pct>97)pct=97;bar.style.width=pct+'%';msg.textContent=p.msg||'Procesando…';
    if(p.listo){clearInterval(poll);bar.style.width='100%';msg.style.color='#34d399';go.disabled=false;go.textContent='🚀 Lanzar otra';}
   });},900);
  }).catch(function(){msg.style.color='#fb7185';msg.textContent='error de conexión';go.disabled=false;go.textContent='🚀 Lanzar campaña';});};
})();
</script>
"""


def _hoy() -> str:
    # zona Argentina (UTC-3) sin depender de tz del server
    return (_dt.datetime.utcnow() - _dt.timedelta(hours=3)).strftime("%Y-%m-%d")


# --- Resumen VACÍO: misma forma que espera el dashboard, todo en cero ---
def resumen_vacio() -> dict:
    h = _hoy()
    ceros_int = ["ordenes", "unidades", "reemb_cantidad", "reemb_monto", "reemb_despachados",
                 "ventas_recompras", "ventas_periodo", "meli_ventas", "meli_unidades", "tot_ordenes"]
    ceros_float = ["facturado", "cobrado", "costo_prod", "envio", "comision", "impuestos",
                   "com_plataforma", "com_pago", "fullfilment", "envios", "envio_prom",
                   "costos_extra", "reemb_perdida", "gan_por_venta", "cpa", "publi_ars",
                   "publi_cuenta", "ganancia", "margen", "roas", "roas_be", "ticket",
                   "tasa_recompra", "facturacion_recompras",
                   "meli_facturado", "meli_cobrado", "meli_comision", "meli_costo", "meli_ganancia",
                   "meli_rent", "meli_fullfilment", "meli_aov", "meli_roas",
                   "tot_facturado", "tot_ganancia", "tot_margen", "tot_costo", "tot_fullfilment",
                   "tot_gan_por_venta", "tot_aov"]
    r = {"fecha": h, "desde": h, "hasta": h, "actualizado": h,
         "moneda": "ARS", "dolar": 1200.0}
    for k in ceros_int:
        r[k] = 0
    for k in ceros_float:
        r[k] = 0.0
    return r


def _blob_vacio() -> dict:
    return {"raw": resumen_vacio(), "prod": [], "ords": []}


# ---------------- Datos reales de Shopify → dashboard ----------------
def _shopify_orders(shop, token, desde, hasta):
    """Trae los pedidos de Shopify del período (paginado por Link header)."""
    out = []
    url = "https://%s/admin/api/2026-07/orders.json" % shop
    params = {"status": "any", "limit": 250,
              "created_at_min": desde + "T00:00:00-03:00",
              "created_at_max": hasta + "T23:59:59-03:00",
              "fields": "id,order_number,name,total_price,current_total_price,financial_status,cancelled_at,line_items,refunds,created_at,shipping_lines,shipping_address"}
    headers = {"X-Shopify-Access-Token": token}
    for _ in range(40):
        r = requests.get(url, headers=headers, params=params, timeout=30)
        if r.status_code != 200:
            break
        out.extend(r.json().get("orders", []))
        link = r.headers.get("Link", "") or r.headers.get("link", "")
        nxt = None
        for part in link.split(","):
            if 'rel="next"' in part and "<" in part and ">" in part:
                nxt = part[part.find("<") + 1:part.find(">")]
        if not nxt:
            break
        url = nxt
        params = None
    return out


_MP_COST_CACHE = {}
TIENDA_PCT = 1.0        # comisión de tienda (Shopify/TN): 1% fijo por venta, no editable
IIBB_PCT = 3.5          # Ingresos Brutos: 3,5% fijo por venta, no editable
ENVIO_DOMICILIO = 9000  # costo promedio de envío a domicilio (Andreani)
ENVIO_SUCURSAL = 6000   # costo promedio de envío a sucursal (Andreani)


def _envio_costo(o) -> int:
    """Costo de envío del pedido según sea domicilio o sucursal (promedios fijos).
    Fallback cuando el pedido todavía no está en Envialo."""
    sl = o.get("shipping_lines") or []
    txt = " ".join(((s.get("title") or "") + " " + (s.get("code") or "")) for s in sl).lower()
    if any(k in txt for k in ("sucursal", "pickup", "pick up", "retiro", "punto")):
        return ENVIO_SUCURSAL
    return ENVIO_DOMICILIO


ENVIALO_BASE = "https://www.envialo.com.ar/api/v1"
ENVIALO_KEYS = DATA_DIR / "envialo_keys.json"   # API key de Envialo por usuario (persistente)
_ENVIALO_CACHE = {}   # email -> (ts, {nº pedido: costo})


def _envialo_keys() -> dict:
    try:
        return _json.loads(ENVIALO_KEYS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _envialo_save_key(email, key) -> None:
    d = _envialo_keys()
    d[str(email)] = key
    ENVIALO_KEYS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


def _envialo_key_de(email):
    import os
    return (_envialo_keys().get(email) or os.getenv("ENVIALO_API_KEY", "")) if email else ""


def _envialo_fetch(key: str) -> dict:
    """{nº de pedido → costo REAL de envío} pegándole a Envialo con esa API key."""
    out = {}
    try:
        for pg in range(40):
            r = requests.get(ENVIALO_BASE + "/orders", headers={"X-API-Key": key},
                             params={"page": pg, "limit": 50}, timeout=30)
            if r.status_code != 200:
                break
            j = r.json()
            for o in j.get("orders", []):
                n = str(o.get("number") or "").strip()
                sc = o.get("shipping_cost")
                if n and sc not in (None, "", 0):
                    try:
                        out[n] = float(sc)
                    except (TypeError, ValueError):
                        pass
            if not j.get("hasMore"):
                break
    except Exception:
        pass
    return out


def _envialo_costos(email=None):
    """{nº de pedido → costo REAL de envío} desde la cuenta Envialo de ESE usuario.
    Vacío si no conectó su API key. Cache 5 min por usuario."""
    import time as _t
    key = _envialo_key_de(email)
    if not key:
        return {}
    c = _ENVIALO_CACHE.get(email)
    if c and (_t.time() - c[0] < 300):
        return c[1]
    out = _envialo_fetch(key)
    _ENVIALO_CACHE[email] = (_t.time(), out)
    return out


@app.get("/envialo/estado")
def envialo_estado():
    email = _user_actual()
    return jsonify({"ok": True, "conectado": bool(email and email in _envialo_keys())})


@app.post("/envialo/conectar")
@limiter.limit("20 per hour")
def envialo_conectar():
    """Guarda la API key de Envialo del usuario. Valida pegándole una vez a la API."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "error": "Tenés que iniciar sesión."}), 401
    data = request.get_json(silent=True) or request.form
    key = (data.get("api_key", "") or "").strip()
    if not key:
        return jsonify({"ok": False, "error": "Pegá tu API key de Envialo."}), 400
    try:
        r = requests.get(ENVIALO_BASE + "/orders", headers={"X-API-Key": key},
                         params={"page": 0, "limit": 1}, timeout=25)
        if r.status_code in (401, 403):
            return jsonify({"ok": False, "error": "La API key no es válida (Envialo la rechazó)."}), 400
        if r.status_code != 200:
            return jsonify({"ok": False, "error": "Envialo no respondió bien. Probá de nuevo."}), 400
    except Exception:
        return jsonify({"ok": False, "error": "No pudimos conectar con Envialo. Probá de nuevo."}), 502
    _envialo_save_key(email, key)
    _ENVIALO_CACHE.pop(email, None)
    return jsonify({"ok": True})


@app.get("/desconectar-envialo")
def desconectar_envialo():
    email = _user_actual()
    if email:
        d = _envialo_keys()
        d.pop(email, None)
        ENVIALO_KEYS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
        _ENVIALO_CACHE.pop(email, None)
    return redirect("/?integ=1")


def _mp_costos(email, desde, hasta):
    """Costos REALES de MercadoPago del período (de la cuenta MP conectada del usuario).
    costo real por pago = transaction_amount − net_received_amount (incluye comisión +
    financiación de cuotas + IVA, todo lo que MP realmente te sacó).
    Devuelve montos y % separando 1 pago (al instante) de cuotas (3 sin interés). None si no hay MP."""
    import os, time as _t
    tk = _mp_tokens().get(email)
    token = tk.get("access_token") if tk else None
    if not token:
        return None
    key = (email, desde, hasta)
    c = _MP_COST_CACHE.get(key)
    if c and (_t.time() - c[0] < 300):
        return c[1]
    ini = desde + "T00:00:00.000-03:00"
    fin = hasta + "T23:59:59.999-03:00"
    vol1 = cost1 = volc = costc = 0.0
    n1 = nc = 0
    offset = 0
    try:
        while True:
            r = requests.get("https://api.mercadopago.com/v1/payments/search",
                             headers={"Authorization": "Bearer " + token},
                             params={"sort": "date_approved", "criteria": "desc",
                                     "range": "date_approved", "begin_date": ini, "end_date": fin,
                                     "status": "approved", "offset": offset, "limit": 100}, timeout=30)
            if r.status_code >= 400:
                return None if offset == 0 else None
            data = r.json()
            res = data.get("results") or []
            for p in res:
                ta = float(p.get("transaction_amount") or 0)
                det = p.get("transaction_details") or {}
                net = float(det.get("net_received_amount") or 0)
                costo = (ta - net) if net else sum(float(f.get("amount") or 0)
                                                   for f in (p.get("fee_details") or []))
                inst = int(p.get("installments") or 1)
                if inst <= 1:
                    vol1 += ta; cost1 += costo; n1 += 1
                else:
                    volc += ta; costc += costo; nc += 1
            offset += 100
            if offset >= (data.get("paging") or {}).get("total", 0) or not res:
                break
    except Exception:
        return None
    vol = vol1 + volc
    costo = cost1 + costc
    out = {"vol": round(vol, 2), "costo": round(costo, 2),
           "pct_total": round(costo / vol * 100, 2) if vol else 0.0,
           "pct_1pago": round(cost1 / vol1 * 100, 2) if vol1 else 0.0,
           "pct_cuotas": round(costc / volc * 100, 2) if volc else 0.0,
           "costo_1pago": round(cost1, 2), "costo_cuotas": round(costc, 2),
           "n_1pago": n1, "n_cuotas": nc}
    _MP_COST_CACHE[key] = (_t.time(), out)
    return out


@app.get("/pf-mp-costos")
def pf_mp_costos():
    """% real que te saca MercadoPago (al instante vs 3 cuotas), para mostrar en Comisiones."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    desde = request.args.get("desde") or _hoy()
    hasta = request.args.get("hasta") or desde
    mp = _mp_costos(email, desde, hasta)
    if mp is None:
        return jsonify({"ok": True, "conectado": False})
    return jsonify({"ok": True, "conectado": True, **mp})


@app.get("/pf-debug-ordenes")
def pf_debug_ordenes():
    """Detalle de costos REALES por pedido (últimos N) para verificar el cálculo pedido-por-pedido."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    n = int(request.args.get("n") or 3)
    desde = request.args.get("desde") or (_dt.date.today() - _dt.timedelta(days=10)).isoformat()
    hasta = request.args.get("hasta") or _hoy()
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token"):
        return jsonify({"ok": True, "shopify": False})
    try:
        orders = _shopify_orders(tk.get("shop"), tk.get("access_token"), desde, hasta)
    except Exception:
        return jsonify({"ok": False, "error": "shopify"})
    orders = [o for o in orders if not o.get("cancelled_at")]
    orders.sort(key=lambda o: int(o.get("order_number") or 0), reverse=True)
    costos = (_costos().get(email) or {})
    emap = _envialo_costos(email)
    pagos = _mp_pagos_lista(email, desde, hasta)
    by_ref, by_amt = {}, {}
    for p in (pagos or []):
        if p["ref"]:
            by_ref.setdefault(str(p["ref"]), []).append(p)
        by_amt.setdefault(p["amount"], []).append(p)
    out = []
    for o in orders[:n]:
        tot = float(o.get("total_price") or o.get("current_total_price") or 0)
        u = sum(int(li.get("quantity") or 0) for li in (o.get("line_items") or []))
        cp = sum(_costo_qty(costos.get(str(li.get("product_id") or "")), int(li.get("quantity") or 0))
                 for li in (o.get("line_items") or []))
        num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
        # envío real o promedio
        env = emap.get(num)
        env_fuente = "envialo" if env is not None else "promedio"
        if env is None:
            env = _envio_costo(o)
        # match MP
        pago = None
        if pagos is not None:
            for ref in (str(o.get("id")), str(o.get("order_number")), num):
                lst = by_ref.get(ref)
                if lst:
                    pago = lst.pop(0); break
            if pago is None:
                lst = by_amt.get(round(tot))
                if lst:
                    pago = lst.pop(0)
        mp_fee = pago["fee"] if pago else 0.0
        mp_neto = pago["net"] if pago else None
        iibb = tot * IIBB_PCT / 100.0
        tienda = tot * TIENDA_PCT / 100.0
        gan = tot - cp - mp_fee - env - iibb - tienda
        out.append({"pedido": num, "total": round(tot, 2), "unidades": u,
                    "costo_prod": round(cp, 2), "mp_fee": round(mp_fee, 2),
                    "mp_neto_recibido": (round(mp_neto, 2) if mp_neto is not None else None),
                    "mp_matcheo": ("ok" if pago else "SIN MATCH"),
                    "envio": round(env, 2), "envio_fuente": env_fuente,
                    "iibb": round(iibb, 2), "tienda": round(tienda, 2),
                    "ganancia": round(gan, 2)})
    return jsonify({"ok": True, "shopify": True, "mp_conectado": pagos is not None, "ordenes": out})


_IMG_CACHE = {}


def _shop_img(shop, token, product_id):
    """URL de la foto de un producto de Shopify (cacheada). '' si no hay."""
    pid = str(product_id or "").strip()
    if not pid:
        return ""
    if pid in _IMG_CACHE:
        return _IMG_CACHE[pid]
    url = ""
    try:
        r = requests.get("https://%s/admin/api/2026-07/products/%s.json" % (shop, pid),
                         headers={"X-Shopify-Access-Token": token},
                         params={"fields": "id,image,images"}, timeout=15)
        if r.status_code == 200:
            p = (r.json() or {}).get("product") or {}
            url = ((p.get("image") or {}).get("src")
                   or ((p.get("images") or [{}])[0] or {}).get("src") or "")
    except Exception:
        pass
    _IMG_CACHE[pid] = url
    return url


# ==================== STOCK (inventario) ====================
STOCK_FILE = DATA_DIR / "stock.json"           # {email: {pid: {nombre,unidad,stock,costo,sku}}}
STOCK_LEDGER = DATA_DIR / "stock_ledger.json"  # {email: {order_key: {delta:{pid:qty}, estado}}}
STOCK_PEDIDOS = DATA_DIR / "stock_pedidos.json"  # {email: [{id,pid,qty,fecha,estado}]}


def _stk_read(path, default):
    try:
        return _json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def _stk_write(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")


def _stock_orders(email, dias=90):
    """Pedidos recientes (TN + Shopify) -> [{key, pid_qty, estado, fecha}].
    estado: paid | pending | refunded | cancelled."""
    desde = (_dt.datetime.utcnow() - _dt.timedelta(days=dias)).date().isoformat()
    out = []
    tk = _tn_tokens().get(email)
    if tk and tk.get("access_token") and tk.get("store_id"):
        store, hdr = tk["store_id"], _tn_headers(tk["access_token"])
        page = 1
        while page <= 14:
            try:
                r = requests.get("%s/%s/orders" % (TN_API, store), headers=hdr,
                                 params={"per_page": 50, "page": page,
                                         "created_at_min": desde + "T00:00:00-03:00",
                                         "fields": "id,payment_status,cancelled_at,created_at,products"},
                                 timeout=30)
                lote = r.json() if r.content else []
            except Exception:
                lote = []
            if not isinstance(lote, list) or not lote:
                break
            for o in lote:
                pq = {}
                for li in (o.get("products") or []):
                    pid = "tn:%s" % li.get("product_id")
                    pq[pid] = pq.get(pid, 0) + int(li.get("quantity") or 0)
                ps = (o.get("payment_status") or "").lower()
                est = ("cancelled" if o.get("cancelled_at")
                       else "refunded" if ps in ("refunded", "voided", "partially_refunded")
                       else "paid" if ps == "paid" else "pending")
                out.append({"key": "tn:%s" % o.get("id"), "pid_qty": pq, "estado": est,
                            "fecha": str(o.get("created_at") or "")[:10]})
            if len(lote) < 50:
                break
            page += 1
    stk = _shop_tokens().get(email)
    if stk and stk.get("access_token") and stk.get("shop"):
        try:
            r = requests.get("https://%s/admin/api/2026-07/orders.json" % stk["shop"],
                             headers={"X-Shopify-Access-Token": stk["access_token"]},
                             params={"status": "any", "limit": 250,
                                     "created_at_min": desde + "T00:00:00-03:00",
                                     "fields": "id,financial_status,cancelled_at,created_at,line_items"},
                             timeout=30)
            for o in (r.json().get("orders", []) if r.status_code == 200 else []):
                pq = {}
                for li in (o.get("line_items") or []):
                    pid = str(li.get("product_id") or "")
                    if pid:
                        pq[pid] = pq.get(pid, 0) + int(li.get("quantity") or 0)
                fs = (o.get("financial_status") or "").lower()
                est = ("cancelled" if o.get("cancelled_at")
                       else "refunded" if fs in ("refunded", "partially_refunded", "voided")
                       else "paid" if fs in ("paid", "partially_paid") else "pending")
                out.append({"key": "sh:%s" % o.get("id"), "pid_qty": pq, "estado": est,
                            "fecha": str(o.get("created_at") or "")[:10]})
        except Exception:
            pass
    return out


def _stock_sync(email):
    """Aplica ventas pagadas (descuenta) y devoluciones/cancelaciones (repone). Idempotente."""
    st = _stk_read(STOCK_FILE, {}); prods = st.get(email) or {}
    if not prods:
        return prods, []
    led = _stk_read(STOCK_LEDGER, {}); ledE = led.get(email) or {}
    orders = _stock_orders(email, 90)
    changed = False
    for o in orders:
        deltas = {pid: q for pid, q in o["pid_qty"].items() if pid in prods and q > 0}
        if not deltas:
            continue
        entry = ledE.get(o["key"])
        if o["estado"] == "paid" and not entry:
            for pid, q in deltas.items():
                prods[pid]["stock"] = int(prods[pid].get("stock", 0)) - q
            ledE[o["key"]] = {"delta": deltas, "estado": "paid", "fecha": o["fecha"]}
            changed = True
        elif o["estado"] in ("refunded", "cancelled") and entry and entry.get("estado") == "paid":
            for pid, q in (entry.get("delta") or {}).items():
                if pid in prods:
                    prods[pid]["stock"] = int(prods[pid].get("stock", 0)) + q
            entry["estado"] = "reverted"
            changed = True
    if changed:
        st[email] = prods; _stk_write(STOCK_FILE, st)
        led[email] = ledE; _stk_write(STOCK_LEDGER, led)
    return prods, orders


def _stock_seed(email, pid):
    """Al empezar a trackear un producto, marca los pedidos ACTUALES como ya procesados
    (no retro-descuenta el histórico; solo cuentan las ventas NUEVAS)."""
    led = _stk_read(STOCK_LEDGER, {}); ledE = led.get(email) or {}
    for o in _stock_orders(email, 120):
        if pid in o["pid_qty"] and o["key"] not in ledE:
            ledE[o["key"]] = {"delta": {}, "estado": "seed"}
    led[email] = ledE; _stk_write(STOCK_LEDGER, led)


def _stock_metrics(email, pid, orders):
    hoy = _dt.datetime.utcnow().date()
    dias = {}
    for o in orders:
        if o["estado"] != "paid":
            continue
        q = o["pid_qty"].get(pid, 0)
        if not q:
            continue
        try:
            f = _dt.date.fromisoformat(o["fecha"])
        except Exception:
            continue
        off = (hoy - f).days
        if 0 <= off < 14:
            dias[off] = dias.get(off, 0) + q
    ventas14 = [dias.get(13 - i, 0) for i in range(14)]
    d14 = sum(ventas14); d7 = sum(ventas14[-7:]); d3 = sum(dias.get(off, 0) for off in range(3))
    # Promedio/día = RITMO RECIENTE y se recalcula en vivo con cada venta (no queda anclado a días
    # viejos flojos). Tomamos el mayor entre el ritmo de los ultimos 3 dias (tu pace de ahora) y el
    # de 7 dias (base semanal): al escalar sube al toque. Si no hubo ventas recientes, cae a 14 dias.
    if d3 or d7:
        rate = round(max(d3 / 3.0, d7 / 7.0))
    else:
        rate = round(d14 / 14) if d14 else 0
    return {"ventas14": ventas14, "d7": d7, "d14": d14, "rate": rate}


@app.get("/pf-stock")
def pf_stock():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "productos": []})
    prods, orders = _stock_sync(email)
    ped = (_stk_read(STOCK_PEDIDOS, {}).get(email) or [])
    out = []
    for pid, p in prods.items():
        m = _stock_metrics(email, pid, orders)
        pend = [x for x in ped if x.get("pid") == pid and x.get("estado") == "proceso"]
        out.append({"id": pid, "nombre": p.get("nombre", ""), "unidad": p.get("unidad", "u"),
                    "stock": int(p.get("stock", 0)), "costo": float(p.get("costo", 0)),
                    "sku": p.get("sku", ""), "ventas14": m["ventas14"], "d7": m["d7"],
                    "d14": m["d14"], "rate": m["rate"], "pendientes": pend})
    return jsonify({"ok": True, "productos": out})


@app.get("/pf-stock-catalogo")
def pf_stock_catalogo():
    """Productos de la tienda para elegir cuál trackear en Stock."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": True, "productos": []})
    cat = []
    for p in _tn_productos(email):
        cat.append({"id": p["id"], "nombre": p["nombre"], "sku": p.get("sku_base", ""),
                    "costo": _costo_num(p.get("costo"))})
    tk = _shop_tokens().get(email)
    if tk and tk.get("access_token") and tk.get("shop"):
        costos = (_costos().get(email) or {})
        try:
            r = requests.get("https://%s/admin/api/2026-07/products.json" % tk["shop"],
                             headers={"X-Shopify-Access-Token": tk["access_token"]},
                             params={"limit": 250, "fields": "id,title,variants"}, timeout=30)
            for pr in (r.json().get("products", []) if r.status_code == 200 else []):
                v = (pr.get("variants") or [{}])[0]
                pid = str(pr.get("id"))
                cat.append({"id": pid, "nombre": pr.get("title") or "",
                            "sku": v.get("sku") or "", "costo": _costo_num(costos.get(pid))})
        except Exception:
            pass
    return jsonify({"ok": True, "productos": cat})


@app.post("/pf-stock-set")
def pf_stock_set():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False}), 401
    d = request.get_json(silent=True) or {}
    pid = str(d.get("pid") or "").strip()
    if not pid:
        return jsonify({"ok": False, "msg": "falta el producto"}), 400
    st = _stk_read(STOCK_FILE, {}); prods = st.get(email) or {}
    prev = prods.get(pid, {}); nuevo = pid not in prods
    prods[pid] = {"nombre": d.get("nombre") or prev.get("nombre", ""),
                  "unidad": d.get("unidad") or prev.get("unidad", "u"),
                  "stock": int(d.get("stock", prev.get("stock", 0))),
                  "costo": float(d.get("costo", prev.get("costo", 0))),
                  "sku": d.get("sku") or prev.get("sku", "")}
    st[email] = prods; _stk_write(STOCK_FILE, st)
    if nuevo:
        _stock_seed(email, pid)
    return jsonify({"ok": True})


@app.post("/pf-stock-pedir")
def pf_stock_pedir():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False}), 401
    d = request.get_json(silent=True) or {}
    pid = str(d.get("pid") or ""); qty = int(d.get("qty") or 0)
    if not pid or qty <= 0:
        return jsonify({"ok": False, "msg": "cantidad inválida"}), 400
    import uuid
    allp = _stk_read(STOCK_PEDIDOS, {}); lst = allp.get(email) or []
    lst.insert(0, {"id": uuid.uuid4().hex[:10], "pid": pid, "qty": qty,
                   "fecha": _dt.datetime.utcnow().date().isoformat(), "estado": "proceso"})
    allp[email] = lst; _stk_write(STOCK_PEDIDOS, allp)
    return jsonify({"ok": True})


@app.post("/pf-stock-depositar")
def pf_stock_depositar():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False}), 401
    d = request.get_json(silent=True) or {}
    pedid = str(d.get("id") or "")
    allp = _stk_read(STOCK_PEDIDOS, {}); lst = allp.get(email) or []
    o = next((x for x in lst if x.get("id") == pedid and x.get("estado") == "proceso"), None)
    if not o:
        return jsonify({"ok": False, "msg": "pedido no encontrado"}), 404
    st = _stk_read(STOCK_FILE, {}); prods = st.get(email) or {}
    if o["pid"] in prods:
        prods[o["pid"]]["stock"] = int(prods[o["pid"]].get("stock", 0)) + int(o["qty"])
        st[email] = prods; _stk_write(STOCK_FILE, st)
    o["estado"] = "depositado"
    allp[email] = lst; _stk_write(STOCK_PEDIDOS, allp)
    return jsonify({"ok": True})


@app.get("/pf-mp-efectivo")
def pf_mp_efectivo():
    """Ventas en EFECTIVO (Rapipago/PagoFacil = payment_type 'ticket') del MercadoPago
    conectado del usuario logueado (ej. NoxaLab). Se abre en el navegador logueado."""
    email = _user_actual()
    if not email:
        return "No hay sesion. Entra a RealProfit primero (logueate) y volve a abrir este link.", 401
    tk = _mp_tokens().get(email)
    token = tk.get("access_token") if tk else None
    if not token:
        return "No tenes MercadoPago conectado en RealProfit (Integraciones -> MercadoPago).", 400
    begin = (_dt.datetime.utcnow() - _dt.timedelta(days=150)).date().isoformat() + "T00:00:00.000-03:00"
    end = _dt.datetime.utcnow().date().isoformat() + "T23:59:59.999-03:00"
    efectivo = []; ultimos = []; offset = 0
    try:
        while offset < 500:
            r = requests.get("https://api.mercadopago.com/v1/payments/search",
                             headers={"Authorization": "Bearer " + token},
                             params={"sort": "date_created", "criteria": "desc",
                                     "range": "date_created", "begin_date": begin, "end_date": end,
                                     "offset": offset, "limit": 50}, timeout=30)
            if r.status_code >= 400:
                if offset == 0:
                    return "MercadoPago rechazo la consulta (%d). Reconecta MP en RealProfit." % r.status_code, 502
                break
            res = r.json().get("results") or []
            if not res:
                break
            for pp in res:
                if len(ultimos) < 8:
                    ultimos.append(pp)
                if (pp.get("payment_type_id") or "") == "ticket":
                    efectivo.append(pp)
            if len(res) < 50:
                break
            offset += 50
    except Exception as e:
        return "Error consultando MP: %s" % e, 500

    def fmt(pp):
        pa = pp.get("payer") or {}; idf = pa.get("identification") or {}; ph = pa.get("phone") or {}
        nom = (str(pa.get("first_name") or "") + " " + str(pa.get("last_name") or "")).strip()
        tel = (str(ph.get("area_code") or "") + str(ph.get("number") or "")).strip()
        dni = (str(idf.get("type") or "") + " " + str(idf.get("number") or "")).strip()
        return ("<tr><td>%s</td><td style='text-align:right'>$%s</td><td>%s</td><td>%s</td><td>%s</td>"
                "<td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>") % (
            (pp.get("date_approved") or pp.get("date_created") or "")[:19],
            "{:,.0f}".format(float(pp.get("transaction_amount") or 0)).replace(",", "."),
            pp.get("status") or "", nom or "-", pa.get("email") or "-", dni or "-", tel or "-",
            pp.get("payment_method_id") or "", pp.get("external_reference") or "-", pp.get("id"))

    cab = ("<tr><th>Fecha</th><th>Monto</th><th>Estado</th><th>Nombre</th><th>Email</th>"
           "<th>DNI</th><th>Tel</th><th>Metodo</th><th>Ref pedido</th><th>MP id</th></tr>")
    if efectivo:
        cuerpo = "<h2>Ventas en EFECTIVO (Rapipago/PagoFacil) &mdash; %d</h2><table>%s%s</table>" % (
            len(efectivo), cab, "".join(fmt(x) for x in efectivo[:30]))
    else:
        cuerpo = ("<h2>No hay ventas en efectivo (ticket) en los ultimos 150 dias.</h2>"
                  "<p style='color:#93a3ba'>Te muestro los ultimos 8 pagos por si el metodo es otro:</p>"
                  "<table>%s%s</table>" % (cab, "".join(fmt(x) for x in ultimos)))
    html = ("<!doctype html><html><head><meta charset=utf-8><title>Ventas efectivo</title>"
            "<style>body{background:#0a0f18;color:#eef3f9;font-family:system-ui,-apple-system;padding:22px}"
            "h1{font-size:19px;margin:0 0 4px}h2{font-size:15px;color:#eef3f9;margin:18px 0 10px}"
            "table{border-collapse:collapse;width:100%;font-size:12.5px;overflow-x:auto;display:block}"
            "th,td{border:1px solid #1c2739;padding:8px 10px;text-align:left;white-space:nowrap}"
            "th{color:#93a3ba;text-transform:uppercase;font-size:10px;letter-spacing:.4px}"
            "td{font-variant-numeric:tabular-nums}</style></head><body>"
            "<h1>MercadoPago &mdash; %s</h1>%s</body></html>") % (email, cuerpo)
    return html


@app.get("/pf-recompras")
def pf_recompras():
    """Recompras (clientes que ya compraron antes) del período. Se pide APARTE del dashboard
    para no frenar la carga (el histórico de 90 días es pesado). Cacheado 10 min por dentro."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    desde = request.args.get("desde") or _hoy()
    hasta = request.args.get("hasta") or desde
    rc = 0
    rf = 0.0
    try:
        c1, f1 = _recompras_periodo(_tn_hist_orders(email, hasta), desde, hasta)
        rc += c1; rf += f1
    except Exception:
        pass
    try:
        c2, f2 = _recompras_periodo(_shop_hist_orders(email, hasta), desde, hasta)
        rc += c2; rf += f2
    except Exception:
        pass
    return jsonify({"ok": True, "recompras": rc, "fact_recompra": round(rf, 2)})


@app.get("/pf-version")
def pf_version():
    """Marcador de versión (sin login) para confirmar que el deploy está fresco."""
    return jsonify({"ok": True, "v": "2026-08-14-recompras34-async"})


@app.get("/pf-diag")
def pf_diag():
    """Diagnóstico (logueado): por qué la facturación puede dar 0. Muestra el desglose de estados
    de las órdenes y lo que calcula el resumen, para el período elegido (por defecto últimos 7 días)."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "motivo": "sin login"})
    try:
        dias = max(1, min(90, int(request.args.get("dias") or 7)))
    except Exception:
        dias = 7
    hasta = _hoy()
    desde = (_dt.date.fromisoformat(hasta) - _dt.timedelta(days=dias - 1)).isoformat()
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token"):
        return jsonify({"ok": True, "shopify_conectado": False,
                        "nota": "No hay tienda Shopify conectada para este usuario → facturación 0."})
    try:
        orders = _shopify_orders(tk.get("shop"), tk.get("access_token"), desde, hasta)
    except Exception as e:
        return jsonify({"ok": False, "error": "shopify", "detalle": str(e)[:120]})
    orders = [o for o in orders if not o.get("cancelled_at")]
    PAG = ("paid", "partially_paid", "refunded", "partially_refunded")
    por_estado, fact_pag, fact_total, n_pag = {}, 0.0, 0.0, 0
    for o in orders:
        st = (o.get("financial_status") or "sin_estado").lower()
        tot = float(o.get("total_price") or o.get("current_total_price") or 0)
        por_estado[st] = por_estado.get(st, 0) + 1
        fact_total += tot
        if st in PAG:
            fact_pag += tot; n_pag += 1
    blob = _shopify_resumen(email, desde, hasta) or {}
    raw = blob.get("raw", {})
    return jsonify({
        "ok": True, "shopify_conectado": True, "tienda": tk.get("shop"),
        "periodo": {"desde": desde, "hasta": hasta, "dias": dias},
        "ordenes_totales_no_canceladas": len(orders),
        "ordenes_por_estado": por_estado,
        "ordenes_PAGADAS": n_pag,
        "facturacion_TODAS": round(fact_total, 2),
        "facturacion_SOLO_PAGADAS": round(fact_pag, 2),
        "resumen_calculado": {
            "tot_facturado": raw.get("tot_facturado"),
            "ordenes": raw.get("ordenes"),
            "ticket": raw.get("ticket"),
            "ganancia": raw.get("ganancia"),
        },
        "explica": "Si 'ordenes_por_estado' tiene casi todo en 'pending' → por eso facturación baja/0 (no cuentan hasta pagarse)."
    })


_TN_ESTADO = {"paid": "Pagado", "pending": "Pendiente", "authorized": "Pendiente",
              "in_process": "Pendiente", "voided": "Anulado", "refunded": "Reembolsado",
              "partially_paid": "Pendiente", "abandoned": "—"}


def _tn_ventas(email, desde, hasta):
    """Ventas de Tiendanube (mismo formato que las filas de Shopify en pf-ventas)."""
    tk = _tn_tokens().get(email)
    if not tk or not tk.get("access_token") or not tk.get("store_id"):
        return None
    out = []
    for o in _tn_orders_raw(email, desde, hasta):        # órdenes cacheadas (compartidas con el dashboard)
        num = str(o.get("number") or "")
        if not num or o.get("cancelled_at"):
            continue
        if (o.get("payment_status") or "").lower() != "paid":   # como Shopify: pendiente NO cuenta
            continue
        tot = float(o.get("total") or 0)
        out.append({"num": num, "origen": "Tiendanube",
                    "estado": _TN_ESTADO.get((o.get("payment_status") or "").lower(), "—"),
                    "fecha": o.get("created_at") or "", "total": round(tot, 2),
                    "neto": round(tot * (1 - TIENDA_PCT / 100.0), 2)})
    return out


@app.get("/pf-ventas")
def pf_ventas():
    """Órdenes reales (Shopify + Tiendanube) de los últimos `dias` días para 'Últimas ventas'.
    Liviano (sin el fetch pesado de MP): neto exacto se ve en el detalle de cada orden."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "ventas": []})
    try:
        dias = max(1, min(90, int(request.args.get("dias") or 7)))
    except Exception:
        dias = 7
    hasta = _hoy()
    desde = (_dt.date.today() - _dt.timedelta(days=dias - 1)).isoformat()
    out = []
    tk = _shop_tokens().get(email)
    if tk and tk.get("access_token"):
        try:
            orders = _shopify_orders(tk.get("shop"), tk.get("access_token"), desde, hasta)
            for o in orders:
                if o.get("cancelled_at"):
                    continue
                tot = float(o.get("total_price") or o.get("current_total_price") or 0)
                num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
                out.append({"num": num, "origen": "Shopify",
                            "estado": _ESTADO_TXT.get((o.get("financial_status") or "").lower(), "—"),
                            "fecha": o.get("created_at") or "", "total": round(tot, 2),
                            "neto": round(tot * (1 - TIENDA_PCT / 100.0), 2)})
        except Exception:
            pass
    tn = _tn_ventas(email, desde, hasta)
    if tn:
        out.extend(tn)
    out.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    return jsonify({"ok": True, "ventas": out[:300]})


@app.get("/pf-orden")
def pf_orden():
    """Detalle de UNA orden para el modal 'Últimas ventas': productos (con foto), cliente,
    medio de pago y desglose financiero real (MP + cuotas + tienda + costo producto + neto)."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    num = (request.args.get("num") or "").strip().replace("#", "")
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token") or not num:
        return jsonify({"ok": False})
    shop, token = tk.get("shop"), tk.get("access_token")
    try:
        r = requests.get("https://%s/admin/api/2026-07/orders.json" % shop,
                         headers={"X-Shopify-Access-Token": token},
                         params={"name": "#" + num, "status": "any", "limit": 1,
                                 "fields": "id,order_number,name,created_at,financial_status,total_price,"
                                           "current_total_price,total_discounts,line_items,customer,"
                                           "contact_email,gateway,payment_gateway_names,shipping_lines,shipping_address"}, timeout=25)
        orders = (r.json() or {}).get("orders") or []
        o = orders[0] if orders else None
    except Exception:
        o = None
    if not o:
        return jsonify({"ok": False, "msg": "no encontrada"})
    tot = float(o.get("total_price") or o.get("current_total_price") or 0)
    desc = float(o.get("total_discounts") or 0)
    costos = (_costos().get(email) or {})
    items, costo_prod = [], 0.0
    for li in (o.get("line_items") or []):
        q = int(li.get("quantity") or 0)
        pid = str(li.get("product_id") or "")
        cu = costos.get(pid)
        costo_prod += _costo_qty(cu, q)
        items.append({"nombre": li.get("title") or li.get("name") or "?", "cantidad": q,
                      "precio": float(li.get("price") or 0), "foto": _shop_img(shop, token, pid)})
    # MP: matcheo el pago de esta orden por fecha/monto para el desglose real.
    fecha = str(o.get("created_at") or "")[:10] or _hoy()
    d0 = fecha; d1 = (_dt.date.fromisoformat(fecha) + _dt.timedelta(days=4)).isoformat() if fecha else _hoy()
    pagos = _mp_pagos_lista(email, d0, d1) if fecha else None
    pago = None
    if pagos:
        for p in pagos:
            if p["ref"] in (str(o.get("id")), str(o.get("order_number")), num):
                pago = p; break
        if pago is None:
            for p in pagos:
                if p["amount"] == round(tot):
                    pago = p; break
    fee_mp = pago["fee_mp"] if pago else 0.0
    fee_cuotas = pago["fee_cuotas"] if pago else 0.0
    inst = pago["inst"] if pago else 1
    tienda = tot * TIENDA_PCT / 100.0
    iibb = tot * IIBB_PCT / 100.0                       # Ingresos Brutos (3,5%)
    # Envío: costo REAL de Envialo si el pedido está ahí; si no, promedio domicilio/sucursal.
    _real_env = _envialo_costos(email).get(num)
    envio = _real_env if _real_env is not None else _envio_costo(o)
    # MP ya descuenta sus comisiones en pago["net"]; si no hay pago, lo estimamos.
    net_mp = pago["net"] if pago else round(tot - fee_mp - fee_cuotas, 2)
    # NETO REAL de la venta: lo que entra por MP menos fee tienda, IIBB, costo de producto y envío.
    neto = net_mp - tienda - iibb - costo_prod - envio
    cust = o.get("customer") or {}
    medio = "Mercado Pago" if pago else (", ".join(o.get("payment_gateway_names") or []) or o.get("gateway") or "—")
    return jsonify({"ok": True, "orden": {
        "num": num, "origen": "Shopify", "estado": _ESTADO_TXT.get((o.get("financial_status") or "").lower(), "—"),
        "fecha": o.get("created_at") or "",
        "cliente": (cust.get("first_name", "") + " " + cust.get("last_name", "")).strip() or (o.get("contact_email") or ""),
        "email": o.get("contact_email") or cust.get("email") or "",
        "medio": medio, "items": items,
        "total": round(tot, 2), "descuento": round(desc, 2),
        "fee_mp": round(fee_mp, 2), "fee_cuotas": round(fee_cuotas, 2), "cuotas": inst,
        "fee_tienda": round(tienda, 2), "iibb": round(iibb, 2), "envio": round(envio, 2),
        "envio_real": _real_env is not None,
        "costo_prod": round(costo_prod, 2), "neto": round(neto, 2)}})


# ==================== DESPACHOS (Andreani + Shopify/TiendaNube) ====================
DESP_STATE = DATA_DIR / "despachos_estado.json"   # {email: {order_num: "enviar"}} — marcados "por enviar"
_SUC_KEYS = ("sucursal", "pickup", "pick up", "pick-up", "retiro", "punto", "agenc", "hop")


def _desp_state(email) -> dict:
    try:
        return (_json.loads(DESP_STATE.read_text(encoding="utf-8"))).get(email, {})
    except Exception:
        return {}


def _desp_save(email, st) -> None:
    try:
        d = _json.loads(DESP_STATE.read_text(encoding="utf-8"))
    except Exception:
        d = {}
    d[email] = st
    DESP_STATE.parent.mkdir(parents=True, exist_ok=True)
    DESP_STATE.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


SKUS_FILE = DATA_DIR / "prod_skus.json"   # {email: {product_id: {"tipo": "xn|spray|fijo", "base": "..."}}}


def _skus_map(email) -> dict:
    try:
        return (_json.loads(SKUS_FILE.read_text(encoding="utf-8"))).get(email, {})
    except Exception:
        return {}


def _skus_save(email, m) -> None:
    try:
        d = _json.loads(SKUS_FILE.read_text(encoding="utf-8"))
    except Exception:
        d = {}
    d[email] = m
    SKUS_FILE.parent.mkdir(parents=True, exist_ok=True)
    SKUS_FILE.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


def _sku_cfg(v) -> dict:
    """Normaliza → {'tipo','base','map'}. tipo: unitario | variable | fijo.
    Compat con lo viejo: xn→unitario, spray→variable, string→fijo."""
    if isinstance(v, dict):
        t = (v.get("tipo") or "unitario")
        t = {"xn": "unitario", "spray": "variable"}.get(t, t)
        if t not in ("unitario", "variable", "fijo"):
            t = "unitario"
        return {"tipo": t, "base": (v.get("base") or ""), "map": (v.get("map") or {})}
    if isinstance(v, str) and v.strip():
        return {"tipo": "fijo", "base": v.strip(), "map": {}}
    return {"tipo": "unitario", "base": "", "map": {}}


def _sku_calc(cfg, u: int) -> str:
    """SKU de la orden según el tipo y la cantidad comprada `u`."""
    c = _sku_cfg(cfg)
    t, b, m = c["tipo"], (c["base"] or "").strip(), (c.get("map") or {})
    if t == "variable":
        s = (str(m.get(str(int(u))) or "")).strip()
        return s if s else ("x%d" % u)            # cantidad sin definir → xN
    if t == "fijo":
        return b
    # unitario: 'xN base' (ej 'x2 Pote')
    return (("x%d %s" % (u, b)).strip()) if b else ("x%d" % u)


@app.post("/pf-sku-set")
def pf_sku_set():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    data = request.get_json(silent=True) or {}
    pid = str(data.get("pid") or "").strip()
    if not pid:
        return jsonify({"ok": False})
    tipo = (data.get("tipo") or "unitario").strip()
    tipo = {"xn": "unitario", "spray": "variable"}.get(tipo, tipo)
    if tipo not in ("unitario", "variable", "fijo"):
        tipo = "unitario"
    base = str(data.get("base") or "").strip()
    smap = {}                                          # variable: {cantidad: sku}
    for k, v in (data.get("map") or {}).items():
        if str(k).isdigit() and str(v or "").strip():
            smap[str(int(k))] = str(v).strip()
    m = _skus_map(email)
    if tipo == "variable":
        if smap:
            m[pid] = {"tipo": "variable", "base": "", "map": smap}
        else:
            m[pid] = {"tipo": "variable", "base": "", "map": {}}
    elif base:
        m[pid] = {"tipo": tipo, "base": base}
    else:
        m.pop(pid, None)                               # sin base y no-variable → limpiar
    _skus_save(email, m)
    return jsonify({"ok": True, "ejemplos": {str(n): _sku_calc(m.get(pid), n) for n in (1, 2, 3)}})


def _es_sucursal_ship(o) -> bool:
    sl = o.get("shipping_lines") or []
    txt = " ".join(((s.get("title") or "") + " " + (s.get("code") or "")) for s in sl).lower()
    return any(k in txt for k in _SUC_KEYS)


def _dni_de(o) -> str:
    for na in (o.get("note_attributes") or []):
        n = (na.get("name") or "").lower()
        if any(k in n for k in ("dni", "documento", "cuil", "cuit")):
            v = "".join(ch for ch in str(na.get("value") or "") if ch.isdigit())
            if v:
                return v
    sa = o.get("shipping_address") or {}
    comp = "".join(ch for ch in str(sa.get("company") or "") if ch.isdigit())
    return comp


def _despachos_orders(email, desde=None, hasta=None):
    """TODOS los pedidos a despachar del usuario: Shopify + Tiendanube juntos, ya
    mapeados a la MISMA forma (para que pasen por el MISMO resolver Andreani, intacto)."""
    out = []
    sh = _despachos_orders_shopify(email, desde, hasta)
    tn = _tiendanube_orders(email, desde, hasta)
    if sh is None and tn is None:
        return None                      # ninguna tienda conectada
    if sh:
        out.extend(sh)
    if tn:
        out.extend(tn)
    out.sort(key=lambda x: int(x["num"]) if str(x["num"]).isdigit() else 0, reverse=True)
    return out


def _tn_shipping(o):
    """El envío real de un pedido de Tiendanube (tipo/opción/punto de retiro)."""
    fs = o.get("fulfillments") or []
    if fs and isinstance(fs, list) and isinstance(fs[0], dict):
        return fs[0].get("shipping") or {}
    return {}


def _tn_suc_nombre(sh):
    """String de la sucursal/HOP que eligió el cliente en Tiendanube. Se arma con la
    opción de envío + los datos del punto de retiro, para que el resolver Andreani lo
    matchee IGUAL que el título de Shopify (no cambia la lógica de sucursales)."""
    partes = [(sh.get("option") or {}).get("name") or ""]
    pd = sh.get("pickup_details") or {}
    for k in ("name", "address", "street", "number", "locality", "city", "zipcode"):
        v = pd.get(k)
        if v:
            partes.append(str(v))
    return " ".join(p for p in partes if p).strip()


def _tn_tel(o):
    """Primer teléfono usable del cliente de Tiendanube."""
    sa = o.get("shipping_address") or {}
    cust = o.get("customer") or {}
    for p in (sa.get("phone"), o.get("contact_phone"), o.get("billing_phone"), cust.get("phone")):
        p = (p or "").strip()
        if p and not p.lower().startswith("no inform") and _re_and.sub(r"\D", "", p):
            return p
    return ""


def _tiendanube_orders(email, desde=None, hasta=None):
    """Pedidos de Tiendanube PAGADOS y NO despachados (por empaquetar + por enviar),
    mapeados a la MISMA forma que _despachos_orders_shopify. Devuelve None si el usuario
    no tiene Tiendanube conectada."""
    tk = _tn_tokens().get(email)
    if not tk or not tk.get("access_token"):
        return None
    store, token = tk.get("store_id"), tk.get("access_token")
    if not store:
        return None
    hdr = _tn_headers(token)
    st = _desp_state(email)
    out, vistos = [], set()
    filtros = [{"payment_status": "paid", "shipping_status": "unpacked"},     # por empaquetar
               {"payment_status": "paid", "shipping_status": "unfulfilled"}]  # packed → por enviar
    for filt in filtros:
        page = 1
        while page <= 20:
            params = {"per_page": 200, "page": page, **filt}
            if desde:
                params["created_at_min"] = desde + "T00:00:00-03:00"
            if hasta:
                params["created_at_max"] = hasta + "T23:59:59-03:00"
            try:
                r = requests.get("%s/%s/orders" % (TN_API, store), headers=hdr,
                                 params=params, timeout=40)
                lote = r.json() if r.content else []
            except Exception:
                lote = []
            if not isinstance(lote, list) or not lote:
                break
            for o in lote:
                num = str(o.get("number") or "")
                if not num or num in vistos:
                    continue
                vistos.add(num)
                if o.get("cancelled_at"):
                    continue
                if (o.get("payment_status") or "") != "paid":
                    continue
                unidades = sum(int(p.get("quantity") or 0) for p in (o.get("products") or []))
                if unidades == 0:                       # ebook-only → no se despacha
                    continue
                sh = _tn_shipping(o)
                es_suc = sh.get("type") == "pickup"
                sa = o.get("shipping_address") or {}
                cust = o.get("customer") or {}
                nombre = (sa.get("name") or o.get("contact_name") or cust.get("name") or "—")
                calle = (str(sa.get("address") or "").strip() + " " + str(sa.get("number") or "").strip()).strip()
                floor = str(sa.get("floor") or "").strip()
                localidad = (sa.get("locality") or sa.get("city") or "").strip()
                cp = str(sa.get("zipcode") or "").strip()
                prov = (sa.get("province") or "").strip()
                tel = _tn_tel(o)
                dni = str(o.get("contact_identification") or cust.get("identification") or "").strip()
                incompleta = (not es_suc) and (not calle or not cp or not tel)
                estado = st.get(num) or "empaquetar"
                out.append({
                    "num": num, "nombre": (nombre or "").strip(),
                    "tipo": "sucursal" if es_suc else "domicilio",
                    "localidad": localidad, "cp": cp, "provincia": prov, "unidades": unidades,
                    "total": round(float(o.get("total") or 0), 2),
                    "tel": tel, "dni": dni, "fecha": o.get("created_at") or o.get("completed_at") or "",
                    "email": o.get("contact_email") or (cust.get("email") or ""),
                    "suc_nombre": _tn_suc_nombre(sh), "calle": calle, "extra": floor,
                    "incompleta": incompleta, "estado": estado,
                })
            if len(lote) < 200:
                break
            page += 1
    out.sort(key=lambda x: int(x["num"]) if str(x["num"]).isdigit() else 0, reverse=True)
    return out


def _despachos_orders_shopify(email, desde=None, hasta=None):
    """Órdenes de Shopify PAGADAS y NO despachadas (para despachar por Andreani).
    Solo entran las PAGADAS (si no está paga, no aparece). Filtra por fecha si se pasa desde/hasta."""
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token"):
        return None
    shop, token = tk.get("shop"), tk.get("access_token")
    params = {"status": "open", "financial_status": "paid",
              "fulfillment_status": "unshipped", "limit": 250,
              "fields": "id,order_number,name,total_price,current_total_price,"
                        "financial_status,fulfillment_status,cancelled_at,line_items,"
                        "created_at,shipping_lines,shipping_address,customer,contact_email,"
                        "note_attributes"}
    if desde:
        params["created_at_min"] = desde + "T00:00:00-03:00"
    if hasta:
        params["created_at_max"] = hasta + "T23:59:59-03:00"
    try:
        r = requests.get("https://%s/admin/api/2026-07/orders.json" % shop,
                         headers={"X-Shopify-Access-Token": token},
                         params=params, timeout=40)
        orders = (r.json() or {}).get("orders") or []
    except Exception:
        return []
    st = _desp_state(email)
    out = []
    for o in orders:
        if o.get("cancelled_at"):
            continue
        if (o.get("financial_status") or "").lower() != "paid":   # doble seguro: solo pagadas
            continue
        num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
        sa = o.get("shipping_address") or {}
        cust = o.get("customer") or {}
        nombre = (sa.get("name") or ((cust.get("first_name", "") + " " + cust.get("last_name", "")).strip())
                  or o.get("contact_email") or "—")
        suc = _es_sucursal_ship(o)
        unidades = sum(int(li.get("quantity") or 0) for li in (o.get("line_items") or []))
        localidad = sa.get("city") or ""
        cp = sa.get("zip") or ""
        prov = sa.get("province") or ""
        tel = sa.get("phone") or (cust.get("phone") or "")
        # Domicilio incompleto (sucursal no necesita dirección de casa).
        incompleta = (not suc) and (not sa.get("address1") or not cp or not tel)
        estado = st.get(num) or "empaquetar"   # exportada / enviada / (default) empaquetar
        out.append({
            "num": num, "nombre": nombre.strip(), "tipo": "sucursal" if suc else "domicilio",
            "localidad": localidad, "cp": cp, "provincia": prov, "unidades": unidades,
            "total": round(float(o.get("total_price") or o.get("current_total_price") or 0), 2),
            "tel": tel, "dni": _dni_de(o), "fecha": o.get("created_at") or "",
            "email": o.get("contact_email") or (cust.get("email") or ""),
            "suc_nombre": " ".join((s.get("title") or "") for s in (o.get("shipping_lines") or [])).strip(),
            "calle": sa.get("address1") or "", "extra": sa.get("address2") or "",
            "incompleta": incompleta, "estado": estado,
        })
    out.sort(key=lambda x: int(x["num"]) if str(x["num"]).isdigit() else 0, reverse=True)
    return out


@app.get("/pf-despachos")
def pf_despachos_list():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "rows": []})
    desde = request.args.get("desde") or None
    hasta = request.args.get("hasta") or None
    rows = _despachos_orders(email, desde, hasta)
    if rows is None:
        return jsonify({"ok": True, "shopify": False, "rows": []})

    def _suma(lst):
        return round(sum(r["total"] for r in lst), 2)
    grp = {e: [r for r in rows if r["estado"] == e] for e in ("empaquetar", "exportada", "enviada")}
    return jsonify({"ok": True, "shopify": True, "rows": rows,
                    "resumen": {
                        "empaquetar": {"n": len(grp["empaquetar"]), "monto": _suma(grp["empaquetar"])},
                        "exportada": {"n": len(grp["exportada"]), "monto": _suma(grp["exportada"])},
                        "enviada": {"n": len(grp["enviada"]), "monto": _suma(grp["enviada"])},
                        "todas": {"n": len(rows), "monto": _suma(rows)}}})


# ============================ FACTURACIÓN ============================
# RealProfit no emite en ARCA (eso vive en METAFY con el certificado AFIP).
# Acá listamos las VENTAS pagadas de la tienda y llevamos el control de cuáles
# están facturadas (marcado manual), con KPIs y export para el contador.
FACT_STATE = DATA_DIR / "factura_estado.json"   # {email: {order_num: {"comp": "", "fecha": ""}}}


def _fact_state(email) -> dict:
    try:
        return (_json.loads(FACT_STATE.read_text(encoding="utf-8"))).get(email, {})
    except Exception:
        return {}


def _fact_save(email, st) -> None:
    try:
        d = _json.loads(FACT_STATE.read_text(encoding="utf-8"))
    except Exception:
        d = {}
    d[email] = st
    FACT_STATE.parent.mkdir(parents=True, exist_ok=True)
    FACT_STATE.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


# ===================== MOVIMIENTOS & SOCIOS =====================
MOV_STATE = DATA_DIR / "movimientos.json"   # {email: {"rows": [...], "seq": int, "_sv": int}}

_MOV_SEED = []            # arranca TODO en 0 (el usuario carga lo real)
MOV_SEED_VERSION = 2      # subir este número resetea los datos sembrados viejos una sola vez


def _mov_all() -> dict:
    try:
        return _json.loads(MOV_STATE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _mov_write(d) -> None:
    MOV_STATE.parent.mkdir(parents=True, exist_ok=True)
    MOV_STATE.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


def _mov_get(email) -> dict:
    d = _mov_all()
    cur = d.get(email)
    # sin datos, o versión de seed vieja → (re)sembrar. Con _MOV_SEED vacío queda en 0.
    if not cur or cur.get("_sv") != MOV_SEED_VERSION:
        rows, seq = [], 0
        for m in _MOV_SEED:
            seq += 1
            r = dict(m); r["id"] = seq
            rows.append(r)
        cur = {"rows": rows, "seq": seq, "_sv": MOV_SEED_VERSION}
        d[email] = cur
        _mov_write(d)
    return cur


def _mov_save(email, cur) -> None:
    d = _mov_all()
    d[email] = cur
    _mov_write(d)


_MOV_MP_CACHE = {}   # {email: (ts, rows)} — cache de los movimientos auto de MP


def _mp_movimientos(email, desde, hasta):
    """Movimientos REALES de Mercado Pago: ingreso = neto que deposita MP por cada venta
    (net_received_amount, sin restarle nada mas), y devolucion = lo reembolsado. Auto."""
    tk = _mp_tokens().get(email)
    token = tk.get("access_token") if tk else None
    if not token:
        return []
    ini = desde + "T00:00:00.000-03:00"
    fin = hasta + "T23:59:59.999-03:00"
    out = []
    offset = 0
    try:
        while offset < 5000:
            r = requests.get("https://api.mercadopago.com/v1/payments/search",
                             headers={"Authorization": "Bearer " + token},
                             params={"sort": "date_approved", "criteria": "desc",
                                     "range": "date_approved", "begin_date": ini, "end_date": fin,
                                     "offset": offset, "limit": 100}, timeout=30)
            if r.status_code >= 400:
                break
            data = r.json(); res = data.get("results") or []
            for pmt in res:
                if pmt.get("status") not in ("approved", "refunded"):
                    continue
                det = pmt.get("transaction_details") or {}
                net = float(det.get("net_received_amount") or 0)
                fecha = str(pmt.get("date_approved") or pmt.get("date_created") or "")[:10]
                ref = (pmt.get("external_reference") or "").strip()
                pid = str(pmt.get("id"))
                if net > 0:
                    out.append({"id": "mp:" + pid, "d": fecha, "clase": "ingreso", "cat": "Venta MP",
                                "desc": ("Pedido #" + ref) if ref else ("Pago MP " + pid),
                                "socio": "marca", "monto": round(net), "auto": True})
                refd = float(pmt.get("transaction_amount_refunded") or 0)
                if refd > 0:
                    out.append({"id": "mpr:" + pid, "d": fecha, "clase": "devolucion", "cat": "Devolucion MP",
                                "desc": ("Reembolso #" + ref) if ref else ("Reembolso " + pid),
                                "socio": "marca", "monto": round(refd), "auto": True})
            offset += 100
            if offset >= (data.get("paging") or {}).get("total", 0) or not res:
                break
    except Exception:
        return out
    return out


@app.get("/pf-movimientos")
def pf_movimientos():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "rows": []})
    cur = _mov_get(email)
    rows = list(cur.get("rows", []))
    # AUTO: neto real de cada venta de Mercado Pago + devoluciones (ultimos 60 dias, cacheado)
    import time as _t
    c = _MOV_MP_CACHE.get(email)
    if c and (_t.time() - c[0] < 180):
        mp_rows = c[1]
    else:
        try:
            hasta = _hoy(); desde = _hoy()   # SOLO HOY (desde las 00:00)
            mp_rows = _mp_movimientos(email, desde, hasta)
        except Exception:
            mp_rows = []
        _MOV_MP_CACHE[email] = (_t.time(), mp_rows)
    rows = rows + mp_rows
    base = cur.get("base")
    if base and base.get("fecha") == _hoy():
        rows.append({"id": "base", "d": base["fecha"], "clase": "ingreso", "cat": "Saldo inicial",
                     "desc": "Saldo en MP al iniciar el dia", "socio": "marca",
                     "monto": base["monto"], "auto": True})
    rows.sort(key=lambda r: str(r.get("d") or ""), reverse=True)
    return jsonify({"ok": True, "rows": rows})


@app.post("/pf-movimientos-add")
def pf_movimientos_add():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    data = request.get_json(silent=True) or {}
    clase = str(data.get("clase", "")).strip()
    if clase not in ("ingreso", "egreso", "aporte", "devolucion"):
        return jsonify({"ok": False, "msg": "tipo inválido"})
    try:
        monto = round(float(data.get("monto") or 0))
    except Exception:
        monto = 0
    if monto <= 0:
        return jsonify({"ok": False, "msg": "monto inválido"})
    socio = str(data.get("socio", "marca")).strip() or "marca"
    if socio not in ("marca", "cristian", "socio"):
        socio = "marca"
    cur = _mov_get(email)
    seq = int(cur.get("seq", 0)) + 1
    row = {"id": seq,
           "d": str(data.get("d", "") or "—")[:10],
           "clase": clase,
           "cat": str(data.get("cat", "") or "Gasto")[:40],
           "desc": str(data.get("desc", "") or "")[:120],
           "socio": socio,
           "monto": monto}
    cur.setdefault("rows", []).append(row)
    cur["seq"] = seq
    _mov_save(email, cur)
    return jsonify({"ok": True, "id": seq})


@app.post("/pf-movimientos-base")
def pf_movimientos_base():
    """Fija el 'saldo inicial' del dia con el saldo ACTUAL de MP que pega el usuario.
    base = saldo_actual - (neto de hoy - devoluciones de hoy) -> asi base + hoy = saldo real."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    data = request.get_json(silent=True) or {}
    raw = str(data.get("saldo") or "").replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        saldo = round(float(raw or 0))
    except Exception:
        saldo = 0
    hoy = _hoy()
    try:
        mp = _mp_movimientos(email, hoy, hoy)
    except Exception:
        mp = []
    hoy_neto = (sum(m["monto"] for m in mp if m["clase"] == "ingreso")
                - sum(m["monto"] for m in mp if m["clase"] == "devolucion"))
    cur = _mov_get(email)
    cur["base"] = {"fecha": hoy, "monto": round(saldo - hoy_neto)}
    _mov_save(email, cur)
    _MOV_MP_CACHE.pop(email, None)
    return jsonify({"ok": True, "base": cur["base"]["monto"], "saldo": saldo})


@app.post("/pf-movimientos-del")
def pf_movimientos_del():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    data = request.get_json(silent=True) or {}
    try:
        mid = int(data.get("id"))
    except Exception:
        return jsonify({"ok": False})
    cur = _mov_get(email)
    cur["rows"] = [r for r in cur.get("rows", []) if int(r.get("id", 0)) != mid]
    _mov_save(email, cur)
    return jsonify({"ok": True})


def _facturacion_orders(email, desde=None, hasta=None):
    """VENTAS pagadas de la tienda (todas, despachadas o no) para facturar/controlar.
    Devuelve None si la tienda no está conectada."""
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token"):
        return None
    shop, token = tk.get("shop"), tk.get("access_token")
    params = {"status": "any", "financial_status": "paid", "limit": 250,
              "fields": "id,order_number,name,total_price,current_total_price,"
                        "financial_status,cancelled_at,line_items,created_at,"
                        "customer,contact_email,billing_address,shipping_address,"
                        "gateway,payment_gateway_names"}
    if desde:
        params["created_at_min"] = desde + "T00:00:00-03:00"
    if hasta:
        params["created_at_max"] = hasta + "T23:59:59-03:00"
    try:
        r = requests.get("https://%s/admin/api/2026-07/orders.json" % shop,
                         headers={"X-Shopify-Access-Token": token},
                         params=params, timeout=40)
        orders = (r.json() or {}).get("orders") or []
    except Exception:
        return []
    st = _fact_state(email)
    out = []
    for o in orders:
        if o.get("cancelled_at"):
            continue
        if (o.get("financial_status") or "").lower() != "paid":
            continue
        num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
        ba = o.get("billing_address") or o.get("shipping_address") or {}
        cust = o.get("customer") or {}
        nombre = (ba.get("name") or ((cust.get("first_name", "") + " " + cust.get("last_name", "")).strip())
                  or o.get("contact_email") or "—")
        names = o.get("payment_gateway_names") or []
        gw = (names[0] if names else (o.get("gateway") or "")).lower()
        if "mercado" in gw:
            medio, prio = "MercadoPago", 2
        elif any(k in gw for k in ("transfer", "bank", "manual", "efect", "cash", "deposit")):
            medio, prio = "Transferencia", 3
        elif any(k in gw for k in ("card", "credit", "debit", "shopify_payments", "stripe", "tarjeta")):
            medio, prio = "Tarjeta", 1
        else:
            medio, prio = ((names[0] if names else "Otro") or "Otro").title(), 2
        fch = (o.get("created_at") or "")[:10]
        fdmy = ("%s/%s/%s" % (fch[8:10], fch[5:7], fch[0:4])) if len(fch) == 10 else ""
        fac = st.get(num) if isinstance(st.get(num), dict) else None
        out.append({
            "num": num, "nombre": nombre.strip(),
            "dni": _dni_de(o), "fecha": o.get("created_at") or "", "fecha_dmy": fdmy,
            "total": round(float(o.get("total_price") or o.get("current_total_price") or 0), 2),
            "email": o.get("contact_email") or (cust.get("email") or ""),
            "localidad": ba.get("city") or "", "cp": ba.get("zip") or "",
            "provincia": ba.get("province") or "",
            "dir_fiscal": ((ba.get("address1") or "") + " " + (ba.get("address2") or "")).strip(),
            "medio": medio, "prio": prio,
            "facturada": bool(fac), "comprobante": (fac or {}).get("comp", ""),
            "emit": (fac or {}).get("emit", ""), "fecha_fact": (fac or {}).get("fecha", ""),
        })
    out.sort(key=lambda x: int(x["num"]) if str(x["num"]).isdigit() else 0, reverse=True)
    return out


@app.get("/pf-facturacion")
def pf_facturacion_list():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "rows": []})
    desde = request.args.get("desde") or None
    hasta = request.args.get("hasta") or None
    _ac0 = _arca_get(email)
    rows = _facturacion_orders(email, desde, hasta)
    if rows is None:
        return jsonify({"ok": True, "shopify": False, "rows": [],
                        "arca": {"nombre": _ac0.get("nombre", ""), "cuit": _ac0.get("cuit", ""),
                                 "pv": _ac0.get("pv", "5"), "vinculado": bool(_ac0.get("cuit"))}})
    # --- Facturación automática: marca las ventas nuevas que ya cumplieron el retraso ---
    auto = _auto_cfg(email)
    if auto.get("on") and auto.get("since"):
        try:
            since = _dt.datetime.fromisoformat(auto["since"])
            ahora = _dt.datetime.now(since.tzinfo) if since.tzinfo else _dt.datetime.now()
            delay = _dt.timedelta(minutes=int(auto.get("delay", 10)))
            nuevas = []
            for r in rows:
                if r["facturada"]:
                    continue
                try:
                    cr = _dt.datetime.fromisoformat((r.get("fecha") or "").replace("Z", "+00:00"))
                    cr_n = cr.replace(tzinfo=None)
                except Exception:
                    continue
                if cr_n > since.replace(tzinfo=None) and (ahora.replace(tzinfo=None) - cr_n) >= delay:
                    nuevas.append(r["num"])
            if nuevas:
                _marcar_facturadas(email, nuevas, auto.get("emit", ""))
                rows = _facturacion_orders(email, desde, hasta) or rows
        except Exception:
            pass
    fact = [r for r in rows if r["facturada"]]
    pend = [r for r in rows if not r["facturada"]]
    tk = _shop_tokens().get(email) or {}
    shopnm = (tk.get("shop") or "").split(".")[0]
    ac = _arca_get(email)

    def _suma(lst):
        return round(sum(r["total"] for r in lst), 2)
    return jsonify({"ok": True, "shopify": True, "rows": rows,
                    "arca": {"nombre": ac.get("nombre") or shopnm or email,
                             "cuit": ac.get("cuit", ""), "pv": ac.get("pv", "5"),
                             "tipo": ac.get("tipo", "Factura C · Monotributo"),
                             "metodo": ac.get("metodo", ""),
                             "vinculado": bool(ac.get("cuit"))},
                    "auto": {"on": bool(auto.get("on")), "delay": int(auto.get("delay", 10))},
                    "resumen": {
                        "ventas": {"n": len(rows), "monto": _suma(rows)},
                        "facturadas": {"n": len(fact), "monto": _suma(fact)},
                        "pendientes": {"n": len(pend), "monto": _suma(pend)}}})


def _marcar_facturadas(email, nums, emit=""):
    """Marca ventas como facturadas asignando comprobante (uso interno / auto)."""
    st = _fact_state(email)
    try:
        seq = int(st.get("_seq", 0))
    except Exception:
        seq = 0
    for n in [str(x) for x in nums]:
        prev = st.get(n) if isinstance(st.get(n), dict) else {}
        comp = prev.get("comp")
        if not comp:
            seq += 1
            comp = "00005-%08d" % seq
        st[n] = {"comp": comp, "emit": emit or prev.get("emit", ""), "fecha": prev.get("fecha", "")}
    st["_seq"] = seq
    _fact_save(email, st)


@app.post("/pf-facturacion-marcar")
def pf_facturacion_marcar():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    data = request.get_json(silent=True) or {}
    nums = [str(n) for n in (data.get("nums") or [])]
    accion = data.get("accion", "facturada")   # facturada / pendiente
    emit = (data.get("emit") or "").strip()     # fecha de emisión elegida (dd/mm/aaaa)
    st = _fact_state(email)
    try:
        seq = int(st.get("_seq", 0))
    except Exception:
        seq = 0
    for n in nums:
        if accion == "pendiente":
            st.pop(n, None)
        else:
            prev = st.get(n) if isinstance(st.get(n), dict) else {}
            comp = prev.get("comp")
            if not comp:
                seq += 1
                comp = "00005-%08d" % seq
            st[n] = {"comp": comp, "emit": emit or prev.get("emit", ""),
                     "fecha": prev.get("fecha", "")}
    st["_seq"] = seq
    _fact_save(email, st)
    return jsonify({"ok": True, "n": len(nums)})


# --- Facturación automática (flag on/off por usuario) ---
AUTO_FACT = DATA_DIR / "auto_fact.json"   # {email: {"on": bool, "delay": 10, "since": iso}}


def _auto_cfg(email) -> dict:
    try:
        return (_json.loads(AUTO_FACT.read_text(encoding="utf-8"))).get(email, {}) or {}
    except Exception:
        return {}


@app.post("/pf-facturacion-auto")
def pf_facturacion_auto():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    data = request.get_json(silent=True) or {}
    try:
        d = _json.loads(AUTO_FACT.read_text(encoding="utf-8"))
    except Exception:
        d = {}
    cur = d.get(email, {}) or {}
    on = bool(data.get("on"))
    d[email] = {"on": on, "delay": int(data.get("delay", 10)),
                "emit": (data.get("emit") or "").strip(),
                "since": _dt.datetime.now().isoformat() if on and not cur.get("on") else cur.get("since", "")}
    AUTO_FACT.parent.mkdir(parents=True, exist_ok=True)
    AUTO_FACT.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return jsonify({"ok": True, "auto": d[email]})


# --- Vinculación ARCA (config por usuario) ---
ARCA_CFG = DATA_DIR / "arca_config.json"   # {email: {cuit, pv, nombre, tipo, metodo, token}}


def _arca_get(email) -> dict:
    try:
        return (_json.loads(ARCA_CFG.read_text(encoding="utf-8"))).get(email, {}) or {}
    except Exception:
        return {}


@app.post("/pf-facturacion-arca")
def pf_facturacion_arca():
    """Vincula ARCA: guarda CUIT + Punto de Venta + método (token de servicio o certificado)."""
    import re
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "msg": "sin sesión"})
    data = request.get_json(silent=True) or {}
    cuit = re.sub(r"\D", "", str(data.get("cuit") or ""))
    pv = re.sub(r"\D", "", str(data.get("pv") or "")) or "5"
    if len(cuit) != 11:
        return jsonify({"ok": False, "msg": "El CUIT debe tener 11 dígitos."})
    try:
        d = _json.loads(ARCA_CFG.read_text(encoding="utf-8"))
    except Exception:
        d = {}
    cuit_fmt = "%s-%s-%s" % (cuit[:2], cuit[2:10], cuit[10:])
    d[email] = {"cuit": cuit_fmt, "pv": pv,
                "nombre": (data.get("nombre") or "").strip(),
                "tipo": (data.get("tipo") or "Factura C · Monotributo").strip(),
                "metodo": (data.get("metodo") or "").strip(),
                "token": (data.get("token") or "").strip()}
    ARCA_CFG.parent.mkdir(parents=True, exist_ok=True)
    ARCA_CFG.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return jsonify({"ok": True, "arca": {k: v for k, v in d[email].items() if k != "token"}})


@app.post("/pf-facturacion-informe")
def pf_facturacion_informe():
    """Exporta un .xlsx de las ventas FACTURADAS del período para el contador."""
    import openpyxl
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "msg": "sin sesión"}), 400
    data = request.get_json(silent=True) or {}
    desde = data.get("desde") or None
    hasta = data.get("hasta") or None
    rows = _facturacion_orders(email, desde, hasta) or []
    fact = [r for r in rows if r["facturada"]]
    if not fact:
        return jsonify({"ok": False, "msg": "No hay ventas marcadas como facturadas en ese período."}), 400
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturación"
    ws.append(["Fecha emisión", "Fecha venta", "Pedido", "Cliente", "CUIT/DNI", "Medio",
               "Localidad", "Provincia", "CP", "Dirección", "Email", "Comprobante", "Total"])
    for r in sorted(fact, key=lambda x: int(x["num"]) if str(x["num"]).isdigit() else 0):
        ws.append([r.get("emit", "") or (r.get("fecha") or "")[:10], r.get("fecha_dmy", "") or (r.get("fecha") or "")[:10],
                   r["num"], r["nombre"], r.get("dni", ""), r.get("medio", ""),
                   r.get("localidad", ""), r.get("provincia", ""), r.get("cp", ""),
                   r.get("dir_fiscal", ""), r.get("email", ""), r.get("comprobante", ""),
                   r["total"]])
    out = DATA_DIR / ("Informe-Facturacion-%s.xlsx" % _dt.datetime.now().strftime("%Y%m%d-%H%M%S"))
    wb.save(str(out))
    return send_file(str(out), as_attachment=True, download_name=out.name)


@app.post("/pf-despachos-marcar")
def pf_despachos_marcar():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    data = request.get_json(silent=True) or {}
    nums = [str(n) for n in (data.get("nums") or [])]
    accion = (data.get("accion") or "exportada").strip()   # exportada / enviada / empaquetar
    st = _desp_state(email)
    for n in nums:
        if accion == "empaquetar":
            st.pop(n, None)
        else:
            st[n] = accion
    _desp_save(email, st)
    return jsonify({"ok": True, "n": len(nums), "accion": accion})


ANDREANI_TPL = RAIZ / "EnvioMasivoExcelPaquetes.xlsx"


def _split_nombre(nm):
    p = (nm or "").strip().split()
    if len(p) <= 1:
        return (nm or ""), ""
    return " ".join(p[:-1]), p[-1]


def _calle_num(dir_):
    import re
    d = (dir_ or "").strip()
    m = re.search(r"(\d+)\s*$", d)
    if m:
        return (d[:m.start()].strip().rstrip(",") or d), m.group(1)
    m2 = re.search(r"\d+", d)
    if m2:
        return (d.replace(m2.group(0), "").strip() or d), m2.group(0)
    return d, ""


# ===== Resolvedor Andreani: sucursal exacta + Prov/Loc/CP oficial (sin login, endpoints públicos) =====
import re as _re_and
import unicodedata as _ud_and

_AND_CFG = {"cpidx": None, "sucs": None}


def _and_norm(s):
    s = _ud_and.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return _re_and.sub(r"[^A-Z0-9 ]", " ", s.upper())


def _and_normP(s):
    s = _ud_and.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return _re_and.sub(r"\s+", " ", s.strip().upper())


def _and_toks(s):
    return [w for w in _and_norm(s).split() if w]


def _and_cp4(x):
    m = _re_and.search(r"(\d{4})", str(x or ""))
    return m.group(1) if m else ""


def _and_split_tel(p):
    d = _re_and.sub(r"\D", "", str(p or ""))
    d = _re_and.sub(r"^0", "", d)
    d = _re_and.sub(r"^549?", "", d)
    if d.startswith("11"):
        return "11", d[2:]
    return d[:3], d[3:]


def _and_cfg(wb):
    if _AND_CFG["cpidx"] is None:
        cpidx, sucs = {}, []
        try:
            for row in wb["Configuracion"].iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    sucs.append(str(row[0]).strip())
                if row and len(row) > 4 and row[4]:
                    v = str(row[4]).strip()
                    cpidx.setdefault(_re_and.sub(r"\D", "", v.split("/")[-1]), []).append(v)
        except Exception:
            pass
        _AND_CFG["cpidx"], _AND_CFG["sucs"] = cpidx, sucs
    return _AND_CFG["cpidx"], _AND_CFG["sucs"]


def _and_pcl(cp, prov, cpidx):
    cands = cpidx.get(_and_cp4(cp), [])
    if not cands:
        return None
    pn = _and_normP(prov)
    mism = [c for c in cands if pn and (pn in _and_normP(c) or _and_normP(c.split("/")[0]) in pn)]
    pool = sorted(mism or cands, key=lambda c: len(c.split("/")[1]) if len(c.split("/")) > 2 else 99)
    return pool[0]


def _and_mid(zeny):
    m = _re_and.search(r"[-–]\s*(.*?)\s*\((.*)\)\s*$", str(zeny))
    return (m.group(1) if m else _re_and.sub(r".*?[-–]\s*", "", str(zeny))).strip()


def _and_suc_excel(zeny, sucs):
    es_hop = "HOP" in str(zeny).upper()
    q = set(_and_toks(_and_mid(zeny)))
    qn = {w for w in q if w.isdigit()}
    qw = {w for w in q if not w.isdigit() and len(w) >= 4}
    best, bs, bnum, bword = None, -1, False, False
    for o in sucs:
        if es_hop != _and_norm(o).startswith("PUNTO ANDREANI HOP"):
            continue
        ts = set(_and_toks(o))
        nums = {w for w in ts if w.isdigit()}
        sc = sum(len(w) for w in (q & ts))
        numok = bool(qn and nums and qn & nums)
        if numok:
            sc += 25
        if sc > bs:
            bs, best, bnum, bword = sc, o, numok, bool(qw & {w for w in ts if len(w) >= 4})
    conf = (bnum and bword) if es_hop else (bs >= 4)
    return best, conf


def _and_suc_live(zeny, cp):
    """HOP exacto con los endpoints PÚBLICOS de Andreani (autosuggest + byCoordenadas). Sin login."""
    hd = {"User-Agent": "Mozilla/5.0", "Referer": "https://www.andreani.com/buscar-sucursal"}
    es_hop = "HOP" in str(zeny).upper()
    addr = _and_mid(zeny)
    qn = {w for w in _and_toks(addr) if w.isdigit()}
    qw = {w for w in _and_toks(addr) if not w.isdigit() and len(w) >= 4}
    if not qn:
        return None
    for q in [f"{addr}, {cp}", addr]:
        try:
            js = requests.get("https://www.andreani.com/api/autosuggest",
                              params={"q": q, "limit": 6}, headers=hd, timeout=12).json()
        except Exception:
            continue
        for it in (js.get("items") or []):
            pos = it.get("position") or {}
            if not pos:
                continue
            try:
                pts = requests.get("https://www.andreani.com/api/sucursales/byCoordenadas",
                                   params={"lat": pos["lat"], "lng": pos["lng"]}, headers=hd, timeout=12).json()
            except Exception:
                continue
            for p in (pts if isinstance(pts, list) else []):
                nom = p.get("descripcion", "")
                if es_hop and "hop" not in nom.lower():
                    continue
                nt = set(_and_toks(nom))
                if (qn & {w for w in nt if w.isdigit()}) and (qw & {w for w in nt if len(w) >= 4}):
                    return nom
    return None


@app.post("/pf-despachos-excel")
def pf_despachos_excel():
    """Genera el Excel de carga masiva de Andreani (2 hojas: A domicilio / A sucursal)
    con los pedidos seleccionados, mapeando los campos de Shopify a la plantilla.
    Resuelve la sucursal al nombre OFICIAL de Andreani y arregla DNI/teléfono/Prov-Loc-CP."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False}), 401
    data = request.get_json(silent=True) or {}
    nums = set(str(n) for n in (data.get("nums") or []))
    if not nums:
        return jsonify({"ok": False, "msg": "sin pedidos"}), 400
    rows = _despachos_orders(email) or []
    sel = [r for r in rows if r["num"] in nums]
    if not sel:
        return jsonify({"ok": False, "msg": "no encontré esos pedidos"}), 404
    tpl = ANDREANI_TPL if ANDREANI_TPL.exists() else Path(_os.path.expanduser("~/Downloads/EnvioMasivoExcelPaquetes.xlsx"))
    if not tpl.exists():
        return jsonify({"ok": False, "msg": "falta la plantilla EnvioMasivoExcelPaquetes.xlsx en el servidor"}), 500
    try:
        import openpyxl
    except Exception:
        return jsonify({"ok": False, "msg": "falta openpyxl en el servidor"}), 500
    try:
        wb = openpyxl.load_workbook(tpl)
        ws_dom = wb["A domicilio"]
        ws_suc = wb["A sucursal"]
    except Exception as e:
        return jsonify({"ok": False, "msg": "no pude abrir la plantilla: %s" % e}), 500
    # Mínimos Andreani (jul-2026): SUMA de lados (alto+ancho+prof) >= 35 cm y peso >= 1 kg.
    # 15+12+10 = 37 cm (cumple con margen). Peso FIJO 1000 g (1 kg) para TODOS los paquetes:
    # no escala con las unidades a propósito (si escalara, Andreani cobra más por peso declarado).
    ALTO, ANCHO, PROF = 15, 12, 10
    PESO = 1000
    cpidx, sucs = _and_cfg(wb)   # lista oficial de la propia plantilla (sucursales + Prov/Loc/CP)
    r_dom = r_suc = 3
    for r in sel:
        nom, ape = _split_nombre(r["nombre"])
        valor = int(round(r["total"]))
        peso = PESO
        dni = str(r.get("dni") or "").strip() or "00000000"   # Andreani exige DNI
        tel_cod, tel_num = _and_split_tel(r.get("tel"))        # teléfono partido en código/número
        email = r.get("email") or ""
        if r["tipo"] == "sucursal":
            # resolver al nombre OFICIAL de Andreani: primero la lista de la plantilla,
            # si no matchea (HOP nuevo), byCoordenadas en vivo (público). Nunca a otra provincia.
            suc_raw = r.get("suc_nombre") or ""
            of, conf = _and_suc_excel(suc_raw, sucs)
            if not conf:
                try:
                    live = _and_suc_live(suc_raw, r.get("cp") or "")
                except Exception:
                    live = None
                if live:
                    of = live
            vals = [None, peso, ALTO, ANCHO, PROF, valor, r["num"], nom, ape, dni,
                    email, tel_cod, tel_num, of or suc_raw]
            for c, v in enumerate(vals, start=1):
                ws_suc.cell(r_suc, c, v)
            r_suc += 1
        else:
            calle, numero = _calle_num(r.get("calle"))
            extra = r.get("extra") or ""
            depto = ""
            if not str(numero).strip() and extra:
                # RealProfit a veces mete el número (o la dirección) en 'extra' → recuperarlo
                m = _re_and.search(r"\b(\d{1,6})\b", str(extra))
                if m:
                    numero = m.group(1)
                    if len(str(calle).strip()) <= 3:
                        pre = _re_and.sub(r"^(calle|av\.?|avenida)\s+", "",
                                          str(extra)[:m.start()].strip(), flags=_re_and.I).strip()
                        if pre:
                            calle = pre
                    extra = ""
            else:
                depto, extra = extra, ""
            pcl = _and_pcl(r.get("cp"), r.get("provincia"), cpidx) or _and_normP(r.get("provincia"))
            vals = [None, peso, ALTO, ANCHO, PROF, valor, r["num"], nom, ape, dni,
                    email, tel_cod, tel_num, calle, numero, extra, depto, pcl, ""]
            for c, v in enumerate(vals, start=1):
                ws_dom.cell(r_dom, c, v)
            r_dom += 1
    import io
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Andreani.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/pf-despachos-sku-sync")
def pf_despachos_sku_sync():
    """Trae los SKU de los productos de la tienda a la sección Productos."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False})
    n = 0
    try:
        tk = _shop_tokens().get(email)
        if tk and tk.get("access_token"):
            r = requests.get("https://%s/admin/api/2026-07/products.json" % tk.get("shop"),
                             headers={"X-Shopify-Access-Token": tk.get("access_token")},
                             params={"limit": 250, "fields": "id,variants"}, timeout=30)
            skus = _skus_map(email)
            for p in ((r.json() or {}).get("products") or []):
                for v in (p.get("variants") or []):
                    sk = (v.get("sku") or "").strip()
                    if sk:
                        skus[str(p.get("id"))] = sk
                        n += 1
                        break
            _skus_save(email, skus)
    except Exception:
        pass
    return jsonify({"ok": True, "n": n})


_SKU_JOBS = {}      # {job_id: {done,total,msg,listo,error,pdf,email,stats}} — progreso del estampado


def _sku_de_items(items, skus):
    """SKU a estampar = qué empaquetar, según la config de Productos del usuario (tipo
    spray/fijo/xN por CADA producto). NO hardcodea VisionPure: usa lo que el usuario cargó.
    items: [(sku_key, cantidad, nombre_producto)]."""
    partes = []
    for key, qty, pname in items:
        if not qty or qty <= 0:
            continue
        cfg = skus.get(str(key))
        if cfg is None:
            cfg = {"tipo": "xn", "base": (pname or "").strip()}   # sin configurar → 'xN nombre'
        s = _sku_calc(cfg, qty)
        if s:
            partes.append(s)
    return " + ".join(partes)


def _sku_nint(texto):
    m = _re_and.search(r'N[°ºo]?\s*Interno\s*:\s*#?\s*(\d+)', texto)
    return m.group(1) if m else ""


def _sku_id(texto):
    m = _re_and.search(r'(?:N[°ºo]?\s*Interno|Id)\s*:\s*#?\s*(\d+)', texto, _re_and.I)
    return m.group(1) if m else ""


def _sku_pedido(texto, nuevo):
    return _sku_nint(texto) if nuevo else _sku_id(texto)


def _sku_estampar_nuevo(pg, sku):
    """Andreani ORIGINAL (carga masiva): SKU a la izquierda del 'Bulto 1 / 1', caja negra, texto blanco."""
    import fitz
    anc = pg.search_for("1 / 1") or pg.search_for("Bulto")
    if anc:
        r = anc[0]; xr = r.x0 - 6.0; y = r.y1 - 3.0
    else:
        xr, y = 157.0, 107.0
    fs = 9.5
    w = fitz.get_text_length(sku, fontname="hebo", fontsize=fs)
    x0 = xr - w; cap = fs * 0.70; pad_x, pad_y = 3.0, 2.0
    caja = fitz.Rect(x0 - pad_x, y - cap - pad_y, xr + pad_x, y + pad_y)
    pg.draw_rect(caja, fill=(0, 0, 0), color=(0, 0, 0), width=0)
    pg.insert_text(fitz.Point(x0, y), sku, fontsize=fs, fontname="hebo", color=(1, 1, 1))


def _sku_estampar_ecom(pg, sku, texto):
    """Formato 'ENCOMIENDA ECOMMERCE': caja negra en el hueco según domicilio/sucursal."""
    import fitz
    anc = (pg.search_for("ID: #") or pg.search_for("Id: #") or pg.search_for("ID:") or pg.search_for("Id:"))
    r = anc[0] if anc else None
    es_suc = bool(r and r.y0 < 90)
    fs = 11.0
    w = fitz.get_text_length(sku, fontname="hebo", fontsize=fs)
    if es_suc:
        x0 = 8.0; y = (r.y1 + 30.0) if r else 113.0
    else:
        x0 = 95.0; y = (r.y0 - 4.0) if r else 98.0
    pad_x, pad_y, cap = 4.0, 3.0, fs * 0.72
    caja = fitz.Rect(x0 - pad_x, y - cap - pad_y, x0 + w + pad_x, y + pad_y)
    pg.draw_rect(caja, fill=(0, 0, 0), color=(0, 0, 0), width=0)
    pg.insert_text(fitz.Point(x0, y), sku, fontsize=fs, fontname="hebo", color=(1, 1, 1))


def _sku_estampar_std(pg, sku):
    """Etiqueta estándar (284×425): recuadro amarillo con el SKU, anclado al 'Id:', lado izquierdo."""
    import fitz
    id_rects = pg.search_for("Id:")
    if id_rects:
        r = id_rects[0]; y1 = r.y0 - 1.5; x0, x1 = 4.0, 186.0
    else:
        x0, x1, y1 = 4.0, 186.0, 100.0
    y0 = y1 - 16.5
    pg.draw_rect(fitz.Rect(x0, y0, x1, y1), color=(0.85, 0.45, 0), fill=(1, 0.9, 0.45), width=1.2)
    pg.insert_text(fitz.Point(x0 + 6, y1 - 4.5), "SKU:  " + sku, fontsize=9, fontname="hebo", color=(0.55, 0.05, 0.05))


def _sku_norm(s):
    """Normaliza para comparar nombres: sin acentos, mayúsculas, solo letras/números."""
    s = _ud_and.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    s = _re_and.sub(r"[^A-Za-z0-9 ]", " ", s).upper()
    return _re_and.sub(r"\s+", " ", s).strip()


def _sku_tokens(nombre):
    """Tokens 'fuertes' de un nombre (>=3 letras) para comparar destinatario ↔ pedido."""
    return {t for t in _sku_norm(nombre).split() if len(t) >= 3 and not t.isdigit()}


def _sku_nombre_coincide(label_nom, pedido_nom):
    """True si el destinatario de la etiqueta y el del pedido comparten nombre/apellido.
    Si en alguno falta el dato, no bloquea (True)."""
    a, b = _sku_tokens(label_nom), _sku_tokens(pedido_nom)
    if not a or not b:
        return True
    return bool(a & b)


def _sku_label_nombre(texto):
    """Nombre del destinatario que figura en la etiqueta (para verificar el match)."""
    m = _re_and.search(r"Destinatario\s*:\s*(.+)", texto, _re_and.I)
    return m.group(1).strip() if m else ""


def _sku_pedidos_map(email):
    """{nº pedido → {'nom': destinatario, 'items': [(sku_key, cantidad, nombre_prod)]}} de las
    tiendas conectadas (Shopify + Tiendanube). Trae los PRODUCTOS de cada pedido para calcular
    el SKU con la config de Productos. Guarda el nombre para verificar el match. RÁPIDO."""
    mapa = {}
    # --- Tiendanube (sku_key = 'tn:<product_id>', igual que en Productos) ---
    tk = _tn_tokens().get(email)
    if tk and tk.get("access_token") and tk.get("store_id"):
        store, hdr = tk["store_id"], _tn_headers(tk["access_token"])
        for page in (1, 2):
            try:
                r = requests.get("%s/%s/orders" % (TN_API, store), headers=hdr, params={
                    "per_page": 200, "page": page, "sort": "-id", "payment_status": "paid",
                    "fields": "number,products,contact_name,shipping_address"}, timeout=30)
                d = r.json() if r.content else []
            except Exception:
                d = []
            if not isinstance(d, list) or not d:
                break
            for o in d:
                items = []
                for p in (o.get("products") or []):
                    nm = p.get("name")
                    if isinstance(nm, dict):
                        nm = nm.get("es") or next(iter(nm.values()), "") if nm else ""
                    items.append(("tn:%s" % p.get("product_id"), int(p.get("quantity") or 0), nm or ""))
                nom = ((o.get("shipping_address") or {}).get("name") or o.get("contact_name") or "")
                mapa[str(o.get("number"))] = {"nom": nom, "items": items}
            if len(d) < 200:
                break
    # --- Shopify (sku_key = '<product_id>') ---
    tks = _shop_tokens().get(email)
    if tks and tks.get("access_token"):
        hasta = _hoy()
        desde = (_dt.date.today() - _dt.timedelta(days=60)).isoformat()
        try:
            for o in _shopify_orders(tks.get("shop"), tks.get("access_token"), desde, hasta):
                num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
                if not num:
                    continue
                items = []
                for li in (o.get("line_items") or []):
                    items.append((str(li.get("product_id") or ""), int(li.get("quantity") or 0),
                                  li.get("title") or li.get("name") or ""))
                sa = o.get("shipping_address") or {}
                cu = o.get("customer") or {}
                nom = (sa.get("name") or ((cu.get("first_name", "") + " " + cu.get("last_name", "")).strip()))
                mapa[num] = {"nom": nom, "items": items}
        except Exception:
            pass
    return mapa


def _sku_hoja_empaquetar(doc, detalle):
    """Hoja final A4 'PARA EMPAQUETAR': cuántas bolsas de cada SKU (genérico, sin hardcodear
    ningún producto). Cuenta cuántas etiquetas comparten el mismo SKU."""
    import fitz
    from collections import Counter
    paquetes = Counter(d["sku"] for d in detalle if d.get("sku"))
    if not paquetes:
        return
    NEG, BLA = (0, 0, 0), (1, 1, 1)
    pg = doc.new_page(width=595, height=842)
    pg.insert_text((50, 92), "PARA EMPAQUETAR", fontname="hebo", fontsize=30, color=NEG)
    pg.draw_line((50, 112), (545, 112), color=NEG, width=1.2)
    y = 175
    total_bolsas = 0
    for k, v in sorted(paquetes.items(), key=lambda x: (-x[1], x[0])):
        total_bolsas += v
        etq = "%dX %s" % (v, "BOLSA" if v == 1 else "BOLSAS")
        fs = 20
        w = fitz.get_text_length(etq, fontname="hebo", fontsize=fs)
        pg.draw_rect(fitz.Rect(50, y - 17, 50 + w + 12, y + 6), color=None, fill=NEG)
        pg.insert_text((56, y), etq, fontname="hebo", fontsize=fs, color=BLA)
        pg.insert_text((50 + w + 24, y), "DE %s" % k, fontname="hebo", fontsize=fs, color=NEG)
        y += 50
        if y > 770:
            pg = doc.new_page(width=595, height=842); y = 90
    y += 10
    pg.draw_line((50, y), (545, y), color=NEG, width=1.2)
    y += 40
    pg.insert_text((55, y), "TOTAL BOLSAS:   %d" % total_bolsas, fontname="hebo", fontsize=18, color=NEG)


@app.post("/pf-despachos-sku")
def pf_despachos_sku():
    """Arranca el estampado en background (para mostrar barra de progreso) y devuelve un job id.
    El front consulta /pf-despachos-sku-progreso y baja el PDF con /pf-despachos-sku-descargar."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False}), 401
    f = request.files.get("pdf") or (next(iter(request.files.values())) if request.files else None)
    if not f:
        return jsonify({"ok": False, "msg": "subí el PDF de etiquetas"}), 400
    try:
        import fitz  # noqa: F401
    except Exception:
        return jsonify({"ok": False, "msg": "falta PyMuPDF en el servidor (esperá el redeploy)"}), 500
    data = f.read()
    import uuid
    job = uuid.uuid4().hex[:12]
    _SKU_JOBS[job] = {"done": 0, "total": 0, "msg": "Leyendo el PDF…", "listo": False,
                      "error": None, "pdf": None, "email": email, "stats": {}}
    threading.Thread(target=_sku_run, args=(job, data, email), daemon=True).start()
    return jsonify({"ok": True, "job": job})


def _sku_run(job, data, email):
    """Procesa el PDF en background, actualizando el progreso del job."""
    st = _SKU_JOBS.get(job)
    try:
        import fitz
        import io
        st["msg"] = "Sincronizando pedidos de tu tienda…"
        skus = _skus_map(email)                 # config de SKU por producto (lo que cargó el usuario)
        mapa = _sku_pedidos_map(email)          # {nº → nombre + productos} de Shopify + Tiendanube
        doc = fitz.open(stream=data, filetype="pdf")
        total = len(doc)
        st["total"] = total
        estampadas = conflicto = sin_pedido = 0
        detalle = []
        orden = []   # (clave, indice de pagina) -> reordenar por SKU al final (x1, x2, x3...)
        for i, pg in enumerate(doc):
            st["done"] = i
            st["msg"] = "Analizando etiqueta %d de %d…" % (i + 1, total)
            texto = pg.get_text()
            nuevo = ("Bulto" in texto) and bool(_re_and.search(r"Peso:\s*\d+\s*Gr", texto, _re_and.I))
            ped = _sku_pedido(texto, nuevo)
            ent = mapa.get(str(ped)) if ped else None
            if not ent:
                sin_pedido += 1
                detalle.append({"pedido": ped or "?", "sku": ""})
                orden.append(((999999, 0, ""), i)); continue
            # Verificación por NOMBRE: si el nº matchea pero el destinatario no → no estampo.
            if not _sku_nombre_coincide(_sku_label_nombre(texto), ent.get("nom", "")):
                conflicto += 1
                detalle.append({"pedido": ped or "?", "sku": "", "conflicto": True})
                orden.append(((999998, 0, ""), i)); continue
            items = ent.get("items") or []
            unidades = sum(int(it[1] or 0) for it in items)
            sku = _sku_de_items(items, skus)
            detalle.append({"pedido": ped or "?", "sku": sku})
            if not sku:
                orden.append(((999997, 0, ""), i)); continue
            if nuevo:
                _sku_estampar_nuevo(pg, sku)
            elif "ENCOMIENDA" in texto:
                _sku_estampar_ecom(pg, sku, texto)
            else:
                _sku_estampar_std(pg, sku)
            estampadas += 1
            orden.append(((unidades, len(sku), sku), i))   # menor a mayor: por unidades, luego SKU
        st["msg"] = "Ordenando etiquetas y armando 'PARA EMPAQUETAR'…"
        orden.sort(key=lambda x: x[0])                      # x1, x2, x3... y agrupa mismos SKU
        nuevo_doc = fitz.open()
        for _clave, idx in orden:
            nuevo_doc.insert_pdf(doc, from_page=idx, to_page=idx)
        _sku_hoja_empaquetar(nuevo_doc, detalle)
        buf = io.BytesIO()
        nuevo_doc.save(buf, garbage=3, deflate=True)
        nuevo_doc.close()
        doc.close()
        st["pdf"] = buf.getvalue()
        st["done"] = total
        st["stats"] = {"total": total, "estampadas": estampadas,
                       "conflicto": conflicto, "sin_pedido": sin_pedido}
        st["msg"] = "¡Listo! %d de %d etiquetas con SKU." % (estampadas, total)
        st["listo"] = True
    except Exception as e:
        st["error"] = str(e)
        st["listo"] = True


@app.get("/pf-despachos-sku-progreso")
def pf_despachos_sku_progreso():
    if not _user_actual():
        return jsonify({"ok": False}), 401
    st = _SKU_JOBS.get((request.args.get("job") or "").strip())
    if not st:
        return jsonify({"ok": False, "msg": "job no encontrado"}), 404
    return jsonify({"ok": True, "done": st["done"], "total": st["total"], "msg": st["msg"],
                    "listo": st["listo"], "error": st["error"], "stats": st.get("stats") or {}})


@app.get("/pf-despachos-sku-descargar")
def pf_despachos_sku_descargar():
    if not _user_actual():
        return jsonify({"ok": False}), 401
    job = (request.args.get("job") or "").strip()
    st = _SKU_JOBS.get(job)
    if not st or not st.get("pdf"):
        return jsonify({"ok": False, "msg": "todavía no está listo"}), 404
    import io
    pdf = st["pdf"]
    _SKU_JOBS.pop(job, None)          # libero memoria una vez descargado
    return send_file(io.BytesIO(pdf), as_attachment=True,
                     download_name="etiquetas-con-sku.pdf", mimetype="application/pdf")


@app.post("/pf-despachos-seg-leer")
def pf_despachos_seg_leer():
    """Lee el PDF de Andreani → N° Interno + seguimiento de cada pedido.
    Pendiente de calibrar con un PDF real → por ahora informa."""
    if not _user_actual():
        return jsonify({"ok": False}), 401
    return jsonify({"ok": False, "msg": "la lectura del PDF se está calibrando con un rótulo real de Andreani"})


@app.post("/pf-despachos-seg-enviar")
def pf_despachos_seg_enviar():
    """Carga el tracking en la tienda (Shopify/TN) y notifica al cliente por mail."""
    if not _user_actual():
        return jsonify({"ok": False}), 401
    return jsonify({"ok": False, "msg": "el envío de seguimiento se conecta en el próximo paso"})


_MP_LISTA_CACHE = {}   # {(email,desde,hasta): (ts, out)} — pagos de MP, cache 60s: se pide en varias secciones


def _mp_pagos_lista(email, desde, hasta):
    """Lista de pagos aprobados de MP del usuario: {ref, amount, net, fee}. None si no hay MP.
    Sirve para MATCHEAR cada pago con su pedido de Shopify (comisión exacta por venta, sin inflar)."""
    import time as _t
    ck = (email, desde, hasta)
    c = _MP_LISTA_CACHE.get(ck)
    if c and (_t.time() - c[0] < 60):     # mismo período pedido de nuevo en <60s → sin re-bajar de MP
        return c[1]
    tk = _mp_tokens().get(email)
    token = tk.get("access_token") if tk else None
    if not token:
        return None
    # Ventana de MP un poco más ancha que el período: un pago puede aprobarse hasta unos días
    # después de creado el pedido (transferencias, cuotas). Los que no matcheen se ignoran.
    try:
        h2 = (_dt.date.fromisoformat(hasta) + _dt.timedelta(days=4)).isoformat()
    except Exception:
        h2 = hasta
    ini = desde + "T00:00:00.000-03:00"
    fin = h2 + "T23:59:59.999-03:00"
    out = []
    offset = 0
    try:
        while True:
            r = requests.get("https://api.mercadopago.com/v1/payments/search",
                             headers={"Authorization": "Bearer " + token},
                             params={"sort": "date_approved", "criteria": "desc",
                                     "range": "date_approved", "begin_date": ini, "end_date": fin,
                                     "status": "approved", "offset": offset, "limit": 100}, timeout=30)
            if r.status_code >= 400:
                return None if offset == 0 else out
            data = r.json()
            res = data.get("results") or []
            for p in res:
                ta = float(p.get("transaction_amount") or 0)
                det = p.get("transaction_details") or {}
                net = float(det.get("net_received_amount") or 0)
                fee = (ta - net) if net else sum(float(f.get("amount") or 0)
                                                 for f in (p.get("fee_details") or []))
                # desglose para el modal: comisión MP (mercadopago+application) vs cuotas (financing)
                fd = p.get("fee_details") or []
                finanz = sum(float(f.get("amount") or 0) for f in fd if f.get("type") == "financing_fee")
                base = sum(float(f.get("amount") or 0) for f in fd if f.get("type") != "financing_fee")
                if not fd:
                    base = fee; finanz = 0.0
                out.append({"ref": (p.get("external_reference") or "").strip(),
                            "amount": round(ta), "net": round(net, 2), "fee": round(fee, 2),
                            "inst": int(p.get("installments") or 1),
                            "fee_mp": round(base, 2), "fee_cuotas": round(finanz, 2),
                            "medio": (p.get("payment_method_id") or p.get("payment_type_id") or "")})
            offset += 100
            if offset >= (data.get("paging") or {}).get("total", 0) or not res:
                break
    except Exception:
        return None
    _MP_LISTA_CACHE[ck] = (_t.time(), out)
    return out


def _norm_txt(s):
    return "".join(ch for ch in str(s or "").lower() if ch.isalnum())


def _norm_tel(s):
    d = "".join(ch for ch in str(s or "") if ch.isdigit())
    return d[-10:] if len(d) >= 10 else d


def _recompras_periodo(orders, desde, hasta):
    """orders: historico [{'fecha','total','email','tel','nombre'}]. Dos ordenes son del MISMO cliente si
    comparten >=2 de {email, telefono, nombre}. Devuelve (cant, facturacion) de las compras del periodo
    [desde,hasta] que NO son la 1ra del cliente (o sea, recompras)."""
    parent = {}
    def find(x):
        parent.setdefault(x, x)
        r = x
        while parent[r] != r:
            r = parent[r]
        while parent[x] != r:
            parent[x], x = r, parent[x]
        return r
    def union(a, b):
        parent[find(a)] = find(b)
    for i, o in enumerate(orders):
        e, ph, n = _norm_txt(o.get("email")), _norm_tel(o.get("tel")), _norm_txt(o.get("nombre"))
        parent.setdefault(("o", i), ("o", i))
        if e and ph:
            union(("o", i), ("k", "ep:%s|%s" % (e, ph)))
        if e and n:
            union(("o", i), ("k", "en:%s|%s" % (e, n)))
        if ph and n:
            union(("o", i), ("k", "pn:%s|%s" % (ph, n)))
    groups = {}
    for i in range(len(orders)):
        groups.setdefault(find(("o", i)), []).append(i)
    cant = 0
    fact = 0.0
    for idxs in groups.values():
        if len(idxs) < 2:
            continue
        idxs.sort(key=lambda i: (orders[i].get("fecha") or ""))
        for j, i in enumerate(idxs):
            if j == 0:
                continue
            f = (orders[i].get("fecha") or "")[:10]
            if desde <= f <= hasta:
                cant += 1
                fact += float(orders[i].get("total") or 0)
    return cant, round(fact, 2)


_TN_HIST_CACHE = {}
_SHOP_HIST_CACHE = {}


def _tn_hist_orders(email, hasta):
    """Ordenes PAGADAS de TN de los ultimos 180 dias hasta `hasta` (cache 10 min) - para recompras."""
    import time as _t
    ck = (email, hasta)
    c = _TN_HIST_CACHE.get(ck)
    if c and (_t.time() - c[0] < 600):
        return c[1]
    try:
        d0 = (_dt.date.fromisoformat(hasta) - _dt.timedelta(days=90)).isoformat()
    except Exception:
        d0 = hasta
    hist = []
    for o in _tn_orders_raw(email, d0, hasta):
        if (o.get("payment_status") or "").lower() != "paid" or o.get("cancelled_at"):
            continue
        cu = o.get("customer") or {}
        sa = o.get("shipping_address") or {}
        hist.append({"fecha": o.get("created_at") or "", "total": float(o.get("total") or 0),
                     "email": o.get("contact_email") or cu.get("email") or "",
                     "tel": o.get("contact_phone") or o.get("billing_phone") or cu.get("phone") or sa.get("phone") or "",
                     "nombre": cu.get("name") or o.get("contact_name") or sa.get("name") or ""})
    _TN_HIST_CACHE[ck] = (_t.time(), hist)
    return hist


def _shop_hist_orders(email, hasta):
    """Ordenes pagadas de Shopify de los ultimos 180 dias hasta `hasta` (cache 10 min) - para recompras."""
    import time as _t, re as _re
    ck = (email, hasta)
    c = _SHOP_HIST_CACHE.get(ck)
    if c and (_t.time() - c[0] < 600):
        return c[1]
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token") or not tk.get("shop"):
        return []
    shop, token = tk["shop"], tk["access_token"]
    try:
        d0 = (_dt.date.fromisoformat(hasta) - _dt.timedelta(days=90)).isoformat()
    except Exception:
        d0 = hasta
    hist = []
    url = "https://%s/admin/api/2026-07/orders.json" % shop
    params = {"status": "any", "financial_status": "paid", "limit": 250,
              "created_at_min": d0 + "T00:00:00-03:00", "created_at_max": hasta + "T23:59:59-03:00",
              "fields": "created_at,total_price,customer,contact_email,email,billing_address,shipping_address"}
    for _ in range(20):
        try:
            r = requests.get(url, headers={"X-Shopify-Access-Token": token}, params=params, timeout=30)
        except Exception:
            break
        if r.status_code != 200:
            break
        for o in (r.json().get("orders") or []):
            cu = o.get("customer") or {}
            ba = o.get("billing_address") or {}
            sa = o.get("shipping_address") or {}
            hist.append({"fecha": o.get("created_at") or "", "total": float(o.get("total_price") or 0),
                         "email": o.get("email") or o.get("contact_email") or cu.get("email") or "",
                         "tel": cu.get("phone") or ba.get("phone") or sa.get("phone") or "",
                         "nombre": ((cu.get("first_name") or "") + " " + (cu.get("last_name") or "")).strip() or ba.get("name") or sa.get("name") or ""})
        nxt = None
        for part in (r.headers.get("Link", "") or "").split(","):
            if 'rel="next"' in part:
                m = _re.search(r"<([^>]+)>", part)
                if m:
                    nxt = m.group(1)
        if not nxt:
            break
        url = nxt
        params = {}
    _SHOP_HIST_CACHE[ck] = (_t.time(), hist)
    return hist


def _shopify_resumen(email, desde, hasta):
    """Arma el 'raw' que espera el dashboard con los pedidos reales de Shopify + costos cargados."""
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token"):
        return None
    shop = tk.get("shop"); token = tk.get("access_token")
    try:
        orders = _shopify_orders(shop, token, desde, hasta)
    except Exception:
        return None
    costos = (_costos().get(email) or {})
    r = resumen_vacio()
    r["fecha"] = desde if desde == hasta else (desde + " a " + hasta)
    r["desde"] = desde; r["hasta"] = hasta
    r["actualizado"] = (_dt.datetime.utcnow() - _dt.timedelta(hours=3)).strftime("%H:%M:%S")
    emap = _envialo_costos(email)   # {nº pedido → costo REAL de envío} (vacío si no conectó Envialo)
    # MP: pagos reales del período para MATCHEAR cada pedido con su pago (comisión exacta).
    pagos = _mp_pagos_lista(email, desde, hasta)
    mp_conectado = pagos is not None
    by_ref, by_amt = {}, {}
    for p in (pagos or []):
        if p["ref"]:
            by_ref.setdefault(str(p["ref"]), []).append(p)
        by_amt.setdefault(p["amount"], []).append(p)
    mp_costo = 0.0
    mp_match = 0
    fact = cobr = costo_prod = reemb_monto = envio_monto = 0.0
    unidades = ordenes = reemb_cant = envio_real = 0
    prodmap = {}
    ords_list = []
    # Solo cuentan las órdenes YA COBRADAS. Efectivo/transferencia pendiente, autorizado sin capturar o
    # 'pending' NO suman (ni facturación, ni órdenes, ni CPA) hasta que se paguen. Al cobrarse pasan a 'paid'
    # y recién ahí entran. Los reembolsos fueron ventas reales cobradas → siguen contando.
    PAGADAS = ("paid", "partially_paid", "refunded", "partially_refunded")
    for o in orders:
        if o.get("cancelled_at"):
            continue
        if (o.get("financial_status") or "").lower() not in PAGADAS:
            continue
        ordenes += 1
        tot = float(o.get("total_price") or o.get("current_total_price") or 0)
        fact += tot
        # Envío: costo REAL de Envialo si el pedido ya está ahí; si no, promedio domicilio/sucursal.
        _num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
        _real = emap.get(_num)
        if _real is not None:
            envio_monto += _real; envio_real += 1
        else:
            envio_monto += _envio_costo(o)
        # MP: matcheo este pedido con su pago real (por referencia; fallback por monto exacto).
        pago = None
        if mp_conectado:
            for ref in (str(o.get("id")), str(o.get("order_number")), _num):
                lst = by_ref.get(ref)
                if lst:
                    pago = lst.pop(0); break
            if pago is None:
                lst = by_amt.get(round(tot))
                if lst:
                    pago = lst.pop(0)
            if pago is not None:
                mp_costo += pago["fee"]; mp_match += 1
        if (o.get("financial_status") or "") in ("paid", "partially_paid", "authorized"):
            cobr += tot
        for li in (o.get("line_items") or []):
            q = int(li.get("quantity") or 0)
            unidades += q
            pid = str(li.get("product_id") or "")
            c = costos.get(pid)
            if c:
                costo_prod += _costo_qty(c, q)
            nm = li.get("title") or "?"
            prodmap[nm] = prodmap.get(nm, 0) + q
        for rf in (o.get("refunds") or []):
            reemb_cant += 1
            for tx in (rf.get("transactions") or []):
                reemb_monto += float(tx.get("amount") or 0)
        # Fila para la tabla "Últimas ventas" (neto = lo real que entró por MP si matcheó, sino el total).
        ords_list.append({"num": _num, "origen": "Shopify",
                          "estado": _ESTADO_TXT.get((o.get("financial_status") or "").lower(), "—"),
                          "fecha": o.get("created_at") or "",
                          "total": round(tot, 2),
                          "neto": round(pago["net"] if pago else tot, 2)})
    # Costos por venta. MercadoPago: comisión REAL matcheada pedido-por-pedido (mp_costo, ya
    # sumado en el loop). Así NO infla con pagos que no son de la tienda. Si no hay MP conectado,
    # cae al % manual. FIJOS: 1% tienda + 3,5% Ingresos Brutos. Envío: real (Envialo) o promedio.
    iibb_monto = fact * IIBB_PCT / 100.0
    tienda_monto = fact * TIENDA_PCT / 100.0
    if not mp_conectado:
        cu = _comis_user(email)
        mp_costo = fact * (cu["mp_comision"] + cu["mp_cuotas"]) * (1 + cu["iva"] / 100.0) / 100.0
    comision_monto = mp_costo + iibb_monto + tienda_monto
    r["mp_costo_real"] = round(mp_costo, 2)
    r["mp_match"] = mp_match            # pedidos que matchearon su pago de MP (comisión exacta)
    r["iibb_monto"] = round(iibb_monto, 2)
    r["tienda_monto"] = round(tienda_monto, 2)
    r["envio_monto"] = round(envio_monto, 2)
    r["envio_real"] = envio_real       # cuántos pedidos usaron el costo REAL de Envialo
    ganancia = fact - costo_prod - comision_monto - envio_monto
    # Break-even: contribución ANTES de ads (lo que queda para pagar publicidad).
    _pre_ads = fact - costo_prod - comision_monto - envio_monto
    r["be_roas"] = round(fact / _pre_ads, 2) if _pre_ads > 0 else 0.0
    r["breakeven_roas"] = r["be_roas"]
    r["be_cpa"] = round(_pre_ads / ordenes, 2) if ordenes else 0.0
    r["breakeven_cpa"] = r["be_cpa"]
    r["ordenes"] = ordenes
    r["ventas_periodo"] = ordenes
    r["unidades"] = unidades
    r["facturado"] = round(fact, 2)
    r["cobrado"] = round(cobr, 2)
    r["costo_prod"] = round(costo_prod, 2)
    r["comision"] = round(comision_monto, 2)
    r["ganancia"] = round(ganancia, 2)
    r["margen"] = round(ganancia / fact * 100, 2) if fact else 0.0
    r["ticket"] = round(fact / ordenes, 2) if ordenes else 0.0
    r["gan_por_venta"] = round(ganancia / ordenes, 2) if ordenes else 0.0
    r["reemb_cantidad"] = reemb_cant
    r["reemb_monto"] = round(reemb_monto, 2)
    # Totales (por ahora solo Shopify, sin MELI)
    r["tot_ordenes"] = ordenes
    r["tot_facturado"] = round(fact, 2)
    r["tot_ganancia"] = round(ganancia, 2)
    r["tot_costo"] = round(costo_prod, 2)
    r["tot_margen"] = round(ganancia / fact * 100, 2) if fact else 0.0
    r["tot_aov"] = round(fact / ordenes, 2) if ordenes else 0.0
    r["tot_gan_por_venta"] = round(ganancia / ordenes, 2) if ordenes else 0.0
    prod = [{"nombre": k, "unidades": v, "facturado": 0.0}
            for k, v in sorted(prodmap.items(), key=lambda x: -x[1])[:10]]
    ords_list.sort(key=lambda x: int(x["num"]) if str(x["num"]).isdigit() else 0, reverse=True)
    return {"raw": r, "prod": prod, "ords": ords_list}


_TN_ORDERS_CACHE = {}   # {(email,desde,hasta): (ts, [orders])} — órdenes crudas de TN, cache 60s


def _tn_orders_raw(email, desde, hasta):
    """Órdenes crudas de TN del período (cache 60s). Las comparten el dashboard y últimas ventas,
    así no se bajan dos veces en la misma carga."""
    import time as _t
    ck = (email, desde, hasta)
    c = _TN_ORDERS_CACHE.get(ck)
    if c and (_t.time() - c[0] < 60):
        return c[1]
    tk = _tn_tokens().get(email)
    if not tk or not tk.get("access_token") or not tk.get("store_id"):
        return []
    store, hdr = tk["store_id"], _tn_headers(tk["access_token"])
    out, vistos, page = [], set(), 1
    while page <= 20:
        try:
            r = requests.get("%s/%s/orders" % (TN_API, store), headers=hdr, params={
                "per_page": 200, "page": page,
                "created_at_min": desde + "T00:00:00-03:00",
                "created_at_max": hasta + "T23:59:59-03:00"}, timeout=40)
            lote = r.json() if r.content else []
        except Exception:
            lote = []
        if not isinstance(lote, list) or not lote:
            break
        for o in lote:
            n = str(o.get("number") or "")
            if n and n not in vistos:
                vistos.add(n); out.append(o)
        if len(lote) < 200:
            break
        page += 1
    _TN_ORDERS_CACHE[ck] = (_t.time(), out)
    return out


def _tn_resumen(email, desde, hasta):
    """Mismo 'raw'/prod/ords que _shopify_resumen pero con los pedidos de Tiendanube."""
    tk = _tn_tokens().get(email)
    if not tk or not tk.get("access_token") or not tk.get("store_id"):
        return None
    orders = [o for o in _tn_orders_raw(email, desde, hasta) if not o.get("cancelled_at")]
    costos = (_costos().get(email) or {})
    pagos = _mp_pagos_lista(email, desde, hasta)
    by_amt = {}
    for p in (pagos or []):
        by_amt.setdefault(p["amount"], []).append(p)
    mp_conectado = pagos is not None
    r = resumen_vacio()
    r["fecha"] = desde if desde == hasta else (desde + " a " + hasta)
    r["desde"] = desde; r["hasta"] = hasta
    r["actualizado"] = (_dt.datetime.utcnow() - _dt.timedelta(hours=3)).strftime("%H:%M:%S")
    fact = cobr = costo_prod = envio_monto = mp_costo = 0.0
    unidades = ordenes = mp_match = 0
    prodmap, ords_list = {}, []
    for o in orders:
        if (o.get("payment_status") or "").lower() != "paid":
            continue
        ordenes += 1
        tot = float(o.get("total") or 0)
        fact += tot; cobr += tot
        envio_monto += ENVIO_SUCURSAL if _tn_shipping(o).get("type") == "pickup" else ENVIO_DOMICILIO
        if mp_conectado:
            lst = by_amt.get(round(tot))
            if lst:
                pago = lst.pop(0); mp_costo += pago["fee"]; mp_match += 1
        for p in (o.get("products") or []):
            q = int(p.get("quantity") or 0); unidades += q
            c = costos.get("tn:%s" % p.get("product_id"))
            if c:
                costo_prod += _costo_qty(c, q)
            nm = p.get("name") or "?"
            if isinstance(nm, dict):
                nm = nm.get("es") or next(iter(nm.values()), "?")
            prodmap[nm] = prodmap.get(nm, 0) + q
        ords_list.append({"num": str(o.get("number") or ""), "origen": "Tiendanube",
                          "estado": "Pagado", "fecha": o.get("created_at") or "",
                          "total": round(tot, 2), "neto": round(tot * (1 - TIENDA_PCT / 100.0), 2)})
    iibb_monto = fact * IIBB_PCT / 100.0
    tienda_monto = fact * TIENDA_PCT / 100.0
    if not mp_conectado:
        cu = _comis_user(email)
        mp_costo = fact * (cu["mp_comision"] + cu["mp_cuotas"]) * (1 + cu["iva"] / 100.0) / 100.0
    comision_monto = mp_costo + iibb_monto + tienda_monto
    ganancia = fact - costo_prod - comision_monto - envio_monto
    r["mp_costo_real"] = round(mp_costo, 2); r["mp_match"] = mp_match
    r["iibb_monto"] = round(iibb_monto, 2); r["tienda_monto"] = round(tienda_monto, 2)
    r["envio_monto"] = round(envio_monto, 2); r["envio_real"] = 0
    _pre = fact - costo_prod - comision_monto - envio_monto
    r["be_roas"] = r["breakeven_roas"] = round(fact / _pre, 2) if _pre > 0 else 0.0
    r["be_cpa"] = r["breakeven_cpa"] = round(_pre / ordenes, 2) if ordenes else 0.0
    r["ordenes"] = r["ventas_periodo"] = r["tot_ordenes"] = ordenes
    r["unidades"] = unidades
    r["facturado"] = r["tot_facturado"] = round(fact, 2)
    r["cobrado"] = round(cobr, 2)
    r["costo_prod"] = r["tot_costo"] = round(costo_prod, 2)
    r["comision"] = round(comision_monto, 2)
    r["ganancia"] = r["tot_ganancia"] = round(ganancia, 2)
    r["margen"] = r["tot_margen"] = round(ganancia / fact * 100, 2) if fact else 0.0
    r["ticket"] = r["tot_aov"] = round(fact / ordenes, 2) if ordenes else 0.0
    r["gan_por_venta"] = r["tot_gan_por_venta"] = round(ganancia / ordenes, 2) if ordenes else 0.0
    r["reemb_cantidad"] = 0; r["reemb_monto"] = 0.0
    prod = [{"nombre": k, "unidades": v, "facturado": 0.0}
            for k, v in sorted(prodmap.items(), key=lambda x: -x[1])[:10]]
    ords_list.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    return {"raw": r, "prod": prod, "ords": ords_list}


def _combinar_resumen(a, b):
    """Suma dos blobs de resumen (Shopify + Tiendanube) y recalcula los ratios."""
    if not a:
        return b
    if not b:
        return a
    ra, rb = a["raw"], b["raw"]
    r = resumen_vacio()
    r["fecha"] = ra.get("fecha"); r["desde"] = ra.get("desde"); r["hasta"] = ra.get("hasta")
    r["actualizado"] = ra.get("actualizado")
    SUM = ["mp_costo_real", "mp_match", "iibb_monto", "tienda_monto", "envio_monto", "envio_real",
           "ordenes", "ventas_periodo", "unidades", "facturado", "cobrado", "costo_prod",
           "comision", "ganancia", "reemb_cantidad", "reemb_monto",
           "tot_ordenes", "tot_facturado", "tot_ganancia", "tot_costo"]
    for k in SUM:
        r[k] = round((ra.get(k) or 0) + (rb.get(k) or 0), 2)
    fact = r["facturado"]; gan = r["ganancia"]; ordn = r["ordenes"]
    r["margen"] = r["tot_margen"] = round(gan / fact * 100, 2) if fact else 0.0
    r["ticket"] = r["tot_aov"] = round(fact / ordn, 2) if ordn else 0.0
    r["gan_por_venta"] = r["tot_gan_por_venta"] = round(gan / ordn, 2) if ordn else 0.0
    _pre = fact - r["costo_prod"] - r["comision"] - r["envio_monto"]
    r["be_roas"] = r["breakeven_roas"] = round(fact / _pre, 2) if _pre > 0 else 0.0
    r["be_cpa"] = r["breakeven_cpa"] = round(_pre / ordn, 2) if ordn else 0.0
    prod = (a.get("prod") or []) + (b.get("prod") or [])
    prod.sort(key=lambda x: -x.get("unidades", 0))
    ords = (a.get("ords") or []) + (b.get("ords") or [])
    ords.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    return {"raw": r, "prod": prod[:10], "ords": ords}


_ORDS_CACHE = {}
_ESTADO_TXT = {"paid": "Pagado", "partially_paid": "Parcial", "pending": "Pendiente",
               "authorized": "Autorizado", "partially_refunded": "Reemb. parcial",
               "refunded": "Reembolsado", "voided": "Anulado"}


def _shopify_ordenes(email, dias=90):
    """Últimas órdenes REALES de Shopify (últimos `dias` días) para la tabla 'Últimas ventas'.
    Cada una: {num, origen, estado, fecha, total, neto}. neto = lo real que entró por MP (matcheado)
    o el total si no matcheó/no es MP. Cache 2 min por usuario."""
    import time as _t
    tk = _shop_tokens().get(email)
    if not tk or not tk.get("access_token"):
        return []
    c = _ORDS_CACHE.get(email)
    if c and (_t.time() - c[0] < 120):
        return c[1]
    hasta = _hoy()
    desde = (_dt.date.today() - _dt.timedelta(days=dias)).isoformat()
    try:
        orders = _shopify_orders(tk.get("shop"), tk.get("access_token"), desde, hasta)
    except Exception:
        return []
    orders = [o for o in orders if not o.get("cancelled_at")]
    pagos = _mp_pagos_lista(email, desde, hasta)   # para el neto real por pedido
    by_ref, by_amt = {}, {}
    for p in (pagos or []):
        if p["ref"]:
            by_ref.setdefault(str(p["ref"]), []).append(p)
        by_amt.setdefault(p["amount"], []).append(p)
    out = []
    for o in orders:
        tot = float(o.get("total_price") or o.get("current_total_price") or 0)
        num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
        pago = None
        if pagos is not None:
            for ref in (str(o.get("id")), str(o.get("order_number")), num):
                lst = by_ref.get(ref)
                if lst:
                    pago = lst.pop(0); break
            if pago is None:
                lst = by_amt.get(round(tot))
                if lst:
                    pago = lst.pop(0)
        neto = pago["net"] if pago else tot
        fs = (o.get("financial_status") or "").lower()
        out.append({"num": num, "origen": "Shopify",
                    "estado": _ESTADO_TXT.get(fs, (fs or "—").title()),
                    "fecha": o.get("created_at") or "",
                    "total": round(tot, 2), "neto": round(neto, 2)})
    out.sort(key=lambda x: int(x["num"]) if str(x["num"]).isdigit() else 0, reverse=True)
    out = out[:200]
    _ORDS_CACHE[email] = (_t.time(), out)
    return out


# ---------------- Login / Registro ----------------
_LOGIN_PAGE = """<!doctype html><html lang=es><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1"><title>RealProfit — Ingresá</title>
<style>
*{box-sizing:border-box} body{margin:0;font-family:system-ui,-apple-system,sans-serif;background:#0b1220;color:#e2e8f0;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}
.card{background:#0f1826;border:1px solid #1e2b3d;border-radius:18px;padding:34px 30px;width:100%;max-width:380px;box-shadow:0 20px 60px rgba(0,0,0,.45)}
.logo{font-size:24px;font-weight:800;text-align:center;color:#fff}
.sub{text-align:center;color:#94a3b8;font-size:13px;margin:4px 0 24px}
.tabs{display:flex;gap:8px;margin-bottom:18px}
.tab{flex:1;padding:10px;border-radius:10px;background:#111c2b;color:#94a3b8;text-align:center;cursor:pointer;font-weight:600;font-size:14px;border:1px solid transparent}
.tab.on{background:#137fec;color:#fff}
label{display:block;font-size:12px;color:#94a3b8;margin:12px 0 6px}
input{width:100%;padding:12px 14px;border-radius:10px;border:1px solid #1e2b3d;background:#0b1220;color:#e2e8f0;font-size:15px;outline:none}
input:focus{border-color:#137fec}
.btn{width:100%;margin-top:20px;padding:13px;border-radius:10px;border:0;background:#137fec;color:#fff;font-weight:700;font-size:15px;cursor:pointer}
.btn:hover{background:#0f6ad0}
.err{color:#f87171;font-size:13px;margin-top:12px;text-align:center;min-height:18px}
</style></head><body>
<div class=card>
 <div class=logo>📊 RealProfit</div>
 <div class=sub>Tu ganancia real, en tiempo real</div>
 <div class=tabs>
  <div class=tab id=tLogin onclick="modo('login')">Ingresar</div>
  <div class=tab id=tReg onclick="modo('registro')">Crear cuenta</div>
 </div>
 <form id=f onsubmit="return enviar(event)">
  <label>Email</label><input id=email type=email autocomplete=email required placeholder="tu@email.com">
  <label>Contraseña</label><input id=password type=password autocomplete=current-password required placeholder="Mínimo 4 caracteres">
  <button class=btn id=submit type=submit>Ingresar</button>
  <div class=err id=err></div>
 </form>
</div>
<script>
var M='login';
function modo(m){M=m;
 document.getElementById('tLogin').classList.toggle('on',m==='login');
 document.getElementById('tReg').classList.toggle('on',m==='registro');
 document.getElementById('submit').textContent=m==='login'?'Ingresar':'Crear cuenta';
 document.getElementById('err').textContent='';}
modo('login');
async function enviar(e){e.preventDefault();
 var err=document.getElementById('err');err.textContent='';
 var fd=new FormData();
 fd.append('email',document.getElementById('email').value);
 fd.append('password',document.getElementById('password').value);
 try{var r=await fetch('/'+(M==='login'?'login':'registro'),{method:'POST',body:fd});
  var j=await r.json();
  if(j.ok){location.href='/';}else{err.textContent=j.msg||'Error';}
 }catch(x){err.textContent='Error de conexión';}
 return false;}
</script></body></html>"""


@app.post("/registro")
@limiter.limit("20 per hour")
def registro():
    email = (request.form.get("email") or "").strip().lower()
    pw = request.form.get("password") or ""
    if not email or "@" not in email or "." not in email:
        return jsonify({"ok": False, "msg": "Ingresá un email válido."})
    if len(pw) < 4:
        return jsonify({"ok": False, "msg": "La contraseña necesita al menos 4 caracteres."})
    us = _users()
    if email in us:
        return jsonify({"ok": False, "msg": "Ya existe una cuenta con ese email. Iniciá sesión."})
    us[email] = {"pass": generate_password_hash(pw), "creado": _hoy()}
    _users_save(us)
    session.clear()
    session["email"] = email
    return jsonify({"ok": True})


@app.post("/login")
@limiter.limit("30 per hour")
def login():
    email = (request.form.get("email") or "").strip().lower()
    pw = request.form.get("password") or ""
    u = _users().get(email)
    if not u or not check_password_hash(u.get("pass", ""), pw):
        return jsonify({"ok": False, "msg": "Email o contraseña incorrectos."})
    session.clear()
    session["email"] = email
    return jsonify({"ok": True})


@app.get("/logout")
def logout():
    session.clear()
    return redirect("/")


# ---------------- Página de Integraciones (Mercado Pago, Mercado Libre, Tiendanube, Shopify, Meta) ----------------
_INTEGRACIONES_PAGE = """<!doctype html><html lang=es><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1"><title>RealProfit — Integraciones</title>
<style>
*{box-sizing:border-box} body{margin:0;font-family:system-ui,-apple-system,sans-serif;background:#0b1220;color:#e2e8f0;min-height:100vh}
.wrap{display:flex;min-height:100vh}
.side{width:230px;flex:none;background:#0f1826;border-right:1px solid #1e2b3d;display:flex;flex-direction:column;padding:20px 14px;position:sticky;top:0;height:100vh}
.brand{font-size:19px;font-weight:800;color:#fff;padding:6px 10px 20px}
.nav a{display:flex;align-items:center;gap:10px;padding:11px 12px;border-radius:10px;color:#94a3b8;text-decoration:none;font-weight:600;font-size:14px;margin-bottom:4px}
.nav a.on{background:#137fec;color:#fff} .nav a:hover:not(.on){background:#111c2b;color:#e2e8f0}
.userbox{margin-top:auto;display:flex;align-items:center;gap:10px;background:#111c2b;border:1px solid #1e2b3d;padding:9px 11px;border-radius:12px}
.av{width:34px;height:34px;border-radius:50%;background:#137fec;display:flex;align-items:center;justify-content:center;font-weight:700;color:#fff}
.main{flex:1;padding:34px 40px;max-width:1000px}
h1{margin:0;font-size:26px;color:#f1f5f9} .lead{color:#94a3b8;margin:6px 0 28px;font-size:14px}
.sechead{font-size:13px;color:#94a3b8;text-transform:uppercase;letter-spacing:.5px;margin:0 0 14px;font-weight:700}
.card{display:flex;align-items:center;gap:16px;background:#0f1826;border:1px solid #1e2b3d;border-radius:14px;padding:18px 20px;margin-bottom:12px}
.ic{width:46px;height:46px;border-radius:12px;flex:none;display:flex;align-items:center;justify-content:center;font-size:22px;font-weight:800}
.info{flex:1;min-width:0} .nm{font-weight:700;font-size:16px;color:#f1f5f9;display:flex;align-items:center;gap:8px}
.tag{font-size:11px;color:#94a3b8;background:#111c2b;border:1px solid #1e2b3d;padding:2px 8px;border-radius:20px;font-weight:600}
.desc{color:#94a3b8;font-size:13px;margin-top:3px}
.right{display:flex;align-items:center;gap:12px;flex:none}
.pill{font-size:12px;font-weight:600;display:flex;align-items:center;gap:6px;color:#94a3b8}
.dot{width:8px;height:8px;border-radius:50%;background:#475569}
.pill.on{color:#34d399}.pill.on .dot{background:#34d399}
.btn{background:#137fec;color:#fff;border:0;border-radius:10px;padding:10px 18px;font-weight:700;font-size:14px;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center;gap:7px}
.btn:hover{background:#0f6ad0}
.btn.soon{background:#1a2536;color:#7b8aa0}
.btn.off{background:#241a10;color:#ffb35a;border:1px solid #4a3a1a}
@media(max-width:720px){.side{display:none}.main{padding:22px 18px}.card{flex-wrap:wrap}}
</style></head><body>
<div class=wrap>
 <div class=side>
  <div class=brand>📊 RealProfit</div>
  <div class=nav>
   <a href="/">▦ Dashboard</a>
   <a href="/integraciones" class=on>⚙ Integraciones</a>
  </div>
  <div class=userbox><div class=av>{{INICIAL}}</div>
   <div style="line-height:1.25;min-width:0"><div style="font-weight:600;font-size:13px;max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{{EMAIL}}</div>
   <a href="/logout" style="color:#94a3b8;font-size:11px;text-decoration:none">Cerrar sesión</a></div></div>
 </div>
 <div class=main>
  <h1>Integraciones</h1>
  <div class=lead>Conectá tus ventas, pagos y anuncios para ver, en un solo lugar, si tu negocio gana o pierde plata.</div>
  <div class=sechead>Plataformas disponibles</div>
  <div id=cards></div>
 </div>
</div>
<script>
var PLAT=[
 {key:'mp',   nm:'Mercado Pago', tag:'Pagos',    desc:'Trae tus pagos y movimientos para calcular tu ganancia real.', col:'#009ee3', ic:'💳', real:true},
 {key:'meli', nm:'Mercado Libre',tag:'Ventas',   desc:'Incluye tus ventas de marketplace en el cálculo de beneficio.', col:'#ffe600', ic:'🛒'},
 {key:'tn',   nm:'Tiendanube',   tag:'Ventas',   desc:'Trae tus ventas online de tu tienda para el beneficio real.', col:'#2d6cdf', ic:'🌩️'},
 {key:'shopify',nm:'Shopify',    tag:'Ventas',   desc:'Conectá tu tienda para centralizar ventas y costos por orden.', col:'#95bf47', ic:'🛍️'},
 {key:'meta', nm:'Meta Ads',     tag:'Anuncios', desc:'Trae tus campañas de Facebook e Instagram Ads.', col:'#0866ff', ic:'📣'},
];
function esc(s){return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}
function render(mpOn){
 var h='';
 PLAT.forEach(function(p){
  var conectado = (p.key==='mp' && mpOn);
  var estado = conectado
    ? '<span class="pill on"><span class=dot></span>Conectado</span>'
    : '<span class=pill><span class=dot></span>No conectado</span>';
  var boton;
  if(p.key==='mp'){
   boton = conectado
     ? '<a class="btn off" href="/desconectar-mp" onclick="window.location.assign(\\'/desconectar-mp\\');return false;">Desconectar</a>'
     : '<a class="btn" href="/conectar-mp" onclick="window.location.assign(\\'/conectar-mp\\');return false;">⚡ Conectar</a>';
  } else {
   boton = '<button class="btn soon" onclick="alert(\\'Muy pronto podés conectar '+esc(p.nm)+'. Lo estamos activando.\\')">Conectar</button>';
  }
  var dark = (p.col==='#ffe600'||p.col==='#95bf47');
  h += '<div class=card>'
    + '<div class=ic style="background:'+p.col+(dark?';color:#0b1220':';color:#fff')+'">'+p.ic+'</div>'
    + '<div class=info><div class=nm>'+esc(p.nm)+' <span class=tag>'+p.tag+'</span></div><div class=desc>'+esc(p.desc)+'</div></div>'
    + '<div class=right>'+estado+boton+'</div></div>';
 });
 document.getElementById('cards').innerHTML=h;
}
render(false);
fetch('/mp/estado').then(function(r){return r.json();}).then(function(j){ render(!!(j&&j.conectado)); }).catch(function(){});
if(new URLSearchParams(location.search).get('conectado')==='1'){ try{history.replaceState({},'','/integraciones');}catch(e){} }
</script></body></html>"""


@app.get("/integraciones")
def integraciones():
    email = _user_actual()
    if not email:
        return redirect("/")
    inicial = (email[0] if email else "?").upper()
    html = _INTEGRACIONES_PAGE.replace("{{EMAIL}}", email).replace("{{INICIAL}}", inicial)
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.get("/desconectar-mp")
def desconectar_mp():
    email = _user_actual()
    if email:
        d = _mp_tokens()
        d.pop(email, None)
        MP_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return redirect("/integraciones")


@app.get("/conectar-meta")
@limiter.limit("30 per hour")
def conectar_meta():
    """Manda al usuario a autorizar Meta (Facebook Login) para leer sus Ads."""
    if not _user_actual():
        return redirect("/")
    cfg = _meta_cfg()
    if not cfg["app_id"]:
        return ("Falta configurar el App ID de Meta (variables en Render).", 400)
    state = _secrets.token_urlsafe(16)
    session["meta_state"] = state
    qs = _url.urlencode({"client_id": cfg["app_id"], "redirect_uri": cfg["redirect_uri"],
                         "state": state, "response_type": "code",
                         "scope": "ads_read,ads_management,read_insights,business_management,pages_show_list,pages_read_engagement,pages_manage_ads,pages_manage_metadata,instagram_basic,instagram_manage_insights"})
    return redirect("https://www.facebook.com/%s/dialog/oauth?%s" % (META_API, qs), code=302)


@app.get("/meta/callback")
@limiter.limit("30 per hour")
def meta_callback():
    """Meta vuelve con 'code'. Lo cambiamos por token y lo hacemos de larga duración (60 días)."""
    cfg = _meta_cfg()
    code = request.args.get("code")
    state = request.args.get("state")
    if not code:
        return ("RealProfit — punto de conexión con Meta. Volvé a la app y usá «Conectar».", 200)
    if not state or state != session.get("meta_state"):
        return ("La conexión no pasó el control de seguridad. Reintentá desde el botón.", 400)
    try:
        r = requests.get("https://graph.facebook.com/%s/oauth/access_token" % META_API,
                         params={"client_id": cfg["app_id"], "redirect_uri": cfg["redirect_uri"],
                                 "client_secret": cfg["app_secret"], "code": code}, timeout=30)
        tok = r.json() if r.content else {}
    except Exception:
        return ("No pudimos conectar con Meta en este momento. Probá de nuevo.", 502)
    if not tok.get("access_token"):
        return ("Meta no autorizó la conexión. Reintentá.", 400)
    # Token de larga duración (60 días) para no reconectar seguido.
    try:
        r2 = requests.get("https://graph.facebook.com/%s/oauth/access_token" % META_API,
                          params={"grant_type": "fb_exchange_token", "client_id": cfg["app_id"],
                                  "client_secret": cfg["app_secret"], "fb_exchange_token": tok["access_token"]},
                          timeout=30)
        j2 = r2.json() if r2.content else {}
        if j2.get("access_token"):
            tok = j2
    except Exception:
        pass
    email = _user_actual()
    if not email:
        return redirect("/")
    _meta_save_token(email, tok)
    session.pop("meta_state", None)
    return redirect("/?integ=1", code=302)


@app.post("/meta/token-manual")
@limiter.limit("30 per hour")
def meta_token_manual():
    """Conexión SIN OAuth: el usuario pega su token (System User o Graph Explorer) y lo guardamos.
    Lo usan el gasto (dashboard) Y el subir ads. Valida contra Meta antes de guardar."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "error": "Entrá a RealProfit primero."}), 401
    data = request.get_json(silent=True) or {}
    token = str(data.get("token") or "").strip()
    if len(token) < 30:
        return jsonify({"ok": False, "error": "Pegá un token válido de Meta."}), 400
    try:
        r = requests.get("https://graph.facebook.com/%s/me/adaccounts" % META_API,
                         params={"access_token": token,
                                 "fields": "account_id,name,currency,account_status", "limit": 500}, timeout=30)
        j = r.json() if r.content else {}
    except Exception:
        return jsonify({"ok": False, "error": "No pude contactar a Meta. Probá de nuevo."}), 502
    if j.get("error"):
        return jsonify({"ok": False, "error": "Token inválido: " + (j["error"].get("message") or "")[:120]}), 400
    cuentas = [{"id": a.get("account_id"),
                "name": a.get("name") or ("Cuenta " + str(a.get("account_id"))),
                "moneda": a.get("currency")} for a in (j.get("data") or [])]
    if not cuentas:
        return jsonify({"ok": False, "error": "El token no ve ninguna cuenta publicitaria (¿le falta ads_read o el acceso a la cuenta?)."}), 400
    d = _meta_tokens(); tk = d.get(email) or {}
    tk["access_token"] = token
    tk["manual"] = True
    if not tk.get("cuenta"):
        tk["cuenta"] = cuentas[0]["id"]
    d[email] = tk
    META_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return jsonify({"ok": True, "cuentas": cuentas, "elegida": tk.get("cuenta")})


@app.get("/meta/estado")
def meta_estado():
    email = _user_actual()
    tk = _meta_tokens().get(email) if email else None
    conectado = bool(tk and tk.get("access_token"))
    return jsonify({"ok": True, "conectado": conectado, "cuenta": (tk or {}).get("cuenta")})


@app.get("/meta/cuentas")
def meta_cuentas():
    """Lista las cuentas publicitarias a las que el usuario dio acceso al conectar."""
    email = _user_actual()
    tk = _meta_tokens().get(email) if email else None
    if not tk or not tk.get("access_token"):
        return jsonify({"ok": True, "cuentas": [], "elegida": None})
    cuentas = []
    try:
        r = requests.get("https://graph.facebook.com/%s/me/adaccounts" % META_API,
                         params={"access_token": tk["access_token"],
                                 "fields": "account_id,name,currency", "limit": 500}, timeout=30)
        for a in (r.json().get("data") or []):
            cuentas.append({"id": a.get("account_id"), "name": a.get("name") or ("Cuenta " + str(a.get("account_id"))),
                            "moneda": a.get("currency")})
    except Exception:
        pass
    return jsonify({"ok": True, "cuentas": cuentas, "elegida": tk.get("cuenta")})


@app.post("/meta/cuenta")
def meta_elegir_cuenta():
    """Guarda cuál cuenta publicitaria mira este usuario (de ahí sale el gasto)."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "error": "login"}), 401
    data = request.get_json(silent=True) or {}
    cid = str(data.get("cuenta") or "").strip()
    d = _meta_tokens(); tk = d.get(email) or {}
    tk["cuenta"] = cid
    d[email] = tk
    META_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return jsonify({"ok": True})


def _meta_spend(email, desde, hasta):
    """Gasto en ads (USD/ARS) de la cuenta elegida, en el período. 0 si no hay cuenta/conexión.

    ATAJO DUEÑO: si el email logueado es el dueño (env META_OWNER_EMAIL), usa directo el
    System User token de METAFY (env META_OWNER_TOKEN) + su cuenta (env META_OWNER_ACT),
    sin OAuth ni App Review. Solo para tu propio usuario."""
    import os
    tk = _meta_tokens().get(email)
    token = tk.get("access_token") if tk else None    # token PEGADO por el usuario (prioridad)
    cuenta = tk.get("cuenta") if tk else None
    es_owner = False
    if not token:                                      # fallback: System User token de env (atajo dueño)
        owner = os.getenv("META_OWNER_EMAIL", "").strip().lower()
        if owner and email and email.strip().lower() == owner and os.getenv("META_OWNER_TOKEN"):
            token = os.getenv("META_OWNER_TOKEN")
            cuenta = os.getenv("META_OWNER_ACT") or cuenta
            es_owner = True
    if not token or not cuenta:
        return 0.0
    acc = str(cuenta).replace("act_", "")
    try:
        r = requests.get("https://graph.facebook.com/%s/act_%s/insights" % (META_API, acc),
                         params={"access_token": token, "fields": "spend,account_currency",
                                 "time_range": _json.dumps({"since": desde, "until": hasta}),
                                 "level": "account"}, timeout=30)
        data = r.json().get("data") or []
        if data:
            spend = float(data[0].get("spend") or 0)
            moneda = (data[0].get("account_currency") or "").upper()
            # Si la cuenta factura en USD y la tienda es en pesos, convierto con el dólar en vivo
            # (mismo criterio que METAFY: USDC/ARS de criptoya). Para tu usuario (owner) o cualquier
            # cuenta USD. Se puede fijar con la env DOLAR_ARS (número, 'blue' o 'cripto').
            if moneda == "USD" or es_owner:
                spend *= _dolar_ars_vivo()
            return spend
    except Exception:
        pass
    return 0.0


_dolar_cache = {"ts": 0, "val": 0.0}


def _dolar_ars_vivo() -> float:
    """USD→ARS en vivo (USDC/ARS de criptoya, ask de Ripio ≈ lo que se paga en ads), cache 10 min.
    Config con env DOLAR_ARS: número fijo (ej 1500), 'blue', o 'cripto'/'arq' (default)."""
    import os, time as _t
    c = _dolar_cache
    if c["val"] and (_t.time() - c["ts"] < 600):
        return c["val"]
    cfg = (os.getenv("DOLAR_ARS") or "cripto").strip().lower()
    try:
        return float(cfg)                    # valor fijo si pusieron un número
    except ValueError:
        pass
    val = 0.0
    try:
        if cfg == "blue":
            val = float(requests.get("https://dolarapi.com/v1/dolares/blue", timeout=10).json().get("venta") or 0)
        else:
            j = requests.get("https://criptoya.com/api/usdc/ars/1", timeout=10).json()
            val = float((j.get("ripioexchange") or {}).get("ask") or 0)
            if not val:
                asks = [float(v.get("ask") or 0) for v in j.values() if isinstance(v, dict) and v.get("ask")]
                val = sum(asks) / len(asks) if asks else 0.0
        if not val:
            val = float(requests.get("https://dolarapi.com/v1/dolares/cripto", timeout=10).json().get("venta") or 0)
    except Exception:
        val = 0.0
    if not val or val < 500:                 # API caída → último bueno o piso sano (nunca 1)
        return c["val"] or 1000.0
    c.update({"ts": _t.time(), "val": val})
    return val


@app.get("/desconectar-meta")
def desconectar_meta():
    email = _user_actual()
    if email:
        d = _meta_tokens()
        d.pop(email, None)
        META_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return redirect("/?integ=1")


# ==================== SUBIR CREATIVOS (Meta Ads) ====================
_ADS_API = "https://graph.facebook.com/v23.0"
_ADS_JOBS = {}
_ADS_UPLOADS = {}   # upload_id -> carpeta temporal con los videos que subió el usuario

# Cuentas configuradas (CP1/NoxaLab). Extensible a más cuentas.
_ADS_CUENTAS = {
    "cp1": {
        "nombre": "CP1 — NoxaLab", "ad_account": "1913715339273327",
        "page": "1175786222292931", "pixel": "1592535622574011", "ig": "17841415440483313",  # IG @noxalab.ar
        "landing": "https://noxalab-arg.myshopify.com/products/noxalab",
        "titulo": "+10.000 Hombres Usan NoxaLab 💪", "subtitulo": "Ultimas unidades", "presupuesto": 35,
        "copy": ("⚡ ¿Sentís que el cuerpo ya no responde como antes?\n\n"
                 "Si venís buscando:\n\n"
                 "\U0001f525 Recuperar tu energía y llegar a la noche con ganas\n"
                 "\U0001f4aa Sentirte más fuerte y seguro de vos mismo\n"
                 "❤️ Apoyar una mejor circulación y vitalidad masculina\n\n"
                 "...no sos el único.\n\n"
                 "Miles de hombres ya están sumando NoxaLab a su rutina diaria.\n"
                 "Solo 1 scoop al día, disuelto en agua.\n"
                 "Fórmula en polvo con 7 activos y NAD+ liposomal.\n\n"
                 "\U0001f447 Tocá \"Comprar Ahora\" y descubrí por qué."),
    },
    "va1": {
        "nombre": "VA1 — VisionPure", "ad_account": "964010428983612",
        "page": "1105184446002428", "pixel": "1237139148560920", "ig": "17841471399362397",  # IG @visionpure.argentina
        "landing": "https://tryvisionpure.shop/productos/visionpure-recupera-la-nitidez-que-perdiste-con-los-anos2/",
        "titulo": "Recuperá la nitidez que perdiste con los años", "subtitulo": "Envío gratis hoy", "presupuesto": 35,
        "copy": ("\U0001f441️ ¿Ves más borroso de cerca o te cuesta manejar de noche?\n\n"
                 "Con los años, los ojos piden una mano.\n\n"
                 "VisionPure es un spray ocular con luteína que:\n"
                 "✨ Ayuda a nutrir tus ojos día a día\n"
                 "\U0001f319 Muchos notaron ver más nítido para leer y manejar\n"
                 "\U0001f4a7 Fácil de usar: un par de gotas, sin pastillas\n\n"
                 "Miles de personas +65 ya lo suman a su rutina.\n\n"
                 "\U0001f447 Tocá \"Comprar Ahora\" y recuperá la nitidez."),
    },
}


_ads_local = threading.local()   # token del usuario que lanzó la subida (por thread), para no cruzar cuentas


def _ads_token():
    # 1) Si el job de subida fijó un token para ESTE thread, usá ese (cuenta correcta, sin cruces).
    t = getattr(_ads_local, "token", None)
    if t:
        return t
    # 2) Request: token que el usuario logueado PEGÓ en la app (meta_tokens.json) → env META_TOKEN.
    try:
        email = _user_actual()
    except Exception:
        email = None
    toks = _meta_tokens()
    if email and (toks.get(email) or {}).get("access_token"):
        return toks[email]["access_token"]
    for v in toks.values():
        if isinstance(v, dict) and v.get("access_token"):
            return v["access_token"]
    return _os.getenv("META_TOKEN") or ""


def _ads_call(method, path, data=None, files=None, params=None):
    p = dict(params or {}); p["access_token"] = _ads_token()
    r = requests.request(method, "%s/%s" % (_ADS_API, path), data=data, files=files, params=p, timeout=120)
    j = r.json() if r.content else {}
    if r.status_code >= 400:
        e = j.get("error", {})
        raise RuntimeError(e.get("error_user_msg") or e.get("message") or ("HTTP %d" % r.status_code))
    return j


def _ads_crear(acct, obj, payload):
    return _ads_call("POST", "act_%s/%s" % (acct, obj),
                     data={k: (_json.dumps(v) if isinstance(v, (dict, list)) else v)
                           for k, v in payload.items()})["id"]


def _ads_subir_video(acct, ruta):
    """Sube un video a /advideos. Grandes (>40MB) por partes (resumable)."""
    size = _os.path.getsize(ruta)
    base = "act_%s/advideos" % acct
    if size <= 40 * 1024 * 1024:
        with open(ruta, "rb") as f:
            return _ads_call("POST", base, files={"source": f})["id"]
    ini = _ads_call("POST", base, data={"upload_phase": "start", "file_size": size})
    sess, vid = ini["upload_session_id"], ini["video_id"]
    so, eo = int(ini["start_offset"]), int(ini["end_offset"])
    with open(ruta, "rb") as f:
        while so < eo:
            f.seek(so); chunk = f.read(eo - so)
            res = _ads_call("POST", base,
                            data={"upload_phase": "transfer", "upload_session_id": sess, "start_offset": so},
                            files={"video_file_chunk": ("chunk", chunk)})
            so, eo = int(res["start_offset"]), int(res["end_offset"])
    _ads_call("POST", base, data={"upload_phase": "finish", "upload_session_id": sess})
    return vid


def _ads_es_imagen(ruta):
    return ruta.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))


def _ads_subir_imagen(acct, ruta):
    """Sube una imagen a /adimages y devuelve su image_hash."""
    with open(ruta, "rb") as f:
        r = _ads_call("POST", "act_%s/adimages" % acct, files={"img": f})
    imgs = r.get("images", {}) or {}
    if not imgs:
        raise RuntimeError("Meta no devolvió el hash de la imagen")
    return next(iter(imgs.values()))["hash"]


def _ads_video_ready(vid):
    import time as _t
    for _ in range(75):
        st = _ads_call("GET", str(vid), params={"fields": "status"}).get("status", {}).get("video_status")
        if st == "ready":
            return True
        if st == "error":
            return False
        _t.sleep(2.5)
    return True


def _ads_thumb(vid):
    import time as _t
    for _ in range(15):
        d = _ads_call("GET", "%s/thumbnails" % vid, params={"fields": "uri,is_preferred"}).get("data", [])
        if d:
            return ([t for t in d if t.get("is_preferred")] or d)[0]["uri"]
        _t.sleep(2)
    try:
        return _ads_call("GET", str(vid), params={"fields": "picture"}).get("picture")
    except Exception:
        return None


def _ads_start_5am():
    now = _dt.datetime.utcnow() - _dt.timedelta(hours=3)
    d = now.date() if now.hour < 5 else (now + _dt.timedelta(days=1)).date()
    return "%sT05:00:00-03:00" % d.isoformat()


def _ads_sched(params):
    """start_time desde el día/hora que eligió el usuario; si no, 5am por defecto."""
    f = (params.get("fecha") or "").strip()
    h = (params.get("hora") or "05:00").strip()
    if f:
        if len(h) != 5:
            h = "05:00"
        return "%sT%s:00-03:00" % (f, h)
    return _ads_start_5am()


def _ads_camp_payload(nombre, cbo, presup, status):
    p = {"name": nombre, "objective": "OUTCOME_SALES", "special_ad_categories": [],
         "buying_type": "AUCTION", "bid_strategy": "LOWEST_COST_WITHOUT_CAP", "status": status}
    if cbo:
        p["daily_budget"] = int(presup) * 100
    return p


def _ads_adset_payload(nombre, campaign_id, pixel, cbo, presup, status, start=None):
    p = {"name": nombre, "campaign_id": campaign_id, "billing_event": "IMPRESSIONS",
         "optimization_goal": "OFFSITE_CONVERSIONS",
         "promoted_object": {"pixel_id": pixel, "custom_event_type": "PURCHASE"},
         "attribution_spec": [{"event_type": "CLICK_THROUGH", "window_days": 7},
                              {"event_type": "VIEW_THROUGH", "window_days": 1}],
         "targeting": {"geo_locations": {"countries": ["AR"]},
                       "targeting_automation": {"advantage_audience": 1}},
         "status": status}
    if not cbo:
        p["daily_budget"] = int(presup) * 100        # ABO: presupuesto por conjunto
    if status == "ACTIVE":
        p["start_time"] = start or _ads_start_5am()
    return p


def _ads_creative_payload(nombre, medio, cfg, ad):
    """medio: {'kind':'video','video_id','thumb'} o {'kind':'image','image_hash'}.
    ad: {copy, titulo, subtitulo, url} (cae a los defaults de la cuenta)."""
    url = (ad.get("url") or "").strip() or cfg["landing"]
    cta = {"type": "SHOP_NOW", "value": {"link": url}}
    title = (ad.get("titulo") or cfg.get("titulo") or "")
    msg = ad.get("copy") or cfg["copy"]
    _sub = (ad.get("subtitulo") or "").strip() or (cfg.get("subtitulo") or "").strip()
    if medio.get("kind") == "image":
        ld = {"image_hash": medio["image_hash"], "link": url, "message": msg,
              "name": title, "call_to_action": cta}
        if _sub:
            ld["description"] = _sub
        oss = {"page_id": cfg["page"], "link_data": ld}
    else:
        vd = {"video_id": medio["video_id"], "message": msg,
              "call_to_action": cta, "title": title}
        if _sub:
            vd["link_description"] = _sub
        if medio.get("thumb"):
            vd["image_url"] = medio["thumb"]
        oss = {"page_id": cfg["page"], "video_data": vd}
    if cfg.get("ig"):
        oss["instagram_user_id"] = cfg["ig"]
    return {"name": nombre, "object_story_spec": oss,
            "degrees_of_freedom_spec": {"creative_features_spec": {"site_extensions": {"enroll_status": "OPT_OUT"}}}}


def _ads_adset_dup(acct, src_id, campaign_id, nombre, pixel, status, start=None):
    """Crea un conjunto NUEVO copiando la config de src_id (segmentación/optimización/pixel/atribución),
    vacío. No copia los ads viejos. Filtra placeholders que Meta rechaza (UNDEFINED/NONE)."""
    F = ("name,billing_event,optimization_goal,targeting,promoted_object,attribution_spec,"
         "destination_type,optimization_sub_event,pacing_type,daily_budget")
    s = _ads_call("GET", str(src_id), params={"fields": F})
    p = {"name": nombre or (s.get("name", "CONJUNTO") + " copia"), "campaign_id": campaign_id,
         "billing_event": s.get("billing_event") or "IMPRESSIONS",
         "optimization_goal": s.get("optimization_goal") or "OFFSITE_CONVERSIONS",
         "targeting": s.get("targeting") or {"geo_locations": {"countries": ["AR"]},
                                             "targeting_automation": {"advantage_audience": 1}},
         "status": status}
    if s.get("promoted_object"):
        p["promoted_object"] = s["promoted_object"]
    elif pixel:
        p["promoted_object"] = {"pixel_id": pixel, "custom_event_type": "PURCHASE"}
    if s.get("attribution_spec"):
        p["attribution_spec"] = s["attribution_spec"]
    if s.get("destination_type") and s["destination_type"] != "UNDEFINED":
        p["destination_type"] = s["destination_type"]
    if s.get("optimization_sub_event") not in (None, "", "NONE"):
        p["optimization_sub_event"] = s["optimization_sub_event"]
    if s.get("pacing_type"):
        p["pacing_type"] = s["pacing_type"]
    if s.get("daily_budget"):
        p["daily_budget"] = s["daily_budget"]        # ABO: copia el presupuesto del conjunto
    if status == "ACTIVE":
        p["start_time"] = start or _ads_start_5am()
    return _ads_crear(acct, "adsets", p)


def _ads_google_creds():
    from google.oauth2 import service_account
    scopes = ["https://www.googleapis.com/auth/drive.readonly"]
    raw = _os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
    if raw:
        return service_account.Credentials.from_service_account_info(_json.loads(raw), scopes=scopes)
    fp = _os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE") or str(RAIZ / "service-account.json")
    return service_account.Credentials.from_service_account_file(fp, scopes=scopes)


def _ads_drive_fid(link):
    m = _re_and.search(r"/folders/([A-Za-z0-9_-]+)", link or "")
    if m:
        return m.group(1)
    m = _re_and.search(r"[?&]id=([A-Za-z0-9_-]+)", link or "")
    return m.group(1) if m else (link or "").strip()


def _ads_drive_listar(link):
    """Lista los videos de la carpeta (sin bajarlos) para el botón 'Buscar videos'."""
    from googleapiclient.discovery import build
    svc = build("drive", "v3", credentials=_ads_google_creds())
    fid = _ads_drive_fid(link)
    q = "'%s' in parents and trashed=false" % fid
    files = svc.files().list(q=q, fields="files(id,name,mimeType,size)", supportsAllDrives=True,
                             includeItemsFromAllDrives=True, pageSize=100).execute().get("files", [])
    vids = [f for f in files if "video" in (f.get("mimeType") or "") or "image" in (f.get("mimeType") or "")
            or f.get("name", "").lower().endswith((".mp4", ".mov", ".m4v", ".jpg", ".jpeg", ".png", ".webp"))]
    vids.sort(key=lambda f: f.get("name", ""))
    return [{"name": f.get("name", ""),
             "mb": round(int(f.get("size") or 0) / 1048576) if f.get("size") else 0} for f in vids]


def _ads_drive_bajar(link, dest_dir):
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    svc = build("drive", "v3", credentials=_ads_google_creds())
    fid = _ads_drive_fid(link)
    q = "'%s' in parents and trashed=false" % fid
    files = svc.files().list(q=q, fields="files(id,name,mimeType)", supportsAllDrives=True,
                             includeItemsFromAllDrives=True, pageSize=100).execute().get("files", [])
    vids = [f for f in files if "video" in (f.get("mimeType") or "") or "image" in (f.get("mimeType") or "")
            or f.get("name", "").lower().endswith((".mp4", ".mov", ".m4v", ".jpg", ".jpeg", ".png", ".webp"))]
    vids.sort(key=lambda f: f.get("name", ""))
    out = []
    for f in vids:
        p = _os.path.join(dest_dir, f["name"])
        req = svc.files().get_media(fileId=f["id"])
        with open(p, "wb") as fh:
            dl = MediaIoBaseDownload(fh, req, chunksize=1024 * 1024 * 8)
            done = False
            while not done:
                _, done = dl.next_chunk()
        out.append(p)
    return out


def _ads_run(job, params):
    """Baja de Drive → sube videos → crea campaña + N conjuntos + ads. Con progreso."""
    _ads_local.token = params.get("_token") or None   # usar el token del usuario que lanzó (cuenta correcta)
    st = _ADS_JOBS.get(job)
    import tempfile, shutil
    tmp = tempfile.mkdtemp(prefix="ads_")
    try:
        cfg = dict(_ADS_CUENTAS.get(params.get("cuenta") or "cp1") or _ADS_CUENTAS["cp1"])
        if (params.get("page") or "").strip():
            cfg["page"] = params["page"].strip()
        if (params.get("pixel") or "").strip():
            cfg["pixel"] = params["pixel"].strip()
        if "ig" in params:
            cfg["ig"] = (params.get("ig") or "").strip()
        acct, pixel = cfg["ad_account"], cfg["pixel"]
        cbo = (params.get("tipo") or "cbo") != "abo"
        presup = int(params.get("presupuesto") or cfg["presupuesto"])
        n_conj = max(1, min(20, int(params.get("conjuntos") or 1)))
        estado = "ACTIVE" if params.get("estado") == "activa" else "PAUSED"
        angulo = (params.get("angulo") or "VARIOS").strip()
        ad = {"copy": (params.get("copy") or "").strip(), "titulo": (params.get("titulo") or "").strip(),
              "subtitulo": (params.get("subtitulo") or "").strip(), "url": (params.get("url") or "").strip()}
        modo_conj = params.get("modo_conjunto") or "nuevo"   # nuevo | dup | usar
        src = params.get("adset_src_id")
        start = _ads_sched(params)                            # día/hora de salida

        up_id = (params.get("upload_id") or "").strip()
        if up_id:
            st["msg"] = "Tomando tus videos…"
            updir = _ADS_UPLOADS.get(up_id)
            rutas = sorted(_os.path.join(updir, f) for f in _os.listdir(updir)) if (updir and _os.path.isdir(updir)) else []
            if not rutas:
                raise RuntimeError("no encontré los videos que subiste (probá subirlos de nuevo)")
        else:
            st["msg"] = "Bajando videos de Drive…"
            rutas = _ads_drive_bajar(params.get("drive", ""), tmp)
            if not rutas:
                raise RuntimeError("no encontré videos en ese Drive (¿está compartido con la service account?)")
        st["total"] = len(rutas) * 2 + 1 + n_conj

        # subir + procesar videos EN PARALELO (mucho más rápido que uno por uno)
        import concurrent.futures as _cf
        medios = [None] * len(rutas)
        _pl = threading.Lock(); _pn = {"n": 0}

        def _prep(idx, ruta):
            if _ads_es_imagen(ruta):
                medio = {"kind": "image", "image_hash": _ads_subir_imagen(acct, ruta)}
            else:
                vid = _ads_subir_video(acct, ruta)
                _ads_video_ready(vid)
                medio = {"kind": "video", "video_id": vid, "thumb": _ads_thumb(vid)}
            with _pl:
                _pn["n"] += 2
                st["done"] = _pn["n"]
                st["msg"] = "Creativos listos %d/%d…" % (_pn["n"] // 2, len(rutas))
            return idx, medio

        with _cf.ThreadPoolExecutor(max_workers=min(6, max(1, len(rutas)))) as _ex:
            _futs = [_ex.submit(_prep, i, r) for i, r in enumerate(rutas)]
            for _f in _cf.as_completed(_futs):
                _idx, _m = _f.result()
                medios[_idx] = _m

        # campaña
        st["done"] = len(rutas) * 2; st["msg"] = "Creando campaña…"
        now = _dt.datetime.utcnow() - _dt.timedelta(hours=3)
        fecha = "%d-%d" % (now.day, now.month)
        if params.get("modo_campana") == "existente" and params.get("campaign_id"):
            campaign_id = params["campaign_id"]
        else:
            campaign_id = _ads_crear(acct, "campaigns",
                                     _ads_camp_payload("%s %s" % (fecha, angulo), cbo, presup, estado))

        # determinar los CONJUNTOS destino
        base = (params.get("conjunto_nombre") or "CONJUNTO").strip() or "CONJUNTO"
        adsets = []
        if modo_conj == "usar" and src:
            adsets = [src]                                   # añadir los ads a un conjunto existente
            n_conj = 1
        else:
            for c in range(n_conj):
                st["done"] = len(rutas) * 2 + 1 + c
                st["msg"] = "Creando conjunto %d de %d…" % (c + 1, n_conj)
                nombre_conj = base if n_conj == 1 else ("%s %d" % (base, c + 1))
                if modo_conj == "dup" and src:              # copia la config de un conjunto existente
                    adsets.append(_ads_adset_dup(acct, src, campaign_id, nombre_conj, pixel, estado, start))
                else:                                        # conjunto nuevo estándar
                    adsets.append(_ads_crear(acct, "adsets",
                                             _ads_adset_payload(nombre_conj, campaign_id, pixel, cbo, presup, estado, start)))

        # ads en cada conjunto (1 por video)
        creados = 0
        for adset_id in adsets:
            for i, medio in enumerate(medios, start=1):
                cid = _ads_crear(acct, "adcreatives", _ads_creative_payload(str(i), medio, cfg, ad))
                _ads_crear(acct, "ads", {"name": str(i), "adset_id": adset_id,
                                         "creative": {"creative_id": cid}, "status": estado})
                creados += 1
        st["done"] = st["total"]
        st["stats"] = {"campaign_id": campaign_id, "conjuntos": len(adsets), "ads": creados,
                       "tipo": "CBO" if cbo else "ABO",
                       "estado": "Programada 5 AM" if estado == "ACTIVE" else "Pausada"}
        st["msg"] = "¡Listo! %d anuncios en %d conjunto(s) (%s)." % (creados, n_conj, st["stats"]["estado"])
        st["listo"] = True
    except Exception as e:
        st["error"] = str(e); st["listo"] = True
    finally:
        try:
            shutil.rmtree(tmp, ignore_errors=True)
        except Exception:
            pass


@app.get("/pf-ads-cuentas")
def pf_ads_cuentas():
    if not _user_actual():
        return jsonify({"ok": False, "cuentas": []})
    tok = _ads_token()
    accesibles = set()                                   # cuentas publicitarias que ve ESTE token
    if tok:
        try:
            r = requests.get("https://graph.facebook.com/%s/me/adaccounts" % META_API,
                             params={"access_token": tok, "fields": "account_id", "limit": 200}, timeout=20)
            accesibles = {str(a.get("account_id")) for a in ((r.json() or {}).get("data") or [])}
        except Exception:
            accesibles = set()
    cs = [{"key": k, "nombre": v["nombre"], "presupuesto": v["presupuesto"], "copy": v["copy"],
           "titulo": v.get("titulo", ""), "subtitulo": v.get("subtitulo", "")}
          for k, v in _ADS_CUENTAS.items()
          if (not accesibles) or str(v["ad_account"]) in accesibles]   # solo las que el token puede usar
    return jsonify({"ok": True, "cuentas": cs, "token": bool(tok)})


@app.get("/pf-ads-identidad")
def pf_ads_identidad():
    """Página / IG / pixel VINCULADOS a la cuenta (para los desplegables)."""
    if not _user_actual():
        return jsonify({"ok": False})
    cfg = _ADS_CUENTAS.get(request.args.get("cuenta") or "cp1") or _ADS_CUENTAS["cp1"]
    acct = cfg["ad_account"]

    def lst(path, fields):
        try:
            return _ads_call("GET", "act_%s/%s" % (acct, path), params={"fields": fields, "limit": 50}).get("data", [])
        except Exception:
            return []

    def lst_raw(path, fields):
        try:
            return _ads_call("GET", path, params={"fields": fields, "limit": 100}).get("data", [])
        except Exception:
            return []
    pixels = [{"id": p["id"], "name": p.get("name", "")} for p in lst("adspixels", "id,name")]
    igs = [{"id": i["id"], "name": "@" + (i.get("username") or "")} for i in lst("instagram_accounts", "id,username")]
    # Páginas: las que conectaste (me/accounts) + las promocionables de la cuenta. Sin nombres hardcodeados.
    pages, seen = [], set()
    for p in (lst_raw("me/accounts", "id,name") + lst("promote_pages", "id,name")):
        pid = p.get("id")
        if pid and pid not in seen:
            seen.add(pid); pages.append({"id": pid, "name": p.get("name", "") or "Página"})
    if cfg.get("page") and cfg["page"] not in seen:                # la del config, con su NOMBRE REAL
        nm = ""
        try:
            nm = (_ads_call("GET", str(cfg["page"]), params={"fields": "name"}) or {}).get("name", "")
        except Exception:
            pass
        pages.insert(0, {"id": cfg["page"], "name": nm or "Página conectada"})
    return jsonify({"ok": True, "pixels": pixels, "igs": igs, "pages": pages,
                    "def": {"page": cfg["page"], "pixel": cfg["pixel"], "ig": cfg.get("ig", "")}})


@app.get("/pf-ads-campanas")
def pf_ads_campanas():
    if not _user_actual():
        return jsonify({"ok": False, "campanas": []})
    cfg = _ADS_CUENTAS.get(request.args.get("cuenta") or "cp1") or _ADS_CUENTAS["cp1"]
    try:
        r = _ads_call("GET", "act_%s/campaigns" % cfg["ad_account"],
                      params={"fields": "name,effective_status,daily_budget", "limit": 100,
                              "effective_status": _json.dumps(["ACTIVE"])})   # solo campañas ACTIVAS
        out = [{"id": c["id"], "name": c.get("name", ""), "cbo": bool(c.get("daily_budget")),
                "presupuesto": int((c.get("daily_budget") or 0)) // 100, "activa": True}
               for c in r.get("data", []) if c.get("effective_status") == "ACTIVE"]
        out.sort(key=lambda x: int(x["id"]) if str(x["id"]).isdigit() else 0, reverse=True)
        return jsonify({"ok": True, "campanas": out})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "campanas": []})


@app.get("/pf-ads-conjuntos")
def pf_ads_conjuntos():
    """Conjuntos (adsets) REALES de una campaña, para elegir cuál duplicar / al que sumar ads."""
    if not _user_actual():
        return jsonify({"ok": False, "conjuntos": []})
    cid = (request.args.get("campaign_id") or "").strip()
    if not cid:
        return jsonify({"ok": False, "conjuntos": []})
    try:
        r = _ads_call("GET", "%s/adsets" % cid,
                      params={"fields": "name,effective_status", "limit": 60})
        out = []
        for s in r.get("data", []):
            try:
                n = len(_ads_call("GET", "%s/ads" % s["id"], params={"fields": "id", "limit": 100}).get("data", []))
            except Exception:
                n = 0
            out.append({"id": s["id"], "name": s.get("name", ""),
                        "activo": s.get("effective_status") == "ACTIVE", "n_ads": n})
        out.sort(key=lambda x: int(x["id"]) if str(x["id"]).isdigit() else 0, reverse=True)
        return jsonify({"ok": True, "conjuntos": out})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "conjuntos": []})


@app.post("/pf-ads-drive-listar")
def pf_ads_drive_listar():
    if not _user_actual():
        return jsonify({"ok": False}), 401
    data = request.get_json(silent=True) or {}
    if not (data.get("drive") or "").strip():
        return jsonify({"ok": False, "msg": "pegá el link de Drive"}), 400
    try:
        vids = _ads_drive_listar(data["drive"])
        if not vids:
            return jsonify({"ok": False, "msg": "no encontré videos (¿compartiste la carpeta con la service account?)"})
        return jsonify({"ok": True, "videos": vids})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.post("/pf-ads-subir")
def pf_ads_subir():
    if not _user_actual():
        return jsonify({"ok": False}), 401
    import tempfile, uuid
    files = request.files.getlist("videos")
    if not files:
        return jsonify({"ok": False, "msg": "elegí al menos un video"}), 400
    up_id = uuid.uuid4().hex[:12]
    d = tempfile.mkdtemp(prefix="adsup_")
    vids = []
    for f in files:
        name = _os.path.basename(f.filename or "video.mp4")
        if not name.lower().endswith((".mp4", ".mov", ".m4v", ".jpg", ".jpeg", ".png", ".webp")):
            continue
        ruta = _os.path.join(d, name)
        f.save(ruta)
        vids.append({"name": name, "mb": round(_os.path.getsize(ruta) / 1048576)})
    if not vids:
        return jsonify({"ok": False, "msg": "esos archivos no son videos ni fotos (.mp4/.mov/.jpg/.png)"}), 400
    vids.sort(key=lambda v: v["name"])
    _ADS_UPLOADS[up_id] = d
    return jsonify({"ok": True, "upload_id": up_id, "videos": vids})


@app.post("/pf-ads-lanzar")
def pf_ads_lanzar():
    if not _user_actual():
        return jsonify({"ok": False}), 401
    if not _ads_token():
        return jsonify({"ok": False, "msg": "falta META_TOKEN en el servidor (Render → Environment)"}), 400
    data = request.get_json(silent=True) or {}
    if not (data.get("drive") or "").strip() and not (data.get("upload_id") or "").strip():
        return jsonify({"ok": False, "msg": "pegá el link de Drive o subí tus videos"}), 400
    import uuid
    job = uuid.uuid4().hex[:12]
    data["_token"] = _ads_token()   # token del usuario logueado AHORA (request) → el thread usa este, no el de otra cuenta
    _ADS_JOBS[job] = {"done": 0, "total": 0, "msg": "Arrancando…", "listo": False, "error": None, "stats": {}}
    threading.Thread(target=_ads_run, args=(job, data), daemon=True).start()
    return jsonify({"ok": True, "job": job})


@app.get("/pf-ads-progreso")
def pf_ads_progreso():
    if not _user_actual():
        return jsonify({"ok": False}), 401
    st = _ADS_JOBS.get((request.args.get("job") or "").strip())
    if not st:
        return jsonify({"ok": False}), 404
    return jsonify({"ok": True, **{k: st[k] for k in ("done", "total", "msg", "listo", "error", "stats")}})


# ---------------- Dashboard ----------------
@app.get("/")
def home():
    email = _user_actual()
    if not email:                       # sin login → pantalla de ingreso
        return Response(_LOGIN_PAGE, mimetype="text/html")
    try:
        html = (RAIZ / "pf.html").read_text(encoding="utf-8")
    except Exception as e:
        return Response(f"<p style='color:#fff;font-family:sans-serif;padding:20px'>"
                        f"No se pudo cargar pf.html: {e}</p>", mimetype="text/html")
    import json as _json
    blob = _json.dumps(_blob_vacio(), ensure_ascii=False)
    # Inyectamos datos VACÍOS (sin esto el dashboard haría fetch y mostraría error).
    if "</head>" in html:
        html = html.replace("</head>", "<script>window.__MFY__=" + blob + ";</script></head>", 1)
    # Caja de usuario abajo a la izquierda (email + cerrar sesión).
    inicial = (email[0] if email else "?").upper()
    userbox = ('<a class="rp-pill" href="/logout" title="Cerrar sesión" style="bottom:16px" '
               'onclick="window.location.assign(\'/logout\');return false;">'
               '<span class="rp-ic" style="background:#137fec;color:#fff;font-weight:700;font-size:14px">'
               + inicial + '</span>'
               '<span class="rp-lbl"><span class="em">' + email + '</span>'
               '<span style="color:#94a3b8;font-size:11px">Cerrar sesión &#8594;</span></span></a>')
    # Dejar solo Dashboard + Integraciones, sacar el logo, botón MP y caja de usuario.
    extra = _SOLO_DASH + userbox
    if "</body>" in html:
        html = html.replace("</body>", extra + "</body>", 1)
    else:
        html = html + extra
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


# ---------------- Endpoints en blanco (para que no rompa nada) ----------------
_PF_CACHE = {}   # (email, desde, hasta) -> (momento, blob) — evita pegarle a Shopify en cada refresco


@app.get("/pf-periodo")
def pf_periodo():
    email = _user_actual()
    desde = request.args.get("desde") or _hoy()
    hasta = request.args.get("hasta") or desde
    if email:
        key = (email, desde, hasta)
        now = _dt.datetime.utcnow()
        c = _PF_CACHE.get(key)
        if c and (now - c[0]).total_seconds() < 60:
            return jsonify({"ok": True, **c[1]})
        blob = None
        if email in _shop_tokens():
            blob = _shopify_resumen(email, desde, hasta)
        tn_blob = _tn_resumen(email, desde, hasta)   # Tiendanube (None si no está conectada)
        if tn_blob:
            blob = _combinar_resumen(blob, tn_blob)
        if blob is None:
            blob = _blob_vacio()
        # Gasto en ads de Meta (cuenta elegida) → INVERSIÓN ADS / ROAS / CPA / ganancia.
        spend = _meta_spend(email, desde, hasta)
        if spend:
            r = blob["raw"]
            fact = r.get("facturado", 0.0)
            ordenes = r.get("ordenes", 0)
            r["publi_ars"] = round(spend, 2)
            r["publi_cuenta"] = round(spend, 2)
            r["ganancia"] = round(r.get("ganancia", fact) - spend, 2)
            r["margen"] = round(r["ganancia"] / fact * 100, 2) if fact else 0.0
            r["roas"] = round(fact / spend, 2) if spend else 0.0
            r["cpa"] = round(spend / ordenes, 2) if ordenes else 0.0
            r["gan_por_venta"] = round(r["ganancia"] / ordenes, 2) if ordenes else 0.0
            r["tot_ganancia"] = r["ganancia"]
            r["tot_margen"] = r["margen"]
        _PF_CACHE[key] = (now, blob)
        return jsonify({"ok": True, **blob})
    return jsonify({"ok": True, **_blob_vacio()})


PROD_COSTOS = DATA_DIR / "prod_costos.json"   # costos por producto que carga cada usuario (persistente)


def _costos() -> dict:
    try:
        return _json.loads(PROD_COSTOS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _costo_qty(cost, qty) -> float:
    """Costo de `qty` unidades.
    - VARIABLE: costo guardado como dict {'1':c1,'2':c2,'3':c3,'4':c4} (costo por cantidad).
      Usa el de esa cantidad exacta; si no está (5+), usa el unitario ('1') × cantidad.
    - UNITARIO/FIJO (número plano): costo*qty (backward-compatible)."""
    if not cost:
        return 0.0
    try:
        q = int(qty or 0)
    except Exception:
        q = 0
    if isinstance(cost, dict):
        if str(q) in cost:
            return float(cost.get(str(q)) or 0)
        return float(cost.get("1") or 0) * q     # cantidad no definida → unitario × cantidad
    try:
        return float(cost) * q
    except Exception:
        return 0.0


def _costo_num(c) -> float:
    """Costo como número para pantallas que muestran 1 solo valor (usa la cantidad '1' si es variable)."""
    if isinstance(c, dict):
        return float(c.get("1") or 0)
    try:
        return float(c or 0)
    except Exception:
        return 0.0


def _guardar_costo(email, pid, costo) -> None:
    d = _costos()
    u = d.get(email) or {}
    if isinstance(costo, dict):                                   # VARIABLE: {'1':c1,'2':c2,...}
        m = {}
        for k, v in costo.items():
            if str(k).isdigit():
                try:
                    fv = float(v or 0)
                except Exception:
                    fv = 0
                if fv > 0:
                    m[str(int(k))] = fv
        if m:
            u[str(pid)] = m
        else:
            u.pop(str(pid), None)
    elif costo and float(costo) > 0:                              # UNITARIO/FIJO (número)
        u[str(pid)] = float(costo)
    else:
        u.pop(str(pid), None)
    d[email] = u
    PROD_COSTOS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


_TN_PROD_CACHE = {}   # {email: (ts, [raw])} — el fetch a TN es lento, lo cacheamos 60s


def _tn_productos_raw(email):
    """Fetch crudo de productos de Tiendanube (cacheado 60s). Solo id/nombre/sku/precio/img."""
    import time as _t
    c = _TN_PROD_CACHE.get(email)
    if c and (_t.time() - c[0] < 60):
        return c[1]
    tk = _tn_tokens().get(email)
    if not tk or not tk.get("access_token") or not tk.get("store_id"):
        return []
    store, hdr = tk["store_id"], _tn_headers(tk["access_token"])
    raw, page = [], 1
    while page <= 10:
        try:
            r = requests.get("%s/%s/products" % (TN_API, store), headers=hdr,
                             params={"per_page": 200, "page": page,
                                     "fields": "id,name,variants,images"}, timeout=30)   # menos payload = más rápido
            lote = r.json() if r.content else []
        except Exception:
            lote = []
        if not isinstance(lote, list) or not lote:
            break
        for p in lote:
            nombre = p.get("name")
            if isinstance(nombre, dict):
                nombre = nombre.get("es") or next(iter(nombre.values()), "") if nombre else ""
            v = (p.get("variants") or [{}])[0]
            imgs = p.get("images") or []
            raw.append({"pid": "tn:%s" % p.get("id"), "nombre": nombre or "",
                        "sku_shopify": v.get("sku") or "", "precio": float(v.get("price") or 0),
                        "img": (imgs[0].get("src") if imgs else "") or ""})
        if len(lote) < 200:
            break
        page += 1
    _TN_PROD_CACHE[email] = (_t.time(), raw)
    return raw


def _tn_productos(email):
    """Productos de TN con costo/SKU aplicados FRESCOS (el fetch va cacheado; las ediciones no)."""
    costos = (_costos().get(email) or {})
    skus = _skus_map(email)
    prods = []
    for p in _tn_productos_raw(email):
        pid = p["pid"]
        g = skus.get(pid)
        cfg = _sku_cfg(g) if g else {"tipo": "fijo", "base": p["sku_shopify"], "map": {}}
        prods.append({
            "id": pid, "nombre": p["nombre"],
            "sku_tipo": cfg["tipo"], "sku_base": cfg["base"], "sku_map": cfg.get("map") or {},
            "sku_ej": _sku_calc(g or cfg, 2),
            "precio": p["precio"], "img": p["img"],
            "costo": costos.get(pid) or 0,
        })
    return prods


@app.get("/pf-productos")
def pf_productos():
    """Productos de las tiendas conectadas (Shopify + Tiendanube). Sin tienda → lista vacía."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": True, "tienda": None, "productos": [], "sin_costo": 0})
    tk = _shop_tokens().get(email)
    prod_tn = _tn_productos(email)
    if (not tk or not tk.get("access_token")) and not prod_tn:
        return jsonify({"ok": True, "tienda": None, "productos": [], "sin_costo": 0})
    if (not tk or not tk.get("access_token")):
        # solo Tiendanube conectada
        sin = sum(1 for x in prod_tn if not x.get("costo"))
        return jsonify({"ok": True, "tienda": "Tiendanube", "productos": prod_tn, "sin_costo": sin})
    shop = tk.get("shop"); token = tk.get("access_token")
    costos = (_costos().get(email) or {})
    skus_guardados = _skus_map(email)
    productos = list(prod_tn)
    try:
        r = requests.get("https://%s/admin/api/2026-07/products.json" % shop,
                         headers={"X-Shopify-Access-Token": token},
                         params={"limit": 250, "fields": "id,title,image,variants,status"}, timeout=30)
        if r.status_code == 200:
            for p in r.json().get("products", []):
                v = (p.get("variants") or [{}])[0]
                pid = p.get("id")
                guardado = skus_guardados.get(str(pid))
                if guardado:
                    cfg = _sku_cfg(guardado)                       # tipo + base que cargó el usuario
                elif v.get("sku"):
                    cfg = {"tipo": "fijo", "base": v.get("sku"), "map": {}}   # fallback: el SKU fijo de Shopify
                else:
                    cfg = {"tipo": "unitario", "base": "", "map": {}}         # default: por cantidad
                productos.append({
                    "id": pid,
                    "nombre": p.get("title") or "",
                    "sku_tipo": cfg["tipo"], "sku_base": cfg["base"], "sku_map": cfg.get("map") or {},
                    "sku_ej": _sku_calc(guardado or cfg, 2),        # ejemplo con 2 unidades para mostrar
                    "precio": float(v.get("price") or 0),
                    "img": (p.get("image") or {}).get("src") or "",
                    "costo": costos.get(str(pid)) or 0,
                })
    except Exception:
        pass
    sin = sum(1 for x in productos if not x.get("costo"))
    tienda = "Shopify + Tiendanube" if prod_tn else "Shopify"
    return jsonify({"ok": True, "tienda": tienda, "shop": shop, "productos": productos, "sin_costo": sin})


@app.post("/pf-guardar-costo")
def pf_guardar_costo():
    """Guarda el costo de UN producto para este usuario."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "error": "login"}), 401
    data = request.get_json(silent=True) or {}
    pid = str(data.get("id") or "")
    if not pid:
        return jsonify({"ok": False, "error": "falta id"}), 400
    if data.get("variable") or isinstance(data.get("costos"), dict):   # VARIABLE → mapa {cantidad: costo}
        m = {}
        for k, v in (data.get("costos") or {}).items():
            try:
                m[str(int(k))] = float(v or 0)
            except Exception:
                pass
        _guardar_costo(email, pid, m)
    else:                                                              # UNITARIO/FIJO → número plano
        try:
            costo = float(data.get("costo") or 0)
        except Exception:
            costo = 0
        _guardar_costo(email, pid, costo)
    return jsonify({"ok": True})


# ---------------- Comisiones (MP + tienda + IVA + IIBB) ----------------
COMIS = DATA_DIR / "comisiones.json"   # config de comisiones por usuario (persistente)
_COMIS_CAMPOS = ["mp_comision", "mp_cuotas", "comision_tienda", "iva", "ingresos_brutos"]


def _comis() -> dict:
    try:
        return _json.loads(COMIS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _comis_user(email) -> dict:
    u = (_comis().get(email) or {})
    return {k: float(u.get(k) or 0) for k in _COMIS_CAMPOS}


def _comis_pct(email) -> float:
    """% total que se lleva cada venta: (MP + cuotas + tienda) con IVA + Ingresos Brutos (sin IVA)."""
    c = _comis_user(email)
    con_iva = (c["mp_comision"] + c["mp_cuotas"] + c["comision_tienda"]) * (1 + c["iva"] / 100.0)
    return round(con_iva + c["ingresos_brutos"], 4)


@app.get("/pf-comisiones")
def pf_comisiones():
    email = _user_actual()
    if not email:
        return jsonify({"ok": True, "comis": {k: 0 for k in _COMIS_CAMPOS}, "pct": 0})
    return jsonify({"ok": True, "comis": _comis_user(email), "pct": _comis_pct(email)})


@app.post("/pf-comisiones")
def pf_guardar_comisiones():
    email = _user_actual()
    if not email:
        return jsonify({"ok": False, "error": "login"}), 401
    data = request.get_json(silent=True) or {}
    d = _comis()
    u = {}
    for k in _COMIS_CAMPOS:
        try:
            u[k] = float(data.get(k) or 0)
        except Exception:
            u[k] = 0.0
    d[email] = u
    COMIS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return jsonify({"ok": True, "pct": _comis_pct(email)})


@app.get("/pf-marketing")
def pf_marketing():
    return jsonify({"ok": True, "mk": {}})



@app.get("/pf-despachos")
def pf_despachos():
    return jsonify({"ok": True, "desp": []})


@app.get("/pf-ventas-nuevas")
def pf_ventas_nuevas():
    """Últimas ventas del Dashboard (Shopify + Tiendanube) del rango pedido."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": True, "ventas": []})
    desde = request.args.get("desde") or (_dt.date.today() - _dt.timedelta(days=6)).isoformat()
    hasta = request.args.get("hasta") or _hoy()
    out = []
    tk = _shop_tokens().get(email)
    if tk and tk.get("access_token"):
        try:
            for o in _shopify_orders(tk.get("shop"), tk.get("access_token"), desde, hasta):
                if o.get("cancelled_at"):
                    continue
                tot = float(o.get("total_price") or o.get("current_total_price") or 0)
                num = str(o.get("order_number") or o.get("name") or "").replace("#", "").strip()
                out.append({"num": num, "origen": "Shopify",
                            "estado": _ESTADO_TXT.get((o.get("financial_status") or "").lower(), "—"),
                            "fecha": o.get("created_at") or "", "total": round(tot, 2),
                            "neto": round(tot * (1 - TIENDA_PCT / 100.0), 2)})
        except Exception:
            pass
    tn = _tn_ventas(email, desde, hasta)
    if tn:
        out.extend(tn)
    out.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    return jsonify({"ok": True, "ventas": out[:300]})


@app.get("/pf-opciones")
def pf_opciones():
    return jsonify({"ok": True, "op": {"pages": [], "cuentas": [], "igs": [], "pixels": [],
                                       "titulo_default": "", "url_default": ""}})


@app.get("/pf-ordenes")
def pf_ordenes():
    return jsonify({"ok": True, "ordenes": []})


@app.get("/conectar-mp")
@limiter.limit("30 per hour")
def conectar_mp():
    """Manda al usuario a la pantalla oficial de MercadoPago para que autorice (OAuth)."""
    if not _user_actual():                # hay que estar logueado para conectar SU cuenta
        return redirect("/")
    cfg = _mp_cfg()
    if not cfg["client_id"]:
        return ("Falta configurar el Client ID de MercadoPago en mp_secrets.json "
                "(seguí la guía).", 400)
    state = _secrets.token_urlsafe(16)
    session["mp_state"] = state
    qs = _url.urlencode({"client_id": cfg["client_id"], "response_type": "code",
                         "platform_id": "mp", "redirect_uri": cfg["redirect_uri"], "state": state})
    return redirect("https://auth.mercadopago.com.ar/authorization?" + qs, code=302)


@app.get("/mp/callback")
@limiter.limit("30 per hour")
def mp_callback():
    """MercadoPago devuelve acá con un 'code'. Lo cambiamos por el token del usuario."""
    cfg = _mp_cfg()
    code = request.args.get("code")
    state = request.args.get("state")
    if not code:
        # Visita SIN código = MercadoPago validando la URL (o alguien la abrió directo).
        # Devolvemos 200 OK para pasar la validación de MP (si damos 400, rechaza la Redirect URL).
        return ("RealProfit — punto de conexión con MercadoPago. "
                "Volvé a la app y usá el botón «Conectar con MercadoPago».", 200)
    # Anti-CSRF: el state tiene que coincidir con el que generamos al iniciar.
    if not state or state != session.get("mp_state"):
        return ("La conexión no pasó el control de seguridad. Reintentá desde el botón.", 400)
    try:
        r = requests.post("https://api.mercadopago.com/oauth/token", json={
            "client_id": cfg["client_id"], "client_secret": cfg["client_secret"],
            "grant_type": "authorization_code", "code": code,
            "redirect_uri": cfg["redirect_uri"]}, timeout=30)
        tok = r.json() if r.content else {}
    except Exception:
        return ("No pudimos conectar con MercadoPago en este momento. Probá de nuevo.", 502)
    if not tok.get("access_token"):
        # No exponemos la respuesta cruda de MP (podría filtrar detalles).
        return ("MercadoPago no autorizó la conexión. Reintentá.", 400)
    email = _user_actual()
    if not email:
        return redirect("/")
    # El token de MP se guarda BAJO EL EMAIL del usuario → cada uno ve solo lo suyo.
    _mp_save_token(email, tok)
    session.pop("mp_state", None)                 # el state es de un solo uso
    return redirect("/?integ=1", code=302)


@app.get("/mp/estado")
def mp_estado():
    """¿ESTE usuario (su email logueado) tiene su MercadoPago conectado? Cada uno ve SOLO lo suyo."""
    email = _user_actual()
    conectado = bool(email and email in _mp_tokens())
    return jsonify({"ok": True, "conectado": conectado})


@app.post("/shopify/byoa-start")
@limiter.limit("30 per hour")
def shopify_byoa_start():
    """OAuth con la app PROPIA de cada tienda (estilo Envialo): recibe shop + Client ID + Client Secret."""
    if not _user_actual():
        return jsonify({"ok": False, "error": "Tenés que iniciar sesión."}), 401
    data = request.get_json(silent=True) or request.form
    shop = _shop_normalizar(data.get("shop", ""))
    cid = (data.get("client_id", "") or "").strip()
    secret = (data.get("client_secret", "") or "").strip()
    if not _shop_valido(shop):
        return jsonify({"ok": False, "error": "Dominio inválido. Usá el formato tutienda.myshopify.com"}), 400
    if not cid or not secret:
        return jsonify({"ok": False, "error": "Faltan el Client ID o el Client Secret."}), 400
    cfg = _shop_cfg()
    state = _secrets.token_urlsafe(16)
    session["shop_state"] = state
    session["shop_dom"] = shop
    session["shop_cid"] = cid
    session["shop_secret"] = secret
    qs = _url.urlencode({"client_id": cid, "scope": cfg["scopes"],
                         "redirect_uri": cfg["redirect_uri"], "state": state})
    return jsonify({"ok": True, "url": "https://" + shop + "/admin/oauth/authorize?" + qs})


@app.get("/conectar-shopify")
@limiter.limit("30 per hour")
def conectar_shopify():
    """Manda al usuario a autorizar SU tienda Shopify (OAuth con la app global de env)."""
    if not _user_actual():
        return redirect("/")
    cfg = _shop_cfg()
    if not cfg["client_id"]:
        return ("Falta configurar el Client ID de Shopify (variables en Render).", 400)
    shop = _shop_normalizar(request.args.get("shop", ""))
    if not _shop_valido(shop):
        return ("Dominio de tienda inválido. Usá el formato tutienda.myshopify.com", 400)
    state = _secrets.token_urlsafe(16)
    session["shop_state"] = state
    session["shop_dom"] = shop
    qs = _url.urlencode({"client_id": cfg["client_id"], "scope": cfg["scopes"],
                         "redirect_uri": cfg["redirect_uri"], "state": state})
    return redirect("https://" + shop + "/admin/oauth/authorize?" + qs, code=302)


@app.get("/shopify/callback")
@limiter.limit("30 per hour")
def shopify_callback():
    """Shopify vuelve acá con 'code' + 'hmac'. Verificamos la firma y cambiamos el code por el token."""
    import hmac as _hmac
    import hashlib as _hashlib
    cfg = _shop_cfg()
    # Si vino por «tu propia app» (BYOA), usamos las claves que pegó el usuario; si no, las de env.
    cid = session.get("shop_cid") or cfg["client_id"]
    secret = session.get("shop_secret") or cfg["client_secret"]
    code = request.args.get("code")
    shop = _shop_normalizar(request.args.get("shop", ""))
    state = request.args.get("state")
    hmac_recibido = request.args.get("hmac", "")
    if not code:
        return ("RealProfit — punto de conexión con Shopify. "
                "Volvé a la app y usá el botón «Conectar».", 200)
    if not state or state != session.get("shop_state"):
        return ("La conexión no pasó el control de seguridad. Reintentá desde el botón.", 400)
    if not _shop_valido(shop):
        return ("Tienda inválida.", 400)
    # Verificar HMAC: firma de todos los params (menos hmac/signature) ordenados, con el client_secret.
    params = {k: v for k, v in request.args.items() if k not in ("hmac", "signature")}
    mensaje = "&".join("%s=%s" % (k, params[k]) for k in sorted(params))
    calc = _hmac.new(secret.encode(), mensaje.encode(), _hashlib.sha256).hexdigest()
    if not hmac_recibido or not _hmac.compare_digest(calc, hmac_recibido):
        return ("La firma de Shopify no es válida. Reintentá.", 400)
    try:
        r = requests.post("https://" + shop + "/admin/oauth/access_token", json={
            "client_id": cid, "client_secret": secret,
            "code": code}, timeout=30)
        tok = r.json() if r.content else {}
    except Exception:
        return ("No pudimos conectar con Shopify en este momento. Probá de nuevo.", 502)
    if not tok.get("access_token"):
        return ("Shopify no autorizó la conexión. Reintentá.", 400)
    email = _user_actual()
    if not email:
        return redirect("/")
    tok["shop"] = shop
    _shop_save_token(email, tok)          # token guardado BAJO EL EMAIL → cada uno ve solo lo suyo
    for k in ("shop_state", "shop_dom", "shop_cid", "shop_secret"):
        session.pop(k, None)
    return redirect("/?integ=1", code=302)


@app.get("/shopify/estado")
def shopify_estado():
    """¿ESTE usuario tiene su Shopify conectado? Devuelve también el dominio de la tienda."""
    email = _user_actual()
    d = _shop_tokens()
    conectado = bool(email and email in d)
    shop = (d.get(email) or {}).get("shop", "") if conectado else ""
    return jsonify({"ok": True, "conectado": conectado, "shop": shop})


# ---------------- Tiendanube (OAuth un-clic, igual que Shopify) ----------------
TIENDANUBE_SECRETS = RAIZ / "tiendanube_secrets.json"     # App ID + Client Secret (dueño de la app)
TIENDANUBE_TOKENS = DATA_DIR / "tiendanube_tokens.json"   # {email: {access_token, store_id}}
TN_API = "https://api.tiendanube.com/v1"
TN_UA = "RealProfit (soporte@realprofitapp.com)"


def _tn_cfg() -> dict:
    import os
    try:
        c = _json.loads(TIENDANUBE_SECRETS.read_text(encoding="utf-8"))
    except Exception:
        c = {}
    return {"app_id": os.getenv("TIENDANUBE_APP_ID") or c.get("app_id", ""),
            "client_secret": os.getenv("TIENDANUBE_CLIENT_SECRET") or c.get("client_secret", "")}


def _tn_tokens() -> dict:
    try:
        return _json.loads(TIENDANUBE_TOKENS.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _tn_save_token(email, data) -> None:
    d = _tn_tokens()
    d[str(email)] = data
    TIENDANUBE_TOKENS.parent.mkdir(parents=True, exist_ok=True)
    TIENDANUBE_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")


def _tn_headers(token) -> dict:
    return {"Authentication": "bearer " + str(token), "User-Agent": TN_UA}


@app.get("/conectar-tiendanube")
@limiter.limit("30 per hour")
def conectar_tiendanube():
    """Manda al usuario a autorizar SU tienda Tiendanube (OAuth con la app de RealProfit)."""
    if not _user_actual():
        return redirect("/")
    cfg = _tn_cfg()
    if not cfg["app_id"]:
        return ("Falta configurar el App ID de Tiendanube (variables en Render).", 400)
    return redirect("https://www.tiendanube.com/apps/%s/authorize" % cfg["app_id"], code=302)


@app.get("/tiendanube/callback")
@limiter.limit("30 per hour")
def tiendanube_callback():
    """Tiendanube vuelve acá con 'code'. Lo cambiamos por el access_token + store_id."""
    cfg = _tn_cfg()
    code = request.args.get("code")
    if not code:
        return ("RealProfit — punto de conexión con Tiendanube. "
                "Volvé a la app y usá el botón «Conectar».", 200)
    try:
        r = requests.post("https://www.tiendanube.com/apps/authorize/token", json={
            "client_id": cfg["app_id"], "client_secret": cfg["client_secret"],
            "grant_type": "authorization_code", "code": code}, timeout=30)
        tok = r.json() if r.content else {}
    except Exception:
        return ("No pudimos conectar con Tiendanube en este momento. Probá de nuevo.", 502)
    if not tok.get("access_token") or not tok.get("user_id"):
        return ("Tiendanube no autorizó la conexión. Reintentá.", 400)
    email = _user_actual()
    if not email:
        return redirect("/")
    _tn_save_token(email, {"access_token": tok["access_token"], "store_id": str(tok["user_id"])})
    return redirect("/?integ=1", code=302)


@app.get("/tiendanube/estado")
def tiendanube_estado():
    """¿ESTE usuario tiene su Tiendanube conectada?"""
    email = _user_actual()
    d = _tn_tokens()
    conectado = bool(email and email in d)
    store = (d.get(email) or {}).get("store_id", "") if conectado else ""
    return jsonify({"ok": True, "conectado": conectado, "store": store})


@app.get("/desconectar-tiendanube")
def desconectar_tiendanube():
    email = _user_actual()
    if email:
        d = _tn_tokens()
        if email in d:
            d.pop(email, None)
            TIENDANUBE_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return redirect("/?integ=1", code=302)


@app.post("/tiendanube/guardar-token")
@limiter.limit("40 per hour")
def tiendanube_guardar_token():
    """Conecta Tiendanube pegando Store ID + Access Token (sin OAuth). Verifica antes de guardar."""
    email = _user_actual()
    if not email:
        return jsonify({"ok": False}), 401
    data = request.get_json(silent=True) or {}
    store = "".join(ch for ch in str(data.get("store_id") or "") if ch.isdigit())
    token = str(data.get("token") or "").strip()
    if not store or not token:
        return jsonify({"ok": False, "msg": "Pegá el Store ID y el token."})
    try:
        r = requests.get("%s/%s/orders" % (TN_API, store),
                         headers=_tn_headers(token), params={"per_page": 1}, timeout=20)
        if r.status_code in (401, 403):
            return jsonify({"ok": False, "msg": "Token o Store ID inválido (Tiendanube lo rechazó)."})
        if r.status_code >= 400:
            return jsonify({"ok": False, "msg": "Tiendanube respondió %s. Revisá los datos." % r.status_code})
    except Exception:
        return jsonify({"ok": False, "msg": "No pude verificar contra Tiendanube ahora. Probá de nuevo."})
    _tn_save_token(email, {"access_token": token, "store_id": store})
    return jsonify({"ok": True})


@app.post("/shopify/guardar-token")
@limiter.limit("40 per hour")
def shopify_guardar_token():
    """Guarda la conexión de Shopify por TOKEN (app propia de cada tienda). Verifica el token antes."""
    if not _user_actual():
        return jsonify({"ok": False, "error": "Tenés que iniciar sesión."}), 401
    data = request.get_json(silent=True) or request.form
    shop = _shop_normalizar(data.get("shop", ""))
    token = (data.get("token", "") or "").strip()
    if not _shop_valido(shop):
        return jsonify({"ok": False, "error": "Dominio inválido. Usá el formato tutienda.myshopify.com"}), 400
    if not token:
        return jsonify({"ok": False, "error": "Falta pegar el token."}), 400
    # Verificar el token contra la tienda (si anda, es válido).
    try:
        r = requests.get("https://" + shop + "/admin/api/2026-07/shop.json",
                         headers={"X-Shopify-Access-Token": token}, timeout=20)
    except Exception:
        return jsonify({"ok": False, "error": "No pudimos contactar la tienda. Revisá el dominio."}), 502
    if r.status_code != 200:
        return jsonify({"ok": False, "error": "El token no es válido para esa tienda (revisá que lo copiaste completo)."}), 400
    email = _user_actual()
    _shop_save_token(email, {"access_token": token, "shop": shop, "modo": "token"})
    return jsonify({"ok": True, "shop": shop})


@app.get("/desconectar-shopify")
def desconectar_shopify():
    email = _user_actual()
    if email:
        d = _shop_tokens()
        d.pop(email, None)
        SHOPIFY_TOKENS.write_text(_json.dumps(d, ensure_ascii=False, indent=1), encoding="utf-8")
    return redirect("/?integ=1")


# Catch-all defensivo: cualquier otro fetch del dashboard responde vacío (no 404, no error).
@app.route("/<path:_ruta>", methods=["GET", "POST"])
def _stub(_ruta):
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8010, debug=False)
